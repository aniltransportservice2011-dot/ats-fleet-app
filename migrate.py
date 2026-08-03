import sqlite3
from openpyxl import load_workbook
import datetime

wb = load_workbook('/mnt/user-data/outputs/Fleet_Master_Tracker.xlsx', data_only=True)
conn = sqlite3.connect('fleet.db')
cur = conn.cursor()

def get_or_create(table, name):
    if not name or not str(name).strip():
        return None
    name = str(name).strip()
    cur.execute(f"SELECT id FROM {table} WHERE name = ?", (name,))
    row = cur.fetchone()
    if row:
        return row[0]
    cur.execute(f"INSERT INTO {table} (name) VALUES (?)", (name,))
    return cur.lastrowid

def get_or_create_vehicle(vno):
    if not vno or not str(vno).strip():
        return None
    vno = str(vno).strip()
    cur.execute("SELECT id FROM vehicles WHERE vehicle_no = ?", (vno,))
    row = cur.fetchone()
    if row:
        return row[0]
    cur.execute("INSERT INTO vehicles (vehicle_no) VALUES (?)", (vno,))
    return cur.lastrowid

def fmt_date(d):
    if isinstance(d, datetime.datetime):
        return d.strftime('%Y-%m-%d')
    return None

# Migrate Trips
ws_t = wb['Trips']
trip_count = 0
for r in range(7, 481):
    lr = ws_t.cell(row=r, column=3).value
    if not lr:
        continue
    date = fmt_date(ws_t.cell(row=r, column=2).value)
    vehicle_id = get_or_create_vehicle(ws_t.cell(row=r, column=4).value)
    party_id = get_or_create('parties', ws_t.cell(row=r, column=6).value)
    fuel_vendor_id = get_or_create('vendors', ws_t.cell(row=r, column=36).value)
    driveradv_vendor_id = get_or_create('vendors', ws_t.cell(row=r, column=38).value)
    cur.execute("""INSERT INTO trips (date, lr_number, vehicle_id, party_id, from_loc, to_loc,
                   quantity, rate, rate_type, billed_amount, fuel_amount, fuel_vendor_id,
                   driver_adv_amount, driver_adv_vendor_id, party_advance, payment_received)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (date, lr, vehicle_id, party_id,
                 ws_t.cell(row=r, column=7).value, ws_t.cell(row=r, column=8).value,
                 ws_t.cell(row=r, column=9).value, ws_t.cell(row=r, column=10).value,
                 ws_t.cell(row=r, column=13).value, ws_t.cell(row=r, column=14).value,
                 ws_t.cell(row=r, column=35).value, fuel_vendor_id,
                 ws_t.cell(row=r, column=37).value, driveradv_vendor_id,
                 ws_t.cell(row=r, column=39).value or 0, ws_t.cell(row=r, column=40).value or 0))
    trip_count += 1

# Migrate Maintenance
ws_m = wb['Maintenance']
maint_count = 0
for r in range(6, 360):
    veh = ws_m.cell(row=r, column=3).value
    if not veh:
        continue
    date = fmt_date(ws_m.cell(row=r, column=2).value)
    vehicle_id = get_or_create_vehicle(veh)
    vendor_id = get_or_create('vendors', ws_m.cell(row=r, column=7).value)
    cur.execute("""INSERT INTO maintenance (date, vehicle_id, category, amount, paid_amount, vendor_id, notes)
                   VALUES (?,?,?,?,?,?,?)""",
                (date, vehicle_id, ws_m.cell(row=r, column=4).value, ws_m.cell(row=r, column=5).value,
                 ws_m.cell(row=r, column=6).value or 0, vendor_id, ws_m.cell(row=r, column=8).value))
    maint_count += 1

# Migrate Party payment log
ws_p = wb['Party']
party_pay_count = 0
for r in range(7, 260):
    name = ws_p.cell(row=r, column=11).value
    if not name:
        continue
    date = fmt_date(ws_p.cell(row=r, column=10).value)
    amount = ws_p.cell(row=r, column=12).value
    party_id = get_or_create('parties', name)
    cur.execute("INSERT INTO payments (date, payment_type, amount, party_id, mode) VALUES (?,?,?,?,?)",
                (date, 'received', amount, party_id, ws_p.cell(row=r, column=13).value))
    party_pay_count += 1

# Migrate Vendor payment log
ws_v = wb['Vendor']
vendor_pay_count = 0
for r in range(7, 40):
    name = ws_v.cell(row=r, column=11).value
    if not name:
        continue
    date = fmt_date(ws_v.cell(row=r, column=10).value)
    amount = ws_v.cell(row=r, column=12).value
    vendor_id = get_or_create('vendors', name)
    cur.execute("INSERT INTO payments (date, payment_type, amount, vendor_id, mode) VALUES (?,?,?,?,?)",
                (date, 'paid', amount, vendor_id, ws_v.cell(row=r, column=13).value))
    vendor_pay_count += 1

conn.commit()
print(f"Migrated: {trip_count} trips, {maint_count} maintenance, {party_pay_count} party payments, {vendor_pay_count} vendor payments")

cur.execute("SELECT COUNT(*) FROM parties")
print("Total unique parties:", cur.fetchone()[0])
cur.execute("SELECT COUNT(*) FROM vendors")
print("Total unique vendors:", cur.fetchone()[0])
cur.execute("SELECT COUNT(*) FROM vehicles")
print("Total unique vehicles:", cur.fetchone()[0])
conn.close()
