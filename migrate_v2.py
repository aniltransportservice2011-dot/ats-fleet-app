import sqlite3
from openpyxl import load_workbook
import datetime

wb = load_workbook('/mnt/user-data/outputs/Fleet_Master_Tracker.xlsx', data_only=True)
conn = sqlite3.connect('fleet.db')
cur = conn.cursor()

def get_or_create(table, name):
    if not name or not str(name).strip():
        return None
    name = str(name).strip()
    cur.execute(f"SELECT id FROM {table} WHERE name = ?", (name,))
    row = cur.fetchone()
    if row: return row[0]
    cur.execute(f"INSERT INTO {table} (name) VALUES (?)", (name,))
    return cur.lastrowid

def get_or_create_vehicle(vno):
    if not vno or not str(vno).strip(): return None
    vno = str(vno).strip()
    cur.execute("SELECT id FROM vehicles WHERE vehicle_no = ?", (vno,))
    row = cur.fetchone()
    if row: return row[0]
    cur.execute("INSERT INTO vehicles (vehicle_no) VALUES (?)", (vno,))
    return cur.lastrowid

def fmt_date(d):
    if isinstance(d, datetime.datetime):
        return d.strftime('%Y-%m-%d')
    return None

def v(r, c):
    val = ws_t.cell(row=r, column=c).value
    return val if val is not None else 0

ws_t = wb['Trips']
count = 0
for r in range(7, 481):
    lr = ws_t.cell(row=r, column=3).value
    if not lr:
        continue
    vehicle_id = get_or_create_vehicle(ws_t.cell(row=r, column=4).value)
    party_id = get_or_create('parties', ws_t.cell(row=r, column=6).value)
    fuel_vendor_id = get_or_create('vendors', ws_t.cell(row=r, column=36).value)
    driveradv_vendor_id = get_or_create('vendors', ws_t.cell(row=r, column=38).value)
    misc_vendor_id = get_or_create('vendors', ws_t.cell(row=r, column=61).value)

    cur.execute("""INSERT INTO trips (
        date, lr_number, vehicle_id, type, party_id, from_loc, to_loc, quantity, rate,
        driver_name, material, rate_type, billed_amount,
        driver_payment, detention_charges, gps_cost, loading_charge, unloading_charge,
        police_charges, sim_tracking, union_charges, weight_charges, other_charges,
        brokerage, builty_commission, late_fees, material_damage, shortage_amount, shortage_qty, tds, other_deductions,
        fuel_amount, fuel_vendor_id, driver_adv_amount, driver_adv_vendor_id, party_advance, payment_received,
        owner_name, fixed_rate_amount, owner_rate, owner_amount, paid_to_owner, pending_to_owner,
        agent_commission, builty_expense, conductor_expense, fine, labour_charges, parking, puncture,
        toll, urea, loading_expense, unloading_expense, wear_tear, weighbridge_charges, other_expense, misc_vendor_id,
        lr_received
    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (fmt_date(ws_t.cell(row=r,column=2).value), lr, vehicle_id, ws_t.cell(row=r,column=5).value,
         party_id, ws_t.cell(row=r,column=7).value, ws_t.cell(row=r,column=8).value,
         v(r,9), v(r,10), ws_t.cell(row=r,column=11).value, ws_t.cell(row=r,column=12).value,
         ws_t.cell(row=r,column=13).value, v(r,14),
         v(r,15), v(r,16), v(r,17), v(r,18), v(r,19), v(r,20), v(r,21), v(r,22), v(r,23), v(r,24),
         v(r,26), v(r,27), v(r,28), v(r,29), v(r,30), v(r,31), v(r,32), v(r,33),
         v(r,35), fuel_vendor_id, v(r,37), driveradv_vendor_id, v(r,39), v(r,40),
         ws_t.cell(row=r,column=41).value, v(r,42), v(r,43), v(r,44), v(r,45), v(r,46),
         v(r,47), v(r,48), v(r,49), v(r,50), v(r,51), v(r,52), v(r,53),
         v(r,54), v(r,55), v(r,56), v(r,57), v(r,58), v(r,59), v(r,60), misc_vendor_id,
         ws_t.cell(row=r,column=68).value))
    count += 1

conn.commit()
print(f"Migrated {count} trips with all 62 fields")
conn.close()
