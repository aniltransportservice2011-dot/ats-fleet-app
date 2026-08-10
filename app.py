from flask import Flask, render_template, request, redirect, url_for, session
import sqlite3
import datetime
import calendar
import os
import requests
from werkzeug.utils import secure_filename
import compliance_service as cs
from providers.echallan_client import fetch_rc

app = Flask(__name__)
app.secret_key = 'fleet-local-app-anil-transport-secret-key-2026'
app.permanent_session_lifetime = datetime.timedelta(days=30)
DB = 'fleet.db'
INSURANCE_UPLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'uploads', 'insurance')
os.makedirs(INSURANCE_UPLOAD_DIR, exist_ok=True)
TOLL_RECEIPT_UPLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'uploads', 'toll_receipts')
os.makedirs(TOLL_RECEIPT_UPLOAD_DIR, exist_ok=True)

def get_db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn

@app.context_processor
def inject_site_logo():
    """Makes the uploaded company logo available to every template as `site_logo_path` (used for
    the browser tab favicon in base.html) without every single route needing to fetch and pass it
    itself — same file Settings > Company Profile already manages, so it updates automatically
    whenever that's changed."""
    try:
        conn = get_db()
        row = conn.execute("SELECT value FROM settings WHERE key='logo_path'").fetchone()
        conn.close()
        return {'site_logo_path': row['value'] if row and row['value'] else None}
    except sqlite3.Error:
        return {'site_logo_path': None}

def _vehicle_active_idle_days(conn, vehicle_id, registration_date, date_from, date_to):
    """Active days = union of calendar days spanned by each trip's start..end_date (a trip stays
    'active' from when it's entered until it's explicitly ended, not just on its start date).
    Idle days = the remainder of the window, anchored no earlier than the vehicle's registration
    date so a newly bought vehicle isn't counted idle for days before it existed."""
    anchor = registration_date if registration_date and registration_date > date_from else date_from
    if anchor > date_to:
        return 0, 0, 0
    total_days_v = (datetime.datetime.strptime(date_to, '%Y-%m-%d').date() -
                     datetime.datetime.strptime(anchor, '%Y-%m-%d').date()).days + 1
    trip_rows = conn.execute("""SELECT date, end_date FROM trips WHERE vehicle_id=? AND date<=?
                                AND (end_date IS NULL OR end_date>=?) ORDER BY date""",
                              (vehicle_id, date_to, anchor)).fetchall()
    active_dates = set()
    for t in trip_rows:
        start = t['date']
        end = t['end_date'] or t['date']
        if end < start:
            end = start
        s = max(start, anchor)
        e = min(end, date_to)
        if s > e:
            continue
        d = datetime.datetime.strptime(s, '%Y-%m-%d').date()
        e_d = datetime.datetime.strptime(e, '%Y-%m-%d').date()
        while d <= e_d:
            active_dates.add(d.isoformat())
            d += datetime.timedelta(days=1)
    active_days = len(active_dates)
    idle_days = max(total_days_v - active_days, 0)
    return active_days, idle_days, total_days_v

def _toll_by_trip(conn, trip_ids):
    """Real Toll Management amount linked to each trip_id, for the (usually few) trips that have
    one — a BOSS/FASTag import almost never carries a trip number, so most trips won't appear
    here. Callers needing a PER-TRIP figure (an invoice line, a customer's profit calc) should
    fall back to that trip's own manual `toll` field when its id isn't in the returned dict,
    rather than treating an unlinked trip as zero toll cost."""
    trip_ids = [t for t in trip_ids if t]
    if not trip_ids:
        return {}
    placeholders = ','.join('?' * len(trip_ids))
    rows = conn.execute(f"SELECT trip_id, SUM(amount) as total FROM toll_entries WHERE trip_id IN ({placeholders}) GROUP BY trip_id",
                         trip_ids).fetchall()
    return {r['trip_id']: r['total'] for r in rows}

def _trip_toll(t, toll_map):
    """Real per-trip toll cost: the linked Toll Management total if this trip has one, else its
    own manual estimate — never both, so the same real-world toll is never counted twice."""
    linked = toll_map.get(t['id'])
    return linked if linked is not None else (t['toll'] or 0)

def _period_financials(conn, date_from, date_to):
    """Revenue/expense/profit breakdown for a date range, using the same charge
    columns as dashboard()/monthly_summary() so the numbers agree app-wide."""
    trips = conn.execute("SELECT * FROM trips WHERE date>=? AND date<=?", (date_from, date_to)).fetchall()
    revenue = sum(t['billed_amount'] or 0 for t in trips)
    fuel = sum(t['fuel_amount'] or 0 for t in trips)
    driver_adv = sum(t['driver_adv_amount'] or 0 for t in trips)
    parking = sum(t['parking'] or 0 for t in trips)
    misc = sum((t['agent_commission'] or 0) + (t['builty_expense'] or 0) + (t['conductor_expense'] or 0) +
               (t['fine'] or 0) + (t['labour_charges'] or 0) + (t['puncture'] or 0) + (t['urea'] or 0) +
               (t['loading_expense'] or 0) + (t['unloading_expense'] or 0) + (t['wear_tear'] or 0) +
               (t['weighbridge_charges'] or 0) + (t['other_expense'] or 0) + (t['permit_charges'] or 0) for t in trips)
    owner_cost = sum((t['owner_fixed_amount'] or 0) if (t['owner_rate_type'] or 'PER_MT') == 'FIXED' else (t['owner_rate'] or 0) * (t['quantity'] or 0)
                      for t in trips if t['type'] == 'Market')
    maint_all = conn.execute("SELECT COALESCE(SUM(amount),0) FROM maintenance WHERE date>=? AND date<=?", (date_from, date_to)).fetchone()[0]
    # Toll Management (Maintenance > Toll) is the source of truth for period-level toll cost now —
    # 'maint' below means "maintenance excluding toll" so a breakdown chart showing both a
    # 'Maintenance' slice and a 'Toll' slice never double-counts the same rupee in both.
    toll = conn.execute("SELECT COALESCE(SUM(amount),0) FROM maintenance WHERE category='Toll' AND date>=? AND date<=?",
                        (date_from, date_to)).fetchone()[0]
    maint = maint_all - toll
    overheads = conn.execute("SELECT COALESCE(SUM(amount),0) FROM overheads WHERE date>=? AND date<=?", (date_from, date_to)).fetchone()[0]
    salaries = conn.execute("SELECT COALESCE(SUM(amount),0) FROM salaries WHERE date>=? AND date<=?", (date_from, date_to)).fetchone()[0]
    total_expenses = fuel + driver_adv + toll + parking + misc + owner_cost + maint + overheads + salaries
    net_profit = revenue - total_expenses
    margin = round(net_profit / revenue * 100, 2) if revenue else 0
    party_ids = set(t['party_id'] for t in trips if t['party_id'])
    return {
        'trips': trips, 'trip_count': len(trips), 'revenue': revenue, 'fuel': fuel, 'driver_adv': driver_adv,
        'toll': toll, 'parking': parking, 'misc': misc, 'owner_cost': owner_cost, 'maint': maint,
        'overheads': overheads, 'salaries': salaries, 'total_expenses': total_expenses,
        'net_profit': net_profit, 'margin': margin, 'party_ids': party_ids,
    }

def _pct_growth(curr, prev):
    if prev:
        return round((curr - prev) / abs(prev) * 100, 2)
    return None

def _shift_period_back(date_from, date_to):
    """Previous period of the same length, immediately before date_from."""
    d_from = datetime.datetime.strptime(date_from, '%Y-%m-%d').date()
    d_to = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()
    length = (d_to - d_from).days + 1
    prev_to = d_from - datetime.timedelta(days=1)
    prev_from = prev_to - datetime.timedelta(days=length - 1)
    return prev_from.isoformat(), prev_to.isoformat()

def _shift_period_year(date_from, date_to):
    def back_a_year(d):
        try:
            return d.replace(year=d.year - 1)
        except ValueError:
            return d - datetime.timedelta(days=365)
    d_from = datetime.datetime.strptime(date_from, '%Y-%m-%d').date()
    d_to = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()
    return back_a_year(d_from).isoformat(), back_a_year(d_to).isoformat()

def _month_bounds(year, month):
    last_day = calendar.monthrange(year, month)[1]
    return datetime.date(year, month, 1).isoformat(), datetime.date(year, month, last_day).isoformat()

def _quarter_bounds(d):
    q_start_month = (d.month - 1) // 3 * 3 + 1
    start = datetime.date(d.year, q_start_month, 1)
    end_month = q_start_month + 2
    last_day = calendar.monthrange(d.year, end_month)[1]
    end = datetime.date(d.year, end_month, last_day)
    return start.isoformat(), end.isoformat()

def _paginate(page_raw, per_page_raw, total_count, per_page_options=(10, 25, 50, 100), default_per_page=50):
    """Shared page/per_page/total_pages math so every list page paginates identically."""
    try:
        per_page = int(per_page_raw) if per_page_raw else default_per_page
    except (TypeError, ValueError):
        per_page = default_per_page
    if per_page not in per_page_options:
        per_page = default_per_page
    total_pages = max(1, (total_count + per_page - 1) // per_page)
    try:
        page = max(1, int(page_raw or 1))
    except (TypeError, ValueError):
        page = 1
    page = min(page, total_pages)
    return page, per_page, total_pages

def _page_tokens(page, total_pages):
    """Windowed page-number list with ellipses, e.g. [1, '...', 6, 7, 8, '...', 42]."""
    if total_pages <= 7:
        return list(range(1, total_pages + 1))
    tokens = [1]
    if page > 3:
        tokens.append('...')
    for p in range(max(2, page - 1), min(total_pages, page + 1) + 1):
        tokens.append(p)
    if page < total_pages - 2:
        tokens.append('...')
    if total_pages not in tokens:
        tokens.append(total_pages)
    return tokens

_LOCATION_ALIASES = {'RKL': 'ROURKELA', 'BBSR': 'BHUBANESWAR', 'SAGJOR': 'SAGJORE'}
def _clean_loc(raw):
    if not raw:
        return ''
    first_word = raw.split(' ')[0] if ' ' in raw else raw
    cleaned = first_word.upper().replace(',', '').strip()
    return _LOCATION_ALIASES.get(cleaned, cleaned)

def get_or_create_party(conn, name):
    if not name or not name.strip():
        return None
    name = name.strip()
    row = conn.execute("SELECT id FROM parties WHERE name = ? COLLATE NOCASE", (name,)).fetchone()
    if row:
        return row[0]
    cur = conn.execute("INSERT INTO parties (name) VALUES (?)", (name,))
    party_id = cur.lastrowid
    # A vendor record with the same name may already exist (e.g. created first via
    # Maintenance/Fuel). Link it here too, so both sides share one combined ledger
    # instead of the org getting a second, disconnected balance.
    vendor_match = conn.execute(
        "SELECT id FROM vendors WHERE name = ? COLLATE NOCASE AND linked_party_id IS NULL", (name,)).fetchone()
    if vendor_match:
        conn.execute("UPDATE vendors SET linked_party_id = ? WHERE id = ?", (party_id, vendor_match[0]))
    return party_id

def get_or_create_vendor(conn, name):
    if not name or not name.strip():
        return None
    name = name.strip()
    row = conn.execute("SELECT id FROM vendors WHERE name = ? COLLATE NOCASE", (name,)).fetchone()
    if row:
        return row[0]
    # If this name already exists as a Party, link this vendor record to that party
    # instead of creating a fully independent one — keeps ledgers combined.
    party_match = conn.execute("SELECT id FROM parties WHERE name = ? COLLATE NOCASE", (name,)).fetchone()
    if party_match:
        existing_link = conn.execute("SELECT id FROM vendors WHERE linked_party_id = ?", (party_match[0],)).fetchone()
        if existing_link:
            return existing_link[0]
        cur = conn.execute("INSERT INTO vendors (name, linked_party_id) VALUES (?,?)", (name, party_match[0]))
        return cur.lastrowid
    cur = conn.execute("INSERT INTO vendors (name) VALUES (?)", (name,))
    return cur.lastrowid

def _accounts_rows(conn):
    """Shared party+vendor balance list used by both the Ledger page and its export —
    identical formula to what accounts()/dashboard() have always used, just factored out."""
    parties_bal = conn.execute("""SELECT p.id, p.name, p.contact, p.email, p.address, p.credit_limit, p.category, p.gstin, p.status,
        (SELECT COALESCE(SUM(billed_amount),0) FROM trips WHERE party_id=p.id) -
        (SELECT COALESCE(SUM(payment_received)+SUM(party_advance),0) FROM trips WHERE party_id=p.id) -
        (SELECT COALESCE(SUM(amount-COALESCE(allocated_amount,0)),0) FROM payments WHERE party_id=p.id AND payment_type='received') +
        COALESCE(p.opening_balance,0) as balance
        FROM parties p ORDER BY p.name""").fetchall()

    vendors_bal = conn.execute("""SELECT v.id, v.name, v.contact, v.email, v.address, v.credit_limit, v.category, v.gstin, v.status,
        (SELECT COALESCE(SUM(m.amount),0) FROM maintenance m WHERE m.vendor_id=v.id) +
        (SELECT COALESCE(SUM(fuel_amount),0) FROM trips WHERE fuel_vendor_id=v.id) +
        (SELECT COALESCE(SUM(driver_adv_amount),0) FROM trips WHERE driver_adv_vendor_id=v.id) +
        (SELECT COALESCE(SUM(CASE WHEN owner_rate_type='FIXED' THEN owner_fixed_amount ELSE owner_rate*quantity END),0) FROM trips WHERE owner_vendor_id=v.id) +
        (SELECT COALESCE(SUM(CASE WHEN item_type='charge' THEN amount ELSE -amount END),0) FROM invoice_items WHERE vendor_id=v.id) -
        (SELECT COALESCE(SUM(m.paid_amount),0) FROM maintenance m WHERE m.vendor_id=v.id) -
        (SELECT COALESCE(SUM(paid_to_owner),0) FROM trips WHERE owner_vendor_id=v.id) -
        (SELECT COALESCE(SUM(amount-COALESCE(allocated_amount,0)),0) FROM payments WHERE vendor_id=v.id AND payment_type='paid') -
        COALESCE(v.opening_balance,0) as balance
        FROM vendors v WHERE v.linked_party_id IS NULL ORDER BY v.name""").fetchall()

    rows = []
    for p in parties_bal:
        rows.append({'id': p['id'], 'name': p['name'], 'role': 'Party', 'balance': p['balance'] or 0,
                      'contact': p['contact'], 'email': p['email'], 'address': p['address'], 'credit_limit': p['credit_limit'],
                      'group': p['category'] or '', 'gstin': p['gstin'], 'status': p['status'] or 'Active'})
    for v in vendors_bal:
        # Vendor's raw balance is costs-minus-paid (positive = we owe them).
        # Negate so positive consistently means "receivable" and negative means "payable", matching parties.
        rows.append({'id': v['id'], 'name': v['name'], 'role': 'Vendor', 'balance': -(v['balance'] or 0),
                      'contact': v['contact'], 'email': v['email'], 'address': v['address'], 'credit_limit': v['credit_limit'],
                      'group': v['category'] or '', 'gstin': v['gstin'], 'status': v['status'] or 'Active'})
    rows.sort(key=lambda r: r['name'])
    return rows

def _filter_accounts_rows(rows, search_f, role_f, group_f, status_f=''):
    if search_f:
        s = search_f.lower()
        rows = [r for r in rows if s in r['name'].lower() or s in (r['contact'] or '').lower() or s in (r['email'] or '').lower()]
    if role_f:
        rows = [r for r in rows if r['role'] == role_f]
    if group_f:
        rows = [r for r in rows if r['group'] == group_f]
    if status_f:
        rows = [r for r in rows if r['status'].lower() == status_f.lower()]
    return rows

@app.route('/accounts')
def accounts():
    conn = get_db()
    search_f = request.args.get('search', '')
    role_f = request.args.get('role', '')
    group_f = request.args.get('group', '')
    status_f = request.args.get('status', '')

    all_rows = _accounts_rows(conn)
    groups = sorted(set(r['group'] for r in all_rows if r['group']))
    total_parties = sum(1 for r in all_rows if r['role'] == 'Party')
    active_parties = sum(1 for r in all_rows if r['role'] == 'Party' and r['status'] == 'Active')
    total_vendors = sum(1 for r in all_rows if r['role'] == 'Vendor')
    active_vendors = sum(1 for r in all_rows if r['role'] == 'Vendor' and r['status'] == 'Active')
    total_receivable = sum(r['balance'] for r in all_rows if r['balance'] > 0)
    receivable_from = sum(1 for r in all_rows if r['balance'] > 0)
    total_payable = sum(-r['balance'] for r in all_rows if r['balance'] < 0)
    payable_to = sum(1 for r in all_rows if r['balance'] < 0)
    net_balance = sum(r['balance'] for r in all_rows)

    rows = _filter_accounts_rows(all_rows, search_f, role_f, group_f, status_f)
    conn.close()

    total_count = len(rows)
    page, per_page, total_pages = _paginate(request.args.get('page'), request.args.get('per_page'), total_count,
                                             per_page_options=(10, 25, 50, 100), default_per_page=10)
    page_rows = rows[(page - 1) * per_page: page * per_page]
    page_tokens = _page_tokens(page, total_pages)

    from urllib.parse import urlencode
    base_params = request.args.to_dict()
    base_params.pop('page', None)
    base_params.pop('per_page', None)
    base_qs = urlencode(base_params)
    # Tabs (All/Parties/Vendors) each set their own `role` — strip any existing role from the
    # querystring they build on top of, otherwise clicking a tab appends a second role= param
    # and request.args.get('role') keeps returning the *first* one, so tabs stop switching.
    tab_params = dict(base_params)
    tab_params.pop('role', None)
    tab_qs = urlencode(tab_params)

    return render_template('home.html', rows=page_rows, total_count=total_count,
                            total_parties=total_parties, active_parties=active_parties,
                            total_vendors=total_vendors, active_vendors=active_vendors,
                            total_receivable=total_receivable, receivable_from=receivable_from,
                            total_payable=total_payable, payable_to=payable_to, net_balance=net_balance,
                            groups=groups, page=page, per_page=per_page, total_pages=total_pages,
                            page_tokens=page_tokens, base_qs=base_qs, tab_qs=tab_qs,
                            f_search=search_f, f_role=role_f, f_group=group_f, f_status=status_f, active='accounts')

@app.route('/accounts/export')
def export_accounts():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from flask import send_file
    import io
    conn = get_db()
    rows = _filter_accounts_rows(_accounts_rows(conn), request.args.get('search', ''),
                                  request.args.get('role', ''), request.args.get('group', ''), request.args.get('status', ''))
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Ledger"
    headers = ["Name", "Type", "Group / Category", "Phone", "Email", "GSTIN", "Receivable", "Payable", "Balance", "Status"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1B2A4A")
    for r_idx, r in enumerate(rows, 2):
        receivable = r['balance'] if r['balance'] > 0 else 0
        payable = -r['balance'] if r['balance'] < 0 else 0
        for c_idx, val in enumerate([r['name'], r['role'], r['group'], r['contact'] or '', r['email'] or '', r['gstin'] or '',
                                      receivable, payable, r['balance'], r['status']], 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='ledger_export.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/accounts/add', methods=['POST'])
def add_account():
    conn = get_db()
    f = request.form
    name = (f.get('name') or '').strip()
    if not name:
        conn.close()
        return redirect(url_for('accounts'))
    role = f.get('role', 'Party')
    contact = f.get('contact') or None
    email = f.get('email') or None
    address = f.get('address') or None
    category = f.get('category') or None
    gstin = f.get('gstin') or None
    credit_limit = float(f.get('credit_limit') or 0) or None
    opening_balance = float(f.get('opening_balance') or 0)
    try:
        if role == 'Vendor':
            conn.execute("""INSERT INTO vendors (name, category, contact, email, address, credit_limit, opening_balance, gstin)
                            VALUES (?,?,?,?,?,?,?,?)""",
                         (name, category, contact, email, address, credit_limit, opening_balance, gstin))
        else:
            conn.execute("""INSERT INTO parties (name, contact, email, address, credit_limit, opening_balance, category, gstin)
                            VALUES (?,?,?,?,?,?,?,?)""",
                         (name, contact, email, address, credit_limit, opening_balance, category, gstin))
        conn.commit()
    except sqlite3.IntegrityError:
        pass
    conn.close()
    return redirect(url_for('accounts'))

@app.route('/accounts/<role>/<int:account_id>/toggle-status', methods=['POST'])
def toggle_account_status(role, account_id):
    conn = get_db()
    table = 'vendors' if role == 'vendor' else 'parties'
    current = conn.execute(f"SELECT status FROM {table} WHERE id=?", (account_id,)).fetchone()
    new_status = 'Inactive' if (current and current['status'] == 'Active') else 'Active'
    conn.execute(f"UPDATE {table} SET status=? WHERE id=?", (new_status, account_id))
    conn.commit()
    conn.close()
    return redirect(url_for('accounts', **request.args.to_dict()))

@app.route('/')
def home():
    return redirect(url_for('dashboard'))

@app.route('/dashboard')
def dashboard():
    conn = get_db()
    date_from = request.args.get('date_from', '2026-04-01')
    date_to = request.args.get('date_to', '2027-03-31')
    vehicle_f = request.args.get('vehicle', '')

    trip_query = "SELECT t.*, v.vehicle_no FROM trips t LEFT JOIN vehicles v ON t.vehicle_id=v.id WHERE t.date>=? AND t.date<=?"
    params = [date_from, date_to]
    if vehicle_f:
        trip_query += " AND v.vehicle_no=?"
        params.append(vehicle_f)
    trips = conn.execute(trip_query, params).fetchall()

    total_revenue = sum(t['billed_amount'] or 0 for t in trips)
    trip_count = len(trips)
    # Toll is deliberately NOT in this sum — Maintenance > Toll (toll_entries/maintenance
    # category='Toll') is the source of truth for toll cost now, folded into maint_total below.
    # Adding the old per-trip manual `toll` field here too would double-count the same real-world
    # toll charge once it's been synced/imported through Toll Management.
    total_charges_paid = sum((t['fuel_amount'] or 0)+(t['driver_adv_amount'] or 0)+
                              (t['agent_commission'] or 0)+(t['builty_expense'] or 0)+(t['conductor_expense'] or 0)+
                              (t['fine'] or 0)+(t['labour_charges'] or 0)+(t['parking'] or 0)+(t['puncture'] or 0)+
                              (t['urea'] or 0)+(t['loading_expense'] or 0)+(t['unloading_expense'] or 0)+
                              (t['wear_tear'] or 0)+(t['weighbridge_charges'] or 0)+(t['other_expense'] or 0) for t in trips)
    adj_revenue = total_revenue - total_charges_paid

    maint_total_query = "SELECT COALESCE(SUM(m.amount),0) FROM maintenance m LEFT JOIN vehicles v ON m.vehicle_id=v.id WHERE m.date>=? AND m.date<=?"
    maint_toll_query = "SELECT COALESCE(SUM(m.amount),0) FROM maintenance m LEFT JOIN vehicles v ON m.vehicle_id=v.id WHERE m.category='Toll' AND m.date>=? AND m.date<=?"
    maint_total_params = [date_from, date_to]
    if vehicle_f:
        maint_total_query += " AND v.vehicle_no=?"
        maint_toll_query += " AND v.vehicle_no=?"
        maint_total_params.append(vehicle_f)
    maint_total = conn.execute(maint_total_query, maint_total_params).fetchone()[0]
    maint_toll_total = conn.execute(maint_toll_query, maint_total_params).fetchone()[0]
    salaries_total = conn.execute("SELECT COALESCE(SUM(amount),0) FROM salaries WHERE date>=? AND date<=?", (date_from, date_to)).fetchone()[0]
    overheads_total = conn.execute("SELECT COALESCE(SUM(amount),0) FROM overheads WHERE date>=? AND date<=?", (date_from, date_to)).fetchone()[0]
    total_expenses = total_charges_paid + maint_total + overheads_total
    total_profit = total_revenue - total_expenses

    all_vehicles = conn.execute("SELECT DISTINCT vehicle_no FROM vehicles WHERE type IS NOT NULL").fetchall()
    running_vnos = set(t['vehicle_no'] for t in trips if t['vehicle_no'])
    running_count = len(running_vnos)
    idle_count = len(all_vehicles) - running_count

    own_vehicles = conn.execute("SELECT id, vehicle_no FROM vehicles WHERE type IN ('Line','Local')").fetchall()
    own_running_ids = set(t['vehicle_id'] for t in trips if t['type'] in ('Line','Local'))
    own_idle_count = sum(1 for v in own_vehicles if v['id'] not in own_running_ids)
    own_vehicle_count = len(own_vehicles)

    fuel_total = sum(t['fuel_amount'] or 0 for t in trips)
    driveradv_total = sum(t['driver_adv_amount'] or 0 for t in trips)
    toll_total = maint_toll_total  # real Toll Management figure, not the old per-trip manual field
    maint_other_total = maint_total - maint_toll_total  # rest of Maintenance, so this chart's slices don't overlap
    other_exp_total = total_charges_paid - fuel_total - driveradv_total

    exp_items = [('Fuel', fuel_total, '#2a78d6'), ('Driver Advance', driveradv_total, '#eb6834'),
                 ('Toll', toll_total, '#eda100'), ('Maintenance', maint_other_total, '#e34948'),
                 ('Salaries', salaries_total, '#1a9c5b'), ('Expenses', overheads_total, '#7a5ad6'),
                 ('Other', other_exp_total, '#4a3aa7')]
    exp_max = max([v for _, v, _ in exp_items], default=1) or 1
    exp_breakdown = [{'label': l, 'value': v, 'color': c, 'pct': round((v/exp_max)*100, 1)} for l, v, c in exp_items]

    # Vehicle-wise revenue, own vehicles only (Line + Local)
    veh_revenue = {}
    for t in trips:
        if t['type'] in ('Line', 'Local') and t['vehicle_no']:
            veh_revenue[t['vehicle_no']] = veh_revenue.get(t['vehicle_no'], 0) + (t['billed_amount'] or 0)
    veh_rev_list = sorted(veh_revenue.items(), key=lambda x: x[1], reverse=True)[:10]
    veh_rev_max = max([v for _, v in veh_rev_list], default=1) or 1
    vehicle_revenue = [{'vehicle_no': vn, 'revenue': rev, 'pct': round((rev/veh_rev_max)*100, 1)} for vn, rev in veh_rev_list]

    parties_bal = conn.execute("""SELECT p.name,
        (SELECT COALESCE(SUM(billed_amount),0) FROM trips WHERE party_id=p.id) -
        (SELECT COALESCE(SUM(payment_received)+SUM(party_advance),0) FROM trips WHERE party_id=p.id) -
        (SELECT COALESCE(SUM(amount-COALESCE(allocated_amount,0)),0) FROM payments WHERE party_id=p.id AND payment_type='received') +
        COALESCE(p.opening_balance,0) as balance
        FROM parties p ORDER BY balance DESC""").fetchall()
    receivables = [r for r in parties_bal if r['balance'] and r['balance'] > 0]
    total_receivables = sum(r['balance'] for r in receivables)

    vendors_bal = conn.execute("""SELECT v.name,
        (SELECT COALESCE(SUM(m.amount),0) FROM maintenance m WHERE m.vendor_id=v.id) +
        (SELECT COALESCE(SUM(fuel_amount),0) FROM trips WHERE fuel_vendor_id=v.id) +
        (SELECT COALESCE(SUM(driver_adv_amount),0) FROM trips WHERE driver_adv_vendor_id=v.id) +
        (SELECT COALESCE(SUM(CASE WHEN item_type='charge' THEN amount ELSE -amount END),0) FROM invoice_items WHERE vendor_id=v.id) -
        (SELECT COALESCE(SUM(m.paid_amount),0) FROM maintenance m WHERE m.vendor_id=v.id) -
        (SELECT COALESCE(SUM(amount-COALESCE(allocated_amount,0)),0) FROM payments WHERE vendor_id=v.id AND payment_type='paid') -
        COALESCE(v.opening_balance,0) as balance
        FROM vendors v ORDER BY balance DESC""").fetchall()
    payables = [v for v in vendors_bal if v['balance'] and v['balance'] > 0]
    total_payables = sum(v['balance'] for v in payables)

    own_vehicle_turnover = sum(t['billed_amount'] or 0 for t in trips if t['type'] in ('Line','Local'))

    market_trips = [t for t in trips if t['type'] == 'Market']
    market_trip_count = len(market_trips)
    market_billed = sum(t['billed_amount'] or 0 for t in market_trips)
    # Owner's actual contracted cost (fixed rate, or owner_rate x quantity) — not what's been paid
    # to them so far, so the margin holds regardless of whether that payment is still pending.
    market_owner_cost = sum(
        (t['owner_fixed_amount'] or 0) if (t['owner_rate_type'] or 'PER_MT') == 'FIXED' else (t['owner_rate'] or 0) * (t['quantity'] or 0)
        for t in market_trips)
    market_vehicle_profit = market_billed - market_owner_cost

    market_veh_data = {}
    for t in market_trips:
        vn = t['vehicle_no'] or 'Unassigned'
        if vn not in market_veh_data:
            market_veh_data[vn] = {'revenue': 0, 'trips': 0}
        market_veh_data[vn]['revenue'] += t['billed_amount'] or 0
        market_veh_data[vn]['trips'] += 1
    market_veh_list = sorted(market_veh_data.items(), key=lambda x: x[1]['revenue'], reverse=True)[:10]
    market_veh_max = max([d['revenue'] for _, d in market_veh_list], default=1) or 1
    market_vehicle_chart = [{'vehicle_no': vn, 'revenue': d['revenue'], 'trips': d['trips'],
                              'pct': round((d['revenue']/market_veh_max)*100, 1)} for vn, d in market_veh_list]

    recent_trips = conn.execute("""SELECT t.date, t.lr_number, v.vehicle_no, t.from_loc, t.to_loc, p.name as party_name,
                                    t.driver_name, t.billed_amount
                                    FROM trips t LEFT JOIN vehicles v ON t.vehicle_id=v.id
                                    LEFT JOIN parties p ON t.party_id=p.id
                                    ORDER BY t.date DESC LIMIT 8""").fetchall()

    vehicles_list = conn.execute("SELECT DISTINCT vehicle_no FROM vehicles WHERE vehicle_no IS NOT NULL ORDER BY vehicle_no").fetchall()

    fy_map = {'2026-04-01|2027-03-31': ('2026-04-01','2027-03-31'), '2025-04-01|2026-03-31': ('2025-04-01','2026-03-31')}
    f_fy = ''
    for key, (fy_from, fy_to) in fy_map.items():
        if date_from == fy_from and date_to == fy_to:
            f_fy = key
            break

    maint_query = """SELECT m.category, SUM(m.amount) as total FROM maintenance m
                     LEFT JOIN vehicles v ON m.vehicle_id=v.id
                     WHERE m.date>=? AND m.date<=?"""
    maint_params = [date_from, date_to]
    if vehicle_f:
        maint_query += " AND v.vehicle_no=?"
        maint_params.append(vehicle_f)
    maint_query += " GROUP BY m.category ORDER BY total DESC LIMIT 8"
    maint_by_category = conn.execute(maint_query, maint_params).fetchall()
    maint_cat_max = max([m['total'] for m in maint_by_category], default=1) or 1
    maint_by_category = [{'category': m['category'], 'total': m['total'],
                           'pct': round((m['total'] / maint_cat_max) * 100, 1)} for m in maint_by_category]

    # Fleet status donut (own vehicles): Running / Idle
    fleet_total = own_vehicle_count or 1
    fleet_running_pct = round((own_vehicle_count - own_idle_count) / fleet_total * 100, 1)

    # Top performing vehicles table (own vehicles, by revenue)
    top_vehicles = []
    for vn, rev in veh_rev_list[:5]:
        vtrips = sum(1 for t in trips if t['vehicle_no'] == vn)
        top_vehicles.append({'vehicle_no': vn, 'trips': vtrips, 'revenue': rev})

    # Real alerts from real data. Compliance alerts are a pure DB read (cs.refresh_compliance
    # inside get_compliance_alerts never calls a provider) — safe on every page load.
    alerts = []
    for r in receivables[:3]:
        alerts.append({'text': f"{r['name']} balance overdue", 'sub': f"₹{r['balance']:,.0f} pending"})
    if own_idle_count > 0:
        alerts.append({'text': f"{own_idle_count} own vehicle(s) currently idle", 'sub': 'Check Idle Tracker for details'})
    for a in cs.get_compliance_alerts(conn)[:5]:
        verb = 'expired' if a['status'] == 'Expired' else f"expires in {a['days_left']}d"
        alerts.append({'text': f"{a['vehicle_no']} — {a['label']} {verb}", 'sub': f"Expiry {a['expiry']} — renew from Vehicles > Edit > Compliance"})

    conn.close()
    return render_template('dashboard.html',
        total_revenue=total_revenue, adj_revenue=adj_revenue, total_charges_paid=total_charges_paid,
        total_expenses=total_expenses, total_profit=total_profit, trip_count=trip_count,
        running_count=running_count, idle_count=idle_count, own_vehicle_turnover=own_vehicle_turnover,
        market_trip_count=market_trip_count, market_vehicle_profit=market_vehicle_profit, market_vehicle_chart=market_vehicle_chart,
        own_idle_count=own_idle_count, own_vehicle_count=own_vehicle_count,
        fuel_total=fuel_total, driveradv_total=driveradv_total, toll_total=toll_total, other_exp_total=other_exp_total,
        maint_total=maint_total, salaries_total=salaries_total,
        receivables=receivables[:5], total_receivables=total_receivables,
        payables=payables[:5], total_payables=total_payables,
        recent_trips=recent_trips, vehicles=vehicles_list,
        f_date_from=date_from, f_date_to=date_to, f_vehicle=vehicle_f, f_fy=f_fy,
        maint_by_category=maint_by_category, exp_breakdown=exp_breakdown, vehicle_revenue=vehicle_revenue,
        fleet_running_pct=fleet_running_pct, top_vehicles=top_vehicles, alerts=alerts, active='dashboard')

@app.route('/vehicles/compliance/sync-all', methods=['POST'])
def vehicles_compliance_sync_all():
    """Syncs every own-fleet vehicle's Fitness/PUC/Permit against the mock providers, then
    goes straight back to the existing All Vehicles tab — no separate page for this."""
    conn = get_db()
    summary = cs.sync_all_vehicles(conn)
    conn.close()
    return redirect(url_for('vehicles_list', tab='all', synced=summary['synced'],
                             changed=summary['changed'], failed=summary['failed']))

@app.route('/vehicles/<int:v_id>/rc-lookup')
def vehicle_rc_lookup(v_id):
    """On-demand, click-to-fetch RC detail view — hits the live eChallan API right now and
    renders every field it returns, read-only. Does not touch the vehicles table itself; the
    same key/data now also feeds the Fitness/PUC/Permit compliance sync (see
    compliance_service.sync_vehicle + providers/echallan_client.py) but this page never
    writes anything — it's a pure read."""
    conn = get_db()
    v = conn.execute("SELECT id, vehicle_no FROM vehicles WHERE id=?", (v_id,)).fetchone()
    if not v:
        conn.close()
        return redirect(url_for('vehicles_list'))
    key_row = conn.execute("SELECT value FROM settings WHERE key='rc_lookup_api_key'").fetchone()
    api_key = key_row['value'] if key_row else ''
    conn.close()
    result = fetch_rc(v['vehicle_no'], api_key)
    return render_template('vehicle_rc_view.html', v=v, result=result, active='vehicles')

@app.route('/vehicles')
def vehicles_list():
    conn = get_db()
    tab = request.args.get('tab', 'all')
    if tab == 'insurance':
        return _maintenance_insurance_tab(conn, template='vehicles_list.html', active='vehicles', base_path='/vehicles')
    if tab == 'permit':
        return _maintenance_category_tab(conn, 'permit', template='vehicles_list.html', active='vehicles', base_path='/vehicles')
    type_f = request.args.get('type', '')
    status_f = request.args.get('status', '')
    vehicle_f = request.args.get('vehicle', '')
    query = """SELECT v.id, v.vehicle_no, v.type, v.registration_date, v.capacity_mt,
               v.insurance_expiry, v.fitness_expiry, v.puc_valid_upto, v.permit_valid_upto,
               v.status, v.body_type, v.notes, v.chassis_number, v.engine_number,
               (SELECT COUNT(*) FROM trips WHERE vehicle_id=v.id) as trip_count,
               (SELECT COALESCE(SUM(billed_amount),0) FROM trips WHERE vehicle_id=v.id) as total_billed
               FROM vehicles v WHERE v.type IS NOT NULL"""
    params = []
    if type_f:
        query += " AND v.type = ?"
        params.append(type_f)
    if status_f:
        query += " AND COALESCE(v.status,'Active') = ?"
        params.append(status_f)
    if vehicle_f:
        query += " AND v.vehicle_no = ?"
        params.append(vehicle_f)
    query += " ORDER BY v.type, v.vehicle_no"
    all_rows = conn.execute(query, params).fetchall()

    # Real, vehicle-level compliance snapshot for the overview cards + sidebar — every count
    # below is derived straight from the (filtered) vehicles table, never fabricated.
    total_vehicles = len(all_rows)
    active_count = sum(1 for r in all_rows if (r['status'] or 'Active') == 'Active')
    maint_count = sum(1 for r in all_rows if r['status'] == 'In Maintenance')
    inactive_count = sum(1 for r in all_rows if r['status'] == 'Inactive')

    def _bucket_counts(field):
        expired = expiring = 0
        for r in all_rows:
            b = _expiry_bucket(r[field])
            if b == 'expired':
                expired += 1
            elif b == 'expiring':
                expiring += 1
        return expired, expiring

    ins_expired, ins_expiring = _bucket_counts('insurance_expiry')
    puc_expired, puc_expiring = _bucket_counts('puc_valid_upto')
    permit_expired, permit_expiring = _bucket_counts('permit_valid_upto')
    fit_expired, fit_expiring = _bucket_counts('fitness_expiry')
    compliance_overview = [
        {'label': 'Insurance', 'expired': ins_expired, 'expiring': ins_expiring},
        {'label': 'PUC', 'expired': puc_expired, 'expiring': puc_expiring},
        {'label': 'Fitness', 'expired': fit_expired, 'expiring': fit_expiring},
        {'label': 'Permits', 'expired': permit_expired, 'expiring': permit_expiring},
    ]

    body_type_counts = {}
    for r in all_rows:
        bt = r['body_type'] or 'Unspecified'
        body_type_counts[bt] = body_type_counts.get(bt, 0) + 1
    body_type_dist = sorted(body_type_counts.items(), key=lambda kv: -kv[1])

    total_count = len(all_rows)
    page, per_page, total_pages = _paginate(request.args.get('page'), request.args.get('per_page'), total_count,
                                             per_page_options=(10, 25, 50, 100), default_per_page=10)
    page_rows = all_rows[(page - 1) * per_page: page * per_page]
    page_tokens = _page_tokens(page, total_pages)
    from urllib.parse import urlencode
    base_params = request.args.to_dict()
    base_params.pop('page', None)
    base_params.pop('per_page', None)
    base_qs = urlencode(base_params)

    all_vehicle_nos = conn.execute("SELECT DISTINCT vehicle_no FROM vehicles WHERE vehicle_no IS NOT NULL ORDER BY vehicle_no").fetchall()

    # Compliance badges (Insurance/Fitness/PUC/Permit) — own fleet only, see compliance_service's
    # module docstring for why Market vehicles are excluded. Keyed by vehicle_id for O(1) lookup
    # in the template; a Market vehicle simply has no entry and renders "—" for every badge.
    compliance_by_vehicle = {c['vehicle_id']: c for c in cs.refresh_compliance(conn)}

    conn.close()
    return render_template('vehicles_list.html', tab='all', rows=page_rows, f_type=type_f, f_status=status_f,
                            f_vehicle=vehicle_f, all_vehicle_nos=all_vehicle_nos, active='vehicles',
                            total_vehicles=total_vehicles, active_count=active_count, maint_count=maint_count,
                            inactive_count=inactive_count, ins_expiring=ins_expiring, puc_expiring=puc_expiring,
                            permit_expiring=permit_expiring, fit_expiring=fit_expiring,
                            compliance_overview=compliance_overview, body_type_dist=body_type_dist,
                            compliance_by_vehicle=compliance_by_vehicle,
                            page=page, total_pages=total_pages, per_page=per_page, page_tokens=page_tokens, base_qs=base_qs)

EMPLOYEE_AVATAR_PALETTE = [('#eaf1fb', '#2a78d6'), ('#fff3e8', 'var(--accent)'), ('#e8f5ee', 'var(--green)'),
                           ('rgba(74,58,167,0.1)', '#4a3aa7'), ('#fdecea', 'var(--red)')]

def _employee_avatar_colors(seed):
    return EMPLOYEE_AVATAR_PALETTE[(seed or 0) % len(EMPLOYEE_AVATAR_PALETTE)]

def _employee_initials(name):
    parts = (name or '').split()
    if not parts:
        return '??'
    if len(parts) == 1:
        return parts[0][:2].upper()
    return (parts[0][0] + parts[1][0]).upper()

def _employee_outstanding_advance(conn, name):
    given = conn.execute("SELECT COALESCE(SUM(amount),0) FROM advances WHERE employee=? COLLATE NOCASE AND type='given'", (name,)).fetchone()[0]
    repaid = conn.execute("SELECT COALESCE(SUM(amount),0) FROM advances WHERE employee=? COLLATE NOCASE AND type='repaid'", (name,)).fetchone()[0]
    return given - repaid

def _employee_attendance_stats(conn, employee_id, date_from, date_to):
    """(present, absent, leave, half_day, working_days, pct) for one employee over an arbitrary
    date range (not just a calendar month) — working_days is every day that got a status marked
    at all (Present/Absent/Leave/Half Day) within the range, not the range's full length, since
    not every day is necessarily marked yet."""
    rows = conn.execute("SELECT status FROM attendance WHERE employee_id=? AND date>=? AND date<=?",
                        (employee_id, date_from, date_to)).fetchall()
    present = sum(1 for r in rows if r['status'] == 'Present')
    absent = sum(1 for r in rows if r['status'] == 'Absent')
    leave = sum(1 for r in rows if r['status'] == 'Leave')
    half = sum(1 for r in rows if r['status'] == 'Half Day')
    working = len(rows)
    pct = round((present + 0.5 * half) / working * 100, 1) if working else None
    return {'present': present, 'absent': absent, 'leave': leave, 'half_day': half,
            'working_days': working, 'pct': pct}

def _current_month_key():
    return datetime.date.today().strftime('%Y-%m')

def _employee_month_calendar(conn, employee_id, month_key):
    """Weeks of 7 day-cells (Mon-Sun, blank-padded) for the attendance drawer's calendar grid,
    each cell tagged with a one-letter status class the template maps to a colour: p/a/l/h/n."""
    year, month = int(month_key[:4]), int(month_key[5:7])
    first = datetime.date(year, month, 1)
    days_in_month = calendar.monthrange(year, month)[1]
    status_by_day = {r['date']: r for r in conn.execute(
        "SELECT date, status, in_time, out_time FROM attendance WHERE employee_id=? AND substr(date,1,7)=?",
        (employee_id, month_key)).fetchall()}
    class_map = {'Present': 'p', 'Absent': 'a', 'Leave': 'l', 'Half Day': 'h'}
    letter_map = {'Present': 'P', 'Absent': 'A', 'Leave': 'L', 'Half Day': 'HD'}
    weeks = []
    week = [None] * first.weekday()  # Monday=0
    for d in range(1, days_in_month + 1):
        date_str = f"{year:04d}-{month:02d}-{d:02d}"
        rec = status_by_day.get(date_str)
        week.append({'day': d, 'date': date_str,
                      'cls': class_map.get(rec['status'], 'n') if rec else 'n',
                      'letter': letter_map.get(rec['status'], '') if rec else '',
                      'status': rec['status'] if rec else None,
                      'in_time': rec['in_time'] if rec else None, 'out_time': rec['out_time'] if rec else None})
        if len(week) == 7:
            weeks.append(week)
            week = []
    if week:
        while len(week) < 7:
            week.append(None)
        weeks.append(week)
    return weeks

@app.route('/salaries')
def salaries_list():
    conn = get_db()
    tab = request.args.get('tab', 'overview')
    if tab == 'salary':
        return _employees_salary_tab(conn)
    if tab == 'attendance':
        return _employees_attendance_tab(conn)
    return _employees_overview_tab(conn)

def _employees_overview_tab(conn):
    default_from = datetime.date.today().replace(day=1).isoformat()
    date_from = request.args.get('date_from') or default_from
    date_to = request.args.get('date_to') or datetime.date.today().isoformat()
    if date_from > date_to:
        date_from, date_to = date_to, date_from
    today = datetime.date.today().isoformat()

    employees = conn.execute("SELECT * FROM employees ORDER BY name").fetchall()
    total_employees = len(employees)
    drivers = sum(1 for e in employees if (e['type'] or '') == 'Driver')
    office_staff = total_employees - drivers

    today_rows = {r['employee_id']: r['status'] for r in
                  conn.execute("SELECT employee_id, status FROM attendance WHERE date=?", (today,)).fetchall()}
    present_today = sum(1 for s in today_rows.values() if s == 'Present')
    on_leave_today = sum(1 for s in today_rows.values() if s == 'Leave')

    period_pcts = []
    salary_status_by_emp = {}
    latest_salary_by_emp = {}
    for e in employees:
        stats = _employee_attendance_stats(conn, e['id'], date_from, date_to)
        if stats['pct'] is not None:
            period_pcts.append(stats['pct'])
        sal_rows = conn.execute("SELECT net_salary FROM salaries WHERE employee_id=? AND date IS NOT NULL AND date>=? AND date<=?",
                                (e['id'], date_from, date_to)).fetchall()
        salary_status_by_emp[e['id']] = 'paid' if sal_rows else None
        latest_salary_by_emp[e['id']] = sum(s['net_salary'] or 0 for s in sal_rows)
    overall_attendance_pct = round(sum(period_pcts) / len(period_pcts), 1) if period_pcts else 0

    period_sal = conn.execute("""SELECT COALESCE(SUM(net_salary),0), COALESCE(SUM(CASE WHEN payment_status='paid' THEN net_salary ELSE 0 END),0)
                                 FROM salaries WHERE date IS NOT NULL AND date>=? AND date<=?""", (date_from, date_to)).fetchone()
    total_payroll = period_sal[0] or 0
    paid_payroll = period_sal[1] or 0
    pending_payroll = total_payroll - paid_payroll
    total_outstanding_advances = sum(_employee_outstanding_advance(conn, e['name']) for e in employees)

    # Department distribution + Drivers vs Office Staff share the same underlying split, shown two ways.
    dept_counts = {}
    for e in employees:
        d = e['role'] or e['type'] or 'Other'
        dept_counts[d] = dept_counts.get(d, 0) + 1
    dept_dist = sorted([{'label': k, 'count': v} for k, v in dept_counts.items()], key=lambda x: -x['count'])
    dept_max = max([d['count'] for d in dept_dist], default=1) or 1

    # Salary distribution — net salary paid per employee within the selected period, top 6.
    sal_dist = sorted([{'name': e['name'], 'amount': latest_salary_by_emp.get(e['id']) or 0} for e in employees
                        if latest_salary_by_emp.get(e['id'])], key=lambda x: -x['amount'])[:6]
    sal_dist_max = max([s['amount'] for s in sal_dist], default=1) or 1

    # 6-month attendance trend (fleet average %, per calendar month — independent of the
    # date_from/date_to filter above, which is a fixed trailing trend not a range selector).
    trend = []
    y_, m_ = datetime.date.today().year, datetime.date.today().month
    for i in range(5, -1, -1):
        mm = m_ - i; yy = y_
        while mm <= 0:
            mm += 12; yy -= 1
        mf, mt = _month_bounds(yy, mm)
        pcts = [p for e in employees if (p := _employee_attendance_stats(conn, e['id'], mf, mt)['pct']) is not None]
        trend.append({'label': calendar.month_abbr[mm], 'pct': round(sum(pcts) / len(pcts), 1) if pcts else 0})
    trend_max = max([t['pct'] for t in trend], default=1) or 1

    rows = []
    for e in employees:
        stats = _employee_attendance_stats(conn, e['id'], date_from, date_to)
        rows.append({
            'id': e['id'], 'name': e['name'], 'employee_code': e['employee_code'] or f"EMP-{e['id']:03d}",
            'role': e['role'] or e['type'] or '—', 'joining_date': e['joining_date'] or '—',
            'attendance_pct': stats['pct'], 'salary_status': salary_status_by_emp.get(e['id']),
            'net_salary': latest_salary_by_emp.get(e['id']),
            'outstanding': _employee_outstanding_advance(conn, e['name']),
            'status': e['status'] or 'Active',
            'initials': _employee_initials(e['name']), 'avatar_colors': _employee_avatar_colors(e['id']),
            # Full record, for the Edit Employee modal (pre-filled from this same row).
            'type': e['type'], 'raw_role': e['role'], 'mobile': e['mobile'], 'email': e['email'],
            'address': e['address'], 'date_of_birth': e['date_of_birth'], 'bank_account': e['bank_account'],
            'ifsc_code': e['ifsc_code'], 'upi_id': e['upi_id'], 'emergency_contact': e['emergency_contact'],
            'aadhaar': e['aadhaar'], 'pan': e['pan'], 'driving_license': e['driving_license'],
            'basic_salary_raw': e['basic_salary'] or 0,
        })

    total_count = len(rows)
    page, per_page, total_pages = _paginate(request.args.get('page'), request.args.get('per_page'), total_count,
                                             per_page_options=(5, 10, 25, 50), default_per_page=5)
    page_rows = rows[(page - 1) * per_page: page * per_page]
    page_tokens = _page_tokens(page, total_pages)
    from urllib.parse import urlencode
    base_params = request.args.to_dict()
    base_params.pop('page', None); base_params.pop('per_page', None)
    base_qs = urlencode(base_params)

    conn.close()
    return render_template('employees.html', tab='overview',
        total_employees=total_employees, drivers=drivers, office_staff=office_staff,
        present_today=present_today, on_leave_today=on_leave_today, overall_attendance_pct=overall_attendance_pct,
        total_payroll=total_payroll, paid_payroll=paid_payroll, pending_payroll=pending_payroll,
        total_outstanding_advances=total_outstanding_advances,
        dept_dist=dept_dist, dept_max=dept_max, sal_dist=sal_dist, sal_dist_max=sal_dist_max,
        trend=trend, trend_max=trend_max, drivers_pct=round(drivers/total_employees*100,1) if total_employees else 0,
        rows=page_rows, total_count=total_count, page=page, total_pages=total_pages, per_page=per_page,
        page_tokens=page_tokens, base_qs=base_qs, f_date_from=date_from, f_date_to=date_to, active='salaries')

def _employees_salary_tab(conn):
    """Salary is a plain summary/overview now — pay salaries, give advances, and record
    repayments all happen on the classic per-employee Ledger page (/employee/<name>), the same
    proven flow this always used. This tab just reports what happened in the selected period."""
    default_from = datetime.date.today().replace(day=1).isoformat()
    date_from = request.args.get('date_from') or default_from
    date_to = request.args.get('date_to') or datetime.date.today().isoformat()
    if date_from > date_to:
        date_from, date_to = date_to, date_from
    role_f = request.args.get('role', '')
    search_f = request.args.get('search', '')

    employees = conn.execute("SELECT * FROM employees ORDER BY name").fetchall()
    if role_f:
        employees = [e for e in employees if (e['role'] or e['type']) == role_f]
    if search_f:
        s = search_f.lower()
        employees = [e for e in employees if s in (e['name'] or '').lower() or s in (e['employee_code'] or '').lower()]

    rows = []
    for e in employees:
        sal_rows = conn.execute("""SELECT net_salary, date FROM salaries
                                   WHERE employee_id=? AND date IS NOT NULL AND date>=? AND date<=? ORDER BY date DESC""",
                                (e['id'], date_from, date_to)).fetchall()
        rows.append({
            'id': e['id'], 'name': e['name'], 'employee_code': e['employee_code'] or f"EMP-{e['id']:03d}",
            'role': e['role'] or e['type'] or '—', 'total_paid': sum(s['net_salary'] or 0 for s in sal_rows),
            'payment_count': len(sal_rows), 'last_paid_date': sal_rows[0]['date'] if sal_rows else None,
            'outstanding': _employee_outstanding_advance(conn, e['name']),
            'initials': _employee_initials(e['name']), 'avatar_colors': _employee_avatar_colors(e['id']),
        })

    total_payroll = sum(r['total_paid'] for r in rows)
    employees_paid = sum(1 for r in rows if r['payment_count'] > 0)
    total_outstanding_advances = sum(r['outstanding'] for r in rows)
    avg_salary = round(total_payroll / employees_paid, 0) if employees_paid else 0

    total_count = len(rows)
    page, per_page, total_pages = _paginate(request.args.get('page'), request.args.get('per_page'), total_count,
                                             per_page_options=(5, 10, 25, 50), default_per_page=5)
    page_rows = rows[(page - 1) * per_page: page * per_page]
    page_tokens = _page_tokens(page, total_pages)
    from urllib.parse import urlencode
    base_params = request.args.to_dict()
    base_params.pop('page', None); base_params.pop('per_page', None)
    base_qs = urlencode(base_params)

    roles = sorted(set((e['role'] or e['type'] or '') for e in conn.execute("SELECT role, type FROM employees").fetchall() if (e['role'] or e['type'])))

    conn.close()
    return render_template('employees.html', tab='salary',
        rows=page_rows, total_count=total_count, total_payroll=total_payroll,
        employees_paid=employees_paid, total_employees_all=len(employees),
        avg_salary=avg_salary, total_outstanding_advances=total_outstanding_advances,
        page=page, total_pages=total_pages, per_page=per_page, page_tokens=page_tokens, base_qs=base_qs,
        f_date_from=date_from, f_date_to=date_to, roles=roles, f_role=role_f, f_search=search_f,
        active='salaries')

def _recent_month_keys(n):
    out = []
    y, m = datetime.date.today().year, datetime.date.today().month
    for i in range(n):
        mm = m - i; yy = y
        while mm <= 0:
            mm += 12; yy -= 1
        out.append({'key': f"{yy:04d}-{mm:02d}", 'label': f"{calendar.month_name[mm]} {yy}"})
    return out

def _employees_attendance_tab(conn):
    default_from = datetime.date.today().replace(day=1).isoformat()
    date_from = request.args.get('date_from') or default_from
    date_to = request.args.get('date_to') or datetime.date.today().isoformat()
    if date_from > date_to:
        date_from, date_to = date_to, date_from
    role_f = request.args.get('role', '')
    status_f = request.args.get('status', '')
    search_f = request.args.get('search', '')
    today = datetime.date.today().isoformat()

    employees = conn.execute("SELECT * FROM employees ORDER BY name").fetchall()
    if role_f:
        employees = [e for e in employees if (e['role'] or e['type']) == role_f]
    if search_f:
        s = search_f.lower()
        employees = [e for e in employees if s in (e['name'] or '').lower() or s in (e['employee_code'] or '').lower()]

    today_rows = {r['employee_id']: r['status'] for r in
                  conn.execute("SELECT employee_id, status FROM attendance WHERE date=?", (today,)).fetchall()}
    present_kpi = sum(1 for s in today_rows.values() if s == 'Present')
    absent_kpi = sum(1 for s in today_rows.values() if s == 'Absent')
    leave_kpi = sum(1 for s in today_rows.values() if s == 'Leave')
    half_kpi = sum(1 for s in today_rows.values() if s == 'Half Day')
    # "Late" isn't its own status (Mark Attendance only offers Present/Absent/Leave/Half Day) — it's
    # a Present row whose logged in_time is after this cutoff, a simple fixed threshold rather than
    # a per-employee shift schedule this app doesn't model.
    LATE_CUTOFF = '09:30'
    late_kpi = conn.execute("SELECT COUNT(*) FROM attendance WHERE date=? AND status='Present' AND in_time IS NOT NULL AND in_time > ?",
                            (today, LATE_CUTOFF)).fetchone()[0]

    rows = []
    all_pcts = []
    for e in employees:
        stats = _employee_attendance_stats(conn, e['id'], date_from, date_to)
        if stats['pct'] is not None:
            all_pcts.append(stats['pct'])
        band = ('Good' if stats['pct'] is None or stats['pct'] >= 95 else
                'Average' if stats['pct'] >= 85 else str(int(stats['pct'])) + '%' if stats['pct'] is not None else '—')
        row = {'id': e['id'], 'name': e['name'], 'employee_code': e['employee_code'] or f"EMP-{e['id']:03d}",
               'role': e['role'] or e['type'] or '—', 'initials': _employee_initials(e['name']),
               'avatar_colors': _employee_avatar_colors(e['id']), 'today_status': today_rows.get(e['id']), **stats, 'band': band}
        if status_f and row['today_status'] != status_f:
            continue
        rows.append(row)
    overall_pct = round(sum(all_pcts) / len(all_pcts), 1) if all_pcts else 0

    total_count = len(rows)
    page, per_page, total_pages = _paginate(request.args.get('page'), request.args.get('per_page'), total_count,
                                             per_page_options=(6, 10, 25, 50), default_per_page=6)
    page_rows = rows[(page - 1) * per_page: page * per_page]
    page_tokens = _page_tokens(page, total_pages)
    from urllib.parse import urlencode
    base_params = request.args.to_dict()
    base_params.pop('page', None); base_params.pop('per_page', None)
    base_qs = urlencode(base_params)

    roles = sorted(set((e['role'] or e['type'] or '') for e in conn.execute("SELECT role, type FROM employees").fetchall() if (e['role'] or e['type'])))
    all_employees_list = [{'id': e['id'], 'name': e['name'], 'employee_code': e['employee_code'] or f"EMP-{e['id']:03d}"} for e in employees]
    # The drawer's calendar is a fixed month grid — it can't render an arbitrary multi-month range,
    # so it shows the calendar month containing the end of the selected period.
    calendar_month_key = date_to[:7]

    for r in page_rows:
        r['calendar'] = _employee_month_calendar(conn, r['id'], calendar_month_key)
        r['recent_history'] = conn.execute(
            "SELECT date, status, in_time, out_time, remarks FROM attendance WHERE employee_id=? ORDER BY date DESC LIMIT 10",
            (r['id'],)).fetchall()

    conn.close()
    return render_template('employees.html', tab='attendance',
        rows=page_rows, total_count=total_count, present_kpi=present_kpi, absent_kpi=absent_kpi,
        leave_kpi=leave_kpi, half_kpi=half_kpi, late_kpi=late_kpi, overall_pct=overall_pct,
        page=page, total_pages=total_pages, per_page=per_page, page_tokens=page_tokens, base_qs=base_qs,
        f_date_from=date_from, f_date_to=date_to, calendar_month_key=calendar_month_key,
        roles=roles, f_role=role_f, f_status=status_f, f_search=search_f,
        all_employees_list=all_employees_list, today=today, active='salaries')

@app.route('/overheads')
def overheads_list():
    conn = get_db()
    cat_f = request.args.get('category', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    query = "SELECT id, date, category, amount, notes, payment_mode, receipt_number FROM overheads WHERE 1=1"
    params = []
    if cat_f:
        query += " AND category = ?"
        params.append(cat_f)
    if date_from:
        query += " AND date >= ?"
        params.append(date_from)
    if date_to:
        query += " AND date <= ?"
        params.append(date_to)
    query += " ORDER BY date DESC"
    rows = conn.execute(query, params).fetchall()
    categories = conn.execute("SELECT DISTINCT category FROM overheads ORDER BY category").fetchall()

    total_amount = conn.execute("SELECT COALESCE(SUM(amount),0) FROM overheads").fetchone()[0]
    total_count = conn.execute("SELECT COUNT(*) FROM overheads").fetchone()[0]
    import datetime
    this_month = datetime.datetime.now().strftime('%Y-%m')
    this_month_amount = conn.execute("SELECT COALESCE(SUM(amount),0) FROM overheads WHERE substr(date,1,7)=?", (this_month,)).fetchone()[0]
    month_count = conn.execute("SELECT COUNT(DISTINCT substr(date,1,7)) FROM overheads").fetchone()[0]
    avg_per_month = total_amount / month_count if month_count > 0 else 0

    conn.close()
    return render_template('overheads_list.html', rows=rows, categories=categories,
                            f_category=cat_f, f_date_from=date_from, f_date_to=date_to,
                            total_amount=total_amount, total_count=total_count,
                            this_month_amount=this_month_amount, avg_per_month=avg_per_month, active='overheads')

@app.route('/trips')
def trips_list():
    conn = get_db()
    vehicle_f = request.args.get('vehicle', '')
    party_f = request.args.get('party', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    status_f = request.args.get('status', '')
    lr_f = request.args.get('lr_received', '')
    lr_number_f = request.args.get('lr_number', '')
    from_f = request.args.get('from_loc', '')
    to_f = request.args.get('to_loc', '')
    type_f = request.args.get('type', '')

    query = """SELECT t.id, t.date, t.lr_number, v.vehicle_no, p.name as party_name,
               t.from_loc, t.to_loc, t.billed_amount, t.lr_received, t.type,
               t.payment_received, t.party_advance,
               t.end_date, t.end_time, t.actual_km, t.shortage_qty, t.shortage_unit, t.shortage_amount, t.shortage_remarks, t.remarks
               FROM trips t
               LEFT JOIN vehicles v ON t.vehicle_id = v.id
               LEFT JOIN parties p ON t.party_id = p.id
               WHERE 1=1"""
    params = []
    if vehicle_f:
        query += " AND v.vehicle_no LIKE ?"
        params.append(f"%{vehicle_f}%")
    if party_f:
        query += " AND p.name LIKE ?"
        params.append(f"%{party_f}%")
    if date_from:
        query += " AND t.date >= ?"
        params.append(date_from)
    if date_to:
        query += " AND t.date <= ?"
        params.append(date_to)
    if status_f == 'active':
        query += " AND t.end_date IS NULL"
    elif status_f == 'completed':
        query += " AND t.end_date IS NOT NULL"
    if lr_f == 'received':
        query += " AND t.lr_received = 'Yes'"
    elif lr_f == 'not_received':
        query += " AND (t.lr_received IS NULL OR t.lr_received != 'Yes')"
    if lr_number_f:
        query += " AND t.lr_number LIKE ?"
        params.append(f"%{lr_number_f}%")
    if from_f:
        query += " AND t.from_loc LIKE ?"
        params.append(f"%{from_f}%")
    if to_f:
        query += " AND t.to_loc LIKE ?"
        params.append(f"%{to_f}%")
    if type_f:
        query += " AND t.type = ?"
        params.append(type_f)

    where_clause = query[query.find("WHERE"):]

    total_shown = conn.execute(f"SELECT COALESCE(SUM(t.billed_amount),0) FROM trips t LEFT JOIN vehicles v ON t.vehicle_id=v.id LEFT JOIN parties p ON t.party_id=p.id {where_clause}", params).fetchone()[0]
    total_count = conn.execute(f"SELECT COUNT(*) FROM trips t LEFT JOIN vehicles v ON t.vehicle_id=v.id LEFT JOIN parties p ON t.party_id=p.id {where_clause}", params).fetchone()[0]
    pending_total = conn.execute(f"SELECT COALESCE(SUM(t.billed_amount-COALESCE(t.payment_received,0)-COALESCE(t.party_advance,0)),0) FROM trips t LEFT JOIN vehicles v ON t.vehicle_id=v.id LEFT JOIN parties p ON t.party_id=p.id {where_clause}", params).fetchone()[0]
    active_count = conn.execute(f"SELECT COUNT(*) FROM trips t LEFT JOIN vehicles v ON t.vehicle_id=v.id LEFT JOIN parties p ON t.party_id=p.id {where_clause} AND t.end_date IS NULL", params).fetchone()[0]
    lr_received_count = conn.execute(f"SELECT COUNT(*) FROM trips t LEFT JOIN vehicles v ON t.vehicle_id=v.id LEFT JOIN parties p ON t.party_id=p.id {where_clause} AND t.lr_received='Yes'", params).fetchone()[0]
    lr_pending_count = total_count - lr_received_count

    page, per_page, total_pages = _paginate(request.args.get('page'), request.args.get('per_page'), total_count,
                                             per_page_options=(10, 25, 50, 100), default_per_page=50)
    offset = (page - 1) * per_page

    query += f" ORDER BY t.date DESC LIMIT {per_page} OFFSET {offset}"

    trips = conn.execute(query, params).fetchall()
    vehicles = conn.execute("SELECT DISTINCT vehicle_no FROM vehicles WHERE vehicle_no IS NOT NULL ORDER BY vehicle_no").fetchall()
    parties = conn.execute("SELECT DISTINCT name FROM parties ORDER BY name").fetchall()
    conn.close()

    from urllib.parse import urlencode
    base_params = request.args.to_dict()
    base_params.pop('page', None)
    base_params.pop('per_page', None)
    base_qs = urlencode(base_params)
    page_tokens = _page_tokens(page, total_pages)

    return render_template('trips_list.html', trips=trips, total_shown=total_shown, total_count=total_count,
                            pending_total=pending_total, active_count=active_count,
                            lr_received_count=lr_received_count, lr_pending_count=lr_pending_count,
                            page=page, total_pages=total_pages, per_page=per_page, base_qs=base_qs, page_tokens=page_tokens,
                            vehicles=vehicles, parties=parties,
                            f_vehicle=vehicle_f, f_party=party_f, f_date_from=date_from, f_date_to=date_to,
                            f_status=status_f, f_lr=lr_f, f_lr_number=lr_number_f, f_from=from_f, f_to=to_f, f_type=type_f, active='trips')

@app.route('/trips/add', methods=['GET', 'POST'])
def add_trip():
    conn = get_db()
    if request.method == 'POST':
        f = request.form
        def get_or_create_vehicle(vno):
            if not vno or not vno.strip():
                return None
            vno = vno.strip()
            row = conn.execute("SELECT id FROM vehicles WHERE vehicle_no = ? COLLATE NOCASE", (vno,)).fetchone()
            if row:
                return row[0]
            cur = conn.execute("INSERT INTO vehicles (vehicle_no) VALUES (?)", (vno,))
            return cur.lastrowid
        def n(key):
            return float(f.get(key) or 0)

        vehicle_id = get_or_create_vehicle(f.get('vehicle_no'))
        party_id = get_or_create_party(conn, f.get('party_name'))
        fuel_vendor_id = get_or_create_vendor(conn, f.get('fuel_vendor'))
        driveradv_vendor_id = get_or_create_vendor(conn, f.get('driver_adv_vendor'))
        owner_vendor_id = get_or_create_vendor(conn, f.get('owner_name')) if f.get('owner_name') else None
        misc_vendor_id = get_or_create_vendor(conn, f.get('misc_vendor'))

        quantity = n('quantity')
        rate = n('rate')
        rate_type = f.get('rate_type')
        fixed_rate_amount = n('fixed_rate_amount')
        owner_rate_type = f.get('owner_rate_type') or 'PER_MT'
        owner_fixed_amount = n('owner_fixed_amount')

        freight = fixed_rate_amount if rate_type == 'FIXED' else quantity * rate

        # Note: driver_payment ("Driver Bata") is no longer collected from this form — Driver Advance
        # already covers driver payments, so this legacy field is left at its schema default (0) for
        # new trips instead of being asked for twice.
        total_charges = (n('detention_charges')+n('gps_cost')+n('loading_charge')+
                          n('unloading_charge')+n('police_charges')+n('sim_tracking')+n('union_charges')+
                          n('weight_charges')+n('other_charges'))
        total_deductions = (n('brokerage')+n('builty_commission')+n('late_fees')+n('material_damage')+
                             n('shortage_amount')+n('tds')+n('other_deductions'))
        billed_amount = freight + total_charges - total_deductions

        # conductor_expense/wear_tear are no longer collected from this form (dropped in favor of
        # Fuel Liters/Fuel Price below) — simply left out of the INSERT so they take their schema
        # default for every new trip, same "stop collecting, don't touch history" pattern as
        # driver_payment above.
        cols = ['date','lr_number','vehicle_id','type','party_id','from_loc','to_loc','quantity','rate',
                'driver_name','material','rate_type','billed_amount',
                'detention_charges','gps_cost','loading_charge','unloading_charge',
                'police_charges','sim_tracking','union_charges','weight_charges','other_charges',
                'brokerage','builty_commission','late_fees','material_damage','shortage_amount','shortage_qty','tds','other_deductions',
                'fuel_amount','fuel_vendor_id','fuel_liters','fuel_price','driver_adv_amount','driver_adv_vendor_id','party_advance','payment_received',
                'owner_name','fixed_rate_amount','owner_rate','owner_rate_type','owner_fixed_amount','paid_to_owner','owner_vendor_id',
                'agent_commission','builty_expense','fine','labour_charges','parking','puncture',
                'toll','urea','loading_expense','unloading_expense','weighbridge_charges','other_expense','misc_vendor_id',
                'lr_received','is_empty']
        vals = [f.get('date'), f.get('lr_number'), vehicle_id, f.get('type'), party_id, f.get('from_loc'), f.get('to_loc'),
                quantity, rate, f.get('driver_name'), f.get('material'), rate_type, billed_amount,
                n('detention_charges'), n('gps_cost'), n('loading_charge'), n('unloading_charge'),
                n('police_charges'), n('sim_tracking'), n('union_charges'), n('weight_charges'), n('other_charges'),
                n('brokerage'), n('builty_commission'), n('late_fees'), n('material_damage'), n('shortage_amount'),
                n('shortage_qty'), n('tds'), n('other_deductions'),
                n('fuel_amount'), fuel_vendor_id, f.get('fuel_liters') or None, n('fuel_price'), n('driver_adv_amount'), driveradv_vendor_id, n('party_advance'), n('payment_received'),
                f.get('owner_name'), fixed_rate_amount, n('owner_rate'), owner_rate_type, owner_fixed_amount, n('paid_to_owner'), owner_vendor_id,
                n('agent_commission'), n('builty_expense'), n('fine'), n('labour_charges'),
                n('parking'), n('puncture'), n('toll'), n('urea'), n('loading_expense'), n('unloading_expense'),
                n('weighbridge_charges'), n('other_expense'), misc_vendor_id,
                f.get('lr_received') or None, 1 if f.get('is_empty') else 0]
        placeholders = ','.join('?' * len(cols))
        cur = conn.execute(f"INSERT INTO trips ({','.join(cols)}) VALUES ({placeholders})", vals)
        _save_trip_custom_items(conn, cur.lastrowid, f)
        conn.commit()
        conn.close()
        return redirect(url_for('trips_list'))
    conn.close()
    vehicles, parties, vendors, combined_names = _get_autocomplete_lists()
    conn2 = get_db()
    employees = conn2.execute("SELECT name FROM employees ORDER BY name").fetchall()
    conn2.close()
    return render_template('trip_form.html', mode='add', t={}, custom_items=[],
                            vehicles=vehicles, parties=parties, vendors=vendors, combined_names=combined_names,
                            employees=employees, return_to='', active='trips')

def _save_trip_custom_items(conn, trip_id, form):
    """Replace a trip's custom "Others" line items (stored in invoice_items, the same table the
    per-trip invoice editor already uses) from the submitted other_desc/other_type/other_rate/other_qty
    rows. Amount stored is rate*qty — the invoice-facing table only keeps the final amount.
    An item can optionally be tagged with a Vendor (other_vendor) — e.g. diesel bought from a
    second fuel vendor mid-trip, or a cash advance handled by a different party — in which case it
    still appears on the trip's invoice exactly as before, but ALSO counts toward what that vendor
    is owed, the same way fuel_amount/driver_adv_amount already do (see _accounts_rows,
    _get_vendor_ledger_entries, and the matching Dashboard/Business Performance queries)."""
    conn.execute("DELETE FROM invoice_items WHERE trip_id=?", (trip_id,))
    descs = form.getlist('other_desc')
    types = form.getlist('other_type')
    rates = form.getlist('other_rate')
    qtys = form.getlist('other_qty')
    vendor_names = form.getlist('other_vendor')
    for i, desc in enumerate(descs):
        desc = (desc or '').strip()
        if not desc:
            continue
        try:
            rate = float(rates[i]) if i < len(rates) and rates[i] else 0
        except ValueError:
            rate = 0
        try:
            qty = float(qtys[i]) if i < len(qtys) and qtys[i] else 1
        except ValueError:
            qty = 1
        item_type = types[i] if i < len(types) and types[i] in ('charge', 'deduction') else 'charge'
        vendor_name = (vendor_names[i].strip() if i < len(vendor_names) and vendor_names[i] else '')
        vendor_id = get_or_create_vendor(conn, vendor_name) if vendor_name else None
        conn.execute("INSERT INTO invoice_items (trip_id, description, amount, item_type, vendor_id) VALUES (?,?,?,?,?)",
                     (trip_id, desc, rate * (qty or 1), item_type, vendor_id))

def _get_autocomplete_lists():
    conn = get_db()
    vehicles = conn.execute("SELECT vehicle_no FROM vehicles WHERE vehicle_no IS NOT NULL ORDER BY vehicle_no").fetchall()
    parties = conn.execute("SELECT name FROM parties ORDER BY name").fetchall()
    vendors = conn.execute("SELECT name FROM vendors ORDER BY name").fetchall()
    combined_names = sorted(set([p['name'] for p in parties] + [v['name'] for v in vendors]))
    conn.close()
    return vehicles, parties, vendors, combined_names

MAINTENANCE_CHECKLIST = ['Engine Oil', 'Oil Filter', 'Air Filter', 'Diesel Filter', 'Greasing',
    'Brake Adjustment', 'Clutch Adjustment', 'Air Leak Fixed', 'Electrical Check', 'Suspension Check',
    'Wheel Alignment', 'Coolant Change', 'Battery Check', 'Lights Check', 'Road Test']

MAINTENANCE_TAB_LABELS = {'tyres': 'Tyres', 'battery': 'Battery', 'insurance': 'Insurance',
                           'permit': 'Permit & Fitness', 'urea': 'Urea'}
BATTERY_TYPES = ['Tubular', 'Flat Plate', 'Maintenance Free', 'SMF']

def _maintenance_classify(category, service_type=None):
    """Which Maintenance tab an entry belongs on. Entries created through the new Add Service form
    always carry service_type, so those are unambiguous; older free-text entries fall back to
    keyword-matching their category — this is the one place that logic lives, reused everywhere
    (Overview cards, category tabs, the Service tab itself) so they never disagree with each other."""
    if service_type:
        return 'Service'
    c = (category or '').lower()
    if 'tyre' in c or 'tire' in c or 'punt' in c or 'punc' in c:
        return 'Tyres'
    if 'battery' in c:
        return 'Battery'
    if 'insur' in c:
        return 'Insurance'
    if 'fitness' in c or 'permit' in c or 'pollution' in c or 'puc' in c:
        return 'Permit & Fitness'
    if 'urea' in c or 'adblue' in c or 'def' in c:
        return 'Urea'
    if 'toll' in c:
        return 'Toll'
    if 'service' in c:
        return 'Service'
    return 'Other'

TYRE_ACTIONS = ['New Tyre Fitted', 'Tyre Replacement', 'Tyre Resole', 'Puncture Repair']
TYRE_POSITIONS = ['FL', 'FR', 'RL1', 'RR1', 'RL2', 'RR2', 'Spare']

@app.route('/maintenance')
def maintenance_list():
    conn = get_db()
    tab = request.args.get('tab', 'overview')
    # Insurance and Permit & Fitness now live on the Vehicles page — send old links there
    # instead of rendering them here, so nothing bookmarked to /maintenance?tab=insurance breaks.
    if tab in ('insurance', 'permit'):
        conn.close()
        return redirect(url_for('vehicles_list', tab=tab))
    if tab == 'service':
        return _maintenance_service_tab(conn)
    if tab == 'tyres':
        return _maintenance_tyres_tab(conn)
    if tab == 'battery':
        return _maintenance_battery_tab(conn)
    if tab == 'urea':
        return _maintenance_urea_tab(conn)
    if tab == 'toll':
        return _maintenance_toll_tab(conn)
    if tab in MAINTENANCE_TAB_LABELS:
        return _maintenance_category_tab(conn, tab)
    return _maintenance_overview_tab(conn)

def _maintenance_overview_tab(conn):
    last_month_end = datetime.date.today().replace(day=1) - datetime.timedelta(days=1)
    default_from, default_to = _month_bounds(last_month_end.year, last_month_end.month)
    date_from = request.args.get('date_from') or default_from
    date_to = request.args.get('date_to') or default_to
    if date_from > date_to:
        date_from, date_to = date_to, date_from

    own_vehicles = conn.execute("SELECT id, vehicle_no, registration_date FROM vehicles WHERE type IN ('Line','Local') ORDER BY vehicle_no").fetchall()
    total_fleet = len(own_vehicles)

    period_entries = conn.execute("""SELECT m.id, m.date, m.vehicle_id, m.category, m.service_type, m.amount, m.paid_amount, v.vehicle_no
                                     FROM maintenance m JOIN vehicles v ON m.vehicle_id=v.id
                                     WHERE v.type IN ('Line','Local') AND m.date>=? AND m.date<=?
                                     ORDER BY m.date DESC""", (date_from, date_to)).fetchall()
    total_cost = sum(e['amount'] or 0 for e in period_entries)
    total_paid = sum(e['paid_amount'] or 0 for e in period_entries)
    total_unpaid = total_cost - total_paid
    vehicles_serviced = len(set(e['vehicle_id'] for e in period_entries))

    # Fitness & Permit are tracked on the Vehicles page's own Compliance system now (with their own
    # renew/sync flow and cost path) — no longer duplicated here as a Maintenance Overview card.
    cat_labels = ['Service', 'Tyres', 'Battery', 'Insurance', 'Urea', 'Toll']
    cat_slugs = {'Service': 'service', 'Tyres': 'tyres', 'Battery': 'battery', 'Insurance': 'insurance',
                 'Urea': 'urea', 'Toll': 'toll'}
    cat_data = {k: {'vehicles': set(), 'count': 0, 'cost': 0} for k in cat_labels}
    for e in period_entries:
        k = _maintenance_classify(e['category'], e['service_type'])
        if k in cat_data:
            cat_data[k]['vehicles'].add(e['vehicle_id'])
            cat_data[k]['count'] += 1
            cat_data[k]['cost'] += e['amount'] or 0
    category_cards = [{'label': k, 'slug': cat_slugs[k], 'vehicles_done': len(cat_data[k]['vehicles']),
                        'count': cat_data[k]['count'], 'cost': cat_data[k]['cost']} for k in cat_labels]

    km_row = conn.execute("""SELECT COALESCE(SUM(t.actual_km),0) as km FROM trips t JOIN vehicles v ON t.vehicle_id=v.id
                             WHERE v.type IN ('Line','Local') AND t.date>=? AND t.date<=? AND t.actual_km IS NOT NULL""",
                           (date_from, date_to)).fetchone()
    total_km = km_row['km'] or 0
    avg_cost_km = round(total_cost / total_km, 2) if total_km > 0 else None

    active_count = 0
    for v in own_vehicles:
        active_days, idle_days, _ = _vehicle_active_idle_days(conn, v['id'], v['registration_date'], date_from, date_to)
        if active_days > 0:
            active_count += 1
    fleet_availability = round(active_count / total_fleet * 100, 1) if total_fleet else 0

    pending_by_vehicle = {}
    for e in period_entries:
        pend = (e['amount'] or 0) - (e['paid_amount'] or 0)
        if pend > 0.01:
            pending_by_vehicle[e['vehicle_no']] = pending_by_vehicle.get(e['vehicle_no'], 0) + pend
    pending_count = len(pending_by_vehicle)

    # Fleet health bands — derived from each vehicle's period maintenance spend relative to the
    # fleet average (no mechanical condition data exists to score against, so this is explicitly
    # a spend-based proxy: heavier-than-average repair spend this period = worse band).
    fleet_avg_cost = (total_cost / total_fleet) if total_fleet else 0
    health_bands = {'Excellent': 0, 'Good': 0, 'Average': 0, 'Poor': 0, 'Critical': 0}
    for v in own_vehicles:
        v_cost = sum(e['amount'] or 0 for e in period_entries if e['vehicle_id'] == v['id'])
        if v_cost <= 0:
            band = 'Excellent'
        else:
            ratio = v_cost / fleet_avg_cost if fleet_avg_cost else 0
            if ratio < 0.5: band = 'Good'
            elif ratio < 1.0: band = 'Average'
            elif ratio < 2.0: band = 'Poor'
            else: band = 'Critical'
        health_bands[band] += 1

    actions = []
    for vno, amt in sorted(pending_by_vehicle.items(), key=lambda x: -x[1])[:4]:
        actions.append({'icon': '\U0001F4B0', 'label': 'Pending Payment', 'vehicle': vno, 'detail': f"₹{amt:,.0f} unpaid", 'kind': 'bad'})
    last_maint = {}
    for row in conn.execute("SELECT vehicle_id, MAX(date) as last_date FROM maintenance WHERE vehicle_id IS NOT NULL GROUP BY vehicle_id").fetchall():
        last_maint[row['vehicle_id']] = row['last_date']
    d2 = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()
    stale = []
    for v in own_vehicles:
        last_date = last_maint.get(v['id'])
        if last_date:
            days_since = (d2 - datetime.datetime.strptime(last_date, '%Y-%m-%d').date()).days
            if days_since > 90:
                stale.append((v['vehicle_no'], days_since))
        else:
            stale.append((v['vehicle_no'], None))
    stale.sort(key=lambda x: (x[1] is None, -(x[1] or 0)))
    for vno, days in stale[:max(0, 6 - len(actions))]:
        if days is None:
            actions.append({'icon': '⚠️', 'label': 'No Maintenance History', 'vehicle': vno, 'detail': 'No records yet', 'kind': 'warn'})
        else:
            actions.append({'icon': '⏱', 'label': 'No Recent Maintenance', 'vehicle': vno, 'detail': f"{days} days since last entry", 'kind': 'warn'})

    recent_entries = period_entries[:8]
    spend_by_vehicle = {}
    for e in period_entries:
        if e['vehicle_no']:
            spend_by_vehicle[e['vehicle_no']] = spend_by_vehicle.get(e['vehicle_no'], 0) + (e['amount'] or 0)
    top_spend = sorted(spend_by_vehicle.items(), key=lambda x: -x[1])[:5]

    trend = []
    end_d = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()
    y_, m_ = end_d.year, end_d.month
    for i in range(5, -1, -1):
        mm = m_ - i
        yy = y_
        while mm <= 0:
            mm += 12
            yy -= 1
        mf, mt = _month_bounds(yy, mm)
        cost = conn.execute("""SELECT COALESCE(SUM(m.amount),0) FROM maintenance m JOIN vehicles v ON m.vehicle_id=v.id
                               WHERE v.type IN ('Line','Local') AND m.date>=? AND m.date<=?""", (mf, mt)).fetchone()[0]
        trend.append({'label': calendar.month_abbr[mm], 'cost': cost})
    trend_max = max([t['cost'] for t in trend], default=1) or 1
    chart_w, chart_h = 640, 150
    n = len(trend)
    for i, t in enumerate(trend):
        t['x'] = round((i / (n - 1) * (chart_w - 40) + 30) if n > 1 else chart_w / 2, 1)
        t['y'] = round(chart_h - (t['cost'] / trend_max * (chart_h - 30)) - 15, 1)

    conn.close()
    return render_template('maintenance.html', tab='overview',
        total_fleet=total_fleet, vehicles_serviced=vehicles_serviced, pending_count=pending_count,
        total_cost=total_cost, total_unpaid=total_unpaid, avg_cost_km=avg_cost_km, fleet_availability=fleet_availability,
        category_cards=category_cards, health_bands=health_bands, actions=actions,
        recent_entries=recent_entries, top_spend=top_spend, trend=trend, trend_max=trend_max, chart_w=chart_w, chart_h=chart_h,
        f_date_from=date_from, f_date_to=date_to, active='maintenance')

UREA_LOW_STOCK_THRESHOLD_L = 200  # default reorder point per location — no Settings field for
                                   # this yet, so it's a documented constant rather than a
                                   # silently-invented one; easy to move into Settings later.

def _maintenance_urea_tab(conn):
    """Urea (AdBlue/DEF) stock ledger — every row is a real transaction (a purchase into stock,
    or a consumption out of it), with a running balance computed live from the ledger, not a
    separately-maintained counter that could drift. Cost only enters the maintenance table (and
    therefore the Overview 'Urea' card) at the moment stock is purchased or a direct/off-stock
    top-up happens — using stock you already paid for isn't a second expense."""
    location_f = request.args.get('location', '')
    supplier_f = request.args.get('supplier', '')
    type_f = request.args.get('type', '')
    search_f = request.args.get('search', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    query = """SELECT u.*, ve.name as supplier_name, v.vehicle_no
               FROM urea_transactions u LEFT JOIN vendors ve ON u.supplier_id=ve.id
               LEFT JOIN vehicles v ON u.vehicle_id=v.id WHERE 1=1"""
    params = []
    if location_f:
        query += " AND u.location=?"; params.append(location_f)
    if supplier_f:
        query += " AND ve.name=?"; params.append(supplier_f)
    if type_f:
        query += " AND u.txn_type=?"; params.append(type_f)
    if search_f:
        query += " AND (u.batch_no LIKE ? OR u.invoice_no LIKE ? OR v.vehicle_no LIKE ?)"
        params += [f"%{search_f}%"] * 3
    if date_from:
        query += " AND u.date>=?"; params.append(date_from)
    if date_to:
        query += " AND u.date<=?"; params.append(date_to)
    query += " ORDER BY u.date DESC, u.id DESC"
    all_rows = conn.execute(query, params).fetchall()

    # KPIs are computed from the *unfiltered* full ledger — a filtered view shouldn't make the
    # stock level look different from reality.
    full_ledger = conn.execute("SELECT * FROM urea_transactions ORDER BY date, id").fetchall()
    current_stock = 0.0
    stock_in_qty = stock_in_value = 0.0
    for t in full_ledger:
        if t['txn_type'] == 'stock_in':
            current_stock += t['quantity_l']
            stock_in_qty += t['quantity_l']
            stock_in_value += t['total_value'] or 0
        elif t['source'] == 'stock':
            current_stock -= t['quantity_l']
    avg_unit_cost = (stock_in_value / stock_in_qty) if stock_in_qty else 0
    total_value = current_stock * avg_unit_cost

    today = datetime.date.today().isoformat()
    month_start = datetime.date.today().replace(day=1).isoformat()
    today_usage = sum(t['quantity_l'] for t in full_ledger if t['txn_type'] == 'stock_out' and t['date'] == today)
    month_usage = sum(t['quantity_l'] for t in full_ledger if t['txn_type'] == 'stock_out' and t['date'] >= month_start)

    # Consumption (L/100km) needs real distance driven — the own fleet's actual_km already
    # logged on trips this month, not an invented figure.
    month_km = conn.execute("""SELECT COALESCE(SUM(t.actual_km),0) FROM trips t JOIN vehicles v ON t.vehicle_id=v.id
                               WHERE v.type IN ('Line','Local') AND t.date>=? AND t.actual_km IS NOT NULL""",
                            (month_start,)).fetchone()[0]
    avg_consumption = round(month_usage / month_km * 100, 2) if month_km else None

    # Low stock is evaluated per location, since that's the level stock is actually held/reordered at.
    locations = sorted({t['location'] for t in full_ledger if t['location']})
    low_stock_alerts = 0
    for loc in locations:
        bal = 0.0
        for t in full_ledger:
            if t['location'] != loc:
                continue
            if t['txn_type'] == 'stock_in':
                bal += t['quantity_l']
            elif t['source'] == 'stock':
                bal -= t['quantity_l']
        if bal < UREA_LOW_STOCK_THRESHOLD_L:
            low_stock_alerts += 1

    # Stock Trend (last 30 days) — running balance sampled at each transaction date.
    trend_cutoff = (datetime.date.today() - datetime.timedelta(days=30)).isoformat()
    trend_rows = []
    running = 0.0
    for t in full_ledger:
        if t['txn_type'] == 'stock_in':
            running += t['quantity_l']
        elif t['source'] == 'stock':
            running -= t['quantity_l']
        if t['date'] >= trend_cutoff:
            trend_rows.append({'date': t['date'], 'balance': running})

    # Top vehicles by usage (last 30 days, stock_out only).
    usage_by_vehicle = {}
    for t in full_ledger:
        if t['txn_type'] == 'stock_out' and t['date'] >= trend_cutoff and t['vehicle_id']:
            vno = next((r['vehicle_no'] for r in all_rows if r['id'] == t['id']), None)
            usage_by_vehicle.setdefault(t['vehicle_id'], [None, 0.0])
            usage_by_vehicle[t['vehicle_id']][1] += t['quantity_l']
    vname_lookup = {r['vehicle_id']: r['vehicle_no'] for r in all_rows if r['vehicle_id']}
    top_vehicles = sorted([{'vehicle_no': vname_lookup.get(vid, '—'), 'qty': q[1]}
                            for vid, q in usage_by_vehicle.items()], key=lambda x: -x['qty'])[:5]

    total_count = len(all_rows)
    page, per_page, total_pages = _paginate(request.args.get('page'), request.args.get('per_page'), total_count,
                                             per_page_options=(10, 25, 50, 100), default_per_page=8)
    page_rows = all_rows[(page - 1) * per_page: page * per_page]
    page_tokens = _page_tokens(page, total_pages)
    from urllib.parse import urlencode
    base_params = request.args.to_dict()
    base_params.pop('page', None)
    base_params.pop('per_page', None)
    base_qs = urlencode(base_params)

    vehicles, parties, vendors, combined_names = _get_autocomplete_lists()
    suppliers = conn.execute("SELECT DISTINCT name FROM vendors ORDER BY name").fetchall()
    conn.close()
    return render_template('maintenance.html', tab='urea',
        rows=page_rows, total_count=total_count, current_stock=current_stock, total_value=total_value,
        avg_unit_cost=avg_unit_cost, today_usage=today_usage, month_usage=month_usage,
        avg_consumption=avg_consumption, low_stock_alerts=low_stock_alerts, locations=locations,
        trend_rows=trend_rows, top_vehicles=top_vehicles,
        page=page, total_pages=total_pages, per_page=per_page, page_tokens=page_tokens, base_qs=base_qs,
        vehicles=vehicles, suppliers=suppliers, combined_names=combined_names,
        f_location=location_f, f_supplier=supplier_f, f_type=type_f, f_search=search_f,
        f_date_from=date_from, f_date_to=date_to, active='maintenance')

TOLL_STATUS_COLORS = {  # background, text — same style as the compliance badges on Vehicles
    'synced':   ('#e8f5ee', 'var(--green)'),
    'approved': ('#eaf1fb', '#2a78d6'),
    'pending':  ('#fff3e8', 'var(--accent)'),
    'rejected': ('#fdecea', 'var(--red)'),
}
# A handful of real Odisha/national-highway toll plazas this fleet's own trips actually pass —
# used only by the mock FASTag sync below, never shown as if it came from a live feed.
FASTAG_MOCK_PLAZAS = [
    ('Chandikhole Toll Plaza', 'NH-16', 'Odisha'), ('Sambalpur Toll Plaza', 'NH-53', 'Odisha'),
    ('Rengali Toll Plaza', 'NH-520', 'Odisha'), ('Khordha Toll Plaza', 'NH-16', 'Odisha'),
    ('Cuttack Toll Plaza', 'NH-16', 'Odisha'), ('Jharsuguda Toll Plaza', 'NH-49', 'Odisha'),
]
_TOLL_IMPORT_STASH = {}  # excel-preview token -> {'rows': [...], 'ts': datetime} — see toll_excel_preview

def _save_toll_receipt(file_storage):
    if not file_storage or not file_storage.filename:
        return None
    filename = secure_filename(file_storage.filename)
    if not filename:
        return None
    unique = f"toll_{int(datetime.datetime.now().timestamp() * 1000)}_{filename}"
    file_storage.save(os.path.join(TOLL_RECEIPT_UPLOAD_DIR, unique))
    return f"uploads/toll_receipts/{unique}"

def _toll_tab_base_context(conn):
    """Everything the Toll tab needs to render — factored out from _maintenance_toll_tab so the
    Excel preview/import routes (which also re-render this same tab, with extra wizard state
    layered on top) build it identically instead of duplicating the query/KPI logic. Caller owns
    the connection (opens and closes it) since the excel routes need it open a bit longer."""
    vehicle_f = request.args.get('vehicle', '')
    source_f = request.args.get('source', '')
    status_f = request.args.get('status', '')
    search_f = request.args.get('search', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    # Toll Management is own-fleet only (Line/Local) — same scope as Compliance and Urea — a
    # hired/Market vehicle's toll isn't this company's cost to track.
    query = """SELECT te.*, v.vehicle_no, t.lr_number
               FROM toll_entries te JOIN vehicles v ON te.vehicle_id=v.id
               LEFT JOIN trips t ON te.trip_id=t.id
               WHERE v.type IN ('Line','Local')"""
    params = []
    if vehicle_f:
        query += " AND v.vehicle_no=?"; params.append(vehicle_f)
    if source_f:
        query += " AND te.source=?"; params.append(source_f)
    if status_f:
        query += " AND te.status=?"; params.append(status_f)
    if search_f:
        query += " AND (te.toll_plaza LIKE ? OR te.highway LIKE ? OR v.vehicle_no LIKE ?)"
        params += [f"%{search_f}%"] * 3
    if date_from:
        query += " AND te.date>=?"; params.append(date_from)
    if date_to:
        query += " AND te.date<=?"; params.append(date_to)
    query += " ORDER BY te.date DESC, COALESCE(te.time,'') DESC, te.id DESC"
    all_rows = conn.execute(query, params).fetchall()

    # KPIs come from the *unfiltered* own-fleet ledger — narrowing the table below with a filter
    # shouldn't make the stat cards lie about the real totals.
    full_ledger = conn.execute("""SELECT te.*, v.vehicle_no FROM toll_entries te
                                  JOIN vehicles v ON te.vehicle_id=v.id
                                  WHERE v.type IN ('Line','Local')""").fetchall()
    today = datetime.date.today().isoformat()
    month_start = datetime.date.today().replace(day=1).isoformat()
    today_rows = [r for r in full_ledger if r['date'] == today]
    month_rows = [r for r in full_ledger if r['date'] >= month_start]
    today_total = sum(r['amount'] or 0 for r in today_rows)
    month_total = sum(r['amount'] or 0 for r in month_rows)
    fastag_rows = [r for r in full_ledger if r['source'] == 'fastag']
    manual_rows = [r for r in full_ledger if r['source'] == 'manual']
    fastag_total = sum(r['amount'] or 0 for r in fastag_rows)
    manual_total = sum(r['amount'] or 0 for r in manual_rows)
    grand_total = fastag_total + manual_total
    fastag_pct = round(fastag_total / grand_total * 100, 1) if grand_total else 0
    manual_pct = round(manual_total / grand_total * 100, 1) if grand_total else 0

    trips_with_toll = len(set(r['trip_id'] for r in full_ledger if r['trip_id']))
    denom = trips_with_toll or len(full_ledger)
    avg_per_trip = round(grand_total / denom, 2) if denom else 0

    plaza_totals = {}
    for r in full_ledger:
        plaza_totals[r['toll_plaza']] = plaza_totals.get(r['toll_plaza'], 0) + (r['amount'] or 0)
    highest_plaza, highest_plaza_amt = max(plaza_totals.items(), key=lambda x: x[1]) if plaza_totals else (None, 0)
    top_plazas = sorted([{'plaza': k, 'amount': v} for k, v in plaza_totals.items()], key=lambda x: -x['amount'])[:6]
    plaza_max = max([p['amount'] for p in top_plazas], default=1) or 1

    # Monthly trend, last 6 months — same month-bucket pattern as the Overview tab's cost trend.
    trend = []
    end_d = datetime.date.today()
    y_, m_ = end_d.year, end_d.month
    for i in range(5, -1, -1):
        mm = m_ - i; yy = y_
        while mm <= 0:
            mm += 12; yy -= 1
        mf, mt = _month_bounds(yy, mm)
        cost = sum(r['amount'] or 0 for r in full_ledger if mf <= r['date'] <= mt)
        trend.append({'label': calendar.month_abbr[mm], 'amount': cost})
    trend_max = max([t['amount'] for t in trend], default=1) or 1

    hwy_totals = {}
    for r in full_ledger:
        hwy = r['highway'] or 'Unspecified'
        hwy_totals[hwy] = hwy_totals.get(hwy, 0) + (r['amount'] or 0)
    hwy_dist = sorted([{'highway': k, 'amount': v, 'pct': round(v / grand_total * 100, 1) if grand_total else 0}
                        for k, v in hwy_totals.items()], key=lambda x: -x['amount'])[:6]

    veh_totals = {}
    for r in full_ledger:
        veh_totals[r['vehicle_no']] = veh_totals.get(r['vehicle_no'], 0) + (r['amount'] or 0)
    top_vehicles = sorted([{'vehicle_no': k, 'amount': v} for k, v in veh_totals.items()], key=lambda x: -x['amount'])[:6]
    veh_max = max([v['amount'] for v in top_vehicles], default=1) or 1

    total_count = len(all_rows)
    page, per_page, total_pages = _paginate(request.args.get('page'), request.args.get('per_page'), total_count,
                                             per_page_options=(10, 25, 50, 100), default_per_page=10)
    page_rows = all_rows[(page - 1) * per_page: page * per_page]
    page_tokens = _page_tokens(page, total_pages)
    from urllib.parse import urlencode
    base_params = request.args.to_dict()
    base_params.pop('page', None); base_params.pop('per_page', None)
    base_qs = urlencode(base_params)

    own_vehicles = conn.execute("SELECT id, vehicle_no FROM vehicles WHERE type IN ('Line','Local') ORDER BY vehicle_no").fetchall()
    last_sync_row = conn.execute("SELECT value FROM settings WHERE key='toll_fastag_last_sync'").fetchone()

    return dict(
        rows=page_rows, total_count=total_count,
        today_total=today_total, today_count=len(today_rows), month_total=month_total, month_count=len(month_rows),
        fastag_total=fastag_total, fastag_count=len(fastag_rows), fastag_pct=fastag_pct,
        manual_total=manual_total, manual_count=len(manual_rows), manual_pct=manual_pct,
        avg_per_trip=avg_per_trip, highest_plaza=highest_plaza, highest_plaza_amt=highest_plaza_amt,
        trend=trend, trend_max=trend_max, top_plazas=top_plazas, plaza_max=plaza_max, hwy_dist=hwy_dist,
        top_vehicles=top_vehicles, veh_max=veh_max, grand_total=grand_total,
        page=page, total_pages=total_pages, per_page=per_page, page_tokens=page_tokens, base_qs=base_qs,
        vehicles=own_vehicles, last_sync_at=(last_sync_row['value'] if last_sync_row else None),
        f_vehicle=vehicle_f, f_source=source_f, f_status=status_f, f_search=search_f,
        f_date_from=date_from, f_date_to=date_to, status_colors=TOLL_STATUS_COLORS)

def _maintenance_toll_tab(conn):
    """Toll Management — own fleet only. Every real toll charge (FASTag-synced or manually
    logged) also lands in the shared `maintenance` table (category='Toll') so it flows into the
    Overview cost rollup and Ledger automatically, same as Urea/Battery/etc. already do."""
    ctx = _toll_tab_base_context(conn)
    conn.close()
    return render_template('maintenance.html', tab='toll', active='maintenance', **ctx)

def _maintenance_category_tab(conn, tab_slug, template='maintenance.html', active='maintenance', base_path='/maintenance'):
    label = MAINTENANCE_TAB_LABELS[tab_slug]
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    vehicle_f = request.args.get('vehicle', '')

    query = """SELECT m.id, m.date, v.vehicle_no, m.category, m.service_type, m.amount, m.paid_amount, ve.name as vendor_name
               FROM maintenance m LEFT JOIN vehicles v ON m.vehicle_id=v.id LEFT JOIN vendors ve ON m.vendor_id=ve.id
               WHERE 1=1"""
    params = []
    if date_from:
        query += " AND m.date>=?"; params.append(date_from)
    if date_to:
        query += " AND m.date<=?"; params.append(date_to)
    if vehicle_f:
        query += " AND v.vehicle_no=?"; params.append(vehicle_f)
    query += " ORDER BY m.date DESC"
    all_rows = conn.execute(query, params).fetchall()
    rows = [r for r in all_rows if _maintenance_classify(r['category'], r['service_type']) == label]

    total_count = len(rows)
    total_amount = sum(r['amount'] or 0 for r in rows)
    paid_total = sum(r['paid_amount'] or 0 for r in rows)
    unpaid_total = total_amount - paid_total
    vehicles_involved = len(set(r['vehicle_no'] for r in rows if r['vehicle_no']))

    page, per_page, total_pages = _paginate(request.args.get('page'), request.args.get('per_page'), total_count,
                                             per_page_options=(10, 25, 50, 100), default_per_page=25)
    page_rows = rows[(page - 1) * per_page: page * per_page]
    page_tokens = _page_tokens(page, total_pages)
    from urllib.parse import urlencode
    base_params = request.args.to_dict()
    base_params.pop('page', None)
    base_params.pop('per_page', None)
    base_qs = urlencode(base_params)

    all_vehicles = conn.execute("SELECT vehicle_no FROM vehicles WHERE vehicle_no IS NOT NULL ORDER BY vehicle_no").fetchall()
    conn.close()
    return render_template(template, tab=tab_slug, tab_label=label,
                            rows=page_rows, total_count=total_count, total_amount=total_amount,
                            paid_total=paid_total, unpaid_total=unpaid_total, vehicles_involved=vehicles_involved,
                            page=page, total_pages=total_pages, per_page=per_page, page_tokens=page_tokens, base_qs=base_qs,
                            f_date_from=date_from, f_date_to=date_to, f_vehicle=vehicle_f, vehicles=all_vehicles,
                            base_path=base_path, active=active)

def _maintenance_tyres_tab(conn):
    vehicle_f = request.args.get('vehicle', '')
    action_f = request.args.get('action', '')
    position_f = request.args.get('position', '')
    vendor_f = request.args.get('vendor', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    search_f = request.args.get('search', '')

    query = """SELECT m.*, v.vehicle_no, ve.name as vendor_name FROM maintenance m
               LEFT JOIN vehicles v ON m.vehicle_id=v.id LEFT JOIN vendors ve ON m.vendor_id=ve.id
               WHERE 1=1"""
    params = []
    if vehicle_f:
        query += " AND v.vehicle_no=?"; params.append(vehicle_f)
    if vendor_f:
        query += " AND ve.name=?"; params.append(vendor_f)
    if date_from:
        query += " AND m.date>=?"; params.append(date_from)
    if date_to:
        query += " AND m.date<=?"; params.append(date_to)
    query += " ORDER BY m.date DESC"
    all_rows = conn.execute(query, params).fetchall()

    rows = [r for r in all_rows if _maintenance_classify(r['category'], r['service_type']) == 'Tyres']
    if action_f:
        rows = [r for r in rows if (r['tyre_action'] or r['category'] or '') == action_f]
    if position_f:
        rows = [r for r in rows if (r['tyre_position'] or '') == position_f]
    if search_f:
        s = search_f.lower()
        rows = [r for r in rows if s in (r['vehicle_no'] or '').lower() or s in (r['tyre_id'] or '').lower()
                or s in (r['tyre_brand'] or '').lower() or s in (r['invoice_no'] or '').lower() or s in (r['vendor_name'] or '').lower()]

    total_count = len(rows)
    total_cost = sum(r['amount'] or 0 for r in rows)
    paid_total = sum(r['paid_amount'] or 0 for r in rows)
    unpaid_total = total_cost - paid_total
    vehicles_covered = len(set(r['vehicle_no'] for r in rows if r['vehicle_no']))
    puncture_count = sum(1 for r in rows if any(k in (r['tyre_action'] or r['category'] or '').lower() for k in ('punt', 'punc')))
    avg_cost = round(total_cost / total_count, 0) if total_count else 0

    action_totals = {}
    for r in rows:
        a = r['tyre_action'] or r['category'] or 'Other'
        action_totals[a] = action_totals.get(a, 0) + 1
    action_breakdown = sorted(action_totals.items(), key=lambda x: -x[1])[:6]

    today = datetime.date.today()
    trend = []
    end_ref = datetime.datetime.strptime(date_to, '%Y-%m-%d').date() if date_to else today
    y_, m_ = end_ref.year, end_ref.month
    for i in range(5, -1, -1):
        mm = m_ - i
        yy = y_
        while mm <= 0:
            mm += 12
            yy -= 1
        mf, mt = _month_bounds(yy, mm)
        month_rows = [r for r in all_rows if _maintenance_classify(r['category'], r['service_type']) == 'Tyres'
                      and r['date'] and mf <= r['date'] <= mt]
        trend.append({'label': calendar.month_abbr[mm], 'cost': sum(r['amount'] or 0 for r in month_rows)})
    trend_max = max([t['cost'] for t in trend], default=1) or 1
    chart_w, chart_h = 400, 140
    n = len(trend)
    for i, t in enumerate(trend):
        t['x'] = round((i / (n - 1) * chart_w) if n > 1 else chart_w / 2, 1)
        t['y'] = round(chart_h - (t['cost'] / trend_max * (chart_h - 20)) - 10, 1)

    cost_by_vehicle = {}
    for r in rows:
        if r['vehicle_no']:
            cost_by_vehicle[r['vehicle_no']] = cost_by_vehicle.get(r['vehicle_no'], 0) + (r['amount'] or 0)
    top_costly = sorted(cost_by_vehicle.items(), key=lambda x: -x[1])[:5]

    page, per_page, total_pages = _paginate(request.args.get('page'), request.args.get('per_page'), total_count,
                                             per_page_options=(10, 25, 50, 100), default_per_page=10)
    page_rows = rows[(page - 1) * per_page: page * per_page]
    page_tokens = _page_tokens(page, total_pages)
    from urllib.parse import urlencode
    base_params = request.args.to_dict()
    base_params.pop('page', None)
    base_params.pop('per_page', None)
    base_qs = urlencode(base_params)

    vehicles, parties, vendors, combined_names = _get_autocomplete_lists()
    tyre_vendors = conn.execute("SELECT DISTINCT name FROM vendors ORDER BY name").fetchall()
    tyre_brands = sorted(set(r['tyre_brand'] for r in all_rows if r['tyre_brand']))

    # Vehicle Tyre Layout: for the picked vehicle, the latest real entry logged against each
    # position — no fabricated fitment/stock data, just "what does the record actually say".
    global_rows = conn.execute("""SELECT m.category, m.service_type, m.date, m.id, v.vehicle_no
                                   FROM maintenance m LEFT JOIN vehicles v ON m.vehicle_id=v.id
                                   WHERE v.vehicle_no IS NOT NULL""").fetchall()
    tyre_last_by_vehicle = {}
    for r in global_rows:
        if _maintenance_classify(r['category'], r['service_type']) == 'Tyres':
            key = (r['date'] or '', r['id'])
            if r['vehicle_no'] not in tyre_last_by_vehicle or key > tyre_last_by_vehicle[r['vehicle_no']]:
                tyre_last_by_vehicle[r['vehicle_no']] = key
    default_lv = max(tyre_last_by_vehicle, key=lambda vn: tyre_last_by_vehicle[vn]) if tyre_last_by_vehicle \
        else (vehicles[0]['vehicle_no'] if vehicles else '')
    lv = request.args.get('lv') or default_lv

    layout_rows = conn.execute("""SELECT m.*, v.vehicle_no, ve.name as vendor_name, ts.tyre_type FROM maintenance m
                                   LEFT JOIN vehicles v ON m.vehicle_id=v.id LEFT JOIN vendors ve ON m.vendor_id=ve.id
                                   LEFT JOIN tyre_stock ts ON ts.maintenance_id=m.id
                                   WHERE v.vehicle_no=? ORDER BY m.date, m.id""", (lv,)).fetchall() if lv else []
    layout_tyre_rows = [r for r in layout_rows if _maintenance_classify(r['category'], r['service_type']) == 'Tyres']
    layout_positions = {}
    for pos in TYRE_POSITIONS:
        matches = [r for r in layout_tyre_rows if (r['tyre_position'] or '') == pos]
        layout_positions[pos] = matches[-1] if matches else None
    positions_recorded = sum(1 for pos in TYRE_POSITIONS if layout_positions[pos])
    vehicle_tyre_spend = sum(r['amount'] or 0 for r in layout_tyre_rows)
    vehicle_last_updated = max((r['date'] for r in layout_tyre_rows if r['date']), default=None)

    # Tyre Stock: tyres bought ahead of use. Installing one links its existing purchase
    # record to a vehicle/position instead of billing the vendor a second time.
    stock_rows = conn.execute("""SELECT ts.*, ve.name as vendor_name, iv.vehicle_no as installed_vehicle_no
                                  FROM tyre_stock ts LEFT JOIN vendors ve ON ts.vendor_id=ve.id
                                  LEFT JOIN vehicles iv ON ts.installed_vehicle_id=iv.id
                                  ORDER BY (ts.status='In Stock') DESC, ts.purchase_date DESC, ts.id DESC""").fetchall()
    stock_in_count = sum(1 for s in stock_rows if s['status'] == 'In Stock')
    stock_new_count = sum(1 for s in stock_rows if s['status'] == 'In Stock' and (s['tyre_type'] or '') == 'New')
    stock_resole_count = sum(1 for s in stock_rows if s['status'] == 'In Stock' and (s['tyre_type'] or '') == 'Resole')
    stock_value = sum(s['purchase_cost'] or 0 for s in stock_rows if s['status'] == 'In Stock')
    conn.close()

    return render_template('maintenance.html', tab='tyres',
        rows=page_rows, total_count=total_count, total_cost=total_cost, unpaid_total=unpaid_total,
        vehicles_covered=vehicles_covered, puncture_count=puncture_count, avg_cost=avg_cost,
        action_breakdown=action_breakdown, trend=trend, trend_max=trend_max, chart_w=chart_w, chart_h=chart_h, top_costly=top_costly,
        page=page, total_pages=total_pages, per_page=per_page, page_tokens=page_tokens, base_qs=base_qs,
        vehicles=vehicles, tyre_vendors=tyre_vendors, combined_names=combined_names,
        tyre_actions=TYRE_ACTIONS, tyre_positions=TYRE_POSITIONS, tyre_brands=tyre_brands,
        lv=lv, layout_positions=layout_positions, positions_recorded=positions_recorded,
        vehicle_tyre_spend=vehicle_tyre_spend, vehicle_last_updated=vehicle_last_updated,
        stock_rows=stock_rows, stock_in_count=stock_in_count, stock_new_count=stock_new_count,
        stock_resole_count=stock_resole_count, stock_value=stock_value,
        f_vehicle=vehicle_f, f_action=action_f, f_position=position_f, f_vendor=vendor_f,
        f_date_from=date_from, f_date_to=date_to, f_search=search_f, active='maintenance')

def _battery_status(b):
    """Good/Weak/Replace Soon/Dead bands off the latest logged health% — the same honest,
    threshold-based approach used for Fleet Health on Overview. status_override lets a real
    'Mark as Dead' action win outright, since that's an asserted fact, not a derived one."""
    if b['status_override']:
        return b['status_override']
    if not b['vehicle_id']:
        return 'In Stock'
    h = b['health_pct']
    if h is None:
        return 'Good'
    if h >= 70:
        return 'Good'
    if h >= 40:
        return 'Weak'
    if h >= 20:
        return 'Replace Soon'
    return 'Dead'

def _battery_warranty_upto(b):
    base = b['install_date'] or b['purchase_date']
    if not base or not b['warranty_months']:
        return None
    try:
        d = datetime.datetime.strptime(base, '%Y-%m-%d').date()
    except ValueError:
        return None
    total_month = d.month - 1 + int(b['warranty_months'])
    year = d.year + total_month // 12
    month = total_month % 12 + 1
    day = min(d.day, calendar.monthrange(year, month)[1])
    return datetime.date(year, month, day).isoformat()

def _battery_enrich(r):
    d = dict(r)
    d['status'] = _battery_status(r)
    wu = _battery_warranty_upto(r)
    d['warranty_upto'] = wu
    d['warranty_days_left'] = None
    if wu:
        try:
            d['warranty_days_left'] = (datetime.datetime.strptime(wu, '%Y-%m-%d').date() - datetime.date.today()).days
        except ValueError:
            pass
    return d

def _next_battery_no(conn):
    row = conn.execute("SELECT battery_no FROM batteries WHERE battery_no LIKE 'BAT-%' ORDER BY id DESC LIMIT 1").fetchone()
    n = 0
    if row and row['battery_no']:
        try:
            n = int(row['battery_no'].split('-')[-1])
        except ValueError:
            n = 0
    return f"BAT-{n + 1:05d}"

def _maintenance_battery_tab(conn):
    vehicle_f = request.args.get('vehicle', '')
    brand_f = request.args.get('brand', '')
    status_f = request.args.get('status', '')
    type_f = request.args.get('battery_type', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    search_f = request.args.get('search', '')

    raw_rows = conn.execute("""SELECT b.*, v.vehicle_no, ve.name as vendor_name FROM batteries b
                               LEFT JOIN vehicles v ON b.vehicle_id=v.id LEFT JOIN vendors ve ON b.vendor_id=ve.id
                               ORDER BY b.id DESC""").fetchall()
    all_rows = [_battery_enrich(r) for r in raw_rows]

    rows = all_rows
    if vehicle_f:
        rows = [r for r in rows if r['vehicle_no'] == vehicle_f]
    if brand_f:
        rows = [r for r in rows if (r['brand'] or '') == brand_f]
    if type_f:
        rows = [r for r in rows if (r['battery_type'] or '') == type_f]
    if status_f:
        rows = [r for r in rows if r['status'] == status_f]
    if date_from:
        rows = [r for r in rows if (r['install_date'] or r['purchase_date'] or '') >= date_from]
    if date_to:
        rows = [r for r in rows if (r['install_date'] or r['purchase_date'] or '') <= date_to]
    if search_f:
        s = search_f.lower()
        rows = [r for r in rows if s in (r['battery_no'] or '').lower() or s in (r['brand'] or '').lower()
                or s in (r['vehicle_no'] or '').lower() or s in (r['serial_no'] or '').lower()]

    total_batteries = len(all_rows)
    in_use = sum(1 for r in all_rows if r['vehicle_id'] and r['status'] == 'Good')
    in_stock = sum(1 for r in all_rows if not r['vehicle_id'] and r['status'] == 'In Stock')
    weak_soon = sum(1 for r in all_rows if r['vehicle_id'] and r['status'] in ('Weak', 'Replace Soon'))
    dead_count = sum(1 for r in all_rows if r['status'] in ('Dead', 'Scrapped'))

    today = datetime.date.today()
    ages = []
    for r in all_rows:
        if r['install_date']:
            try:
                d = datetime.datetime.strptime(r['install_date'], '%Y-%m-%d').date()
                ages.append((today - d).days / 30.44)
            except ValueError:
                pass
    avg_life_months = round(sum(ages) / len(ages)) if ages else None

    status_counts = {'Good': 0, 'Weak': 0, 'Replace Soon': 0, 'Dead': 0, 'In Stock': 0}
    for r in all_rows:
        status_counts[r['status']] = status_counts.get(r['status'], 0) + 1

    warranty_list = [r for r in all_rows if r['warranty_upto'] and r['warranty_days_left'] is not None
                      and r['warranty_days_left'] >= 0 and r['status'] not in ('Dead', 'Scrapped')]
    warranty_list.sort(key=lambda r: r['warranty_upto'])
    warranty_list = warranty_list[:6]

    low_health = sorted([r for r in all_rows if r['health_pct'] is not None and r['health_pct'] < 70 and r['vehicle_id']],
                         key=lambda r: r['health_pct'])[:6]

    this_year = str(today.year)
    total_cost_year = sum(r['purchase_price'] or 0 for r in all_rows if (r['purchase_date'] or '').startswith(this_year))
    replaced_year = sum(1 for r in all_rows if r['status'] == 'Dead' and (r['last_checked_date'] or r['install_date'] or '').startswith(this_year))
    cost_per_battery = round(total_cost_year / total_batteries) if total_batteries else 0

    total_count = len(rows)
    page, per_page, total_pages = _paginate(request.args.get('page'), request.args.get('per_page'), total_count,
                                             per_page_options=(10, 25, 50, 100), default_per_page=10)
    page_rows = rows[(page - 1) * per_page: page * per_page]
    page_tokens = _page_tokens(page, total_pages)
    from urllib.parse import urlencode
    base_params = request.args.to_dict()
    base_params.pop('page', None)
    base_params.pop('per_page', None)
    base_qs = urlencode(base_params)

    vehicles, parties, vendors, combined_names = _get_autocomplete_lists()
    battery_vendors = conn.execute("SELECT DISTINCT name FROM vendors ORDER BY name").fetchall()
    battery_brands = sorted(set(r['brand'] for r in all_rows if r['brand']))
    stock_batteries = [r for r in all_rows if not r['vehicle_id'] and r['status'] == 'In Stock']

    checks_by_battery = {}
    if page_rows:
        ids = [r['id'] for r in page_rows]
        placeholders = ','.join('?' * len(ids))
        for c in conn.execute(f"SELECT * FROM battery_checks WHERE battery_id IN ({placeholders}) ORDER BY date DESC, id DESC", ids).fetchall():
            checks_by_battery.setdefault(c['battery_id'], []).append(c)
    conn.close()

    return render_template('maintenance.html', tab='battery',
        rows=page_rows, total_count=total_count,
        total_batteries=total_batteries, in_use=in_use, in_stock=in_stock, weak_soon=weak_soon, dead_count=dead_count,
        avg_life_months=avg_life_months, status_counts=status_counts, warranty_list=warranty_list, low_health=low_health,
        total_cost_year=total_cost_year, replaced_year=replaced_year, cost_per_battery=cost_per_battery,
        page=page, total_pages=total_pages, per_page=per_page, page_tokens=page_tokens, base_qs=base_qs,
        vehicles=vehicles, battery_vendors=battery_vendors, combined_names=combined_names,
        battery_brands=battery_brands, battery_types=BATTERY_TYPES, stock_batteries=stock_batteries,
        checks_by_battery=checks_by_battery,
        f_vehicle=vehicle_f, f_brand=brand_f, f_status=status_f, f_battery_type=type_f,
        f_date_from=date_from, f_date_to=date_to, f_search=search_f, active='maintenance')

INSURANCE_TYPES = ['Comprehensive', 'Third Party', 'Transit Insurance']

def _expiry_bucket(date_str):
    """'expired' / 'expiring' (within 30 days) / 'ok' / None for any yyyy-mm-dd expiry-style date
    field — the same 30-day window used for insurance policies, applied to vehicle-level dates
    (insurance_expiry, puc_valid_upto, permit_valid_upto, fitness_expiry) for the Vehicles overview."""
    if not date_str:
        return None
    try:
        d = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return None
    today = datetime.date.today()
    if d < today:
        return 'expired'
    if (d - today).days <= 30:
        return 'expiring'
    return 'ok'

def _insurance_status(p):
    """Active/Expiring Soon/Expired are always date-derived so they can't go stale; 'Cancelled'
    is the one real-world fact that has to be asserted rather than computed, so it wins outright."""
    if p['status_override'] == 'Cancelled':
        return 'Cancelled'
    if not p['expiry_date']:
        return 'Active'
    try:
        exp = datetime.datetime.strptime(p['expiry_date'], '%Y-%m-%d').date()
    except ValueError:
        return 'Active'
    today = datetime.date.today()
    if exp < today:
        return 'Expired'
    if (exp - today).days <= 30:
        return 'Expiring Soon'
    return 'Active'

def _insurance_enrich(p):
    d = dict(p)
    d['status'] = _insurance_status(p)
    d['days_to_expiry'] = None
    if p['expiry_date']:
        try:
            exp = datetime.datetime.strptime(p['expiry_date'], '%Y-%m-%d').date()
            d['days_to_expiry'] = (exp - datetime.date.today()).days
        except ValueError:
            pass
    return d

def _save_insurance_doc(file_storage, prefix):
    if not file_storage or not file_storage.filename:
        return None
    filename = secure_filename(file_storage.filename)
    if not filename:
        return None
    unique = f"{prefix}_{int(datetime.datetime.now().timestamp() * 1000)}_{filename}"
    file_storage.save(os.path.join(INSURANCE_UPLOAD_DIR, unique))
    return f"uploads/insurance/{unique}"

def _maintenance_insurance_tab(conn, template='maintenance.html', active='maintenance', base_path='/maintenance'):
    vehicle_f = request.args.get('vehicle', '')
    status_f = request.args.get('status', '')
    type_f = request.args.get('type', '')
    insurer_f = request.args.get('insurer', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    raw_rows = conn.execute("""SELECT ip.*, v.vehicle_no, ve.name as insurer_name FROM insurance_policies ip
                               LEFT JOIN vehicles v ON ip.vehicle_id=v.id LEFT JOIN vendors ve ON ip.insurer_id=ve.id
                               ORDER BY ip.id DESC""").fetchall()
    all_rows = [_insurance_enrich(r) for r in raw_rows]

    # Insurance entries logged the old free-text way (before this tab existed) live only in
    # `maintenance`, with no policy_number/expiry/etc — surface them too so this list's totals
    # never disagree with Overview's, which reads `maintenance` directly. They're shown honestly
    # as "Not Tracked" (no expiry date exists to derive a real status from) and are read-only here.
    legacy_maint = conn.execute("""SELECT m.id, m.date, m.category, m.service_type, m.amount, m.paid_amount, m.notes,
                                   m.vendor_id, v.vehicle_no, v.id as veh_id, ve.name as insurer_name
                                   FROM maintenance m LEFT JOIN vehicles v ON m.vehicle_id=v.id
                                   LEFT JOIN vendors ve ON m.vendor_id=ve.id
                                   WHERE m.id NOT IN (SELECT maintenance_id FROM insurance_policies WHERE maintenance_id IS NOT NULL)
                                   """).fetchall()
    for m in legacy_maint:
        if _maintenance_classify(m['category'], m['service_type']) != 'Insurance':
            continue
        all_rows.append({
            'id': None, 'vehicle_id': m['veh_id'], 'vehicle_no': m['vehicle_no'],
            'insurance_type': None, 'insurer_id': m['vendor_id'], 'insurer_name': m['insurer_name'],
            'policy_number': None, 'start_date': m['date'], 'expiry_date': None,
            'premium_amount': m['amount'], 'idv': None, 'ncb_pct': None, 'gst_included': None,
            'agent_name': None, 'agent_contact': None, 'agent_email': None, 'reminder_days': None,
            'notes': m['notes'], 'status_override': None, 'policy_doc_path': None, 'invoice_doc_path': None,
            'rc_doc_path': None, 'maintenance_id': m['id'], 'is_legacy': True,
            'status': 'Not Tracked', 'days_to_expiry': None,
        })

    rows = all_rows
    if vehicle_f:
        rows = [r for r in rows if r['vehicle_no'] == vehicle_f]
    if status_f:
        rows = [r for r in rows if r['status'] == status_f]
    if type_f:
        rows = [r for r in rows if (r['insurance_type'] or '') == type_f]
    if insurer_f:
        rows = [r for r in rows if (r['insurer_name'] or '') == insurer_f]
    if date_from:
        rows = [r for r in rows if (r['start_date'] or '') >= date_from]
    if date_to:
        rows = [r for r in rows if (r['start_date'] or '') <= date_to]

    total_policies = len(all_rows)
    active_count = sum(1 for r in all_rows if r['status'] == 'Active')
    expiring_count = sum(1 for r in all_rows if r['status'] == 'Expiring Soon')
    expired_count = sum(1 for r in all_rows if r['status'] == 'Expired')
    active_pct = round(active_count / total_policies * 100, 1) if total_policies else 0

    today = datetime.date.today()
    this_year = str(today.year)
    total_premium_year = sum(r['premium_amount'] or 0 for r in all_rows if (r['start_date'] or '').startswith(this_year))
    avg_premium = round(sum(r['premium_amount'] or 0 for r in all_rows) / total_policies) if total_policies else 0

    own_fleet_count = conn.execute("SELECT COUNT(*) FROM vehicles WHERE type IN ('Line','Local')").fetchone()[0]
    active_vehicle_ids = set(r['vehicle_id'] for r in all_rows if r['status'] == 'Active' and r['vehicle_id'])
    coverage_pct = round(len(active_vehicle_ids) / own_fleet_count * 100) if own_fleet_count else 0
    total_idv = sum(r['idv'] or 0 for r in all_rows if r['status'] == 'Active')

    total_count = len(rows)
    page, per_page, total_pages = _paginate(request.args.get('page'), request.args.get('per_page'), total_count,
                                             per_page_options=(10, 25, 50, 100), default_per_page=8)
    page_rows = rows[(page - 1) * per_page: page * per_page]
    page_tokens = _page_tokens(page, total_pages)
    from urllib.parse import urlencode
    base_params = request.args.to_dict()
    base_params.pop('page', None)
    base_params.pop('per_page', None)
    base_qs = urlencode(base_params)

    vehicles, parties, vendors, combined_names = _get_autocomplete_lists()
    insurers = conn.execute("SELECT DISTINCT name FROM vendors ORDER BY name").fetchall()
    conn.close()

    return render_template(template, tab='insurance',
        rows=page_rows, total_count=total_count, total_policies=total_policies,
        active_count=active_count, expiring_count=expiring_count, expired_count=expired_count, active_pct=active_pct,
        total_premium_year=total_premium_year, avg_premium=avg_premium, coverage_pct=coverage_pct,
        active_vehicle_count=len(active_vehicle_ids), own_fleet_count=own_fleet_count, total_idv=total_idv,
        page=page, total_pages=total_pages, per_page=per_page, page_tokens=page_tokens, base_qs=base_qs,
        vehicles=vehicles, insurers=insurers, combined_names=combined_names, insurance_types=INSURANCE_TYPES,
        f_vehicle=vehicle_f, f_status=status_f, f_type=type_f, f_insurer=insurer_f,
        f_date_from=date_from, f_date_to=date_to, base_path=base_path, active=active)

def _maintenance_service_tab(conn):
    vehicle_f = request.args.get('vehicle', '')
    workshop_f = request.args.get('workshop', '')
    stype_f = request.args.get('service_type', '')
    status_f = request.args.get('status', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    search_f = request.args.get('search', '')

    query = """SELECT m.*, v.vehicle_no, ve.name as vendor_name FROM maintenance m
               LEFT JOIN vehicles v ON m.vehicle_id=v.id LEFT JOIN vendors ve ON m.vendor_id=ve.id
               WHERE 1=1"""
    params = []
    if vehicle_f:
        query += " AND v.vehicle_no=?"; params.append(vehicle_f)
    if workshop_f:
        query += " AND ve.name=?"; params.append(workshop_f)
    if date_from:
        query += " AND m.date>=?"; params.append(date_from)
    if date_to:
        query += " AND m.date<=?"; params.append(date_to)
    query += " ORDER BY m.date DESC"
    all_rows = conn.execute(query, params).fetchall()

    rows = [r for r in all_rows if _maintenance_classify(r['category'], r['service_type']) == 'Service']
    if stype_f:
        rows = [r for r in rows if (r['service_type'] or r['category'] or '') == stype_f]
    if status_f:
        rows = [r for r in rows if (r['status'] or 'Completed') == status_f]
    if search_f:
        s = search_f.lower()
        rows = [r for r in rows if s in (r['vehicle_no'] or '').lower() or s in (r['service_type'] or r['category'] or '').lower()
                or s in (r['invoice_no'] or '').lower() or s in (r['vendor_name'] or '').lower()]

    total_count = len(rows)
    total_cost = sum(r['amount'] or 0 for r in rows)
    open_jobs = sum(1 for r in rows if (r['status'] or 'Completed') == 'Open')
    completed_jobs = sum(1 for r in rows if (r['status'] or 'Completed') == 'Completed')
    breakdown_count = sum(1 for r in rows if 'breakdown' in (r['service_type'] or r['category'] or '').lower())
    avg_cost = round(total_cost / total_count, 0) if total_count else 0

    today = datetime.date.today()
    due_soon = 0
    for r in rows:
        if r['next_service_date']:
            try:
                nd = datetime.datetime.strptime(r['next_service_date'], '%Y-%m-%d').date()
                if today <= nd <= today + datetime.timedelta(days=7):
                    due_soon += 1
            except ValueError:
                pass

    ids = [r['id'] for r in rows]
    items_by_id = {}
    if ids:
        placeholders = ','.join('?' * len(ids))
        for it in conn.execute(f"SELECT * FROM maintenance_items WHERE maintenance_id IN ({placeholders})", ids).fetchall():
            items_by_id.setdefault(it['maintenance_id'], []).append(it)

    type_totals = {}
    for r in rows:
        t = r['service_type'] or r['category'] or 'Other'
        type_totals[t] = type_totals.get(t, 0) + 1
    type_breakdown = sorted(type_totals.items(), key=lambda x: -x[1])[:6]

    trend = []
    end_ref = datetime.datetime.strptime(date_to, '%Y-%m-%d').date() if date_to else today
    y_, m_ = end_ref.year, end_ref.month
    for i in range(5, -1, -1):
        mm = m_ - i
        yy = y_
        while mm <= 0:
            mm += 12
            yy -= 1
        mf, mt = _month_bounds(yy, mm)
        month_rows = [r for r in all_rows if _maintenance_classify(r['category'], r['service_type']) == 'Service'
                      and r['date'] and mf <= r['date'] <= mt]
        trend.append({'label': calendar.month_abbr[mm], 'cost': sum(r['amount'] or 0 for r in month_rows)})
    trend_max = max([t['cost'] for t in trend], default=1) or 1
    chart_w, chart_h = 400, 140
    n = len(trend)
    for i, t in enumerate(trend):
        t['x'] = round((i / (n - 1) * chart_w) if n > 1 else chart_w / 2, 1)
        t['y'] = round(chart_h - (t['cost'] / trend_max * (chart_h - 20)) - 10, 1)

    cost_by_vehicle = {}
    for r in rows:
        if r['vehicle_no']:
            cost_by_vehicle[r['vehicle_no']] = cost_by_vehicle.get(r['vehicle_no'], 0) + (r['amount'] or 0)
    top_costly = sorted(cost_by_vehicle.items(), key=lambda x: -x[1])[:5]

    page, per_page, total_pages = _paginate(request.args.get('page'), request.args.get('per_page'), total_count,
                                             per_page_options=(10, 25, 50, 100), default_per_page=10)
    page_rows = rows[(page - 1) * per_page: page * per_page]
    page_tokens = _page_tokens(page, total_pages)
    from urllib.parse import urlencode
    base_params = request.args.to_dict()
    base_params.pop('page', None)
    base_params.pop('per_page', None)
    base_qs = urlencode(base_params)

    vehicles, parties, vendors, combined_names = _get_autocomplete_lists()
    workshops = conn.execute("SELECT DISTINCT name FROM vendors ORDER BY name").fetchall()
    service_types = sorted(set((r['service_type'] or r['category'] or '') for r in all_rows
                                if _maintenance_classify(r['category'], r['service_type']) == 'Service' and (r['service_type'] or r['category'])))
    conn.close()

    return render_template('maintenance.html', tab='service',
        rows=page_rows, items_by_id=items_by_id, total_count=total_count, total_cost=total_cost,
        open_jobs=open_jobs, completed_jobs=completed_jobs, breakdown_count=breakdown_count, due_soon=due_soon, avg_cost=avg_cost,
        type_breakdown=type_breakdown, trend=trend, trend_max=trend_max, chart_w=chart_w, chart_h=chart_h, top_costly=top_costly,
        page=page, total_pages=total_pages, per_page=per_page, page_tokens=page_tokens, base_qs=base_qs,
        vehicles=vehicles, workshops=workshops, combined_names=combined_names, service_types=service_types,
        checklist_items=MAINTENANCE_CHECKLIST,
        f_vehicle=vehicle_f, f_workshop=workshop_f, f_service_type=stype_f, f_status=status_f,
        f_date_from=date_from, f_date_to=date_to, f_search=search_f, active='maintenance')

@app.route('/maintenance/add', methods=['GET', 'POST'])
def add_maintenance():
    conn = get_db()
    if request.method == 'POST':
        f = request.form
        def get_or_create_vehicle(vno):
            if not vno or not vno.strip():
                return None
            vno = vno.strip()
            row = conn.execute("SELECT id FROM vehicles WHERE vehicle_no = ? COLLATE NOCASE", (vno,)).fetchone()
            if row:
                return row[0]
            cur = conn.execute("INSERT INTO vehicles (vehicle_no) VALUES (?)", (vno,))
            return cur.lastrowid
        vehicle_id = get_or_create_vehicle(f.get('vehicle_no'))
        vendor_id = get_or_create_vendor(conn, f.get('vendor_name'))
        conn.execute("""INSERT INTO maintenance (date, vehicle_id, category, amount, paid_amount, vendor_id, notes)
                        VALUES (?,?,?,?,?,?,?)""",
                     (f.get('date'), vehicle_id, f.get('category'), float(f.get('amount') or 0),
                      float(f.get('paid_amount') or 0), vendor_id, f.get('notes')))
        conn.commit()
        conn.close()
        return redirect(url_for('maintenance_list'))
    conn.close()
    vehicles, parties, vendors, combined_names = _get_autocomplete_lists()
    return render_template('add_maintenance.html', vehicles=vehicles, vendors=vendors, combined_names=combined_names, active='maintenance')

def _save_service_items(conn, service_id, form):
    conn.execute("DELETE FROM maintenance_items WHERE maintenance_id=?", (service_id,))
    names = form.getlist('item_name')
    cats = form.getlist('item_category')
    qtys = form.getlist('item_qty')
    units = form.getlist('item_unit')
    rates = form.getlist('item_rate')
    for i, name in enumerate(names):
        name = (name or '').strip()
        if not name:
            continue
        qty = float(qtys[i]) if i < len(qtys) and qtys[i] else 0
        rate = float(rates[i]) if i < len(rates) and rates[i] else 0
        cat = cats[i] if i < len(cats) else ''
        unit = units[i] if i < len(units) else ''
        conn.execute("""INSERT INTO maintenance_items (maintenance_id, item_name, category, qty, unit, rate, amount)
                        VALUES (?,?,?,?,?,?,?)""", (service_id, name, cat, qty, unit, rate, qty * rate))

@app.route('/maintenance/service/add', methods=['POST'])
def add_service():
    conn = get_db()
    f = request.form
    def get_or_create_vehicle(vno):
        if not vno or not vno.strip():
            return None
        vno = vno.strip()
        row = conn.execute("SELECT id FROM vehicles WHERE vehicle_no = ? COLLATE NOCASE", (vno,)).fetchone()
        if row:
            return row[0]
        cur = conn.execute("INSERT INTO vehicles (vehicle_no) VALUES (?)", (vno,))
        return cur.lastrowid

    vehicle_id = get_or_create_vehicle(f.get('vehicle_no'))
    # Workshop/Vendor uses the exact same combined party+vendor lookup as every other vendor field
    # in the app, so a workshop typed here links up with that same organization's ledger elsewhere.
    vendor_id = get_or_create_vendor(conn, f.get('vendor_name'))
    amount = float(f.get('amount') or 0)
    paid_amount = float(f.get('paid_amount') or 0)
    checklist = ','.join(request.form.getlist('checklist'))
    cur = conn.execute("""INSERT INTO maintenance (date, vehicle_id, category, service_type, amount, paid_amount, vendor_id, notes,
                          km_reading, next_due_km, next_service_date, invoice_no, invoice_date, status, checklist_done)
                          VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (f.get('date'), vehicle_id, f.get('service_type'), f.get('service_type'), amount, paid_amount, vendor_id,
         f.get('notes') or None, float(f.get('km_reading') or 0) or None, float(f.get('next_due_km') or 0) or None,
         f.get('next_service_date') or None, f.get('invoice_no') or None, f.get('invoice_date') or None,
         f.get('status') or 'Completed', checklist or None))
    service_id = cur.lastrowid
    _save_service_items(conn, service_id, request.form)
    conn.commit()
    conn.close()
    return redirect(url_for('maintenance_list', tab='service'))

@app.route('/maintenance/service/edit/<int:m_id>', methods=['POST'])
def edit_service(m_id):
    conn = get_db()
    f = request.form
    def get_or_create_vehicle(vno):
        if not vno or not vno.strip():
            return None
        vno = vno.strip()
        row = conn.execute("SELECT id FROM vehicles WHERE vehicle_no = ? COLLATE NOCASE", (vno,)).fetchone()
        if row:
            return row[0]
        cur = conn.execute("INSERT INTO vehicles (vehicle_no) VALUES (?)", (vno,))
        return cur.lastrowid

    vehicle_id = get_or_create_vehicle(f.get('vehicle_no'))
    vendor_id = get_or_create_vendor(conn, f.get('vendor_name'))
    amount = float(f.get('amount') or 0)
    paid_amount = float(f.get('paid_amount') or 0)
    checklist = ','.join(request.form.getlist('checklist'))
    conn.execute("""UPDATE maintenance SET date=?, vehicle_id=?, category=?, service_type=?, amount=?, paid_amount=?, vendor_id=?,
                    notes=?, km_reading=?, next_due_km=?, next_service_date=?, invoice_no=?, invoice_date=?, status=?, checklist_done=?
                    WHERE id=?""",
        (f.get('date'), vehicle_id, f.get('service_type'), f.get('service_type'), amount, paid_amount, vendor_id,
         f.get('notes') or None, float(f.get('km_reading') or 0) or None, float(f.get('next_due_km') or 0) or None,
         f.get('next_service_date') or None, f.get('invoice_no') or None, f.get('invoice_date') or None,
         f.get('status') or 'Completed', checklist or None, m_id))
    _save_service_items(conn, m_id, request.form)
    conn.commit()
    conn.close()
    return redirect(url_for('maintenance_list', tab='service'))

@app.route('/maintenance/tyre/add', methods=['POST'])
def add_tyre():
    conn = get_db()
    f = request.form
    def get_or_create_vehicle(vno):
        if not vno or not vno.strip():
            return None
        vno = vno.strip()
        row = conn.execute("SELECT id FROM vehicles WHERE vehicle_no = ? COLLATE NOCASE", (vno,)).fetchone()
        if row:
            return row[0]
        cur = conn.execute("INSERT INTO vehicles (vehicle_no) VALUES (?)", (vno,))
        return cur.lastrowid

    vehicle_id = get_or_create_vehicle(f.get('vehicle_no'))

    stock_id = f.get('stock_id')
    if stock_id:
        # Reusing a tyre already bought into stock — update its existing purchase/ledger
        # row in place rather than billing the supplier a second time for the same tyre.
        stock = conn.execute("SELECT * FROM tyre_stock WHERE id=? AND status='In Stock'", (stock_id,)).fetchone()
        if stock:
            # The ledger date stays the purchase date — the cost was incurred then, not now —
            # matching the Tyre Stock table's own Install action.
            km_reading = float(f.get('km_reading') or 0) or None
            conn.execute("UPDATE maintenance SET vehicle_id=?, tyre_position=?, tyre_action=?, km_reading=? WHERE id=?",
                (vehicle_id, f.get('tyre_position') or None, f.get('tyre_action') or 'New Tyre Fitted',
                 km_reading, stock['maintenance_id']))
            conn.execute("""UPDATE tyre_stock SET status='Installed', installed_vehicle_id=?, installed_position=?,
                            installed_date=? WHERE id=?""", (vehicle_id, f.get('tyre_position') or None, f.get('date'), stock['id']))
            conn.commit()
            conn.close()
            return redirect(url_for('maintenance_list', tab='tyres'))

    # Same combined party+vendor lookup used everywhere else — a tyre supplier typed here
    # links up with that same organization's ledger, so the cost flows straight into it.
    vendor_id = get_or_create_vendor(conn, f.get('vendor_name'))
    amount = float(f.get('amount') or 0)
    paid_amount = float(f.get('paid_amount') or 0)
    conn.execute("""INSERT INTO maintenance (date, vehicle_id, category, amount, paid_amount, vendor_id, notes,
                    km_reading, invoice_no, invoice_date, tyre_action, tyre_id, tyre_brand, tyre_position, status)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (f.get('date'), vehicle_id, 'Tyres', amount, paid_amount, vendor_id, f.get('notes') or None,
         float(f.get('km_reading') or 0) or None, f.get('invoice_no') or None, f.get('invoice_date') or None,
         f.get('tyre_action') or None, f.get('tyre_id') or None, f.get('tyre_brand') or None,
         f.get('tyre_position') or None, 'Completed'))
    conn.commit()
    conn.close()
    return redirect(url_for('maintenance_list', tab='tyres'))

@app.route('/maintenance/tyre/edit/<int:m_id>', methods=['POST'])
def edit_tyre(m_id):
    conn = get_db()
    f = request.form
    def get_or_create_vehicle(vno):
        if not vno or not vno.strip():
            return None
        vno = vno.strip()
        row = conn.execute("SELECT id FROM vehicles WHERE vehicle_no = ? COLLATE NOCASE", (vno,)).fetchone()
        if row:
            return row[0]
        cur = conn.execute("INSERT INTO vehicles (vehicle_no) VALUES (?)", (vno,))
        return cur.lastrowid

    vehicle_id = get_or_create_vehicle(f.get('vehicle_no'))
    vendor_id = get_or_create_vendor(conn, f.get('vendor_name'))
    amount = float(f.get('amount') or 0)
    paid_amount = float(f.get('paid_amount') or 0)
    conn.execute("""UPDATE maintenance SET date=?, vehicle_id=?, amount=?, paid_amount=?, vendor_id=?, notes=?,
                    km_reading=?, invoice_no=?, invoice_date=?, tyre_action=?, tyre_id=?, tyre_brand=?, tyre_position=?
                    WHERE id=?""",
        (f.get('date'), vehicle_id, amount, paid_amount, vendor_id, f.get('notes') or None,
         float(f.get('km_reading') or 0) or None, f.get('invoice_no') or None, f.get('invoice_date') or None,
         f.get('tyre_action') or None, f.get('tyre_id') or None, f.get('tyre_brand') or None,
         f.get('tyre_position') or None, m_id))
    conn.commit()
    conn.close()
    return redirect(url_for('maintenance_list', tab='tyres'))

@app.route('/maintenance/tyre/stock/add', methods=['POST'])
def add_tyre_stock():
    conn = get_db()
    f = request.form
    # A stock purchase is its own maintenance/ledger entry (no vehicle yet) — installing it later
    # updates this same row instead of billing the supplier a second time for the same tyre.
    vendor_id = get_or_create_vendor(conn, f.get('vendor_name'))
    cost = float(f.get('purchase_cost') or 0)
    paid_amount = float(f.get('paid_amount') or 0)
    cur = conn.execute("""INSERT INTO maintenance (date, vehicle_id, category, amount, paid_amount, vendor_id, notes,
                          invoice_no, tyre_action, tyre_id, tyre_brand, status)
                          VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
        (f.get('purchase_date'), None, 'Tyres', cost, paid_amount, vendor_id, f.get('notes') or None,
         f.get('invoice_no') or None, 'Stock Purchase', f.get('tyre_id') or None, f.get('brand') or None, 'Completed'))
    maintenance_id = cur.lastrowid
    conn.execute("""INSERT INTO tyre_stock (maintenance_id, tyre_id, brand, tyre_type, purchase_date, purchase_cost,
                    vendor_id, invoice_no, status, notes) VALUES (?,?,?,?,?,?,?,?,?,?)""",
        (maintenance_id, f.get('tyre_id') or None, f.get('brand') or None, f.get('tyre_type') or 'New',
         f.get('purchase_date'), cost, vendor_id, f.get('invoice_no') or None, 'In Stock', f.get('notes') or None))
    conn.commit()
    conn.close()
    return redirect(url_for('maintenance_list', tab='tyres'))

@app.route('/maintenance/tyre/stock/<int:stock_id>/install', methods=['POST'])
def install_tyre_stock(stock_id):
    conn = get_db()
    f = request.form
    def get_or_create_vehicle(vno):
        if not vno or not vno.strip():
            return None
        vno = vno.strip()
        row = conn.execute("SELECT id FROM vehicles WHERE vehicle_no = ? COLLATE NOCASE", (vno,)).fetchone()
        if row:
            return row[0]
        cur = conn.execute("INSERT INTO vehicles (vehicle_no) VALUES (?)", (vno,))
        return cur.lastrowid

    stock = conn.execute("SELECT * FROM tyre_stock WHERE id=?", (stock_id,)).fetchone()
    if stock and stock['status'] == 'In Stock':
        vehicle_id = get_or_create_vehicle(f.get('vehicle_no'))
        install_date = f.get('install_date') or stock['purchase_date']
        km_reading = float(f.get('km_reading') or 0) or None
        conn.execute("UPDATE maintenance SET vehicle_id=?, tyre_position=?, tyre_action=?, km_reading=? WHERE id=?",
            (vehicle_id, f.get('position'), 'New Tyre Fitted', km_reading, stock['maintenance_id']))
        conn.execute("""UPDATE tyre_stock SET status='Installed', installed_vehicle_id=?, installed_position=?,
                        installed_date=? WHERE id=?""", (vehicle_id, f.get('position'), install_date, stock_id))
        conn.commit()
    conn.close()
    return redirect(url_for('maintenance_list', tab='tyres'))

@app.route('/maintenance/tyre/stock/<int:stock_id>/scrap', methods=['POST'])
def scrap_tyre_stock(stock_id):
    conn = get_db()
    conn.execute("UPDATE tyre_stock SET status='Scrapped' WHERE id=? AND status='In Stock'", (stock_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('maintenance_list', tab='tyres'))

@app.route('/maintenance/tyre/stock/<int:stock_id>/delete', methods=['POST'])
def delete_tyre_stock(stock_id):
    conn = get_db()
    stock = conn.execute("SELECT * FROM tyre_stock WHERE id=?", (stock_id,)).fetchone()
    if stock and stock['status'] == 'In Stock':
        conn.execute("DELETE FROM maintenance WHERE id=?", (stock['maintenance_id'],))
        conn.execute("DELETE FROM tyre_stock WHERE id=?", (stock_id,))
        conn.commit()
    conn.close()
    return redirect(url_for('maintenance_list', tab='tyres'))

@app.route('/maintenance/battery/add', methods=['POST'])
def add_battery():
    conn = get_db()
    f = request.form
    def get_or_create_vehicle(vno):
        if not vno or not vno.strip():
            return None
        vno = vno.strip()
        row = conn.execute("SELECT id FROM vehicles WHERE vehicle_no = ? COLLATE NOCASE", (vno,)).fetchone()
        if row:
            return row[0]
        cur = conn.execute("INSERT INTO vehicles (vehicle_no) VALUES (?)", (vno,))
        return cur.lastrowid

    mode = f.get('mode', 'install')
    vehicle_id = get_or_create_vehicle(f.get('vehicle_no')) if mode == 'install' else None
    # Same purchase-record-becomes-ledger-entry pattern as Tyres: the cost posts once, here,
    # whether the battery goes straight onto a vehicle or sits in stock first.
    vendor_id = get_or_create_vendor(conn, f.get('vendor_name'))
    price = float(f.get('purchase_price') or 0)
    battery_no = _next_battery_no(conn)
    health = float(f.get('health_pct') or 0) or None
    voltage = float(f.get('voltage') or 0) or None
    temp_c = float(f.get('temp_c') or 0) or None

    cur = conn.execute("""INSERT INTO maintenance (date, vehicle_id, category, amount, paid_amount, vendor_id, notes, invoice_no, status)
                          VALUES (?,?,?,?,?,?,?,?,?)""",
        (f.get('purchase_date'), vehicle_id, 'Battery', price, float(f.get('paid_amount') or 0), vendor_id,
         f.get('notes') or None, f.get('invoice_no') or None, 'Completed'))
    maintenance_id = cur.lastrowid

    install_date = f.get('install_date') if mode == 'install' else None
    cur2 = conn.execute("""INSERT INTO batteries (battery_no, brand, model, capacity_ah, battery_type, voltage_rating, serial_no,
                           vehicle_id, installed_location, install_date, purchase_date, purchase_price, vendor_id, invoice_no,
                           warranty_months, maintenance_id, health_pct, voltage, temp_c, last_checked_date, notes)
                           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (battery_no, f.get('brand') or None, f.get('model') or None, float(f.get('capacity_ah') or 0) or None,
         f.get('battery_type') or None, f.get('voltage_rating') or None, f.get('serial_no') or None,
         vehicle_id, f.get('installed_location') or None, install_date, f.get('purchase_date'), price, vendor_id,
         f.get('invoice_no') or None, int(f.get('warranty_months') or 0) or None, maintenance_id,
         health, voltage, temp_c, install_date or f.get('purchase_date'), f.get('notes') or None))
    battery_id = cur2.lastrowid

    conn.execute("INSERT INTO battery_checks (battery_id, date, event) VALUES (?,?,?)",
                 (battery_id, f.get('purchase_date'), 'Purchased'))
    if mode == 'install' and install_date:
        conn.execute("INSERT INTO battery_checks (battery_id, date, event, health_pct, voltage, temp_c) VALUES (?,?,?,?,?,?)",
                     (battery_id, install_date, f"Installed in {f.get('vehicle_no')}", health, voltage, temp_c))
    conn.commit()
    conn.close()
    return redirect(url_for('maintenance_list', tab='battery'))

@app.route('/maintenance/battery/<int:battery_id>/install', methods=['POST'])
def install_battery(battery_id):
    conn = get_db()
    f = request.form
    def get_or_create_vehicle(vno):
        if not vno or not vno.strip():
            return None
        vno = vno.strip()
        row = conn.execute("SELECT id FROM vehicles WHERE vehicle_no = ? COLLATE NOCASE", (vno,)).fetchone()
        if row:
            return row[0]
        cur = conn.execute("INSERT INTO vehicles (vehicle_no) VALUES (?)", (vno,))
        return cur.lastrowid

    battery = conn.execute("SELECT * FROM batteries WHERE id=?", (battery_id,)).fetchone()
    if battery and not battery['vehicle_id']:
        vehicle_id = get_or_create_vehicle(f.get('vehicle_no'))
        install_date = f.get('install_date')
        conn.execute("UPDATE batteries SET vehicle_id=?, installed_location=?, install_date=?, last_checked_date=? WHERE id=?",
                     (vehicle_id, f.get('installed_location') or None, install_date, install_date, battery_id))
        if battery['maintenance_id']:
            conn.execute("UPDATE maintenance SET vehicle_id=? WHERE id=?", (vehicle_id, battery['maintenance_id']))
        conn.execute("INSERT INTO battery_checks (battery_id, date, event) VALUES (?,?,?)",
                     (battery_id, install_date, f"Installed in {f.get('vehicle_no')}"))
        conn.commit()
    conn.close()
    return redirect(url_for('maintenance_list', tab='battery'))

@app.route('/maintenance/battery/<int:battery_id>/check', methods=['POST'])
def check_battery(battery_id):
    conn = get_db()
    f = request.form
    health = float(f.get('health_pct') or 0) or None
    voltage = float(f.get('voltage') or 0) or None
    temp_c = float(f.get('temp_c') or 0) or None
    date = f.get('date') or datetime.date.today().isoformat()
    conn.execute("UPDATE batteries SET health_pct=?, voltage=?, temp_c=?, last_checked_date=? WHERE id=?",
                 (health, voltage, temp_c, date, battery_id))
    conn.execute("INSERT INTO battery_checks (battery_id, date, event, health_pct, voltage, temp_c, remarks) VALUES (?,?,?,?,?,?,?)",
                 (battery_id, date, 'Health Check', health, voltage, temp_c, f.get('remarks') or None))
    conn.commit()
    conn.close()
    return redirect(url_for('maintenance_list', tab='battery'))

@app.route('/maintenance/battery/<int:battery_id>/dead', methods=['POST'])
def mark_battery_dead(battery_id):
    conn = get_db()
    today = datetime.date.today().isoformat()
    conn.execute("UPDATE batteries SET status_override='Dead' WHERE id=?", (battery_id,))
    conn.execute("INSERT INTO battery_checks (battery_id, date, event) VALUES (?,?,?)", (battery_id, today, 'Marked Dead'))
    conn.commit()
    conn.close()
    return redirect(url_for('maintenance_list', tab='battery'))

@app.route('/maintenance/battery/<int:battery_id>/delete', methods=['POST'])
def delete_battery(battery_id):
    conn = get_db()
    battery = conn.execute("SELECT * FROM batteries WHERE id=?", (battery_id,)).fetchone()
    if battery and not battery['vehicle_id']:
        if battery['maintenance_id']:
            conn.execute("DELETE FROM maintenance WHERE id=?", (battery['maintenance_id'],))
        conn.execute("DELETE FROM battery_checks WHERE battery_id=?", (battery_id,))
        conn.execute("DELETE FROM batteries WHERE id=?", (battery_id,))
        conn.commit()
    conn.close()
    return redirect(url_for('maintenance_list', tab='battery'))

@app.route('/maintenance/urea/add', methods=['POST'])
def add_urea():
    """One endpoint for all 3 Add Urea modes (mirrors the Tyre/Battery unified-modal pattern):
    'stock_in' (pure stock purchase — the real cost event), 'stock_out' (consume already-paid-for
    stock for a vehicle — no new cost), 'direct' (buy and use in the same moment, bypassing
    tracked stock — its own real cost event, like Tyre's 'New Tyre' mode)."""
    conn = get_db()
    f = request.form
    def get_or_create_vehicle(vno):
        if not vno or not vno.strip():
            return None
        vno = vno.strip()
        row = conn.execute("SELECT id FROM vehicles WHERE vehicle_no = ? COLLATE NOCASE", (vno,)).fetchone()
        if row:
            return row[0]
        cur = conn.execute("INSERT INTO vehicles (vehicle_no) VALUES (?)", (vno,))
        return cur.lastrowid

    mode = f.get('mode')
    date = f.get('date') or datetime.date.today().isoformat()
    location = f.get('location') or None
    notes = f.get('notes') or None
    qty = float(f.get('quantity_l') or 0)
    unit_price = float(f.get('unit_price') or 0)
    total_value = round(qty * unit_price, 2)
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    if mode == 'stock_in':
        supplier_id = get_or_create_vendor(conn, f.get('supplier_name'))
        cur_m = conn.execute("""INSERT INTO maintenance (date, vehicle_id, category, amount, paid_amount,
                                vendor_id, notes, invoice_no) VALUES (?,?,?,?,?,?,?,?)""",
                              (date, None, 'Urea', total_value, float(f.get('paid_amount') or 0),
                               supplier_id, notes, f.get('invoice_no')))
        conn.execute("""INSERT INTO urea_transactions (date, txn_type, source, batch_no, supplier_id,
                        invoice_no, quantity_l, unit_price, total_value, location, notes, maintenance_id, created_at)
                        VALUES (?, 'stock_in', 'stock', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                     (date, f.get('batch_no'), supplier_id, f.get('invoice_no'), qty, unit_price,
                      total_value, location, notes, cur_m.lastrowid, now))
    elif mode == 'stock_out':
        vehicle_id = get_or_create_vehicle(f.get('vehicle_no'))
        conn.execute("""INSERT INTO urea_transactions (date, txn_type, source, batch_no, vehicle_id,
                        quantity_l, unit_price, total_value, location, odometer_km, notes, created_at)
                        VALUES (?, 'stock_out', 'stock', ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                     (date, f.get('batch_no'), vehicle_id, qty, unit_price, total_value, location,
                      float(f.get('odometer_km')) if f.get('odometer_km') else None, notes, now))
    elif mode == 'direct':
        vehicle_id = get_or_create_vehicle(f.get('vehicle_no'))
        supplier_id = get_or_create_vendor(conn, f.get('supplier_name')) if f.get('supplier_name') else None
        cur_m = conn.execute("""INSERT INTO maintenance (date, vehicle_id, category, amount, paid_amount,
                                vendor_id, notes, invoice_no) VALUES (?,?,?,?,?,?,?,?)""",
                              (date, vehicle_id, 'Urea', total_value, float(f.get('paid_amount') or 0),
                               supplier_id, notes, f.get('invoice_no')))
        conn.execute("""INSERT INTO urea_transactions (date, txn_type, source, supplier_id, invoice_no,
                        vehicle_id, quantity_l, unit_price, total_value, location, odometer_km, notes,
                        maintenance_id, created_at)
                        VALUES (?, 'stock_out', 'direct', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                     (date, supplier_id, f.get('invoice_no'), vehicle_id, qty, unit_price, total_value,
                      location, float(f.get('odometer_km')) if f.get('odometer_km') else None, notes,
                      cur_m.lastrowid, now))
    conn.commit()
    conn.close()
    return redirect(url_for('maintenance_list', tab='urea'))

@app.route('/maintenance/urea/<int:txn_id>/edit', methods=['GET', 'POST'])
def edit_urea(txn_id):
    conn = get_db()
    if request.method == 'POST':
        f = request.form
        qty = float(f.get('quantity_l') or 0)
        unit_price = float(f.get('unit_price') or 0)
        total_value = round(qty * unit_price, 2)
        conn.execute("""UPDATE urea_transactions SET date=?, batch_no=?, quantity_l=?, unit_price=?,
                        total_value=?, location=?, odometer_km=?, notes=? WHERE id=?""",
                     (f.get('date'), f.get('batch_no'), qty, unit_price, total_value, f.get('location'),
                      float(f.get('odometer_km')) if f.get('odometer_km') else None, f.get('notes'), txn_id))
        # Keep the linked cost entry (if any) in sync so Overview's Urea card never disagrees
        # with what this ledger shows.
        txn = conn.execute("SELECT maintenance_id FROM urea_transactions WHERE id=?", (txn_id,)).fetchone()
        if txn and txn['maintenance_id']:
            conn.execute("UPDATE maintenance SET date=?, amount=?, notes=? WHERE id=?",
                         (f.get('date'), total_value, f.get('notes'), txn['maintenance_id']))
        conn.commit()
        conn.close()
        return redirect(url_for('maintenance_list', tab='urea'))
    conn.close()
    return redirect(url_for('maintenance_list', tab='urea'))

@app.route('/maintenance/urea/<int:txn_id>/delete', methods=['POST'])
def delete_urea(txn_id):
    conn = get_db()
    txn = conn.execute("SELECT maintenance_id FROM urea_transactions WHERE id=?", (txn_id,)).fetchone()
    if txn and txn['maintenance_id']:
        conn.execute("DELETE FROM maintenance WHERE id=?", (txn['maintenance_id'],))
    conn.execute("DELETE FROM urea_transactions WHERE id=?", (txn_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('maintenance_list', tab='urea'))

@app.route('/maintenance/toll/add', methods=['POST'])
def add_toll():
    conn = get_db()
    f = request.form
    row = conn.execute("SELECT id FROM vehicles WHERE vehicle_no = ? COLLATE NOCASE AND type IN ('Line','Local')",
                        ((f.get('vehicle_no') or '').strip(),)).fetchone()
    vehicle_id = row['id'] if row else None
    if not vehicle_id:
        # Toll Management is own-fleet only — unlike Urea/Battery, this never silently creates a
        # new vehicle record from a typo'd or hired vehicle number.
        conn.close()
        return redirect(url_for('maintenance_list', tab='toll'))

    trip_id = None
    lr = (f.get('trip_lr') or '').strip()
    if lr:
        lr_key = lr.split('|')[0].strip()  # datalist value is "LR_NUMBER | vehicle_no | date"
        trow = conn.execute("SELECT id FROM trips WHERE lr_number = ? COLLATE NOCASE ORDER BY date DESC LIMIT 1", (lr_key,)).fetchone()
        trip_id = trow['id'] if trow else None

    date = f.get('date') or datetime.date.today().isoformat()
    time_ = f.get('time') or None
    amount = float(f.get('amount') or 0)
    source = f.get('source') if f.get('source') in ('fastag', 'manual') else 'manual'
    payment_mode = f.get('payment_mode') or None
    status = 'synced' if source == 'fastag' else 'pending'
    notes = f.get('notes') or None
    receipt_path = _save_toll_receipt(request.files.get('receipt'))
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    cur_m = conn.execute("""INSERT INTO maintenance (date, vehicle_id, category, amount, paid_amount, notes)
                            VALUES (?,?,?,?,?,?)""", (date, vehicle_id, 'Toll', amount, amount, notes))
    conn.execute("""INSERT INTO toll_entries (date, time, vehicle_id, trip_id, toll_plaza, highway, state,
                    amount, source, payment_mode, status, receipt_path, notes, maintenance_id, created_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                 (date, time_, vehicle_id, trip_id, f.get('toll_plaza'), f.get('highway') or None,
                  f.get('state') or None, amount, source, payment_mode, status, receipt_path, notes,
                  cur_m.lastrowid, now))
    conn.commit()
    conn.close()
    return redirect(url_for('maintenance_list', tab='toll'))

@app.route('/maintenance/toll/<int:entry_id>/delete', methods=['POST'])
def delete_toll(entry_id):
    conn = get_db()
    e = conn.execute("SELECT maintenance_id FROM toll_entries WHERE id=?", (entry_id,)).fetchone()
    if e and e['maintenance_id']:
        conn.execute("DELETE FROM maintenance WHERE id=?", (e['maintenance_id'],))
    conn.execute("DELETE FROM toll_entries WHERE id=?", (entry_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('maintenance_list', tab='toll'))

@app.route('/maintenance/toll/<int:entry_id>/approve', methods=['POST'])
def approve_toll(entry_id):
    conn = get_db()
    conn.execute("UPDATE toll_entries SET status='approved' WHERE id=? AND status='pending'", (entry_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('maintenance_list', tab='toll'))

@app.route('/maintenance/toll/<int:entry_id>/reject', methods=['POST'])
def reject_toll(entry_id):
    conn = get_db()
    conn.execute("UPDATE toll_entries SET status='rejected' WHERE id=? AND status='pending'", (entry_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('maintenance_list', tab='toll'))

@app.route('/maintenance/toll/sync-fastag', methods=['POST'])
def sync_fastag_toll():
    """Mock FASTag sync — no live FASTag/bank API is wired up yet, so this fills in a small batch
    of plausible-looking new transactions the same deterministic-mock way the Compliance providers
    do (providers/*.py): reseeded once per hour, so clicking again inside the same hour is a no-op
    instead of piling up duplicate rows — an earlier mock provider polluted the live DB twice before
    this guard existed (see compliance_service.py). Swap this for a real BOSS/FASTag statement feed
    (or just use the Excel import above, which already reads real BOSS exports) once one exists."""
    import random
    conn = get_db()
    vehicles = conn.execute("SELECT id, vehicle_no FROM vehicles WHERE type IN ('Line','Local') ORDER BY vehicle_no").fetchall()
    if not vehicles:
        conn.close()
        return redirect(url_for('maintenance_list', tab='toll'))
    now = datetime.datetime.now()
    # Anchored to the top of the hour (not the exact click time) so every field the ref_no is
    # built from — including txn_time below — is byte-identical across repeated clicks inside the
    # same hour. Deriving txn_time from wall-clock `now` instead of this anchor was the bug: it
    # made ref_no change every second, so the dedup check below never matched and every click kept
    # inserting a fresh-looking "new" batch — this already happened live, see the cleanup note in
    # the route's docstring history / conversation log.
    hour_anchor = now.replace(minute=0, second=0, microsecond=0)
    rnd = random.Random(int(now.strftime('%Y%m%d%H')))
    inserted = 0
    for _ in range(rnd.randint(2, 5)):
        v = rnd.choice(vehicles)
        plaza, hwy, state = rnd.choice(FASTAG_MOCK_PLAZAS)
        amount = rnd.choice([65, 85, 110, 145, 175, 210, 280, 350])
        txn_time = hour_anchor - datetime.timedelta(minutes=rnd.randint(5, 600))
        ref_no = f"FT{txn_time.strftime('%y%m%d%H%M%S')}{v['id']}"
        if conn.execute("SELECT 1 FROM toll_entries WHERE reference_no=?", (ref_no,)).fetchone():
            continue
        cur_m = conn.execute("""INSERT INTO maintenance (date, vehicle_id, category, amount, paid_amount)
                                VALUES (?,?,?,?,?)""", (txn_time.strftime('%Y-%m-%d'), v['id'], 'Toll', amount, amount))
        conn.execute("""INSERT INTO toll_entries (date, time, vehicle_id, toll_plaza, highway, state, amount,
                        source, payment_mode, status, reference_no, maintenance_id, created_at)
                        VALUES (?,?,?,?,?,?,?, 'fastag', 'FASTag Wallet', 'synced', ?, ?, ?)""",
                     (txn_time.strftime('%Y-%m-%d'), txn_time.strftime('%H:%M'), v['id'], plaza, hwy, state,
                      amount, ref_no, cur_m.lastrowid, now.strftime('%Y-%m-%d %H:%M:%S')))
        inserted += 1
    conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES ('toll_fastag_last_sync', ?)",
                 (now.strftime('%Y-%m-%d %H:%M:%S'),))
    conn.commit()
    conn.close()
    return redirect(url_for('maintenance_list', tab='toll', synced=inserted))

@app.route('/maintenance/toll/template')
def toll_excel_template():
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    import io
    wb = Workbook()
    ws = wb.active
    ws.title = 'Toll Entries'
    headers = ['Vehicle Number', 'Trip Number', 'Date', 'Time', 'Toll Plaza', 'Highway', 'State', 'Amount',
               'Source', 'Payment Mode', 'Reference Number', 'Notes']
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='1A1A1A')
    ws.append(['OD14AD0117', 'TRIP-1524', '2026-08-07', '10:35', 'Chandikhole Toll Plaza', 'NH-16', 'Odisha',
                425, 'FASTag', '', 'FT260807103512', ''])
    ws.append(['OD14AE1122', '', '2026-08-06', '23:05', 'Durg Expressway Toll', 'NH-53', 'Chhattisgarh',
                560, 'Manual', 'Cash', '', ''])
    widths = [16, 14, 12, 10, 26, 10, 14, 10, 10, 14, 18, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='Toll_Entry_Template.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# Transaction-type keywords that mean "money moved but a vehicle did NOT cross a toll plaza" —
# wallet top-ups, internal transfers between accounts, recoveries/reversals of an earlier
# misdirected transfer, refunds, manual balance adjustments, and the statement's own running-total
# rows. None of these are a toll charge and must never be counted as one.
def _parse_toll_excel(file_storage, conn):
    """Auto-detects and parses either of two supported layouts:
      1. A BlackBuck BOSS account statement export (Transaction Time / Nature (C/D) / Amount /
         From/To / Description / Truck Number / TransactionId / Opening Balance / Closing Balance).
         BlackBuck's Top-Up Plan makes the SAME toll amount show up three times in this ledger —
         once as the real charge, and twice more as bookkeeping around it:
           (a) Debit, From/To=Wallet, Description contains "FasTag Recharge" — the vehicle's
               actual toll cost. This is the ONLY row type counted as toll spend.
           (b) Credit, From/To=BlackBuck, Description contains "BlackBuck transfer" — a temporary
               advance funding (a), not a second toll. Always ignored.
           (c) Debit, From/To=BOSS Account, Description contains "Recovery of Amount transferred"
               — BOSS recovering the advance from (b) back out. Also not a toll. Always ignored.
         Also ignored: UPI Top-Up, Wallet Transfer, Refund, Adjustment, Opening/Closing Balance
         rows, and any row with no Truck Number — none of those are a vehicle's toll cost either.
      2. The plain manual template this page's 'Download Template' link provides (Vehicle Number /
         Trip Number / Date / Time / Toll Plaza / Highway / State / Amount / Source / Payment Mode /
         Reference Number / Notes).
    Returns (fmt, parsed_rows, skipped_count, skipped_amount) — parsed_rows only contains rows that
    passed the toll/non-toll classification; everything else (BlackBuck advances, BOSS recoveries,
    top-ups, transfers, refunds, balance rows) is excluded outright, never entering Vehicle Toll
    Cost / Monthly Toll Reports / Fleet Analytics, matching 'never count these as toll expense'."""
    import openpyxl
    wb = openpyxl.load_workbook(file_storage, data_only=True)
    ws = wb.active

    own_vehicles = {v['vehicle_no'].upper(): v['id'] for v in
                     conn.execute("SELECT id, vehicle_no FROM vehicles WHERE type IN ('Line','Local')").fetchall()}
    # Keyed (not just a set) so a re-import can tell whether a duplicate TransactionId's row
    # actually changed since it was saved — a re-upload with a corrected amount/date/etc. should
    # offer to override the existing record, not just get silently skipped as "already imported".
    existing_by_ref = {r['reference_no']: dict(r) for r in conn.execute(
        "SELECT id, reference_no, vehicle_id, date, time, toll_plaza, amount FROM toll_entries WHERE reference_no IS NOT NULL AND reference_no != ''").fetchall()}
    existing_keys = {(r['vehicle_id'], r['date'], r['toll_plaza'], round(r['amount'] or 0, 2)) for r in
                      conn.execute("SELECT vehicle_id, date, toll_plaza, amount FROM toll_entries").fetchall()}
    # Rows already seen earlier IN THIS SAME FILE — grows as we classify, so a repeated
    # TransactionId within one upload is caught too, not just duplicates of already-saved data.
    seen_refs_in_batch = set()

    def _dup_and_changed(ref_no, vehicle_id, date_str, time_str, plaza, amount):
        """Returns (is_dup, changed, existing_id) for a row's reference number."""
        existing = existing_by_ref.get(ref_no) if ref_no else None
        if existing:
            changed = (existing['vehicle_id'] != vehicle_id or existing['date'] != date_str or
                       existing['time'] != time_str or existing['toll_plaza'] != plaza or
                       round(existing['amount'] or 0, 2) != round(amount, 2))
            return True, changed, existing['id']
        if ref_no and ref_no in seen_refs_in_batch:
            return True, False, None  # dup within this same file — nothing saved to diff against
        return False, False, None

    boss_header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=True), 1):
        cells = [str(c).strip().lower() if c is not None else '' for c in row]
        if 'transaction time' in cells and any('nature' in c for c in cells):
            boss_header_row = i
            break

    parsed = []
    skipped_count = 0
    skipped_amount = 0.0

    if boss_header_row:
        fmt = 'boss'
        for row in ws.iter_rows(min_row=boss_header_row + 1, values_only=True):
            if row is None or row[0] is None:
                continue
            txn_time_raw, nature, amount_raw, from_to, desc, truck_no = row[0], row[1], row[2], row[3], row[4], row[5]
            txn_id = row[6] if len(row) > 6 else None
            nature = str(nature).strip().lower() if nature else ''
            desc = str(desc or '').strip()
            desc_l = desc.lower()
            from_to_l = str(from_to or '').strip().lower()
            truck_no = str(truck_no or '').strip().upper()
            try:
                amount = float(amount_raw or 0)
            except (TypeError, ValueError):
                amount = 0

            # The ONE allow-listed shape of a real toll charge — every other combination
            # (BlackBuck advance credits, BOSS recovery debits, top-ups, transfers, refunds,
            # adjustments, balance rows, anything with no truck number) is excluded, not just
            # hidden from the total.
            is_toll_charge = (nature == 'debit' and from_to_l == 'wallet' and 'fastag recharge' in desc_l
                               and bool(truck_no) and amount > 0)
            if not is_toll_charge:
                skipped_count += 1
                if nature == 'debit':
                    skipped_amount += amount  # Recovery-type debits still counted as excluded spend
                continue

            try:
                dt = datetime.datetime.strptime(str(txn_time_raw).strip(), '%d %b %y %I:%M %p')
                date_str, time_str = dt.strftime('%Y-%m-%d'), dt.strftime('%H:%M')
            except (ValueError, TypeError):
                date_str, time_str = '', ''
            vehicle_id = own_vehicles.get(truck_no)
            ref_no = str(txn_id or '').strip()
            toll_plaza = desc or 'FASTag Recharge'

            entry = {'vehicle_no': truck_no, 'vehicle_id': vehicle_id, 'trip_no': '', 'date': date_str,
                     'time': time_str, 'toll_plaza': toll_plaza, 'highway': '', 'state': '',
                     'amount': amount, 'source': 'fastag', 'payment_mode': 'FASTag Wallet',
                     'reference_no': ref_no, 'notes': desc}
            errs = []
            if not vehicle_id: errs.append('Vehicle not in Fleet ERP')
            if not date_str: errs.append('Unrecognised transaction time')
            entry['errors'] = errs
            # Duplicate key is TransactionId alone, per spec — if it already exists (in the DB, or
            # earlier in this same file), skip it rather than ever double-counting a toll. If the
            # existing saved row's own fields differ from this one, flag it as 'changed' so the
            # import step can offer to override instead of just leaving the stale data in place.
            is_dup, changed, existing_id = _dup_and_changed(ref_no, vehicle_id, date_str, time_str, toll_plaza, amount)
            entry['is_dup'] = is_dup
            entry['changed'] = changed
            entry['existing_id'] = existing_id
            entry['valid'] = (len(errs) == 0)
            if entry['valid'] and not entry['is_dup'] and ref_no:
                seen_refs_in_batch.add(ref_no)
            parsed.append(entry)
    else:
        fmt = 'template'
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
        header = [str(c).strip().lower() if c is not None else '' for c in header_row]
        colmap = {h: i for i, h in enumerate(header)}
        def get(row, key):
            idx = colmap.get(key)
            return row[idx] if idx is not None and idx < len(row) else None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or all(c in (None, '') for c in row):
                continue
            vno = str(get(row, 'vehicle number') or '').strip()
            trip_no = str(get(row, 'trip number') or '').strip()
            date_val = get(row, 'date')
            date_str = date_val.strftime('%Y-%m-%d') if hasattr(date_val, 'strftime') else str(date_val or '').strip()
            time_val = get(row, 'time')
            time_str = time_val.strftime('%H:%M') if hasattr(time_val, 'strftime') else str(time_val or '').strip()
            plaza = str(get(row, 'toll plaza') or '').strip()
            highway = str(get(row, 'highway') or '').strip()
            state = str(get(row, 'state') or '').strip()
            try:
                amount = float(get(row, 'amount') or 0)
            except (TypeError, ValueError):
                amount = 0
            source_raw = str(get(row, 'source') or 'Manual').strip()
            source = 'fastag' if source_raw.lower().startswith('fast') else 'manual'
            payment_mode = str(get(row, 'payment mode') or '').strip()
            ref_no = str(get(row, 'reference number') or '').strip()
            notes = str(get(row, 'notes') or '').strip()

            # Rule: zero/blank amount is ignored outright, same as the BOSS-format branch.
            if amount <= 0:
                skipped_count += 1
                continue

            vehicle_id = own_vehicles.get(vno.upper())
            entry = {'vehicle_no': vno, 'vehicle_id': vehicle_id, 'trip_no': trip_no, 'date': date_str,
                     'time': time_str, 'toll_plaza': plaza, 'highway': highway, 'state': state, 'amount': amount,
                     'source': source, 'payment_mode': payment_mode, 'reference_no': ref_no, 'notes': notes}
            errs = []
            if not vno: errs.append('Missing vehicle number')
            elif not vehicle_id: errs.append('Vehicle not in Fleet ERP')
            if not date_str: errs.append('Missing date')
            if not plaza: errs.append('Missing toll plaza (Warning)')  # non-blocking, see below
            blocking_errs = [e for e in errs if not e.endswith('(Warning)')]
            legacy_key = (vehicle_id, date_str, plaza, round(amount, 2))
            is_dup, changed, existing_id = _dup_and_changed(ref_no, vehicle_id, date_str, time_str, plaza, amount)
            if not is_dup and vehicle_id and date_str and plaza and legacy_key in existing_keys:
                is_dup = True  # no reference number to compare — same old best-effort key match
            entry['errors'] = errs
            entry['is_dup'] = is_dup
            entry['changed'] = changed
            entry['existing_id'] = existing_id
            entry['valid'] = (len(blocking_errs) == 0)
            if entry['valid'] and not entry['is_dup'] and ref_no:
                seen_refs_in_batch.add(ref_no)
            parsed.append(entry)

    return fmt, parsed, skipped_count, skipped_amount

@app.route('/maintenance/toll/excel/preview', methods=['POST'])
def toll_excel_preview():
    import uuid
    conn = get_db()
    file_storage = request.files.get('excel_file')
    if not file_storage or not file_storage.filename:
        conn.close()
        return redirect(url_for('maintenance_list', tab='toll'))
    try:
        fmt, parsed, credit_excl_count, credit_excl_amount = _parse_toll_excel(file_storage, conn)
    except Exception:
        ctx = _toll_tab_base_context(conn)
        conn.close()
        return render_template('maintenance.html', tab='toll', active='maintenance',
            excel_error="Couldn't read that file — make sure it's a .xlsx export from BOSS FASTag, or the downloaded template.",
            **ctx)

    token = uuid.uuid4().hex
    _TOLL_IMPORT_STASH[token] = {'rows': parsed, 'ts': datetime.datetime.now()}
    stale = [k for k, v in _TOLL_IMPORT_STASH.items() if (datetime.datetime.now() - v['ts']).total_seconds() > 1800]
    for k in stale:
        _TOLL_IMPORT_STASH.pop(k, None)

    valid_count = sum(1 for r in parsed if r['valid'] and not r['is_dup'])
    dup_count = sum(1 for r in parsed if r['is_dup'])
    changed_count = sum(1 for r in parsed if r['is_dup'] and r.get('changed'))
    invalid_count = sum(1 for r in parsed if not r['valid'])
    valid_amount = sum(r['amount'] for r in parsed if r['valid'] and not r['is_dup'])

    ctx = _toll_tab_base_context(conn)
    conn.close()
    return render_template('maintenance.html', tab='toll', active='maintenance',
        excel_token=token, excel_format=fmt, excel_rows=parsed[:10], excel_row_count=len(parsed),
        excel_valid_count=valid_count, excel_dup_count=dup_count, excel_changed_count=changed_count,
        excel_invalid_count=invalid_count, excel_valid_amount=valid_amount, excel_credit_excl_count=credit_excl_count,
        excel_credit_excl_amount=credit_excl_amount, **ctx)

@app.route('/maintenance/toll/excel/import', methods=['POST'])
def toll_excel_import():
    token = request.form.get('token')
    auto_link_trips = request.form.get('auto_link_trips') == 'on'
    auto_approve = request.form.get('auto_approve') == 'on'
    # A duplicate TransactionId is NEVER re-inserted as a second row — that would double-count a
    # toll (or a recharge) that already exists. The only choice a re-import offers is whether a
    # duplicate whose own fields actually changed since it was saved (amount corrected, wrong
    # vehicle fixed, etc.) should overwrite the existing record, or be left alone.
    override_changed = request.form.get('override_changed') == 'on'
    stash = _TOLL_IMPORT_STASH.pop(token, None)
    conn = get_db()
    imported = 0
    updated = 0
    if stash:
        now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        for row in stash['rows']:
            if not row['valid']:
                continue
            trip_id = None
            if auto_link_trips and row.get('trip_no'):
                trow = conn.execute("SELECT id FROM trips WHERE lr_number=? COLLATE NOCASE ORDER BY date DESC LIMIT 1",
                                     (row['trip_no'],)).fetchone()
                trip_id = trow['id'] if trow else None

            if row['is_dup']:
                if row.get('changed') and override_changed and row.get('existing_id'):
                    existing = conn.execute("SELECT maintenance_id FROM toll_entries WHERE id=?",
                                             (row['existing_id'],)).fetchone()
                    conn.execute("""UPDATE toll_entries SET date=?, time=?, vehicle_id=?, trip_id=?, toll_plaza=?,
                                    highway=?, state=?, amount=?, payment_mode=?, notes=? WHERE id=?""",
                                 (row['date'], row['time'] or None, row['vehicle_id'], trip_id, row['toll_plaza'],
                                  row['highway'] or None, row['state'] or None, row['amount'],
                                  row['payment_mode'] or None, row['notes'] or None, row['existing_id']))
                    if existing and existing['maintenance_id']:
                        conn.execute("""UPDATE maintenance SET date=?, vehicle_id=?, amount=?, paid_amount=?, notes=?
                                        WHERE id=?""",
                                     (row['date'], row['vehicle_id'], row['amount'], row['amount'],
                                      row['notes'] or None, existing['maintenance_id']))
                    updated += 1
                continue  # unchanged duplicate, or changed-but-override-not-requested: leave as-is

            status = 'synced' if row['source'] == 'fastag' else ('approved' if auto_approve else 'pending')
            cur_m = conn.execute("""INSERT INTO maintenance (date, vehicle_id, category, amount, paid_amount, notes)
                                    VALUES (?,?,?,?,?,?)""",
                                  (row['date'], row['vehicle_id'], 'Toll', row['amount'], row['amount'], row['notes'] or None))
            conn.execute("""INSERT INTO toll_entries (date, time, vehicle_id, trip_id, toll_plaza, highway, state,
                            amount, source, payment_mode, status, reference_no, notes, maintenance_id, created_at)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                         (row['date'], row['time'] or None, row['vehicle_id'], trip_id, row['toll_plaza'],
                          row['highway'] or None, row['state'] or None, row['amount'], row['source'],
                          row['payment_mode'] or None, status, row['reference_no'] or None, row['notes'] or None,
                          cur_m.lastrowid, now))
            imported += 1
        conn.commit()
    conn.close()
    return redirect(url_for('maintenance_list', tab='toll', imported=imported, updated=updated))

@app.route('/maintenance/insurance/add', methods=['POST'])
def add_insurance():
    conn = get_db()
    f = request.form
    def get_or_create_vehicle(vno):
        if not vno or not vno.strip():
            return None
        vno = vno.strip()
        row = conn.execute("SELECT id FROM vehicles WHERE vehicle_no = ? COLLATE NOCASE", (vno,)).fetchone()
        if row:
            return row[0]
        cur = conn.execute("INSERT INTO vehicles (vehicle_no) VALUES (?)", (vno,))
        return cur.lastrowid

    vehicle_id = get_or_create_vehicle(f.get('vehicle_no'))
    # Insurers are companies we pay premiums to, same as any other vendor — reusing the vendor
    # table means the premium lands straight in that insurer's ledger, same as every other cost.
    insurer_id = get_or_create_vendor(conn, f.get('insurer_name'))
    premium = float(f.get('premium_amount') or 0)

    cur = conn.execute("""INSERT INTO maintenance (date, vehicle_id, category, amount, paid_amount, vendor_id, notes, status)
                          VALUES (?,?,?,?,?,?,?,?)""",
        (f.get('start_date'), vehicle_id, 'Insurance', premium, premium, insurer_id, f.get('notes') or None, 'Completed'))
    maintenance_id = cur.lastrowid

    policy_doc = _save_insurance_doc(request.files.get('policy_doc'), 'policy')
    invoice_doc = _save_insurance_doc(request.files.get('invoice_doc'), 'invoice')
    rc_doc = _save_insurance_doc(request.files.get('rc_doc'), 'rc')
    status_override = 'Cancelled' if f.get('policy_status') == 'Cancelled' else None

    conn.execute("""INSERT INTO insurance_policies (vehicle_id, insurance_type, insurer_id, policy_number, start_date,
                    expiry_date, premium_amount, idv, ncb_pct, gst_included, agent_name, agent_contact, agent_email,
                    reminder_days, notes, status_override, policy_doc_path, invoice_doc_path, rc_doc_path, maintenance_id)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (vehicle_id, f.get('insurance_type'), insurer_id, f.get('policy_number') or None, f.get('start_date'),
         f.get('expiry_date'), premium, float(f.get('idv') or 0) or None, float(f.get('ncb_pct') or 0) or None,
         f.get('gst_included') or None, f.get('agent_name') or None, f.get('agent_contact') or None,
         f.get('agent_email') or None, int(f.get('reminder_days') or 30), f.get('notes') or None, status_override,
         policy_doc, invoice_doc, rc_doc, maintenance_id))

    # Keep the vehicle's own insurance_expiry (shown on the Vehicles page) in sync with this policy.
    if vehicle_id and f.get('expiry_date'):
        conn.execute("UPDATE vehicles SET insurance_expiry=? WHERE id=?", (f.get('expiry_date'), vehicle_id))
    conn.commit()
    conn.close()
    return redirect(url_for('vehicles_list', tab='insurance'))

@app.route('/maintenance/insurance/<int:policy_id>/edit', methods=['POST'])
def edit_insurance(policy_id):
    conn = get_db()
    f = request.form
    policy = conn.execute("SELECT * FROM insurance_policies WHERE id=?", (policy_id,)).fetchone()
    if not policy:
        conn.close()
        return redirect(url_for('vehicles_list', tab='insurance'))

    def get_or_create_vehicle(vno):
        if not vno or not vno.strip():
            return None
        vno = vno.strip()
        row = conn.execute("SELECT id FROM vehicles WHERE vehicle_no = ? COLLATE NOCASE", (vno,)).fetchone()
        if row:
            return row[0]
        cur = conn.execute("INSERT INTO vehicles (vehicle_no) VALUES (?)", (vno,))
        return cur.lastrowid

    vehicle_id = get_or_create_vehicle(f.get('vehicle_no'))
    insurer_id = get_or_create_vendor(conn, f.get('insurer_name'))
    premium = float(f.get('premium_amount') or 0)

    policy_doc = _save_insurance_doc(request.files.get('policy_doc'), 'policy') or policy['policy_doc_path']
    invoice_doc = _save_insurance_doc(request.files.get('invoice_doc'), 'invoice') or policy['invoice_doc_path']
    rc_doc = _save_insurance_doc(request.files.get('rc_doc'), 'rc') or policy['rc_doc_path']
    status_override = 'Cancelled' if f.get('policy_status') == 'Cancelled' else None

    conn.execute("""UPDATE insurance_policies SET vehicle_id=?, insurance_type=?, insurer_id=?, policy_number=?,
                    start_date=?, expiry_date=?, premium_amount=?, idv=?, ncb_pct=?, gst_included=?, agent_name=?,
                    agent_contact=?, agent_email=?, reminder_days=?, notes=?, status_override=?, policy_doc_path=?,
                    invoice_doc_path=?, rc_doc_path=? WHERE id=?""",
        (vehicle_id, f.get('insurance_type'), insurer_id, f.get('policy_number') or None, f.get('start_date'),
         f.get('expiry_date'), premium, float(f.get('idv') or 0) or None, float(f.get('ncb_pct') or 0) or None,
         f.get('gst_included') or None, f.get('agent_name') or None, f.get('agent_contact') or None,
         f.get('agent_email') or None, int(f.get('reminder_days') or 30), f.get('notes') or None, status_override,
         policy_doc, invoice_doc, rc_doc, policy_id))

    if policy['maintenance_id']:
        conn.execute("UPDATE maintenance SET date=?, vehicle_id=?, amount=?, paid_amount=?, vendor_id=?, notes=? WHERE id=?",
            (f.get('start_date'), vehicle_id, premium, premium, insurer_id, f.get('notes') or None, policy['maintenance_id']))
    if vehicle_id and f.get('expiry_date'):
        conn.execute("UPDATE vehicles SET insurance_expiry=? WHERE id=?", (f.get('expiry_date'), vehicle_id))
    conn.commit()
    conn.close()
    return redirect(url_for('vehicles_list', tab='insurance'))

@app.route('/maintenance/insurance/<int:policy_id>/delete', methods=['POST'])
def delete_insurance(policy_id):
    conn = get_db()
    policy = conn.execute("SELECT * FROM insurance_policies WHERE id=?", (policy_id,)).fetchone()
    if policy:
        if policy['maintenance_id']:
            conn.execute("DELETE FROM maintenance WHERE id=?", (policy['maintenance_id'],))
        conn.execute("DELETE FROM insurance_policies WHERE id=?", (policy_id,))
        conn.commit()
    conn.close()
    return redirect(url_for('vehicles_list', tab='insurance'))

@app.route('/maintenance/insurance/legacy/<int:maintenance_id>/convert', methods=['POST'])
def convert_legacy_insurance(maintenance_id):
    """Fills in the missing policy number/dates/etc. on an old free-text Insurance entry —
    reuses that same maintenance/ledger row (so the cost isn't billed twice) and just attaches
    the new structured insurance_policies record to it, same as every other 'add' flow here."""
    conn = get_db()
    f = request.form
    existing = conn.execute("SELECT id FROM insurance_policies WHERE maintenance_id=?", (maintenance_id,)).fetchone()
    if existing:
        conn.close()
        return redirect(url_for('vehicles_list', tab='insurance'))

    m = conn.execute("SELECT * FROM maintenance WHERE id=?", (maintenance_id,)).fetchone()
    if not m:
        conn.close()
        return redirect(url_for('vehicles_list', tab='insurance'))

    def get_or_create_vehicle(vno):
        if not vno or not vno.strip():
            return None
        vno = vno.strip()
        row = conn.execute("SELECT id FROM vehicles WHERE vehicle_no = ? COLLATE NOCASE", (vno,)).fetchone()
        if row:
            return row[0]
        cur = conn.execute("INSERT INTO vehicles (vehicle_no) VALUES (?)", (vno,))
        return cur.lastrowid

    vehicle_id = get_or_create_vehicle(f.get('vehicle_no')) or m['vehicle_id']
    insurer_id = get_or_create_vendor(conn, f.get('insurer_name'))
    premium = float(f.get('premium_amount') or 0)
    status_override = 'Cancelled' if f.get('policy_status') == 'Cancelled' else None

    policy_doc = _save_insurance_doc(request.files.get('policy_doc'), 'policy')
    invoice_doc = _save_insurance_doc(request.files.get('invoice_doc'), 'invoice')
    rc_doc = _save_insurance_doc(request.files.get('rc_doc'), 'rc')

    conn.execute("""INSERT INTO insurance_policies (vehicle_id, insurance_type, insurer_id, policy_number, start_date,
                    expiry_date, premium_amount, idv, ncb_pct, gst_included, agent_name, agent_contact, agent_email,
                    reminder_days, notes, status_override, policy_doc_path, invoice_doc_path, rc_doc_path, maintenance_id)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (vehicle_id, f.get('insurance_type'), insurer_id, f.get('policy_number') or None, f.get('start_date') or m['date'],
         f.get('expiry_date'), premium, float(f.get('idv') or 0) or None, float(f.get('ncb_pct') or 0) or None,
         f.get('gst_included') or None, f.get('agent_name') or None, f.get('agent_contact') or None,
         f.get('agent_email') or None, int(f.get('reminder_days') or 30), f.get('notes') or m['notes'], status_override,
         policy_doc, invoice_doc, rc_doc, maintenance_id))

    conn.execute("UPDATE maintenance SET vehicle_id=?, date=?, vendor_id=?, amount=?, paid_amount=? WHERE id=?",
        (vehicle_id, f.get('start_date') or m['date'], insurer_id, premium, premium, maintenance_id))
    if vehicle_id and f.get('expiry_date'):
        conn.execute("UPDATE vehicles SET insurance_expiry=? WHERE id=?", (f.get('expiry_date'), vehicle_id))
    conn.commit()
    conn.close()
    return redirect(url_for('vehicles_list', tab='insurance'))

def _filter_entries_by_date(entries, date_from, date_to):
    if date_from:
        entries = [e for e in entries if (e['date'] or '') >= date_from]
    if date_to:
        entries = [e for e in entries if (e['date'] or '') <= date_to]
    return entries

def _filter_ledger_entries(entries, date_from, date_to, kind_f='', vtype_f=''):
    """Same date filtering as before, plus the new Transaction Type (kind) and vehicle Type filters —
    purely narrowing which already-computed entries are shown; none of their debit/credit/balance
    math is touched here."""
    entries = _filter_entries_by_date(entries, date_from, date_to)
    if kind_f:
        entries = [e for e in entries if e.get('kind') == kind_f]
    if vtype_f:
        entries = [e for e in entries if e.get('vehicle_type') == vtype_f]
    return entries

def _ledger_export_entries(all_entries):
    """Applies the same date/Transaction-Type/vehicle-Type/tab filters the page view uses, so an
    export always matches whatever the user currently has filtered — reads straight from the
    querystring since every export route is a plain GET with the same filter param names as the page."""
    entries = _filter_ledger_entries(all_entries, request.args.get('date_from', ''), request.args.get('date_to', ''),
                                      request.args.get('txn_type', ''), request.args.get('type', ''))
    entries = _tab_filter_entries(entries, request.args.get('tab', ''))
    return entries

def _tab_filter_entries(entries, tab_f):
    if tab_f == 'receivables':
        return [e for e in entries if e['debit'] > 0]
    if tab_f == 'payables':
        return [e for e in entries if e['credit'] > 0]
    return entries

def _ledger_page_context(all_entries):
    """Shared filter/tab/pagination handling for both party_ledger() and vendor_ledger()."""
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    kind_f = request.args.get('txn_type', '')
    vtype_f = request.args.get('type', '')
    tab_f = request.args.get('tab', '')

    entries = _filter_ledger_entries(all_entries, date_from, date_to, kind_f, vtype_f)
    entries = _tab_filter_entries(entries, tab_f)

    total_count = len(entries)
    page, per_page, total_pages = _paginate(request.args.get('page'), request.args.get('per_page'), total_count,
                                             per_page_options=(10, 25, 50, 100), default_per_page=10)
    page_entries = entries[(page - 1) * per_page: page * per_page]
    page_tokens = _page_tokens(page, total_pages)

    from urllib.parse import urlencode
    base_params = request.args.to_dict()
    base_params.pop('page', None)
    base_params.pop('per_page', None)
    base_qs = urlencode(base_params)
    tab_params = dict(base_params)
    tab_params.pop('tab', None)
    tab_qs = urlencode(tab_params)

    return {
        'entries': page_entries, 'total_count': total_count, 'page': page, 'per_page': per_page,
        'total_pages': total_pages, 'page_tokens': page_tokens, 'base_qs': base_qs, 'tab_qs': tab_qs,
        'f_date_from': date_from, 'f_date_to': date_to, 'f_txn_type': kind_f, 'f_type': vtype_f, 'f_tab': tab_f or 'all',
    }

@app.route('/ledger/party/<int:party_id>')
def party_ledger(party_id):
    conn = get_db()
    party = conn.execute("SELECT * FROM parties WHERE id=?", (party_id,)).fetchone()
    pending_trips_raw = conn.execute("""SELECT id, date, lr_number, from_loc, to_loc, billed_amount,
                                    (billed_amount - COALESCE(payment_received,0) - COALESCE(party_advance,0)) as pending
                                    FROM trips WHERE party_id=?
                                    AND (billed_amount - COALESCE(payment_received,0) - COALESCE(party_advance,0)) > 0.01
                                    ORDER BY date""", (party_id,)).fetchall()
    conn.close()
    pending_trips = [{'id': t['id'], 'date': t['date'], 'lr_number': t['lr_number'], 'from_loc': t['from_loc'],
                       'to_loc': t['to_loc'], 'billed_amount': t['billed_amount'] or 0, 'pending': t['pending'],
                       'paid': (t['billed_amount'] or 0) - t['pending']} for t in pending_trips_raw]
    total_pending_trips = sum(t['pending'] for t in pending_trips)
    all_entries = _get_party_ledger_entries(party_id)
    final_balance = all_entries[0]['balance'] if all_entries else 0
    ctx = _ledger_page_context(all_entries)

    return render_template('ledger.html', name=party['name'], role='Party',
                            final_balance=final_balance, payment_url=f'/payment/party/{party_id}', export_url=f'/ledger/party/{party_id}/export',
                            entity_id=party_id, entity_type='party', address=party['address'], contact=party['contact'],
                            email=party['email'], credit_limit=party['credit_limit'], since_date=party['since_date'],
                            opening_balance=party['opening_balance'], opening_balance_date=party['opening_balance_date'],
                            gstin=party['gstin'], category=party['category'], status=party['status'] or 'Active',
                            pending_trips=pending_trips, total_pending_trips=total_pending_trips,
                            return_to=request.full_path, active='accounts', **ctx)

@app.route('/ledger/vendor/<int:vendor_id>')
def vendor_ledger(vendor_id):
    conn = get_db()
    vendor = conn.execute("SELECT * FROM vendors WHERE id=?", (vendor_id,)).fetchone()
    if vendor and vendor['linked_party_id']:
        conn.close()
        return redirect(url_for('party_ledger', party_id=vendor['linked_party_id']))
    pending_trips_raw = conn.execute("""SELECT t.id, t.date, t.lr_number, t.from_loc, t.to_loc, v.vehicle_no,
                                    (CASE WHEN t.owner_rate_type='FIXED' THEN t.owner_fixed_amount ELSE COALESCE(t.owner_rate,0)*COALESCE(t.quantity,0) END) as billed_amount,
                                    (CASE WHEN t.owner_rate_type='FIXED' THEN t.owner_fixed_amount ELSE COALESCE(t.owner_rate,0)*COALESCE(t.quantity,0) END - COALESCE(t.paid_to_owner,0)) as pending
                                    FROM trips t LEFT JOIN vehicles v ON t.vehicle_id=v.id WHERE t.owner_vendor_id=?
                                    AND (CASE WHEN t.owner_rate_type='FIXED' THEN t.owner_fixed_amount ELSE COALESCE(t.owner_rate,0)*COALESCE(t.quantity,0) END - COALESCE(t.paid_to_owner,0)) > 0.01
                                    ORDER BY t.date""", (vendor_id,)).fetchall()
    conn.close()
    pending_trips = [{'id': t['id'], 'date': t['date'], 'lr_number': t['lr_number'], 'from_loc': t['from_loc'],
                       'to_loc': t['to_loc'], 'vehicle_no': t['vehicle_no'], 'billed_amount': t['billed_amount'] or 0,
                       'pending': t['pending'], 'paid': (t['billed_amount'] or 0) - t['pending']} for t in pending_trips_raw]
    total_pending_trips = sum(t['pending'] for t in pending_trips)
    all_entries = _get_vendor_ledger_entries(vendor_id)
    final_balance = all_entries[0]['balance'] if all_entries else 0
    ctx = _ledger_page_context(all_entries)

    return render_template('ledger.html', name=vendor['name'], role='Vendor',
                            final_balance=final_balance, payment_url=f'/payment/vendor/{vendor_id}', export_url=f'/ledger/vendor/{vendor_id}/export',
                            entity_id=vendor_id, entity_type='vendor', address=vendor['address'], contact=vendor['contact'],
                            email=vendor['email'], credit_limit=vendor['credit_limit'], since_date=vendor['since_date'],
                            opening_balance=vendor['opening_balance'], opening_balance_date=vendor['opening_balance_date'],
                            gstin=vendor['gstin'], category=vendor['category'], status=vendor['status'] or 'Active',
                            pending_trips=pending_trips, total_pending_trips=total_pending_trips,
                            return_to=request.full_path, active='accounts', **ctx)

def _export_ledger_entries(name, entries):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask import send_file
    import io
    wb = Workbook()
    ws = wb.active
    ws.title = "Ledger"
    headers = ["Date","Type","Detail","Ref / Document","Debit","Credit","Balance"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1B2A4A")
    # Every column wraps within its own cell instead of Excel's default single-line display, which
    # was crowding/overlapping neighbouring columns whenever Detail (or anything else) ran long.
    wrap_top = Alignment(wrap_text=True, vertical='top')
    wrap_top_right = Alignment(wrap_text=True, vertical='top', horizontal='right')
    for r_idx, e in enumerate(entries, 2):
        ws.cell(row=r_idx, column=1, value=e['date']).alignment = wrap_top
        ws.cell(row=r_idx, column=2, value=e.get('kind', '')).alignment = wrap_top
        ws.cell(row=r_idx, column=3, value=e['detail']).alignment = wrap_top
        ws.cell(row=r_idx, column=4, value=e.get('ref', '')).alignment = wrap_top
        ws.cell(row=r_idx, column=5, value=e['debit'] or None).alignment = wrap_top_right
        ws.cell(row=r_idx, column=6, value=e['credit'] or None).alignment = wrap_top_right
        ws.cell(row=r_idx, column=7, value=e['balance']).alignment = wrap_top_right
    widths = {'A': 12, 'B': 14, 'C': 46, 'D': 20, 'E': 14, 'F': 14, 'G': 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    safe_name = "".join(c for c in name if c.isalnum() or c in " _-")[:40]
    return send_file(buf, as_attachment=True, download_name=f'ledger_{safe_name}.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def _lr_label(lr_number, trip_id):
    """LR numbers in this data are sometimes stored already prefixed with '#' — avoid stacking a second one."""
    val = lr_number or trip_id
    val = str(val)
    return val if val.startswith('#') else f"#{val}"

def _payment_base_detail(p, label):
    parts = [label]
    if p['mode']:
        parts.append(f"({p['mode']})")
    detail = ' '.join(parts)
    if p['reference_id']:
        detail += f" — Ref: {p['reference_id']}"
    if p['remarks']:
        detail += f" | {p['remarks']}"
    return detail

def _get_party_ledger_entries(party_id):
    conn = get_db()
    party = conn.execute("SELECT opening_balance, opening_balance_date, since_date FROM parties WHERE id=?", (party_id,)).fetchone()
    trips = conn.execute("""SELECT id, date, lr_number, from_loc, to_loc, billed_amount, payment_received, party_advance, type
                             FROM trips WHERE party_id=? ORDER BY date""", (party_id,)).fetchall()
    # How much of each trip's payment_received came from a ledger payment allocation (vs. being
    # recorded directly on the trip itself) — so the trip's own row isn't inflated with money
    # that actually arrived, and should be shown, on a later payment date.
    trip_alloc = {}
    for row in conn.execute("""SELECT pa.trip_id, SUM(pa.amount) as amt FROM payment_allocations pa
                               JOIN trips t ON pa.trip_id=t.id WHERE t.party_id=? GROUP BY pa.trip_id""", (party_id,)).fetchall():
        trip_alloc[row['trip_id']] = row['amt'] or 0
    # Real invoice number for a trip, if one's been generated yet — shown in Ref/Document ahead
    # of the LR number, since that's the more official document once it exists.
    trip_invoice_no = {}
    for row in conn.execute("SELECT trip_id, invoice_number FROM invoices WHERE trip_id IS NOT NULL").fetchall():
        if row['invoice_number']:
            trip_invoice_no[row['trip_id']] = row['invoice_number']
    payments = conn.execute("""SELECT id, date, amount, allocated_amount, mode, reference_id, remarks FROM payments
                                WHERE party_id=? AND payment_type='received' ORDER BY date""", (party_id,)).fetchall()
    linked_vendor = conn.execute("SELECT id FROM vendors WHERE linked_party_id=?", (party_id,)).fetchone()
    entries = []
    if party and party['opening_balance']:
        ob = party['opening_balance']
        entries.append({'date': party['opening_balance_date'] or party['since_date'] or '', 'detail': 'Opening Balance',
                         'debit': max(ob, 0), 'credit': max(-ob, 0), 'kind': 'Opening Balance', 'ref': '', 'vehicle_type': ''})
    for t in trips:
        original_received = (t['payment_received'] or 0) - trip_alloc.get(t['id'], 0)
        # Short place names (same _clean_loc used on Route Analytics), not the full stored address
        # — a trip's from/to can be a 200+ character pasted address, which has no business being
        # spelled out in full on every ledger row.
        entries.append({'date': t['date'], 'detail': f"Trip: {_lr_label(t['lr_number'], t['id'])} — {_clean_loc(t['from_loc'])} → {_clean_loc(t['to_loc'])}",
                         'debit': t['billed_amount'] or 0, 'credit': original_received + (t['party_advance'] or 0),
                         'kind': 'Trip Bill', 'ref': trip_invoice_no.get(t['id']) or t['lr_number'] or '', 'vehicle_type': t['type'] or '',
                         'link': url_for('trip_view', trip_id=t['id'])})
    for p in payments:
        base_detail = _payment_base_detail(p, 'Payment received')
        allocs = conn.execute("""SELECT t.lr_number, pa.amount FROM payment_allocations pa
                                 JOIN trips t ON pa.trip_id=t.id WHERE pa.payment_id=? ORDER BY pa.id""", (p['id'],)).fetchall()
        for a in allocs:
            entries.append({'date': p['date'], 'detail': f"{base_detail} — Applied to {a['lr_number'] or 'trip'}",
                             'debit': 0, 'credit': a['amount'], 'kind': 'Payment In', 'ref': p['reference_id'] or '', 'vehicle_type': '',
                             'link': url_for('payment_view', payment_id=p['id'])})
        leftover = (p['amount'] or 0) - (p['allocated_amount'] or 0)
        if leftover > 0.004 or not allocs:
            entries.append({'date': p['date'], 'detail': base_detail, 'debit': 0, 'credit': max(leftover, 0),
                             'kind': 'Payment In', 'ref': p['reference_id'] or '', 'vehicle_type': '',
                             'link': url_for('payment_view', payment_id=p['id'])})
    conn.close()
    if linked_vendor:
        # Same organization also acts as a vendor (fuel/maintenance/owner-hire) — pull those in too,
        # so this one ledger reflects everything, instead of splitting across two disconnected records.
        vendor_entries = _get_vendor_ledger_entries(linked_vendor['id'])
        for ve in vendor_entries:
            entries.append({'date': ve['date'], 'detail': ve['detail'] + ' (vendor side)',
                             'debit': ve['debit'], 'credit': ve['credit'],
                             'kind': ve.get('kind', 'Expense Adj.'), 'ref': ve.get('ref', ''), 'vehicle_type': ve.get('vehicle_type', ''),
                             'link': ve.get('link')})
    entries.sort(key=lambda x: x['date'] or '')
    balance = 0
    for e in entries:
        balance += e['debit'] - e['credit']
        e['balance'] = balance
    entries.reverse()
    return entries

def _get_vendor_ledger_entries(vendor_id):
    conn = get_db()
    vendor = conn.execute("SELECT opening_balance, opening_balance_date, since_date FROM vendors WHERE id=?", (vendor_id,)).fetchone()
    maint = conn.execute("""SELECT m.id, m.date, m.category, m.service_type, m.tyre_action, m.tyre_id, m.invoice_no,
                             m.amount, m.paid_amount, v.vehicle_no, b.battery_no, ip.policy_number
                             FROM maintenance m LEFT JOIN vehicles v ON m.vehicle_id=v.id
                             LEFT JOIN batteries b ON b.maintenance_id=m.id
                             LEFT JOIN insurance_policies ip ON ip.maintenance_id=m.id
                             WHERE m.vendor_id=? ORDER BY m.date""", (vendor_id,)).fetchall()
    fuel = conn.execute("""SELECT t.id, t.date, t.fuel_amount, t.type, v.vehicle_no
                           FROM trips t LEFT JOIN vehicles v ON t.vehicle_id=v.id
                           WHERE t.fuel_vendor_id=? ORDER BY t.date""", (vendor_id,)).fetchall()
    adv = conn.execute("""SELECT t.id, t.date, t.driver_adv_amount, t.type, v.vehicle_no
                          FROM trips t LEFT JOIN vehicles v ON t.vehicle_id=v.id
                          WHERE t.driver_adv_vendor_id=? ORDER BY t.date""", (vendor_id,)).fetchall()
    owner_trips = conn.execute("""SELECT t.id, t.date, t.lr_number, t.rate_type, t.fixed_rate_amount, t.owner_rate,
                                  t.owner_rate_type, t.owner_fixed_amount,
                                  t.quantity, t.paid_to_owner, t.type, v.vehicle_no
                                  FROM trips t LEFT JOIN vehicles v ON t.vehicle_id=v.id
                                  WHERE t.owner_vendor_id=? ORDER BY t.date""", (vendor_id,)).fetchall()
    # Trip "Others" items explicitly tagged with this vendor (e.g. a second fuel top-up or an
    # advance handled by someone other than the trip's usual fuel/advance vendor) — same table the
    # trip's own invoice reads from, just also attributed to a vendor here.
    other_items = conn.execute("""SELECT ii.description, ii.amount, ii.item_type, t.date, t.lr_number, t.id as trip_id,
                                  v.vehicle_no, t.type FROM invoice_items ii JOIN trips t ON ii.trip_id=t.id
                                  LEFT JOIN vehicles v ON t.vehicle_id=v.id
                                  WHERE ii.vendor_id=? ORDER BY t.date""", (vendor_id,)).fetchall()
    # Ledger-allocated portion of each owner-hire trip's paid_to_owner (see party-side comment above).
    trip_alloc = {}
    for row in conn.execute("""SELECT pa.trip_id, SUM(pa.amount) as amt FROM payment_allocations pa
                               JOIN trips t ON pa.trip_id=t.id WHERE t.owner_vendor_id=? GROUP BY pa.trip_id""", (vendor_id,)).fetchall():
        trip_alloc[row['trip_id']] = row['amt'] or 0
    # Real invoice number for a trip, if one's been generated — same lookup the party side uses.
    # Fuel/Driver Advance/owner-hire rows below identify themselves by vehicle number first (the
    # more useful identifier day-to-day); when a trip has no vehicle logged, the invoice number is
    # shown instead so the row is never left with no way to trace it back to a real trip.
    trip_invoice_no = {}
    for row in conn.execute("SELECT trip_id, invoice_number FROM invoices WHERE trip_id IS NOT NULL").fetchall():
        if row['invoice_number']:
            trip_invoice_no[row['trip_id']] = row['invoice_number']
    payments = conn.execute("""SELECT id, date, amount, allocated_amount, mode, reference_id, remarks FROM payments
                                WHERE vendor_id=? AND payment_type='paid' ORDER BY date""", (vendor_id,)).fetchall()
    entries = []
    if vendor and vendor['opening_balance']:
        ob = vendor['opening_balance']
        entries.append({'date': vendor['opening_balance_date'] or vendor['since_date'] or '', 'detail': 'Opening Balance',
                         'debit': max(ob, 0), 'credit': max(-ob, 0), 'kind': 'Opening Balance', 'ref': '', 'vehicle_type': ''})
    for m in maint:
        # Just what's needed to identify the entry — category (Battery/Tyre/Service/etc.), which
        # vehicle, and the one specific reference that applies (a tyre position, a battery number,
        # a policy number — never more than one, since a row is only ever one of those). The
        # invoice number goes in Ref, not piled into Detail — that column already exists for it.
        label = m['service_type'] or m['tyre_action'] or m['category'] or 'Maintenance'
        detail = f"Maintenance: {label}" + (f" — {m['vehicle_no']}" if m['vehicle_no'] else '')
        specific_ref = (f"Tyre {m['tyre_id']}" if m['tyre_id'] else
                        m['battery_no'] if m['battery_no'] else
                        f"Policy {m['policy_number']}" if m['policy_number'] else '')
        if specific_ref:
            detail += f" ({specific_ref})"
        entries.append({'date': m['date'], 'detail': detail,
                         'debit': m['paid_amount'] or 0, 'credit': m['amount'] or 0,
                         'kind': 'Expense Adj.', 'ref': m['invoice_no'] or '', 'vehicle_type': '',
                         'link': url_for('maintenance_view', m_id=m['id'])})
    for f in fuel:
        ident = f['vehicle_no'] or trip_invoice_no.get(f['id'], '')
        detail = 'Fuel' + (f" — {ident}" if ident else '')
        entries.append({'date': f['date'], 'detail': detail, 'debit': 0, 'credit': f['fuel_amount'] or 0,
                         'kind': 'Expense Adj.', 'ref': trip_invoice_no.get(f['id'], ''), 'vehicle_type': f['type'] or '',
                         'link': url_for('trip_view', trip_id=f['id'])})
    for a in adv:
        ident = a['vehicle_no'] or trip_invoice_no.get(a['id'], '')
        detail = 'Driver Advance' + (f" — {ident}" if ident else '')
        entries.append({'date': a['date'], 'detail': detail, 'debit': 0, 'credit': a['driver_adv_amount'] or 0,
                         'kind': 'Expense Adj.', 'ref': trip_invoice_no.get(a['id'], ''), 'vehicle_type': a['type'] or '',
                         'link': url_for('trip_view', trip_id=a['id'])})
    for o in owner_trips:
        owed = o['owner_fixed_amount'] if (o['owner_rate_type'] or 'PER_MT')=='FIXED' else (o['owner_rate'] or 0) * (o['quantity'] or 0)
        if owed:
            original_paid = (o['paid_to_owner'] or 0) - trip_alloc.get(o['id'], 0)
            detail = f"Trip: {_lr_label(o['lr_number'], o['id'])} — vehicle hire"
            ident = o['vehicle_no'] or trip_invoice_no.get(o['id'], '')
            if ident:
                detail += f" ({ident})"
            entries.append({'date': o['date'], 'detail': detail,
                             'debit': original_paid, 'credit': owed,
                             'kind': 'Trip Bill', 'ref': trip_invoice_no.get(o['id']) or o['lr_number'] or '', 'vehicle_type': o['type'] or '',
                             'link': url_for('trip_view', trip_id=o['id'])})
    for it in other_items:
        ident = it['vehicle_no'] or trip_invoice_no.get(it['trip_id'], '')
        detail = f"Trip: {_lr_label(it['lr_number'], it['trip_id'])} — {it['description']}" + (f" ({ident})" if ident else '')
        amt = it['amount'] or 0
        # 'charge' = vendor supplied something, we owe them more (credit side, same convention as
        # Fuel/Driver Advance above); 'deduction' reduces what's owed (debit side).
        entries.append({'date': it['date'], 'detail': detail,
                         'debit': amt if it['item_type'] == 'deduction' else 0,
                         'credit': amt if it['item_type'] == 'charge' else 0,
                         'kind': 'Expense Adj.', 'ref': trip_invoice_no.get(it['trip_id']) or it['lr_number'] or '',
                         'vehicle_type': it['type'] or '', 'link': url_for('trip_view', trip_id=it['trip_id'])})
    for p in payments:
        base_detail = _payment_base_detail(p, 'Payment made')
        allocs = conn.execute("""SELECT t.lr_number, pa.amount FROM payment_allocations pa
                                 JOIN trips t ON pa.trip_id=t.id WHERE pa.payment_id=? ORDER BY pa.id""", (p['id'],)).fetchall()
        for a in allocs:
            entries.append({'date': p['date'], 'detail': f"{base_detail} — Applied to {a['lr_number'] or 'trip'}",
                             'debit': a['amount'], 'credit': 0, 'kind': 'Payment Out', 'ref': p['reference_id'] or '', 'vehicle_type': '',
                             'link': url_for('payment_view', payment_id=p['id'])})
        leftover = (p['amount'] or 0) - (p['allocated_amount'] or 0)
        if leftover > 0.004 or not allocs:
            entries.append({'date': p['date'], 'detail': base_detail, 'debit': max(leftover, 0), 'credit': 0,
                             'kind': 'Payment Out', 'ref': p['reference_id'] or '', 'vehicle_type': '',
                             'link': url_for('payment_view', payment_id=p['id'])})
    conn.close()
    entries.sort(key=lambda x: x['date'] or '')
    balance = 0
    for e in entries:
        balance += e['debit'] - e['credit']
        e['balance'] = balance
    entries.reverse()
    return entries

@app.route('/ledger/party/<int:party_id>/export')
def export_party_ledger(party_id):
    conn = get_db()
    party = conn.execute("SELECT * FROM parties WHERE id=?", (party_id,)).fetchone()
    conn.close()
    entries = _get_party_ledger_entries(party_id)
    entries = _ledger_export_entries(entries)
    return _export_ledger_entries(party['name'], entries)

@app.route('/ledger/vendor/<int:vendor_id>/export')
def export_vendor_ledger(vendor_id):
    conn = get_db()
    vendor = conn.execute("SELECT * FROM vendors WHERE id=?", (vendor_id,)).fetchone()
    conn.close()
    entries = _get_vendor_ledger_entries(vendor_id)
    entries = _ledger_export_entries(entries)
    return _export_ledger_entries(vendor['name'], entries)

def _export_ledger_pdf(name, entries, role='', contact='', email='', address=''):
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from flask import send_file
    from xml.sax.saxutils import escape as esc
    import io

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.6*inch)
    styles = getSampleStyleSheet()
    company_style = ParagraphStyle('C', parent=styles['Title'], fontSize=17, textColor=colors.HexColor('#1B2A4A'), alignment=1)
    sub_style = ParagraphStyle('S', parent=styles['Normal'], fontSize=8.5, textColor=colors.HexColor('#5A6B8C'), alignment=1)
    title_style = ParagraphStyle('T', parent=styles['Heading2'], fontSize=13, textColor=colors.HexColor('#1B2A4A'))
    label_style = ParagraphStyle('L', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#5A6B8C'))
    # Every text column below is wrapped in this instead of passed as a raw string — a raw string
    # in a reportlab Table cell never wraps, it just overflows into the next column (that was the
    # "Detail overriding the other columns" bug), so a long trip detail line or party name needs
    # this to stay inside its own column and grow the row's height instead.
    cell_style = ParagraphStyle('CE', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#1A1A1A'), leading=10)
    num_style = ParagraphStyle('CN', parent=cell_style, alignment=2)  # right-aligned, same wrap behaviour
    def cell(text):
        return Paragraph(esc(str(text)), cell_style) if text else ''
    def num_cell(text):
        return Paragraph(esc(str(text)), num_style) if text else ''

    story = [Paragraph("ANIL TRANSPORT SERVICE", company_style),
             Paragraph("Head Off.: Shop No. D/8, Nirmal Market Power House Road, Rourkela - 769001", sub_style),
             Paragraph("GSTIN No.: 21ABDPL6110E1ZG &nbsp;|&nbsp; Mob. +91 9437246272", sub_style),
             Spacer(1, 4)]
    line_table = Table([['']], colWidths=[7*inch], rowHeights=[2])
    line_table.setStyle(TableStyle([('LINEBELOW', (0,0), (-1,-1), 1.5, colors.HexColor('#1B2A4A'))]))
    story.append(line_table)
    story.append(Spacer(1, 14))

    story.append(Paragraph(f"Statement of Account — {esc(name)} ({esc(role)})", title_style))
    contact_line = " &nbsp;|&nbsp; ".join([esc(p) for p in [contact, email, address] if p])
    if contact_line:
        story.append(Paragraph(contact_line, label_style))
    story.append(Spacer(1, 14))

    rows = [['Date', 'Type', 'Detail', 'Ref', 'Debit (Rs.)', 'Credit (Rs.)', 'Balance (Rs.)']]
    for e in entries:
        rows.append([cell(e['date'] or ''), cell(e.get('kind', '')), cell(e['detail'] or ''), cell(e.get('ref', '')),
                     num_cell(f"{e['debit']:,.0f}") if e['debit'] else '', num_cell(f"{e['credit']:,.0f}") if e['credit'] else '',
                     num_cell(f"{e['balance']:,.0f}")])
    t = Table(rows, colWidths=[0.75*inch, 0.85*inch, 2.15*inch, 0.85*inch, 0.85*inch, 0.85*inch, 0.9*inch], repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1B2A4A')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ALIGN', (4,0), (6,-1), 'RIGHT'),
        ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#DDE3EC')),
        ('TOPPADDING', (0,0), (-1,-1), 5), ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(t)
    doc.build(story)
    buf.seek(0)
    safe_name = "".join(c for c in name if c.isalnum() or c in " _-")[:40]
    return send_file(buf, as_attachment=True, download_name=f'ledger_{safe_name}.pdf', mimetype='application/pdf')

@app.route('/ledger/party/<int:party_id>/export/pdf')
def export_party_ledger_pdf(party_id):
    conn = get_db()
    party = conn.execute("SELECT * FROM parties WHERE id=?", (party_id,)).fetchone()
    conn.close()
    entries = _get_party_ledger_entries(party_id)
    entries = _ledger_export_entries(entries)
    return _export_ledger_pdf(party['name'], entries, role='Party', contact=party['contact'] or '', email=party['email'] or '', address=party['address'] or '')

@app.route('/ledger/vendor/<int:vendor_id>/export/pdf')
def export_vendor_ledger_pdf(vendor_id):
    conn = get_db()
    vendor = conn.execute("SELECT * FROM vendors WHERE id=?", (vendor_id,)).fetchone()
    conn.close()
    entries = _get_vendor_ledger_entries(vendor_id)
    entries = _ledger_export_entries(entries)
    return _export_ledger_pdf(vendor['name'], entries, role='Vendor', contact=vendor['contact'] or '', email=vendor['email'] or '', address=vendor['address'] or '')

@app.route('/payment/party/<int:party_id>', methods=['GET', 'POST'])
def add_party_payment(party_id):
    conn = get_db()
    party = conn.execute("SELECT * FROM parties WHERE id=?", (party_id,)).fetchone()
    if request.method == 'POST':
        f = request.form
        total_amount = float(f.get('amount') or 0)
        cur = conn.execute("INSERT INTO payments (date, payment_type, amount, party_id, mode, reference_id, remarks) VALUES (?,?,?,?,?,?,?)",
                            (f.get('date'), 'received', total_amount, party_id, f.get('mode'),
                             f.get('reference_id') or None, f.get('remarks') or None))
        payment_id = cur.lastrowid
        # Pending is computed live (billed_amount - payment_received - party_advance) since it
        # isn't a stored column; allocation fills each selected trip's balance in submitted
        # order (oldest first) until the payment amount runs out.
        trip_ids = f.getlist('trip_ids')
        remaining = total_amount
        allocated_total = 0
        for tid in trip_ids:
            if remaining <= 0.004:
                break
            trip = conn.execute("SELECT billed_amount, payment_received, party_advance FROM trips WHERE id=?", (tid,)).fetchone()
            if not trip:
                continue
            pending = (trip['billed_amount'] or 0) - (trip['payment_received'] or 0) - (trip['party_advance'] or 0)
            if pending <= 0.004:
                continue
            alloc = min(pending, remaining)
            conn.execute("UPDATE trips SET payment_received = COALESCE(payment_received,0) + ? WHERE id=?", (alloc, tid))
            conn.execute("INSERT INTO payment_allocations (payment_id, trip_id, amount) VALUES (?,?,?)", (payment_id, tid, alloc))
            remaining -= alloc
            allocated_total += alloc
        conn.execute("UPDATE payments SET allocated_amount=? WHERE id=?", (allocated_total, payment_id))
        conn.commit()
        conn.close()
        return redirect(url_for('party_ledger', party_id=party_id))
    conn.close()
    return render_template('add_payment.html', name=party['name'], role='Party', action_url=f'/payment/party/{party_id}', active='accounts')

@app.route('/payment/vendor/<int:vendor_id>', methods=['GET', 'POST'])
def add_vendor_payment(vendor_id):
    conn = get_db()
    vendor = conn.execute("SELECT * FROM vendors WHERE id=?", (vendor_id,)).fetchone()
    if request.method == 'POST':
        f = request.form
        total_amount = float(f.get('amount') or 0)
        cur = conn.execute("INSERT INTO payments (date, payment_type, amount, vendor_id, mode, reference_id, remarks) VALUES (?,?,?,?,?,?,?)",
                            (f.get('date'), 'paid', total_amount, vendor_id, f.get('mode'),
                             f.get('reference_id') or None, f.get('remarks') or None))
        payment_id = cur.lastrowid
        trip_ids = f.getlist('trip_ids')
        remaining = total_amount
        allocated_total = 0
        for tid in trip_ids:
            if remaining <= 0.004:
                break
            trip = conn.execute("""SELECT (CASE WHEN owner_rate_type='FIXED' THEN owner_fixed_amount ELSE COALESCE(owner_rate,0)*COALESCE(quantity,0) END) as owed,
                                   paid_to_owner FROM trips WHERE id=?""", (tid,)).fetchone()
            if not trip:
                continue
            pending = (trip['owed'] or 0) - (trip['paid_to_owner'] or 0)
            if pending <= 0.004:
                continue
            alloc = min(pending, remaining)
            conn.execute("UPDATE trips SET paid_to_owner = COALESCE(paid_to_owner,0) + ? WHERE id=?", (alloc, tid))
            conn.execute("INSERT INTO payment_allocations (payment_id, trip_id, amount) VALUES (?,?,?)", (payment_id, tid, alloc))
            remaining -= alloc
            allocated_total += alloc
        conn.execute("UPDATE payments SET allocated_amount=? WHERE id=?", (allocated_total, payment_id))
        conn.commit()
        conn.close()
        return redirect(url_for('vendor_ledger', vendor_id=vendor_id))
    conn.close()
    return render_template('add_payment.html', name=vendor['name'], role='Vendor', action_url=f'/payment/vendor/{vendor_id}', active='accounts')

@app.route('/payments/<int:payment_id>/view')
def payment_view(payment_id):
    """Read-only payment detail — opened from a Ledger row's Payment In/Out entry in the slide-in
    side panel. Payments have no dedicated edit page (they're recorded once via Add Payment and
    never revised), so this is purely a look-don't-touch summary plus the trips it was allocated
    against, with a link back to that party/vendor's full ledger."""
    conn = get_db()
    p = conn.execute("SELECT * FROM payments WHERE id=?", (payment_id,)).fetchone()
    if not p:
        conn.close()
        return redirect(url_for('accounts'))
    entity_name = None
    entity_type = None
    entity_id = None
    if p['party_id']:
        row = conn.execute("SELECT name FROM parties WHERE id=?", (p['party_id'],)).fetchone()
        entity_name = row['name'] if row else None
        entity_type, entity_id = 'party', p['party_id']
    elif p['vendor_id']:
        row = conn.execute("SELECT name FROM vendors WHERE id=?", (p['vendor_id'],)).fetchone()
        entity_name = row['name'] if row else None
        entity_type, entity_id = 'vendor', p['vendor_id']
    allocations = conn.execute("""SELECT pa.amount, t.lr_number, t.id as trip_id, t.from_loc, t.to_loc
                                  FROM payment_allocations pa JOIN trips t ON pa.trip_id=t.id
                                  WHERE pa.payment_id=? ORDER BY pa.id""", (payment_id,)).fetchall()
    conn.close()
    leftover = (p['amount'] or 0) - (p['allocated_amount'] or 0)
    return render_template('payment_view.html', p=p, entity_name=entity_name, entity_type=entity_type,
                            entity_id=entity_id, allocations=allocations, leftover=max(leftover, 0), active='accounts')

@app.route('/salaries/add', methods=['GET', 'POST'])
def add_salary():
    import datetime
    conn = get_db()
    if request.method == 'POST':
        f = request.form
        tx_date = f.get('date')
        month_label = datetime.datetime.strptime(tx_date, '%Y-%m-%d').strftime('%b %Y') if tx_date else ''
        now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        emp_name = f.get('employee')
        emp_type = f.get('employee_type')
        existing = conn.execute("SELECT id FROM employees WHERE name=? COLLATE NOCASE", (emp_name,)).fetchone()
        if existing:
            conn.execute("UPDATE employees SET type=? WHERE id=?", (emp_type, existing[0]))
        else:
            conn.execute("INSERT INTO employees (name, type) VALUES (?,?)", (emp_name, emp_type))
        conn.execute("INSERT INTO salaries (employee, month, amount, date, created_at) VALUES (?,?,?,?,?)",
                     (emp_name, month_label, float(f.get('amount') or 0), tx_date, now))
        conn.commit()
        conn.close()
        return redirect(url_for('salaries_list'))
    employees = conn.execute("SELECT name, type FROM employees ORDER BY name").fetchall()
    conn.close()
    return render_template('add_salary.html', employees=employees, active='salaries')

@app.route('/overheads/add', methods=['GET', 'POST'])
def add_overhead():
    conn = get_db()
    if request.method == 'POST':
        f = request.form
        conn.execute("""INSERT INTO overheads (date, category, amount, notes, payment_mode, receipt_number)
                        VALUES (?,?,?,?,?,?)""",
                     (f.get('date'), f.get('category'), float(f.get('amount') or 0), f.get('notes'),
                      f.get('payment_mode'), f.get('receipt_number')))
        conn.commit()
        conn.close()
        return redirect(url_for('overheads_list'))
    conn.close()
    return render_template('add_overhead.html', active='overheads')

@app.route('/trips/delete/<int:trip_id>', methods=['POST'])
def delete_trip(trip_id):
    conn = get_db()
    conn.execute("DELETE FROM trips WHERE id=?", (trip_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('trips_list'))

@app.route('/trips/<int:trip_id>/end', methods=['POST'])
def end_trip(trip_id):
    conn = get_db()
    trip = conn.execute("SELECT id, end_date FROM trips WHERE id=?", (trip_id,)).fetchone()
    if not trip or trip['end_date']:
        # Already ended (or missing) — no further edits allowed, silently ignore.
        conn.close()
        return redirect(url_for('trips_list', **request.args.to_dict()))

    f = request.form
    def n(key):
        return float(f.get(key) or 0)

    end_date = f.get('end_date')
    freight_amount = n('freight_amount')
    shortage_qty = n('shortage_qty')
    shortage_amount = n('shortage_amount')
    new_billed_amount = freight_amount - shortage_amount

    conn.execute("""UPDATE trips SET end_date=?, end_time=?, actual_km=?, lr_received=?,
                    shortage_qty=?, shortage_unit=?, shortage_amount=?, shortage_date=?, shortage_remarks=?,
                    remarks=?, billed_amount=?
                    WHERE id=?""",
                 (end_date, f.get('end_time') or None, n('actual_km') or None, f.get('lr_received') or None,
                  shortage_qty or None, f.get('shortage_unit') or None, shortage_amount,
                  end_date if shortage_amount else None, f.get('shortage_remarks') or None,
                  f.get('remarks') or None, new_billed_amount, trip_id))
    conn.commit()
    conn.close()
    return redirect(url_for('trips_list', **request.args.to_dict()))

@app.route('/maintenance/delete/<int:m_id>', methods=['POST'])
def delete_maintenance(m_id):
    conn = get_db()
    conn.execute("DELETE FROM maintenance WHERE id=?", (m_id,))
    conn.execute("DELETE FROM maintenance_items WHERE maintenance_id=?", (m_id,))
    conn.commit()
    conn.close()
    tab = request.args.get('tab') or request.form.get('tab') or 'service'
    if tab in ('insurance', 'permit'):
        return redirect(url_for('vehicles_list', tab=tab))
    return redirect(url_for('maintenance_list', tab=tab))

@app.route('/maintenance/edit/<int:m_id>', methods=['GET', 'POST'])
def edit_maintenance(m_id):
    conn = get_db()
    if request.method == 'POST':
        f = request.form
        def get_or_create_vehicle(vno):
            if not vno or not vno.strip():
                return None
            vno = vno.strip()
            row = conn.execute("SELECT id FROM vehicles WHERE vehicle_no = ? COLLATE NOCASE", (vno,)).fetchone()
            if row:
                return row[0]
            cur = conn.execute("INSERT INTO vehicles (vehicle_no) VALUES (?)", (vno,))
            return cur.lastrowid
        vehicle_id = get_or_create_vehicle(f.get('vehicle_no'))
        vendor_id = get_or_create_vendor(conn, f.get('vendor_name'))
        conn.execute("""UPDATE maintenance SET date=?, vehicle_id=?, category=?, amount=?, paid_amount=?, vendor_id=?, notes=?
                        WHERE id=?""",
                     (f.get('date'), vehicle_id, f.get('category'), float(f.get('amount') or 0),
                      float(f.get('paid_amount') or 0), vendor_id, f.get('notes'), m_id))
        conn.commit()
        conn.close()
        return_to = f.get('return_to')
        return redirect(return_to) if return_to else redirect(url_for('maintenance_list'))

    m = conn.execute("""SELECT mt.*, v.vehicle_no, ve.name as vendor_name FROM maintenance mt
                        LEFT JOIN vehicles v ON mt.vehicle_id=v.id
                        LEFT JOIN vendors ve ON mt.vendor_id=ve.id WHERE mt.id=?""", (m_id,)).fetchone()
    vehicles, parties, vendors, combined_names = _get_autocomplete_lists()
    conn.close()
    return render_template('edit_maintenance.html', m=m, vehicles=vehicles, combined_names=combined_names,
                            return_to=request.args.get('return_to', ''), active='maintenance')

@app.route('/maintenance/<int:m_id>/view')
def maintenance_view(m_id):
    """Read-only maintenance detail — opened from a Ledger row in the slide-in side panel instead
    of routing to the editable maintenance form."""
    conn = get_db()
    m = conn.execute("""SELECT mt.*, v.vehicle_no, ve.name as vendor_name FROM maintenance mt
                        LEFT JOIN vehicles v ON mt.vehicle_id=v.id
                        LEFT JOIN vendors ve ON mt.vendor_id=ve.id WHERE mt.id=?""", (m_id,)).fetchone()
    conn.close()
    if not m:
        return redirect(url_for('maintenance_list'))
    return render_template('maintenance_view.html', m=m, active='maintenance')

@app.route('/salaries/delete/<int:s_id>', methods=['POST'])
def delete_salary(s_id):
    conn = get_db()
    conn.execute("DELETE FROM salary_items WHERE salary_id=?", (s_id,))
    conn.execute("DELETE FROM salaries WHERE id=?", (s_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('salaries_list'))

@app.route('/employees/<int:employee_id>/edit', methods=['POST'])
def edit_employee(employee_id):
    conn = get_db()
    f = request.form
    conn.execute("""UPDATE employees SET name=?, type=?, role=?, mobile=?, email=?, address=?, joining_date=?,
                    date_of_birth=?, bank_account=?, ifsc_code=?, upi_id=?, emergency_contact=?, aadhaar=?, pan=?,
                    driving_license=?, basic_salary=? WHERE id=?""",
                 (f.get('name'), f.get('type'), f.get('role'), f.get('mobile') or None, f.get('email') or None,
                  f.get('address') or None, f.get('joining_date') or None, f.get('date_of_birth') or None,
                  f.get('bank_account') or None, f.get('ifsc_code') or None, f.get('upi_id') or None,
                  f.get('emergency_contact') or None, f.get('aadhaar') or None, f.get('pan') or None,
                  f.get('driving_license') or None, float(f.get('basic_salary') or 0), employee_id))
    conn.commit()
    conn.close()
    return redirect(url_for('salaries_list', tab=request.form.get('return_tab') or 'overview'))

@app.route('/employees/<int:employee_id>/deactivate', methods=['POST'])
def deactivate_employee(employee_id):
    conn = get_db()
    row = conn.execute("SELECT status FROM employees WHERE id=?", (employee_id,)).fetchone()
    new_status = 'Inactive' if (row and row['status'] != 'Inactive') else 'Active'
    conn.execute("UPDATE employees SET status=? WHERE id=?", (new_status, employee_id))
    conn.commit()
    conn.close()
    return redirect(url_for('salaries_list'))

@app.route('/salaries/process-payroll', methods=['POST'])
def process_payroll():
    """Bulk-generates this month's salary record for every active employee who doesn't already
    have one — starting from their on-file basic_salary, no allowances/deductions yet (those get
    added via Edit Salary afterward). Never overwrites a month that's already been generated.
    Employees with no basic_salary on file are skipped rather than given a Rs.0 record that would
    just look broken — Edit Employee is where that gets set, and the redirect flags exactly who
    was skipped so it's obvious why they didn't get a payslip this run."""
    conn = get_db()
    month_key = request.form.get('month') or _current_month_key()
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    employees = conn.execute("SELECT id, name, basic_salary FROM employees WHERE status='Active' OR status IS NULL").fetchall()
    created = 0
    skipped_no_salary = []
    for e in employees:
        existing = conn.execute("SELECT id FROM salaries WHERE employee_id=? AND month_key=?", (e['id'], month_key)).fetchone()
        if existing:
            continue
        basic = e['basic_salary'] or 0
        if basic <= 0:
            skipped_no_salary.append(e['name'])
            continue
        conn.execute("""INSERT INTO salaries (employee, month, amount, date, created_at, employee_id, month_key,
                        basic_salary, gross_salary, total_deductions, advance_recovery, net_salary, payment_status)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 0, 0, ?, 'pending')""",
                     (e['name'], month_key, basic, None, now, e['id'], month_key, basic, basic, basic))
        created += 1
    conn.commit()
    conn.close()
    return redirect(url_for('salaries_list', tab='salary', month=month_key, processed=created,
                             skipped=','.join(skipped_no_salary) if skipped_no_salary else None))

@app.route('/employees/<int:employee_id>/salary/<month_key>/save', methods=['POST'])
def save_employee_salary(employee_id, month_key):
    """Upsert this employee's salary for this month — replaces its allowance/deduction line items
    wholesale with what was submitted (simplest correct way to handle add/remove rows in one form)
    and recomputes gross/net from them, rather than trusting a client-side total."""
    conn = get_db()
    f = request.form
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    basic = float(f.get('basic_salary') or 0)

    descs = f.getlist('item_desc')
    amounts = f.getlist('item_amount')
    types = f.getlist('item_type')
    items = []
    for i, desc in enumerate(descs):
        desc = (desc or '').strip()
        if not desc:
            continue
        try:
            amt = float(amounts[i]) if i < len(amounts) and amounts[i] else 0
        except ValueError:
            amt = 0
        item_type = types[i] if i < len(types) and types[i] in ('allowance', 'deduction') else 'allowance'
        items.append((desc, amt, item_type))

    allowances_total = sum(a for _, a, t in items if t == 'allowance')
    deductions_total = sum(a for _, a, t in items if t == 'deduction')
    advance_recovery = float(f.get('advance_recovery') or 0)
    gross = basic + allowances_total
    net = gross - deductions_total - advance_recovery

    existing = conn.execute("SELECT id, payment_status FROM salaries WHERE employee_id=? AND month_key=?",
                            (employee_id, month_key)).fetchone()
    if existing:
        salary_id = existing['id']
        conn.execute("""UPDATE salaries SET basic_salary=?, gross_salary=?, total_deductions=?, advance_recovery=?,
                        net_salary=?, amount=?, month=?, remarks=? WHERE id=?""",
                     (basic, gross, deductions_total, advance_recovery, net, net, month_key, f.get('remarks') or None, salary_id))
        conn.execute("DELETE FROM salary_items WHERE salary_id=?", (salary_id,))
    else:
        emp_name = conn.execute("SELECT name FROM employees WHERE id=?", (employee_id,)).fetchone()['name']
        cur = conn.execute("""INSERT INTO salaries (employee, month, amount, date, created_at, employee_id, month_key,
                              basic_salary, gross_salary, total_deductions, advance_recovery, net_salary, payment_status, remarks)
                              VALUES (?,?,?,?,?,?,?,?,?,?,?,?, 'pending', ?)""",
                     (emp_name, month_key, net, None, now, employee_id, month_key, basic, gross,
                      deductions_total, advance_recovery, net, f.get('remarks') or None))
        salary_id = cur.lastrowid
    for desc, amt, item_type in items:
        conn.execute("INSERT INTO salary_items (salary_id, item_type, description, amount) VALUES (?,?,?,?)",
                     (salary_id, item_type, desc, amt))
    conn.commit()
    conn.close()
    return redirect(url_for('salaries_list', tab='salary', month=month_key))

@app.route('/employees/<int:employee_id>/advance/add', methods=['POST'])
def add_employee_advance(employee_id):
    """Give Advance / Record Repayment, callable from any Employees-module drawer — same
    'advances' table the original standalone Advances page already used, so both stay in sync."""
    conn = get_db()
    f = request.form
    emp = conn.execute("SELECT name FROM employees WHERE id=?", (employee_id,)).fetchone()
    if emp:
        now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        conn.execute("INSERT INTO advances (employee, date, amount, type, notes, created_at) VALUES (?,?,?,?,?,?)",
                     (emp['name'], f.get('date') or datetime.date.today().isoformat(), float(f.get('amount') or 0),
                      f.get('type') if f.get('type') in ('given', 'repaid') else 'given', f.get('notes') or None, now))
        conn.commit()
    conn.close()
    return redirect(url_for('salaries_list', tab=f.get('return_tab') or 'overview'))

@app.route('/salaries/<int:salary_id>/mark-paid', methods=['POST'])
def mark_salary_paid(salary_id):
    conn = get_db()
    now = datetime.datetime.now()
    conn.execute("""UPDATE salaries SET payment_status='paid', payment_date=?, payment_mode=?, transaction_id=?, paid_by=?
                    WHERE id=?""",
                 (request.form.get('payment_date') or now.strftime('%Y-%m-%d'), request.form.get('payment_mode') or None,
                  request.form.get('transaction_id') or None, request.form.get('paid_by') or 'Admin', salary_id))
    conn.commit()
    conn.close()
    return redirect(url_for('salaries_list', tab='salary'))

@app.route('/salaries/bulk-mark-paid', methods=['POST'])
def bulk_mark_salary_paid():
    conn = get_db()
    ids = request.form.getlist('salary_ids')
    now = datetime.datetime.now().strftime('%Y-%m-%d')
    for sid in ids:
        conn.execute("UPDATE salaries SET payment_status='paid', payment_date=? WHERE id=?", (now, sid))
    conn.commit()
    conn.close()
    return redirect(url_for('salaries_list', tab='salary'))

@app.route('/attendance/mark', methods=['POST'])
def mark_attendance():
    """Single-employee, single-date attendance entry — UNIQUE(employee_id, date) means re-marking
    the same day just overwrites it instead of creating a second row."""
    conn = get_db()
    f = request.form
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn.execute("""INSERT INTO attendance (employee_id, date, status, in_time, out_time, remarks, marked_by, created_at)
                    VALUES (?,?,?,?,?,?,?,?)
                    ON CONFLICT(employee_id, date) DO UPDATE SET status=excluded.status, in_time=excluded.in_time,
                    out_time=excluded.out_time, remarks=excluded.remarks, marked_by=excluded.marked_by""",
                 (f.get('employee_id'), f.get('date'), f.get('status'), f.get('in_time') or None, f.get('out_time') or None,
                  f.get('remarks') or None, 'Admin', now))
    conn.commit()
    conn.close()
    return redirect(url_for('salaries_list', tab='attendance', month=(f.get('date') or '')[:7]))

@app.route('/attendance/bulk-mark', methods=['POST'])
def bulk_mark_attendance():
    """Bulk toolbar action — marks the same status for every selected employee on one date, e.g.
    marking a whole crew Present for today in one click."""
    conn = get_db()
    f = request.form
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    date = f.get('date') or datetime.date.today().isoformat()
    status = f.get('status')
    emp_ids = f.getlist('employee_ids')
    for eid in emp_ids:
        conn.execute("""INSERT INTO attendance (employee_id, date, status, marked_by, created_at) VALUES (?,?,?,?,?)
                        ON CONFLICT(employee_id, date) DO UPDATE SET status=excluded.status, marked_by=excluded.marked_by""",
                     (eid, date, status, 'Admin', now))
    conn.commit()
    conn.close()
    return redirect(url_for('salaries_list', tab='attendance', month=date[:7]))

@app.route('/salaries/<int:salary_id>/payslip')
def download_payslip(salary_id):
    from flask import send_file
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from xml.sax.saxutils import escape as esc
    import io

    conn = get_db()
    sal = conn.execute("""SELECT s.*, e.name, e.employee_code, e.role, e.type, e.bank_account, e.ifsc_code
                          FROM salaries s JOIN employees e ON s.employee_id=e.id WHERE s.id=?""", (salary_id,)).fetchone()
    if not sal:
        conn.close()
        return redirect(url_for('salaries_list', tab='salary'))
    items = conn.execute("SELECT item_type, description, amount FROM salary_items WHERE salary_id=?", (salary_id,)).fetchall()
    conn.close()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=0.6*inch, bottomMargin=0.6*inch)
    styles = getSampleStyleSheet()
    company_style = ParagraphStyle('C', parent=styles['Title'], fontSize=17, textColor=colors.HexColor('#1B2A4A'), alignment=1)
    title_style = ParagraphStyle('T', parent=styles['Heading2'], fontSize=13, textColor=colors.HexColor('#1B2A4A'))
    label_style = ParagraphStyle('L', parent=styles['Normal'], fontSize=9.5, textColor=colors.HexColor('#5A6B8C'))

    story = [Paragraph("ANIL TRANSPORT SERVICE", company_style), Spacer(1, 6),
             Paragraph(f"Payslip — {sal['month_key']}", title_style), Spacer(1, 10),
             Paragraph(f"<b>{esc(sal['name'])}</b> ({esc(sal['employee_code'] or '')}) — {esc(sal['role'] or sal['type'] or '')}", label_style),
             Spacer(1, 14)]

    rows = [['Basic Salary', f"Rs. {sal['basic_salary'] or 0:,.0f}"]]
    for it in items:
        prefix = '+ ' if it['item_type'] == 'allowance' else '- '
        rows.append([esc(it['description']), f"{prefix}Rs. {it['amount']:,.0f}"])
    if sal['advance_recovery']:
        rows.append(['Advance Recovery', f"- Rs. {sal['advance_recovery']:,.0f}"])
    rows.append(['NET SALARY', f"Rs. {sal['net_salary'] or 0:,.0f}"])
    t = Table(rows, colWidths=[4.5*inch, 2.4*inch])
    t.setStyle(TableStyle([
        ('FONTSIZE', (0,0), (-1,-1), 9.5), ('ALIGN', (1,0), (1,-1), 'RIGHT'),
        ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#DDE3EC')),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#EAF6EE')), ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,-1), (-1,-1), 11), ('TOPPADDING', (0,0), (-1,-1), 5), ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(t)
    story.append(Spacer(1, 14))
    story.append(Paragraph(f"Payment Status: <b>{esc((sal['payment_status'] or 'pending').title())}</b>" +
                            (f" &nbsp;|&nbsp; Paid on {esc(sal['payment_date'])}" if sal['payment_date'] else ''), label_style))
    if sal['bank_account']:
        story.append(Paragraph(f"Bank A/C: {esc(sal['bank_account'])} &nbsp;|&nbsp; IFSC: {esc(sal['ifsc_code'] or '')}", label_style))
    doc.build(story)
    buf.seek(0)
    safe_name = "".join(c for c in sal['name'] if c.isalnum() or c in " _-")[:30]
    return send_file(buf, as_attachment=True, download_name=f'payslip_{safe_name}_{sal["month_key"]}.pdf', mimetype='application/pdf')

@app.route('/overheads/delete/<int:o_id>', methods=['POST'])
def delete_overhead(o_id):
    conn = get_db()
    conn.execute("DELETE FROM overheads WHERE id=?", (o_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('overheads_list'))

@app.route('/vehicles/add', methods=['GET', 'POST'])
def add_vehicle():
    conn = get_db()
    if request.method == 'POST':
        f = request.form
        vno = (f.get('vehicle_no') or '').strip()
        vtype = f.get('type')
        existing = conn.execute("SELECT id FROM vehicles WHERE vehicle_no=?", (vno,)).fetchone()
        if existing:
            conn.execute("""UPDATE vehicles SET type=?, registration_date=?, capacity_mt=?,
                            insurance_expiry=?, fitness_expiry=?, puc_valid_upto=?, permit_valid_upto=?,
                            status=?, body_type=?, chassis_number=?, engine_number=?, notes=? WHERE id=?""",
                         (vtype, f.get('registration_date'), f.get('capacity_mt') or None,
                          f.get('insurance_expiry'), f.get('fitness_expiry'), f.get('puc_valid_upto'),
                          f.get('permit_valid_upto'), f.get('status') or 'Active', f.get('body_type'),
                          f.get('chassis_number') or None, f.get('engine_number') or None,
                          f.get('notes'), existing[0]))
        else:
            conn.execute("""INSERT INTO vehicles (vehicle_no, type, registration_date, capacity_mt,
                            insurance_expiry, fitness_expiry, puc_valid_upto, permit_valid_upto,
                            status, body_type, chassis_number, engine_number, notes) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                         (vno, vtype, f.get('registration_date'), f.get('capacity_mt') or None,
                          f.get('insurance_expiry'), f.get('fitness_expiry'), f.get('puc_valid_upto'),
                          f.get('permit_valid_upto'), f.get('status') or 'Active', f.get('body_type'),
                          f.get('chassis_number') or None, f.get('engine_number') or None,
                          f.get('notes')))
        conn.commit()
        conn.close()
        return redirect(url_for('vehicles_list'))
    conn.close()
    return render_template('add_vehicle.html', active='vehicles')

@app.route('/trips/export')
def export_trips():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from flask import send_file
    import io
    conn = get_db()
    vehicle_f = request.args.get('vehicle', '')
    party_f = request.args.get('party', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    status_f = request.args.get('status', '')
    lr_f = request.args.get('lr_received', '')
    lr_number_f = request.args.get('lr_number', '')
    query = """SELECT t.date, t.lr_number, v.vehicle_no, p.name as party_name,
               t.from_loc, t.to_loc, t.billed_amount, t.lr_received, t.end_date
               FROM trips t LEFT JOIN vehicles v ON t.vehicle_id = v.id
               LEFT JOIN parties p ON t.party_id = p.id WHERE 1=1"""
    params = []
    if vehicle_f: query += " AND v.vehicle_no LIKE ?"; params.append(f"%{vehicle_f}%")
    if party_f: query += " AND p.name LIKE ?"; params.append(f"%{party_f}%")
    if date_from: query += " AND t.date >= ?"; params.append(date_from)
    if date_to: query += " AND t.date <= ?"; params.append(date_to)
    if status_f == 'active': query += " AND t.end_date IS NULL"
    elif status_f == 'completed': query += " AND t.end_date IS NOT NULL"
    if lr_f == 'received': query += " AND t.lr_received = 'Yes'"
    elif lr_f == 'not_received': query += " AND (t.lr_received IS NULL OR t.lr_received != 'Yes')"
    if lr_number_f: query += " AND t.lr_number LIKE ?"; params.append(f"%{lr_number_f}%")
    query += " ORDER BY t.date DESC"
    rows = conn.execute(query, params).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Trips"
    headers = ["Date","LR Number","Vehicle","Party","From","To","Billed Amount","LR Received","Trip Status"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1B2A4A")
    for r_idx, row in enumerate(rows, 2):
        vals = list(row)[:-2] + ['Yes' if row['lr_received'] == 'Yes' else 'No', 'Completed' if row['end_date'] else 'Active']
        for c_idx, val in enumerate(vals, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='trips_export.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/trips/export/pdf')
def export_trips_pdf():
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from flask import send_file
    import io
    conn = get_db()
    vehicle_f = request.args.get('vehicle', '')
    party_f = request.args.get('party', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    status_f = request.args.get('status', '')
    lr_f = request.args.get('lr_received', '')
    lr_number_f = request.args.get('lr_number', '')
    query = """SELECT t.date, t.lr_number, v.vehicle_no, p.name as party_name,
               t.from_loc, t.to_loc, t.billed_amount, t.lr_received, t.end_date
               FROM trips t LEFT JOIN vehicles v ON t.vehicle_id = v.id
               LEFT JOIN parties p ON t.party_id = p.id WHERE 1=1"""
    params = []
    if vehicle_f: query += " AND v.vehicle_no LIKE ?"; params.append(f"%{vehicle_f}%")
    if party_f: query += " AND p.name LIKE ?"; params.append(f"%{party_f}%")
    if date_from: query += " AND t.date >= ?"; params.append(date_from)
    if date_to: query += " AND t.date <= ?"; params.append(date_to)
    if status_f == 'active': query += " AND t.end_date IS NULL"
    elif status_f == 'completed': query += " AND t.end_date IS NOT NULL"
    if lr_f == 'received': query += " AND t.lr_received = 'Yes'"
    elif lr_f == 'not_received': query += " AND (t.lr_received IS NULL OR t.lr_received != 'Yes')"
    if lr_number_f: query += " AND t.lr_number LIKE ?"; params.append(f"%{lr_number_f}%")
    query += " ORDER BY t.date DESC"
    rows = conn.execute(query, params).fetchall()
    conn.close()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('T', parent=styles['Heading1'], fontSize=15, textColor=colors.HexColor('#1B2A4A'))
    story = [Paragraph("Trips — Filtered View", title_style), Spacer(1, 10)]

    table_rows = [['Date', 'LR Number', 'Vehicle', 'Party', 'From', 'To', 'Billed (Rs.)', 'LR Recv.', 'Status']]
    total_billed = 0
    for r in rows:
        total_billed += r['billed_amount'] or 0
        table_rows.append([r['date'] or '', r['lr_number'] or '', r['vehicle_no'] or '', r['party_name'] or '',
                            r['from_loc'] or '', r['to_loc'] or '', f"{r['billed_amount'] or 0:,.0f}",
                            'Yes' if r['lr_received'] == 'Yes' else 'No',
                            'Completed' if r['end_date'] else 'Active'])
    t = Table(table_rows, colWidths=[0.75*inch, 0.9*inch, 0.8*inch, 1.4*inch, 1.5*inch, 1.5*inch, 0.9*inch, 0.7*inch, 0.8*inch], repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1B2A4A')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('ALIGN', (6,0), (6,-1), 'RIGHT'),
        ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#DDE3EC')),
        ('TOPPADDING', (0,0), (-1,-1), 4), ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(t)
    story.append(Spacer(1, 10))
    story.append(Paragraph(f"Total Trips: {len(rows)} &nbsp;&nbsp;|&nbsp;&nbsp; Total Billed: Rs. {total_billed:,.0f}",
                            ParagraphStyle('F', parent=styles['Normal'], fontSize=10)))
    doc.build(story)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='trips_export.pdf', mimetype='application/pdf')

@app.route('/maintenance/export')
def export_maintenance():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from flask import send_file
    import io
    conn = get_db()
    vehicle_f = request.args.get('vehicle', '')
    vendor_f = request.args.get('vendor', '')
    category_f = request.args.get('category', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    query = """SELECT m.date, v.vehicle_no, m.category, m.amount, ve.name as vendor_name
               FROM maintenance m LEFT JOIN vehicles v ON m.vehicle_id = v.id
               LEFT JOIN vendors ve ON m.vendor_id = ve.id WHERE 1=1"""
    params = []
    if vehicle_f: query += " AND v.vehicle_no LIKE ?"; params.append(f"%{vehicle_f}%")
    if vendor_f: query += " AND ve.name LIKE ?"; params.append(f"%{vendor_f}%")
    if category_f: query += " AND m.category = ?"; params.append(category_f)
    if date_from: query += " AND m.date >= ?"; params.append(date_from)
    if date_to: query += " AND m.date <= ?"; params.append(date_to)
    query += " ORDER BY m.date DESC"
    rows = conn.execute(query, params).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Maintenance"
    headers = ["Date","Vehicle","Category","Amount","Vendor"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1B2A4A")
    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='maintenance_export.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/maintenance/export/pdf')
def export_maintenance_pdf():
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from flask import send_file
    import io
    conn = get_db()
    vehicle_f = request.args.get('vehicle', '')
    vendor_f = request.args.get('vendor', '')
    category_f = request.args.get('category', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    query = """SELECT m.date, v.vehicle_no, m.category, m.amount, ve.name as vendor_name
               FROM maintenance m LEFT JOIN vehicles v ON m.vehicle_id = v.id
               LEFT JOIN vendors ve ON m.vendor_id = ve.id WHERE 1=1"""
    params = []
    if vehicle_f: query += " AND v.vehicle_no LIKE ?"; params.append(f"%{vehicle_f}%")
    if vendor_f: query += " AND ve.name LIKE ?"; params.append(f"%{vendor_f}%")
    if category_f: query += " AND m.category = ?"; params.append(category_f)
    if date_from: query += " AND m.date >= ?"; params.append(date_from)
    if date_to: query += " AND m.date <= ?"; params.append(date_to)
    query += " ORDER BY m.date DESC"
    rows = conn.execute(query, params).fetchall()
    conn.close()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('T', parent=styles['Heading1'], fontSize=15, textColor=colors.HexColor('#1B2A4A'))
    story = [Paragraph("Maintenance — Filtered View", title_style), Spacer(1, 10)]

    table_rows = [['Date', 'Vehicle', 'Category', 'Vendor', 'Amount (Rs.)']]
    total_amount = 0
    for r in rows:
        total_amount += r['amount'] or 0
        table_rows.append([r['date'] or '', r['vehicle_no'] or '', r['category'] or '', r['vendor_name'] or '—',
                            f"{r['amount'] or 0:,.0f}"])
    t = Table(table_rows, colWidths=[1.1*inch, 1.3*inch, 2*inch, 2*inch, 1.3*inch], repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1B2A4A')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (4,0), (4,-1), 'RIGHT'),
        ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#DDE3EC')),
        ('TOPPADDING', (0,0), (-1,-1), 5), ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(t)
    story.append(Spacer(1, 10))
    story.append(Paragraph(f"Total Entries: {len(rows)} &nbsp;&nbsp;|&nbsp;&nbsp; Total Amount: Rs. {total_amount:,.0f}",
                            ParagraphStyle('F', parent=styles['Normal'], fontSize=10)))
    doc.build(story)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='maintenance_export.pdf', mimetype='application/pdf')

@app.route('/advances')
def advances_list():
    conn = get_db()
    rows = conn.execute("""SELECT employee,
        SUM(CASE WHEN type='given' THEN amount ELSE 0 END) as total_given,
        SUM(CASE WHEN type='repaid' THEN amount ELSE 0 END) as total_repaid,
        SUM(CASE WHEN type='given' THEN amount ELSE -amount END) as outstanding
        FROM advances GROUP BY employee ORDER BY employee""").fetchall()
    entries = conn.execute("SELECT employee, date, amount, type, notes FROM advances ORDER BY date DESC").fetchall()
    conn.close()
    return render_template('advances_list.html', rows=rows, entries=entries, active='salaries')

@app.route('/advances/add', methods=['GET', 'POST'])
def add_advance():
    import datetime
    conn = get_db()
    if request.method == 'POST':
        f = request.form
        now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        conn.execute("INSERT INTO advances (employee, date, amount, type, notes, created_at) VALUES (?,?,?,?,?,?)",
                     (f.get('employee'), f.get('date'), float(f.get('amount') or 0), f.get('type'), f.get('notes'), now))
        conn.commit()
        conn.close()
        return redirect(url_for('advances_list'))
    conn.close()
    return render_template('add_advance.html', active='salaries')

@app.route('/vehicles/edit/<int:vehicle_id>', methods=['GET', 'POST'])
def edit_vehicle(vehicle_id):
    conn = get_db()
    if request.method == 'POST':
        f = request.form
        conn.execute("""UPDATE vehicles SET vehicle_no=?, type=?, registration_date=?, capacity_mt=?,
                        insurance_expiry=?, fitness_expiry=?, puc_valid_upto=?, permit_valid_upto=?,
                        status=?, body_type=?, chassis_number=?, engine_number=?, notes=? WHERE id=?""",
                     (f.get('vehicle_no'), f.get('type'), f.get('registration_date'), f.get('capacity_mt') or None,
                      f.get('insurance_expiry'), f.get('fitness_expiry'), f.get('puc_valid_upto'),
                      f.get('permit_valid_upto'), f.get('status') or 'Active', f.get('body_type'),
                      f.get('chassis_number') or None, f.get('engine_number') or None,
                      f.get('notes'), vehicle_id))
        conn.commit()
        conn.close()
        return redirect(url_for('vehicles_list'))
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    conn.close()
    return render_template('edit_vehicle.html', v=vehicle, active='vehicles')

@app.route('/vehicles/toggle_status/<int:vehicle_id>', methods=['POST'])
def toggle_vehicle_status(vehicle_id):
    """Quick Active/Inactive flip from the vehicle list row — a safer everyday action than
    deleting a vehicle outright, which would also wipe its trip/maintenance history."""
    conn = get_db()
    v = conn.execute("SELECT status FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    if v:
        new_status = 'Inactive' if (v['status'] or 'Active') == 'Active' else 'Active'
        conn.execute("UPDATE vehicles SET status=? WHERE id=?", (new_status, vehicle_id))
        conn.commit()
    conn.close()
    tab = request.args.get('tab') or 'all'
    return redirect(url_for('vehicles_list', tab=tab))

@app.route('/vehicles/delete/<int:vehicle_id>', methods=['POST'])
def delete_vehicle(vehicle_id):
    conn = get_db()
    conn.execute("DELETE FROM vehicles WHERE id=?", (vehicle_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('vehicles_list'))

@app.route('/vehicles/<int:vehicle_id>/compliance/<ctype>/renew', methods=['POST'])
def vehicle_compliance_renew(vehicle_id, ctype):
    """Manual renewal — an office user typing in a new certificate/permit number and expiry
    directly, independent of any provider sync. source='manual' distinguishes this from a
    provider-synced record without overwriting whatever the last sync found."""
    if ctype not in ('fitness', 'puc', 'permit'):
        return redirect(url_for('vehicles_list'))
    f = request.form
    conn = get_db()
    now = cs._now()
    existing = conn.execute("SELECT id FROM vehicle_compliance WHERE vehicle_id=? AND compliance_type=?",
                             (vehicle_id, ctype)).fetchone()
    if existing:
        conn.execute("""UPDATE vehicle_compliance SET document_number=?, valid_upto=?, source='manual',
                        sync_status='Not Synced', updated_at=? WHERE id=?""",
                     (f.get('document_number') or None, f.get('valid_upto') or None, now, existing['id']))
    else:
        conn.execute("""INSERT INTO vehicle_compliance
                        (vehicle_id, compliance_type, document_number, valid_upto, source,
                         sync_status, created_at, updated_at)
                        VALUES (?,?,?,?,'manual','Not Synced',?,?)""",
                     (vehicle_id, ctype, f.get('document_number') or None, f.get('valid_upto') or None, now, now))
    # Keep the vehicle's own quick-glance column in sync too, same pattern already used for
    # Insurance's expiry — so the plain date field in the Edit Vehicle form never disagrees
    # with what Compliance shows.
    col = {'fitness': 'fitness_expiry', 'puc': 'puc_valid_upto', 'permit': 'permit_valid_upto'}[ctype]
    if f.get('valid_upto'):
        conn.execute(f"UPDATE vehicles SET {col}=? WHERE id=?", (f.get('valid_upto'), vehicle_id))
    conn.commit()
    conn.close()
    return redirect(url_for('vehicles_list', tab='all'))

@app.route('/vehicles/<int:vehicle_id>/compliance/sync', methods=['POST'])
def vehicle_compliance_sync(vehicle_id):
    """Manual per-vehicle 'Sync Now' — calls the same mock providers the nightly job uses,
    just for one vehicle, so a user doesn't have to wait for 2 AM to see it work."""
    conn = get_db()
    cs.sync_vehicle(conn, vehicle_id)
    conn.commit()
    conn.close()
    return redirect(url_for('vehicles_list', tab='all'))

def _employee_ledger_entries(conn, employee_name, date_from='', date_to=''):
    """Opening Balance / Salary Paid / Advance Given / Advance Repaid, running-balance ledger for
    one employee (by name, matching how `salaries`/`advances` are still keyed) — the same proven
    formula the standalone Employee Ledger page has always used, now reusable from the Employees
    module's drawers too instead of only living on that separate page."""
    sal_query = "SELECT date, amount, created_at FROM salaries WHERE employee=? COLLATE NOCASE"
    sal_params = [employee_name]
    if date_from: sal_query += " AND date >= ?"; sal_params.append(date_from)
    if date_to: sal_query += " AND date <= ?"; sal_params.append(date_to)
    salaries = conn.execute(sal_query, sal_params).fetchall()

    adv_query = "SELECT date, amount, type, notes, created_at FROM advances WHERE employee=? COLLATE NOCASE"
    adv_params = [employee_name]
    if date_from: adv_query += " AND date >= ?"; adv_params.append(date_from)
    if date_to: adv_query += " AND date <= ?"; adv_params.append(date_to)
    advances = conn.execute(adv_query, adv_params).fetchall()
    emp_row = conn.execute("SELECT opening_balance, opening_balance_date FROM employees WHERE name=? COLLATE NOCASE",
                           (employee_name,)).fetchone()

    entries = []
    if emp_row and emp_row['opening_balance']:
        ob = emp_row['opening_balance']
        entries.append({'date': emp_row['opening_balance_date'] or '', 'entry_type': 'Opening Balance',
                         'debit': max(ob, 0), 'credit': max(-ob, 0), 'notes': 'Carried over balance',
                         'created_at': '', 'affects_advance': True})
    for s in salaries:
        if s['amount']:
            entries.append({'date': s['date'], 'entry_type': 'Salary Paid', 'debit': 0, 'credit': s['amount'] or 0,
                             'notes': '', 'created_at': s['created_at'], 'affects_advance': False})
    for a in advances:
        if a['type'] == 'given':
            entries.append({'date': a['date'], 'entry_type': 'Advance Given', 'debit': a['amount'] or 0, 'credit': 0,
                             'notes': a['notes'] or '', 'created_at': a['created_at'], 'affects_advance': True})
        else:
            entries.append({'date': a['date'], 'entry_type': 'Advance Repaid', 'debit': 0, 'credit': a['amount'] or 0,
                             'notes': a['notes'] or '', 'created_at': a['created_at'], 'affects_advance': True})
    entries.sort(key=lambda e: e['date'] or '')

    advance_balance = 0
    for e in entries:
        if e['affects_advance']:
            advance_balance += e['debit'] - e['credit']
        e['running_advance_balance'] = advance_balance
    entries.reverse()

    total_salary_paid = sum(e['credit'] for e in entries if e['entry_type'] == 'Salary Paid')
    return {'entries': entries, 'advance_balance': advance_balance, 'total_salary_paid': total_salary_paid,
            'opening_balance': emp_row['opening_balance'] if emp_row else 0,
            'opening_balance_date': emp_row['opening_balance_date'] if emp_row else ''}

@app.route('/employee/<employee>', methods=['GET', 'POST'])
def employee_ledger(employee):
    import datetime
    conn = get_db()
    if request.method == 'POST':
        f = request.form
        from_tab = f.get('from_tab', 'overview')
        entry_kind = f.get('entry_kind')
        tx_date = f.get('date')
        amount = float(f.get('amount') or 0)
        now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        if entry_kind == 'salary':
            month_label = datetime.datetime.strptime(tx_date, '%Y-%m-%d').strftime('%b %Y') if tx_date else ''
            month_key = tx_date[:7] if tx_date else ''
            emp_row = conn.execute("SELECT id, basic_salary FROM employees WHERE name=? COLLATE NOCASE", (employee,)).fetchone()
            emp_id = emp_row['id'] if emp_row else None
            basic = emp_row['basic_salary'] if emp_row and emp_row['basic_salary'] else amount
            conn.execute("""INSERT INTO salaries
                             (employee, month, amount, date, created_at,
                              employee_id, month_key, basic_salary, gross_salary, total_deductions,
                              advance_recovery, net_salary, payment_status, payment_date, payment_mode, remarks)
                             VALUES (?,?,?,?,?, ?,?,?,?,0, 0,?,'paid',?,?,?)""",
                         (employee, month_label, amount, tx_date, now,
                          emp_id, month_key, basic, amount, amount, tx_date,
                          f.get('payment_mode') or '', f.get('notes') or ''))
        else:
            conn.execute("INSERT INTO advances (employee, date, amount, type, notes, created_at) VALUES (?,?,?,?,?,?)",
                         (employee, tx_date, amount, entry_kind, f.get('notes'), now))
        conn.commit()
        conn.close()
        return redirect(url_for('employee_ledger', employee=employee, from_tab=from_tab))

    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    from_tab = request.args.get('from_tab', 'overview')
    led = _employee_ledger_entries(conn, employee, date_from, date_to)
    conn.close()
    return render_template('employee_ledger.html', employee=employee, entries=led['entries'],
                            advance_balance=led['advance_balance'], total_salary_paid=led['total_salary_paid'],
                            opening_balance=led['opening_balance'],
                            opening_balance_date=led['opening_balance_date'],
                            f_date_from=date_from, f_date_to=date_to, from_tab=from_tab, active='salaries')

@app.route('/employee/<employee>/opening-balance', methods=['POST'])
def update_employee_opening_balance(employee):
    conn = get_db()
    f = request.form
    from_tab = f.get('from_tab', 'overview')
    conn.execute("UPDATE employees SET opening_balance=?, opening_balance_date=? WHERE name=? COLLATE NOCASE",
                 (float(f.get('opening_balance') or 0), f.get('opening_balance_date') or None, employee))
    conn.commit()
    conn.close()
    return redirect(url_for('employee_ledger', employee=employee, from_tab=from_tab))

@app.route('/driver-performance')
def driver_performance():
    return redirect(url_for('performance', **request.args.to_dict()))

def _trend_chart_coords(monthly, value_key):
    """Shared pixel-coordinate helper for the small SVG trend charts on the Performance page."""
    chart_w, chart_h, pad_l, pad_r, pad_t, pad_b = 700, 220, 34, 20, 20, 34
    plot_w, plot_h, n = chart_w - pad_l - pad_r, chart_h - pad_t - pad_b, len(monthly)
    chart_bottom = pad_t + plot_h
    vmax = max([m[value_key] for m in monthly], default=1) or 1
    for i, m in enumerate(monthly):
        m['x'] = round(pad_l + (i * plot_w / (n - 1) if n > 1 else 0), 1)
        m['y'] = round(pad_t + plot_h - (m[value_key] / vmax * plot_h), 1)
    y_ticks = []
    for k in range(5):
        tick_val = round(vmax * k / 4)
        y_ticks.append({'value': tick_val, 'y': round(pad_t + plot_h - (tick_val / vmax * plot_h), 1)})
    return chart_bottom, y_ticks

@app.route('/performance')
def performance():
    conn = get_db()
    date_from = request.args.get('date_from', '2026-04-01')
    date_to = request.args.get('date_to', '2026-07-31')

    # ---------- Driver Performance (same calculation as the original Driver Performance page) ----------
    # Toll excluded from driver-level cost — Toll Management (FASTag/BOSS) tracks cost per
    # vehicle/wallet, not per driver, so there's no real figure to attribute here (a shared
    # vehicle's toll can't be split by who happened to be driving that trip).
    driver_query = """SELECT driver_name,
        COUNT(*) as trip_count,
        SUM(billed_amount) as total_billed,
        SUM(fuel_amount) as total_fuel,
        SUM(COALESCE(driver_payment,0)+COALESCE(detention_charges,0)+COALESCE(other_expense,0)) as total_other_costs,
        MAX(date) as last_trip
        FROM trips WHERE driver_name IS NOT NULL AND driver_name != ''"""
    dparams = []
    if date_from:
        driver_query += " AND date >= ?"; dparams.append(date_from)
    if date_to:
        driver_query += " AND date <= ?"; dparams.append(date_to)
    driver_query += " GROUP BY driver_name ORDER BY total_billed DESC"
    draw = conn.execute(driver_query, dparams).fetchall()

    driver_rows = []
    for r in draw:
        profit = (r['total_billed'] or 0) - (r['total_fuel'] or 0) - (r['total_other_costs'] or 0)
        driver_rows.append({'name': r['driver_name'], 'trip_count': r['trip_count'],
                             'total_billed': r['total_billed'] or 0, 'total_fuel': r['total_fuel'] or 0,
                             'profit': profit, 'last_trip': r['last_trip']})

    total_drivers = len(driver_rows)
    driver_total_trips = sum(r['trip_count'] for r in driver_rows)
    driver_total_billed = sum(r['total_billed'] for r in driver_rows)
    driver_total_fuel = sum(r['total_fuel'] for r in driver_rows)
    driver_total_profit = sum(r['profit'] for r in driver_rows)
    driver_avg_trips = round(driver_total_trips / total_drivers, 1) if total_drivers else 0
    top5_drivers = sorted(driver_rows, key=lambda r: r['trip_count'], reverse=True)[:5]
    top5_drivers_max = max([d['trip_count'] for d in top5_drivers], default=1) or 1
    for d in top5_drivers:
        d['pct'] = round(d['trip_count'] / top5_drivers_max * 100, 1)

    dmonth_q = """SELECT substr(date,1,7) as month, COUNT(*) as trips, COALESCE(SUM(billed_amount),0) as billed
                  FROM trips WHERE driver_name IS NOT NULL AND driver_name != ''"""
    if date_from:
        dmonth_q += " AND date >= ?"
    if date_to:
        dmonth_q += " AND date <= ?"
    dmonth_q += " GROUP BY month ORDER BY month"
    driver_monthly = [{'label': m['month'], 'trips': m['trips'], 'billed': m['billed']}
                       for m in conn.execute(dmonth_q, dparams).fetchall()]
    driver_chart_bottom, driver_y_ticks = _trend_chart_coords(driver_monthly, 'trips')
    driver_trend_max = max([m['trips'] for m in driver_monthly], default=1) or 1

    # ---------- Vehicle Performance (new, mirrors the driver calculation style) ----------
    # Toll excluded here too — the per-vehicle maint_cost query below already pulls every
    # maintenance row for this vehicle (no category filter), which now includes Toll Management's
    # real cost; adding trips.toll on top of that would double-count the same rupee.
    vehicle_query = """SELECT v.id, v.vehicle_no, v.type,
        COUNT(t.id) as trip_count,
        COALESCE(SUM(t.billed_amount),0) as total_billed,
        COALESCE(SUM(t.fuel_amount),0) as total_fuel,
        COALESCE(SUM(COALESCE(t.driver_payment,0)+COALESCE(t.detention_charges,0)+COALESCE(t.other_expense,0)),0) as total_other_costs,
        MAX(t.date) as last_trip
        FROM trips t JOIN vehicles v ON t.vehicle_id=v.id WHERE v.type IN ('Line','Local')"""
    vparams = []
    if date_from:
        vehicle_query += " AND t.date >= ?"; vparams.append(date_from)
    if date_to:
        vehicle_query += " AND t.date <= ?"; vparams.append(date_to)
    vehicle_query += " GROUP BY v.id ORDER BY total_billed DESC"
    vraw = conn.execute(vehicle_query, vparams).fetchall()

    vehicle_rows = []
    for r in vraw:
        maint_q = "SELECT COALESCE(SUM(amount),0) FROM maintenance WHERE vehicle_id=?"
        maint_params = [r['id']]
        if date_from:
            maint_q += " AND date >= ?"; maint_params.append(date_from)
        if date_to:
            maint_q += " AND date <= ?"; maint_params.append(date_to)
        maint_cost = conn.execute(maint_q, maint_params).fetchone()[0]
        profit = (r['total_billed'] or 0) - (r['total_fuel'] or 0) - (r['total_other_costs'] or 0) - maint_cost
        vehicle_rows.append({'name': r['vehicle_no'], 'type': r['type'], 'trip_count': r['trip_count'],
                              'total_billed': r['total_billed'], 'total_fuel': r['total_fuel'],
                              'maint_cost': maint_cost, 'profit': profit, 'last_trip': r['last_trip']})

    total_vehicles = len(vehicle_rows)
    vehicle_total_trips = sum(r['trip_count'] for r in vehicle_rows)
    vehicle_total_billed = sum(r['total_billed'] for r in vehicle_rows)
    vehicle_total_fuel = sum(r['total_fuel'] for r in vehicle_rows)
    vehicle_total_maint = sum(r['maint_cost'] for r in vehicle_rows)
    vehicle_total_profit = sum(r['profit'] for r in vehicle_rows)
    vehicle_avg_trips = round(vehicle_total_trips / total_vehicles, 1) if total_vehicles else 0
    top5_vehicles = sorted(vehicle_rows, key=lambda r: r['trip_count'], reverse=True)[:5]
    top5_vehicles_max = max([v['trip_count'] for v in top5_vehicles], default=1) or 1
    for v in top5_vehicles:
        v['pct'] = round(v['trip_count'] / top5_vehicles_max * 100, 1)

    vmonth_q = """SELECT substr(t.date,1,7) as month, COUNT(*) as trips, COALESCE(SUM(t.billed_amount),0) as billed
                  FROM trips t JOIN vehicles v ON t.vehicle_id=v.id WHERE v.type IN ('Line','Local')"""
    if date_from:
        vmonth_q += " AND t.date >= ?"
    if date_to:
        vmonth_q += " AND t.date <= ?"
    vmonth_q += " GROUP BY month ORDER BY month"
    vehicle_monthly = [{'label': m['month'], 'trips': m['trips'], 'billed': m['billed']}
                        for m in conn.execute(vmonth_q, vparams).fetchall()]
    vehicle_chart_bottom, vehicle_y_ticks = _trend_chart_coords(vehicle_monthly, 'trips')
    vehicle_trend_max = max([m['trips'] for m in vehicle_monthly], default=1) or 1

    conn.close()

    # ---------- Pagination (client-independent, shared page-size convention) ----------
    per_page_raw = request.args.get('per_page')
    d_page, per_page, d_total_pages = _paginate(request.args.get('d_page'), per_page_raw, total_drivers,
                                                 per_page_options=(5, 10, 25, 50), default_per_page=5)
    driver_page_rows = driver_rows[(d_page - 1) * per_page: d_page * per_page]

    v_page, per_page, v_total_pages = _paginate(request.args.get('v_page'), per_page_raw, total_vehicles,
                                                 per_page_options=(5, 10, 25, 50), default_per_page=5)
    vehicle_page_rows = vehicle_rows[(v_page - 1) * per_page: v_page * per_page]
    d_page_tokens = _page_tokens(d_page, d_total_pages)
    v_page_tokens = _page_tokens(v_page, v_total_pages)

    from urllib.parse import urlencode
    base_params = request.args.to_dict()
    for key in ('d_page', 'v_page', 'per_page', 'tab'):
        base_params.pop(key, None)
    base_qs = urlencode(base_params)

    return render_template('performance.html', base_qs=base_qs,
        driver_rows=driver_page_rows, total_drivers=total_drivers, driver_total_trips=driver_total_trips,
        driver_total_billed=driver_total_billed, driver_total_fuel=driver_total_fuel, driver_total_profit=driver_total_profit,
        driver_avg_trips=driver_avg_trips, top5_drivers=top5_drivers,
        driver_monthly=driver_monthly, driver_trend_max=driver_trend_max, driver_chart_bottom=driver_chart_bottom, driver_y_ticks=driver_y_ticks,
        d_page=d_page, d_total_pages=d_total_pages, d_page_tokens=d_page_tokens,
        vehicle_rows=vehicle_page_rows, total_vehicles=total_vehicles, vehicle_total_trips=vehicle_total_trips,
        vehicle_total_billed=vehicle_total_billed, vehicle_total_fuel=vehicle_total_fuel, vehicle_total_maint=vehicle_total_maint,
        vehicle_total_profit=vehicle_total_profit, vehicle_avg_trips=vehicle_avg_trips, top5_vehicles=top5_vehicles,
        vehicle_monthly=vehicle_monthly, vehicle_trend_max=vehicle_trend_max, vehicle_chart_bottom=vehicle_chart_bottom, vehicle_y_ticks=vehicle_y_ticks,
        v_page=v_page, v_total_pages=v_total_pages, v_page_tokens=v_page_tokens,
        per_page=per_page, f_date_from=date_from, f_date_to=date_to, active='performance')

@app.route('/performance/export')
def export_performance():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from flask import send_file
    import io
    conn = get_db()
    date_from = request.args.get('date_from', '2026-04-01')
    date_to = request.args.get('date_to', '2026-07-31')

    driver_query = """SELECT driver_name, COUNT(*) as trip_count, SUM(billed_amount) as total_billed,
        SUM(fuel_amount) as total_fuel,
        SUM(COALESCE(driver_payment,0)+COALESCE(detention_charges,0)+COALESCE(other_expense,0)) as total_other_costs,
        MAX(date) as last_trip
        FROM trips WHERE driver_name IS NOT NULL AND driver_name != ''"""
    dparams = []
    if date_from:
        driver_query += " AND date >= ?"; dparams.append(date_from)
    if date_to:
        driver_query += " AND date <= ?"; dparams.append(date_to)
    driver_query += " GROUP BY driver_name ORDER BY total_billed DESC"
    driver_raw = conn.execute(driver_query, dparams).fetchall()

    # Toll excluded from both queries above/below — Toll Management can't be attributed per driver,
    # and the per-vehicle maint_cost query already carries it (unfiltered maintenance sum), so
    # adding trips.toll here would either be unattributable or double-counted.
    vehicle_query = """SELECT v.id, v.vehicle_no, v.type, COUNT(t.id) as trip_count,
        COALESCE(SUM(t.billed_amount),0) as total_billed, COALESCE(SUM(t.fuel_amount),0) as total_fuel,
        COALESCE(SUM(COALESCE(t.driver_payment,0)+COALESCE(t.detention_charges,0)+COALESCE(t.other_expense,0)),0) as total_other_costs,
        MAX(t.date) as last_trip
        FROM trips t JOIN vehicles v ON t.vehicle_id=v.id WHERE v.type IN ('Line','Local')"""
    vparams = []
    if date_from:
        vehicle_query += " AND t.date >= ?"; vparams.append(date_from)
    if date_to:
        vehicle_query += " AND t.date <= ?"; vparams.append(date_to)
    vehicle_query += " GROUP BY v.id ORDER BY total_billed DESC"
    vehicle_raw = conn.execute(vehicle_query, vparams).fetchall()

    vehicle_rows = []
    for r in vehicle_raw:
        maint_q = "SELECT COALESCE(SUM(amount),0) FROM maintenance WHERE vehicle_id=?"
        maint_params = [r['id']]
        if date_from:
            maint_q += " AND date >= ?"; maint_params.append(date_from)
        if date_to:
            maint_q += " AND date <= ?"; maint_params.append(date_to)
        maint_cost = conn.execute(maint_q, maint_params).fetchone()[0]
        profit = (r['total_billed'] or 0) - (r['total_fuel'] or 0) - (r['total_other_costs'] or 0) - maint_cost
        vehicle_rows.append((r['vehicle_no'], r['type'], r['trip_count'], r['total_billed'] or 0,
                              r['total_fuel'] or 0, maint_cost, profit, r['last_trip'] or ''))
    conn.close()

    navy = "1B2A4A"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=navy)

    wb = Workbook()
    ws = wb.active
    ws.title = "Drivers"
    headers = ["Driver", "Trips", "Total Billed", "Fuel Cost", "Other Costs", "Profit", "Last Trip"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h); c.font = header_font; c.fill = header_fill
    for r_idx, r in enumerate(driver_raw, 2):
        profit = (r['total_billed'] or 0) - (r['total_fuel'] or 0) - (r['total_other_costs'] or 0)
        row = (r['driver_name'], r['trip_count'], r['total_billed'] or 0, r['total_fuel'] or 0,
               r['total_other_costs'] or 0, profit, r['last_trip'] or '')
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    ws2 = wb.create_sheet("Vehicles")
    headers2 = ["Vehicle No", "Type", "Trips", "Total Billed", "Fuel Cost", "Maintenance Cost", "Profit", "Last Trip"]
    for i, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=i, value=h); c.font = header_font; c.fill = header_fill
    for r_idx, row in enumerate(vehicle_rows, 2):
        for c_idx, val in enumerate(row, 1):
            ws2.cell(row=r_idx, column=c_idx, value=val)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='performance_export.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/employees/add', methods=['GET', 'POST'])
def add_employee():
    conn = get_db()
    if request.method == 'POST':
        f = request.form
        name = (f.get('name') or '').strip()
        etype = f.get('type')
        existing = conn.execute("SELECT id FROM employees WHERE name=? COLLATE NOCASE", (name,)).fetchone()
        if not existing:
            cur = conn.execute("""INSERT INTO employees (name, type, employee_code, role, mobile, email, address,
                                  joining_date, date_of_birth, bank_account, ifsc_code, upi_id, emergency_contact,
                                  aadhaar, pan, driving_license, status, basic_salary)
                                  VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?, 'Active', ?)""",
                                (name, etype, f.get('employee_code') or None, f.get('role') or etype,
                                 f.get('mobile') or None, f.get('email') or None, f.get('address') or None,
                                 f.get('joining_date') or None, f.get('date_of_birth') or None,
                                 f.get('bank_account') or None, f.get('ifsc_code') or None, f.get('upi_id') or None,
                                 f.get('emergency_contact') or None, f.get('aadhaar') or None, f.get('pan') or None,
                                 f.get('driving_license') or None, float(f.get('basic_salary') or 0)))
            new_id = cur.lastrowid
            if not f.get('employee_code'):
                conn.execute("UPDATE employees SET employee_code=? WHERE id=?",
                             (f"{'DR' if etype=='Driver' else 'ST'}-{new_id:03d}", new_id))
            conn.commit()
        # 'next' lets the Add Employee modal (which lives on the tabbed Employees page) return the
        # admin to that page instead of the old standalone employee_ledger detail view.
        conn.close()
        if request.form.get('next') == 'employees':
            return redirect(url_for('salaries_list', tab='overview'))
        return redirect(url_for('employee_ledger', employee=name))
    conn.close()
    return render_template('add_employee.html', active='salaries')

# Standard GST state-code prefix (first 2 digits of a GSTIN) → state name. Public, fixed mapping —
# used only to derive a "State" display from a real GSTIN already on file, never fabricated data.
GST_STATE_CODES = {
    '01':'Jammu & Kashmir','02':'Himachal Pradesh','03':'Punjab','04':'Chandigarh','05':'Uttarakhand',
    '06':'Haryana','07':'Delhi','08':'Rajasthan','09':'Uttar Pradesh','10':'Bihar','11':'Sikkim',
    '12':'Arunachal Pradesh','13':'Nagaland','14':'Manipur','15':'Mizoram','16':'Tripura','17':'Meghalaya',
    '18':'Assam','19':'West Bengal','20':'Jharkhand','21':'Odisha','22':'Chhattisgarh','23':'Madhya Pradesh',
    '24':'Gujarat','25':'Daman & Diu','26':'Dadra & Nagar Haveli','27':'Maharashtra','28':'Andhra Pradesh (Old)',
    '29':'Karnataka','30':'Goa','31':'Lakshadweep','32':'Kerala','33':'Tamil Nadu','34':'Puducherry',
    '35':'Andaman & Nicobar','36':'Telangana','37':'Andhra Pradesh','38':'Ladakh',
}
def _gstin_state(gstin):
    if not gstin or len(gstin) < 2:
        return ''
    return GST_STATE_CODES.get(gstin[:2], '')

@app.route('/invoice-center')
def invoice_center():
    conn = get_db()
    invoice_type = request.args.get('invoice_type', 'party')
    party_id = request.args.get('party_id', '')
    vendor_id = request.args.get('vendor_id', '')
    vehicle_f = request.args.get('vehicle', '')
    lr_f = request.args.get('lr_number', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    status_f = request.args.get('invoice_status', '')

    parties = conn.execute("SELECT id, name, contact, gstin FROM parties ORDER BY name").fetchall()
    vendors = conn.execute("SELECT id, name, contact, linked_party_id, gstin FROM vendors ORDER BY name").fetchall()

    # Combined owner picker for Market Vehicle Invoices: every vendor, plus every party not already
    # linked to a vendor — so any organization can be picked as an owner regardless of which
    # table it currently lives in. A party-only pick has no vendor row yet, so resolve/create
    # one and redirect to a plain vendor_id before anything downstream touches it.
    linked_party_ids = {v['linked_party_id'] for v in vendors if v['linked_party_id']}
    owner_options = [{'ref': f"v:{v['id']}", 'name': v['name']} for v in vendors]
    owner_options += [{'ref': f"p:{p['id']}", 'name': p['name']} for p in parties if p['id'] not in linked_party_ids]
    owner_options.sort(key=lambda o: o['name'].lower())

    if invoice_type == 'vehicle_owner' and vendor_id.startswith('p:'):
        p_row = conn.execute("SELECT name FROM parties WHERE id=?", (vendor_id[2:],)).fetchone()
        resolved_id = get_or_create_vendor(conn, p_row['name']) if p_row else None
        conn.commit()
        conn.close()
        return redirect(url_for('invoice_center', invoice_type=invoice_type, vendor_id=resolved_id or ''))
    elif vendor_id.startswith('v:'):
        vendor_id = vendor_id[2:]

    selected_party = None
    selected_vendor = None
    trips = []
    # Which trips already sit in some previously-generated invoice batch — real, derived from
    # invoice_batch_trips, used for the Invoice Status filter/badge (not a stored flag on trips).
    invoiced_trip_ids = {r['trip_id'] for r in conn.execute("SELECT DISTINCT trip_id FROM invoice_batch_trips").fetchall()}

    if invoice_type in ('party', 'tax', 'bill') and party_id:
        selected_party = conn.execute("SELECT * FROM parties WHERE id=?", (party_id,)).fetchone()
        query = """SELECT t.id, t.lr_number, v.vehicle_no, t.from_loc, t.to_loc, t.date, t.billed_amount
                   FROM trips t LEFT JOIN vehicles v ON t.vehicle_id=v.id
                   WHERE t.party_id=?"""
        params = [party_id]
        if vehicle_f:
            query += " AND v.vehicle_no=?"; params.append(vehicle_f)
        if lr_f:
            query += " AND t.lr_number LIKE ?"; params.append(f"%{lr_f}%")
        if date_from:
            query += " AND t.date>=?"; params.append(date_from)
        if date_to:
            query += " AND t.date<=?"; params.append(date_to)
        query += " ORDER BY t.date DESC LIMIT 200"
        trips = conn.execute(query, params).fetchall()
    elif invoice_type == 'vehicle_owner' and vendor_id:
        selected_vendor = conn.execute("SELECT * FROM vendors WHERE id=?", (vendor_id,)).fetchone()
        query = """SELECT t.id, t.lr_number, v.vehicle_no, t.from_loc, t.to_loc, t.date,
                   CASE WHEN t.owner_rate_type='FIXED' THEN t.owner_fixed_amount ELSE t.owner_rate*t.quantity END as billed_amount
                   FROM trips t LEFT JOIN vehicles v ON t.vehicle_id=v.id
                   WHERE t.owner_vendor_id=?"""
        params = [vendor_id]
        if vehicle_f:
            query += " AND v.vehicle_no=?"; params.append(vehicle_f)
        if lr_f:
            query += " AND t.lr_number LIKE ?"; params.append(f"%{lr_f}%")
        if date_from:
            query += " AND t.date>=?"; params.append(date_from)
        if date_to:
            query += " AND t.date<=?"; params.append(date_to)
        query += " ORDER BY t.date DESC LIMIT 200"
        trips = conn.execute(query, params).fetchall()

    # Attach each trip's own real "Others" line items (charges/deductions logged on the trip
    # itself) so they can be shown — and individually excluded — while building this invoice.
    trip_rows = []
    for t in trips:
        is_invoiced = t['id'] in invoiced_trip_ids
        if status_f == 'invoiced' and not is_invoiced:
            continue
        if status_f == 'not_invoiced' and is_invoiced:
            continue
        items = conn.execute("SELECT id, description, amount, item_type FROM invoice_items WHERE trip_id=?", (t['id'],)).fetchall()
        trip_rows.append({'t': t, 'is_invoiced': is_invoiced, 'trip_items': [dict(i) for i in items]})

    entity_gstin = (selected_vendor['gstin'] if selected_vendor else (selected_party['gstin'] if selected_party else '')) or ''
    entity_state = _gstin_state(entity_gstin)
    invoice_settings = _get_invoice_settings(conn)

    vehicles = conn.execute("SELECT DISTINCT vehicle_no FROM vehicles WHERE vehicle_no IS NOT NULL ORDER BY vehicle_no").fetchall()
    conn.close()
    import datetime as _dt
    today_str = _dt.date.today().strftime('%Y-%m-%d')
    return render_template('invoice_center.html', invoice_type=invoice_type, parties=parties, vendors=vendors,
                            owner_options=owner_options,
                            selected_party=selected_party, selected_vendor=selected_vendor, trip_rows=trip_rows,
                            vehicles=vehicles, entity_gstin=entity_gstin, entity_state=entity_state,
                            invoice_settings=invoice_settings, today=today_str,
                            f_party_id=party_id, f_vendor_id=vendor_id, f_vehicle=vehicle_f, f_lr=lr_f,
                            f_date_from=date_from, f_date_to=date_to, f_status=status_f, active='invoices')

@app.route('/invoice-center/review', methods=['POST'])
def invoice_center_review():
    trip_ids = request.form.getlist('trip_ids')
    invoice_type = request.form.get('invoice_type')
    party_id = request.form.get('party_id') or None
    vendor_id = request.form.get('vendor_id') or None

    conn = get_db()
    placeholders = ','.join('?' * len(trip_ids))
    trips = conn.execute(f"""SELECT t.*, v.vehicle_no FROM trips t LEFT JOIN vehicles v ON t.vehicle_id=v.id
                             WHERE t.id IN ({placeholders})""", trip_ids).fetchall() if trip_ids else []

    entity = None
    if invoice_type == 'vehicle_owner' and vendor_id:
        entity = conn.execute("SELECT * FROM vendors WHERE id=?", (vendor_id,)).fetchone()
    elif party_id:
        entity = conn.execute("SELECT * FROM parties WHERE id=?", (party_id,)).fetchone()

    bank_name = conn.execute("SELECT value FROM settings WHERE key='bank_name'").fetchone()['value']
    account_number = conn.execute("SELECT value FROM settings WHERE key='account_number'").fetchone()['value']
    ifsc_code = conn.execute("SELECT value FROM settings WHERE key='ifsc_code'").fetchone()['value']
    account_holder = conn.execute("SELECT value FROM settings WHERE key='account_holder'").fetchone()['value']
    rcm_clause = conn.execute("SELECT value FROM settings WHERE key='rcm_clause'").fetchone()['value']
    conn.close()

    line_items = []
    for t in trips:
        if invoice_type == 'vehicle_owner':
            freight = t['owner_fixed_amount'] if (t['owner_rate_type'] or 'PER_MT')=='FIXED' else (t['owner_rate'] or 0) * (t['quantity'] or 0)
            already_given = (t['paid_to_owner'] or 0) + (t['fuel_amount'] or 0) + (t['driver_adv_amount'] or 0)
            net = freight - already_given
        else:
            freight = t['fixed_rate_amount'] if t['rate_type']=='FIXED' else (t['quantity'] or 0) * (t['rate'] or 0)
            net = t['billed_amount'] or 0
        line_items.append({'trip': t, 'freight': freight, 'net': net})

    total_freight = sum(li['freight'] for li in line_items)
    total_trips = len(line_items)

    import datetime
    today = datetime.datetime.now().strftime('%Y-%m-%d')
    return render_template('invoice_review.html', line_items=line_items, entity=entity, invoice_type=invoice_type,
                            party_id=party_id, vendor_id=vendor_id, total_freight=total_freight, total_trips=total_trips,
                            bank_name=bank_name, account_number=account_number, ifsc_code=ifsc_code,
                            account_holder=account_holder, rcm_clause=rcm_clause, today=today, active='invoices')

def number_to_words_inr(n):
    n = int(round(n))
    ones = ['', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven', 'Eight', 'Nine', 'Ten',
            'Eleven', 'Twelve', 'Thirteen', 'Fourteen', 'Fifteen', 'Sixteen', 'Seventeen', 'Eighteen', 'Nineteen']
    tens = ['', '', 'Twenty', 'Thirty', 'Forty', 'Fifty', 'Sixty', 'Seventy', 'Eighty', 'Ninety']
    def two_digit(num):
        if num < 20:
            return ones[num]
        return tens[num // 10] + (' ' + ones[num % 10] if num % 10 else '')
    def three_digit(num):
        if num >= 100:
            return ones[num // 100] + ' Hundred' + (' ' + two_digit(num % 100) if num % 100 else '')
        return two_digit(num)
    if n == 0:
        return 'Zero'
    parts = []
    crore = n // 10000000; n %= 10000000
    lakh = n // 100000; n %= 100000
    thousand = n // 1000; n %= 1000
    hundred = n
    if crore: parts.append(two_digit(crore) + ' Crore')
    if lakh: parts.append(two_digit(lakh) + ' Lakh')
    if thousand: parts.append(two_digit(thousand) + ' Thousand')
    if hundred: parts.append(three_digit(hundred))
    return ' '.join(parts)

INVOICE_SETTINGS_KEYS = ['company_name', 'address', 'gstin', 'pan', 'phone', 'email', 'bank_name', 'account_holder',
                         'account_number', 'ifsc_code', 'branch', 'rcm_clause', 'cgst_rate', 'sgst_rate',
                         'invoice_prefix', 'next_invoice_number', 'logo_path']
LOGO_UPLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'uploads', 'logo')
os.makedirs(LOGO_UPLOAD_DIR, exist_ok=True)

def _get_invoice_settings(conn):
    s = {}
    for key in INVOICE_SETTINGS_KEYS:
        row = conn.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
        s[key] = row['value'] if row else ''
    return s

@app.route('/settings/logo/upload', methods=['POST'])
def upload_company_logo():
    """Company logo shown on every generated invoice PDF — uploaded once here, reused
    everywhere _build_invoice_pdf runs. Same save-and-record-a-relative-path pattern as
    Insurance's document uploads."""
    f = request.files.get('logo')
    conn = get_db()
    if f and f.filename:
        filename = secure_filename(f.filename)
        ext = os.path.splitext(filename)[1].lower()
        if ext in ('.png', '.jpg', '.jpeg', '.gif', '.webp'):
            unique = f"logo_{int(datetime.datetime.now().timestamp() * 1000)}{ext}"
            f.save(os.path.join(LOGO_UPLOAD_DIR, unique))
            rel_path = f"uploads/logo/{unique}"
            conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES ('logo_path', ?)", (rel_path,))
            conn.commit()
    conn.close()
    return redirect(url_for('settings_page', tab='company'))

def _build_invoice_pdf(trips, invoice_type, entity, s, invoice_number, invoice_date, due_date, payment_status, remarks,
                        cgst_rate=0, sgst_rate=0, extra_loading=0, extra_other=0, tds_rate=0, extra_items=None, toll_map=None):
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.utils import ImageReader
    import io, os

    # Per-trip freight and charge breakdown
    line_items = []
    for t in trips:
        if invoice_type == 'vehicle_owner':
            freight = t['owner_fixed_amount'] if (t['owner_rate_type'] or 'PER_MT')=='FIXED' else (t['owner_rate'] or 0) * (t['quantity'] or 0)
        else:
            freight = t['fixed_rate_amount'] if t['rate_type']=='FIXED' else (t['quantity'] or 0) * (t['rate'] or 0)
        line_items.append({
            'trip': t, 'freight': freight,
            'loading': t['loading_charge'] or 0, 'unloading': t['unloading_charge'] or 0,
            # Real Toll Management amount for this trip if it has one linked, else the trip's own
            # manual estimate — never both, so the same real toll isn't billed twice.
            'permit': t['permit_charges'] or 0, 'toll': _trip_toll(t, toll_map or {}),
            'weighment': t['weight_charges'] or 0, 'driver_bata': t['driver_payment'] or 0,
            'gps': t['gps_cost'] or 0, 'other': t['other_charges'] or 0,
            'fuel_given': t['fuel_amount'] or 0, 'driver_adv_given': t['driver_adv_amount'] or 0,
            'paid_to_owner': t['paid_to_owner'] or 0,
        })

    total_freight = sum(li['freight'] for li in line_items)
    total_loading = sum(li['loading'] for li in line_items) + extra_loading
    total_unloading = sum(li['unloading'] for li in line_items)
    total_permit = sum(li['permit'] for li in line_items)
    total_toll = sum(li['toll'] for li in line_items)
    total_weighment = sum(li['weighment'] for li in line_items)
    total_driver_bata = sum(li['driver_bata'] for li in line_items)
    total_gps = sum(li['gps'] for li in line_items)
    total_other = sum(li['other'] for li in line_items) + extra_other

    # Named ad-hoc charges/deductions (per-trip "Others" items + Invoice Center extra rows), kept
    # separate from the anonymous Other Charges total above so each one can be printed on the
    # invoice by its own description (e.g. "Detention") rather than disappearing into one lump sum.
    extra_items = extra_items or []
    named_extra_total = sum((it['amount'] or 0) if it['item_type'] == 'charge' else -(it['amount'] or 0) for it in extra_items)

    freight_and_charges = total_freight + total_loading + total_unloading + total_permit + total_toll
    additional_charges = total_weighment + total_driver_bata + total_gps + total_other + named_extra_total
    sub_total = freight_and_charges + additional_charges

    cgst_amount = round(sub_total * cgst_rate / 100, 2)
    sgst_amount = round(sub_total * sgst_rate / 100, 2)
    # TDS was previously recorded on every invoice batch (tds_rate) but never actually applied to
    # the printed total — this makes it real. Gated on tds_rate>0 (and passed 0 by any caller that
    # doesn't set it) so an un-set TDS rate changes nothing about the total or the layout below.
    tds_amount = round(sub_total * tds_rate / 100, 2) if tds_rate else 0
    pre_round = sub_total + cgst_amount + sgst_amount - tds_amount
    grand_total = round(pre_round)
    round_off = round(grand_total - pre_round, 2)

    total_fuel_given = sum(li['fuel_given'] for li in line_items)
    total_driver_adv_given = sum(li['driver_adv_given'] for li in line_items)
    total_paid_to_owner = sum(li['paid_to_owner'] for li in line_items)
    total_already_given = total_fuel_given + total_driver_adv_given + total_paid_to_owner
    net_payable = grand_total - total_already_given if invoice_type == 'vehicle_owner' else grand_total

    total_weight = sum(li['trip']['quantity'] or 0 for li in line_items)
    total_vehicles = len(set(li['trip']['vehicle_no'] for li in line_items if li['trip']['vehicle_no']))
    dates = [li['trip']['date'] for li in line_items if li['trip']['date']]
    trip_period = f"{min(dates)} to {max(dates)}" if dates else ''

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=0.4*inch, bottomMargin=0.4*inch, leftMargin=0.45*inch, rightMargin=0.45*inch)
    # Every full-width element below (header, item table, totals box, footer...) is sized off
    # this one constant so their left/right edges always line up with each other — no more
    # tables that are each "approximately" full width but a few points off from their neighbours.
    CONTENT_W = 6.9 * inch
    # The Additional Charges box and the Sub Total/Grand Total box sit side by side and share this
    # same gap treatment as Bill To / Trip Details, so the two never read as one fused block.
    COMBO_GAP = 0.15 * inch
    COMBO_L = 4.2 * inch - COMBO_GAP / 2
    COMBO_R = 2.7 * inch - COMBO_GAP / 2
    styles = getSampleStyleSheet()
    BLACK = colors.HexColor('#1A1A1A')
    GREY = colors.HexColor('#5A5A5A')
    LINE = colors.HexColor('#333333')
    LIGHTBG = colors.HexColor('#EFF3FA')
    company_style = ParagraphStyle('C', parent=styles['Title'], fontSize=16, textColor=BLACK, alignment=0, leading=19)
    tagline_style = ParagraphStyle('TL', parent=styles['Normal'], fontSize=7.5, textColor=GREY, leading=10)
    sub_style = ParagraphStyle('S', parent=styles['Normal'], fontSize=8, textColor=GREY, leading=11)
    label_style = ParagraphStyle('L', parent=styles['Normal'], fontSize=8.5, textColor=GREY, leading=12)
    title_box_style = ParagraphStyle('TB', parent=styles['Heading2'], fontSize=14, textColor=BLACK, alignment=1)
    section_head_style = ParagraphStyle('SH', parent=styles['Normal'], fontSize=8.5, textColor=BLACK, fontName='Helvetica-Bold')
    # Table-cell text that can genuinely run long in real data (route names, material
    # descriptions, LR numbers) — wrapped in Paragraph so it wraps inside its column instead of
    # overflowing into the next cell, which is the exact "long text overriding other text" bug
    # plain strings in a reportlab Table cell cause.
    cell_style = ParagraphStyle('CE', parent=styles['Normal'], fontSize=7.3, textColor=BLACK, leading=9)
    cell_label_style = ParagraphStyle('CL', parent=styles['Normal'], fontSize=8.3, textColor=GREY, leading=10.5)
    cell_val_style = ParagraphStyle('CV', parent=styles['Normal'], fontSize=8.3, textColor=BLACK, leading=10.5)
    def cell(text):
        """Wrap any table-cell value that might be long real data (route names, material,
        LR numbers) so ReportLab wraps it instead of letting it spill into the next column."""
        return Paragraph(esc(text) if text not in (None, '') else '—', cell_style)

    def esc(text):
        """Escape real user-entered text (names, addresses, remarks) before it goes inside a
        Paragraph's mini-XML markup — an unescaped '&' or '<' in a real company/party name
        (e.g. "R&B Transport") would otherwise break the PDF build or render garbled."""
        from xml.sax.saxutils import escape
        return escape(str(text)) if text else ''

    invoice_titles = {'party': 'INVOICE', 'vehicle_owner': 'FREIGHT BILL', 'tax': 'INVOICE', 'bill': 'INVOICE'}
    is_single = len(line_items) == 1

    def section_box(title, rows_data, col_widths, header_bg=LIGHTBG):
        head = Table([[Paragraph(f"<b>{title}</b>", section_head_style)]], colWidths=[sum(col_widths)])
        head.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,-1), header_bg), ('BOX', (0,0), (-1,-1), 0.7, LINE),
                                    ('TOPPADDING', (0,0), (-1,-1), 5), ('BOTTOMPADDING', (0,0), (-1,-1), 5), ('LEFTPADDING', (0,0), (-1,-1), 8)]))
        body = Table(rows_data, colWidths=col_widths)
        body.setStyle(TableStyle([
            ('BOX', (0,0), (-1,-1), 0.7, LINE), ('LINEBELOW', (0,0), (-1,-2), 0.3, colors.HexColor('#DDDDDD')),
            ('FONTSIZE', (0,0), (-1,-1), 8.5), ('TOPPADDING', (0,0), (-1,-1), 5), ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LEFTPADDING', (0,0), (-1,-1), 8),
        ]))
        return [head, body]

    # ---------- Header ----------
    company_block = [
        Paragraph(f"<b>{esc(s['company_name'])}</b>", company_style),
        Paragraph("SAFE | RELIABLE | ON TIME", tagline_style),
        Spacer(1, 3),
        Paragraph(esc(s['address']), sub_style),
    ]
    if invoice_type == 'tax':
        company_block.append(Paragraph(f"GSTIN: {esc(s['gstin'])} &nbsp;&nbsp; PAN: {esc(s['pan'])}", sub_style))
    company_block.append(Paragraph(f"&#9742; {esc(s['phone'])} &nbsp;&nbsp; &#9993; {esc(s['email'])}", sub_style))

    # Company logo, if one's been uploaded (Settings > Company Profile) — sized modestly and
    # placed beside the company name, never dominating the header. Falls back silently to the
    # text-only header if no logo is set, or the stored file has gone missing on disk.
    logo_path = s.get('logo_path')
    logo_full_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', logo_path) if logo_path else None
    if logo_full_path and os.path.isfile(logo_full_path):
        try:
            # Fully decode the pixel data now (not just read the header/dimensions) so a
            # truncated or corrupt upload is caught here and falls back cleanly, rather than
            # crashing later inside doc.build() once it's too late to recover gracefully.
            from PIL import Image as PILImage
            with PILImage.open(logo_full_path) as pil_img:
                pil_img.load()
                iw, ih = pil_img.size
            max_h = 0.45 * inch
            logo_img = Image(logo_full_path, width=max_h * iw / ih, height=max_h)
            header_left_w = CONTENT_W - 2.4*inch
            header_left = Table([[logo_img, company_block]], colWidths=[max_h * iw / ih + 8, header_left_w - (max_h * iw / ih + 8)])
            header_left.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP'), ('LEFTPADDING', (0, 0), (0, 0), 0),
                                              ('RIGHTPADDING', (0, 0), (0, 0), 8), ('LEFTPADDING', (1, 0), (1, 0), 0)]))
        except Exception:
            header_left = company_block  # a corrupt/unreadable logo file never blocks invoice generation
    else:
        header_left = company_block

    # Big bold open title (no heavy box), matching the reference's "INVOICE" treatment — a thin
    # rule under the title separates it from the metadata rather than a full border box.
    big_title_style = ParagraphStyle('BT', parent=styles['Heading1'], fontSize=18, textColor=BLACK,
                                      alignment=2, leading=20, fontName='Helvetica-Bold', spaceAfter=0)
    meta_label_style = ParagraphStyle('ML', parent=styles['Normal'], fontSize=8, textColor=GREY, alignment=2, leading=11)
    meta_val_style = ParagraphStyle('MV', parent=styles['Normal'], fontSize=9.5, textColor=BLACK, alignment=2,
                                     leading=12, fontName='Helvetica-Bold')

    status_color = colors.HexColor('#2E7D32') if payment_status == 'PAID' else colors.HexColor('#B8860B')
    status_bg = colors.HexColor('#E8F5E9') if payment_status == 'PAID' else colors.HexColor('#FFF8E1')
    status_table = Table([[payment_status]], colWidths=[1.1*inch])
    status_table.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,-1), status_bg), ('TEXTCOLOR', (0,0), (-1,-1), status_color),
                                       ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'), ('FONTSIZE', (0,0), (-1,-1), 8.5),
                                       ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('TOPPADDING', (0,0), (-1,-1), 4), ('BOTTOMPADDING', (0,0), (-1,-1), 4)]))

    invoice_box_rows = [
        [Paragraph(invoice_titles.get(invoice_type, 'INVOICE'), big_title_style)],
        [Spacer(1, 6)],
        [Paragraph('INVOICE NO.', meta_label_style)],
        [Paragraph(esc(invoice_number), meta_val_style)],
        [Paragraph(f"Date: {esc(invoice_date)}", meta_label_style)],
        [Spacer(1, 4)],
        [status_table],
    ]
    invoice_box = Table(invoice_box_rows, colWidths=[2.4*inch])
    invoice_box.setStyle(TableStyle([
        ('LINEBELOW', (0,0), (-1,0), 1.4, BLACK), ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
        ('TOPPADDING', (0,0), (-1,-1), 1), ('BOTTOMPADDING', (0,0), (-1,-1), 1),
    ]))

    header_table = Table([[header_left, invoice_box]], colWidths=[CONTENT_W - 2.4*inch, 2.4*inch])
    header_table.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP')]))
    story = [header_table, Spacer(1, 10)]

    # ---------- Bill To + Trip Details/Summary ----------
    bill_name_lines = [f"<b>{esc(entity['name'])}</b>" if entity else '—']
    if entity and entity['address']:
        bill_name_lines.append(esc(entity['address']))
    # A vehicle owner bill runs the opposite direction of every other invoice type — the owner is
    # who WE owe money to, not who owes us — so "BILL TO" (implying they're being asked to pay)
    # is backwards here. "PAY TO" makes the direction of money correct at a glance.
    bill_box_label = 'PAYABLE TO' if invoice_type == 'vehicle_owner' else 'BILL TO'
    def make_bill_box(width):
        body = Table([[Paragraph('<br/>'.join(bill_name_lines), cell_val_style)]], colWidths=[width - 0.2*inch])
        body.setStyle(TableStyle([('TOPPADDING',(0,0),(-1,-1),4), ('BOTTOMPADDING',(0,0),(-1,-1),4)]))
        return section_box(bill_box_label, [[body]], [width])

    if is_single:
        t = line_items[0]['trip']
        col_gap = 0.15 * inch  # visible breathing room between Bill To and Trip Details, not flush-joined
        half_w = (CONTENT_W - col_gap) / 2
        trip_rows = [
            ['LR Number', ':', cell(t['lr_number'])], ['Trip Date', ':', t['date'] or ''],
            ['Vehicle No.', ':', cell(t['vehicle_no'])],
            ['From', ':', cell(t['from_loc'])], ['To', ':', cell(t['to_loc'])],
            ['Material', ':', cell(t['material'])], ['Weight', ':', f"{t['quantity'] or 0} MT"],
        ]
        trip_body = Table(trip_rows, colWidths=[1.1*inch, 0.15*inch, half_w - 1.25*inch - 0.2*inch])
        trip_body.setStyle(TableStyle([('FONTSIZE',(0,0),(-1,-1),8.3), ('TOPPADDING',(0,0),(-1,-1),2.5), ('BOTTOMPADDING',(0,0),(-1,-1),2.5),
                                        ('TEXTCOLOR',(0,0),(0,-1),GREY)]))
        trip_box = section_box('TRIP DETAILS', [[trip_body]], [half_w])
        two_col = Table([[make_bill_box(half_w), '', trip_box]], colWidths=[half_w, col_gap, half_w])
        two_col.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'TOP'), ('LEFTPADDING',(1,0),(1,-1),0), ('RIGHTPADDING',(1,0),(1,-1),0)]))
        story.append(two_col)
    else:
        summary_rows = [[
            Paragraph(f"Trip Period<br/><b>{trip_period}</b>", label_style),
            Paragraph(f"Total Trips<br/><b>{len(line_items)}</b>", label_style),
            Paragraph(f"Total Vehicles<br/><b>{total_vehicles}</b>", label_style),
            Paragraph(f"Total Weight<br/><b>{total_weight:.3f} MT</b>", label_style),
        ]]
        summary_body = Table(summary_rows, colWidths=[CONTENT_W/4]*4)
        summary_body.setStyle(TableStyle([('TOPPADDING',(0,0),(-1,-1),4), ('BOTTOMPADDING',(0,0),(-1,-1),4)]))
        summary_box = section_box('TRIP SUMMARY', [[summary_body]], [CONTENT_W])
        story.extend(make_bill_box(CONTENT_W))
        story.append(Spacer(1, 8))
        story.extend(summary_box)
    story.append(Spacer(1, 12))

    # ---------- Freight Details (single) / Trips table (multi) ----------
    if is_single:
        li = line_items[0]
        t = li['trip']
        # Owner invoices ("vehicle_owner") are billed against the vehicle owner's own Charge Type
        # (owner_rate_type/owner_fixed_amount), which is tracked independently of the party's —
        # everything else (party invoices, tax invoices) uses the party's rate_type as before.
        if invoice_type == 'vehicle_owner':
            is_fixed = (t['owner_rate_type'] or 'PER_MT') == 'FIXED'
            rate_basis = 'Per Trip' if is_fixed else f"{t['quantity']} MT"
            rate_val = t['owner_fixed_amount'] if is_fixed else t['owner_rate']
        else:
            is_fixed = t['rate_type'] == 'FIXED'
            rate_basis = 'Per Trip' if is_fixed else f"{t['quantity']} MT"
            rate_val = t['fixed_rate_amount'] if is_fixed else t['rate']
        freight_desc = 'Freight Charges' + (' (Fixed)' if is_fixed else '')
        freight_rows = [['#', 'DESCRIPTION', 'RATE (Rs.)', 'QTY / BASIS', 'AMOUNT (Rs.)'],
                         ['1', freight_desc, f"{rate_val or 0:,.2f}", rate_basis, f"{li['freight']:,.2f}"]]
        n = 2
        for label, val in [('Loading Charges', li['loading']), ('Unloading Charges', li['unloading']),
                            ('Permit Charges', li['permit']), ('Toll Charges (Estimate)', li['toll'])]:
            if val:
                freight_rows.append([str(n), label, f"{val:,.2f}", 'Per Trip', f"{val:,.2f}"])
                n += 1
        ft = Table(freight_rows, colWidths=[0.3*inch, 2.5*inch, 1.3*inch, 1.5*inch, 1.3*inch], repeatRows=1)
        ft.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), BLACK), ('TEXTCOLOR', (0,0), (-1,0), colors.white), ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8.3), ('ALIGN', (2,0), (4,-1), 'RIGHT'), ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#DDDDDD')),
            ('TOPPADDING', (0,0), (-1,-1), 4.5), ('BOTTOMPADDING', (0,0), (-1,-1), 4.5),
        ]))
        story.append(ft)
        tot_row = Table([['TOTAL FREIGHT & CHARGES', f"Rs. {freight_and_charges:,.2f}"]], colWidths=[5.6*inch, 1.3*inch])
        tot_row.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),LIGHTBG), ('FONTNAME',(0,0),(-1,-1),'Helvetica-Bold'),
                                      ('FONTSIZE',(0,0),(-1,-1),9), ('ALIGN',(1,0),(1,-1),'RIGHT'), ('BOX',(0,0),(-1,-1),0.5,LINE),
                                      ('TOPPADDING',(0,0),(-1,-1),5), ('BOTTOMPADDING',(0,0),(-1,-1),5)]))
        story.append(tot_row)
        story.append(Spacer(1, 12))

        addl_rows = [['#', 'DESCRIPTION', 'BASIS', 'AMOUNT (Rs.)']]
        n = 1
        for label, val in [('Weighment Charges', li['weighment']), ('Driver Bata', li['driver_bata']),
                            ('GPS Charges', li['gps']), ('Other Charges', li['other'])]:
            if val:
                addl_rows.append([str(n), label, 'Per Trip', f"{val:,.2f}"])
                n += 1
        for it in extra_items:
            addl_rows.append([str(n), cell(it['desc']), 'Deduction' if it['item_type'] == 'deduction' else 'Charge',
                               f"- {it['amount']:,.2f}" if it['item_type'] == 'deduction' else f"{it['amount']:,.2f}"])
            n += 1
        if len(addl_rows) > 1:
            addl_left = Table(addl_rows, colWidths=[0.3*inch, 1.9*inch, 1*inch, COMBO_L - 3.2*inch])
            addl_left.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,0), colors.HexColor('#EEEEEE')), ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
                ('FONTSIZE',(0,0),(-1,-1),8.3), ('ALIGN',(3,0),(3,-1),'RIGHT'), ('GRID',(0,0),(-1,-1),0.3,colors.HexColor('#DDDDDD')),
                ('TOPPADDING',(0,0),(-1,-1),4), ('BOTTOMPADDING',(0,0),(-1,-1),4),
            ]))
            addl_box_head = Table([[Paragraph('<b>ADDITIONAL CHARGES</b>', section_head_style)]], colWidths=[COMBO_L])
            addl_box_head.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),LIGHTBG), ('BOX',(0,0),(-1,-1),0.5,LINE),
                                                ('TOPPADDING',(0,0),(-1,-1),5), ('BOTTOMPADDING',(0,0),(-1,-1),5), ('LEFTPADDING',(0,0),(-1,-1),8)]))
            addl_tot = Table([['TOTAL ADDITIONAL CHARGES', f"Rs. {additional_charges:,.2f}"]], colWidths=[COMBO_L - 1*inch, 1*inch])
            addl_tot.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),LIGHTBG), ('FONTNAME',(0,0),(-1,-1),'Helvetica-Bold'),
                                           ('FONTSIZE',(0,0),(-1,-1),8.5), ('ALIGN',(1,0),(1,-1),'RIGHT'), ('BOX',(0,0),(-1,-1),0.5,LINE),
                                           ('TOPPADDING',(0,0),(-1,-1),4), ('BOTTOMPADDING',(0,0),(-1,-1),4)]))

            sub_total_lines = [['Sub Total', f"Rs. {sub_total:,.2f}"]]
            if invoice_type == 'tax':
                sub_total_lines.append([f'CGST ({cgst_rate:g}%)', f"Rs. {cgst_amount:,.2f}"])
                sub_total_lines.append([f'SGST ({sgst_rate:g}%)', f"Rs. {sgst_amount:,.2f}"])
            if tds_rate:
                sub_total_lines.append([f'TDS ({tds_rate:g}%)', f"- Rs. {tds_amount:,.2f}"])
            if invoice_type == 'tax' or tds_rate:
                sub_total_lines.append(['Round Off', f"Rs. {round_off:,.2f}"])
            sub_total_lines.append(['GRAND TOTAL (Rs.)', f"{grand_total:,.2f}"])
            # One bordered box with every row (Sub Total through Grand Total) instead of a plain
            # floating list plus a separately-boxed final row — same "one continuous box, distinct
            # rows" pattern as TOTAL FREIGHT & CHARGES above it.
            right_stack = Table(sub_total_lines, colWidths=[COMBO_R - 1.1*inch, 1.1*inch])
            right_stack.setStyle(TableStyle([
                ('FONTSIZE',(0,0),(-1,-1),8.5), ('ALIGN',(1,0),(1,-1),'RIGHT'),
                ('TOPPADDING',(0,0),(-1,-1),4), ('BOTTOMPADDING',(0,0),(-1,-1),4), ('LEFTPADDING',(0,0),(-1,-1),8),
                ('BOX',(0,0),(-1,-1),0.5,LINE), ('LINEBELOW',(0,0),(-1,-2),0.3,colors.HexColor('#DDDDDD')),
                ('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold'), ('FONTSIZE',(0,-1),(-1,-1),10.5), ('BACKGROUND',(0,-1),(-1,-1),LIGHTBG),
            ]))

            left_col = [addl_box_head, addl_left, addl_tot]
            combo = Table([[left_col, '', right_stack]], colWidths=[COMBO_L, COMBO_GAP, COMBO_R])
            combo.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'TOP'), ('LEFTPADDING',(1,0),(1,-1),0), ('RIGHTPADDING',(1,0),(1,-1),0)]))
            story.append(combo)
        else:
            sub_total_lines = [['Sub Total', f"Rs. {sub_total:,.2f}"]]
            if invoice_type == 'tax':
                sub_total_lines.append([f'CGST ({cgst_rate:g}%)', f"Rs. {cgst_amount:,.2f}"])
                sub_total_lines.append([f'SGST ({sgst_rate:g}%)', f"Rs. {sgst_amount:,.2f}"])
            if tds_rate:
                sub_total_lines.append([f'TDS ({tds_rate:g}%)', f"- Rs. {tds_amount:,.2f}"])
            if invoice_type == 'tax' or tds_rate:
                sub_total_lines.append(['Round Off', f"Rs. {round_off:,.2f}"])
            sub_total_lines.append(['GRAND TOTAL (Rs.)', f"{grand_total:,.2f}"])
            tt = Table(sub_total_lines, colWidths=[5.6*inch, 1.3*inch])
            tt.setStyle(TableStyle([('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold'), ('FONTSIZE',(0,-1),(-1,-1),10.5),
                                     ('BACKGROUND',(0,-1),(-1,-1),LIGHTBG), ('BOX',(0,0),(-1,-1),0.5,LINE),
                                     ('LINEBELOW',(0,0),(-1,-2),0.3,colors.HexColor('#DDDDDD')),
                                     ('ALIGN',(1,0),(1,-1),'RIGHT'), ('FONTSIZE',(0,0),(-2,-2),8.5),
                                     ('TOPPADDING',(0,0),(-1,-1),4), ('BOTTOMPADDING',(0,0),(-1,-1),4),
                                     ('LEFTPADDING',(0,0),(-1,-1),8)]))
            story.append(tt)
    else:
        item_rows = [['#', 'LR NUMBER', 'TRIP DATE', 'VEHICLE NO.', 'FROM', 'TO', 'MATERIAL', 'WEIGHT (MT)', 'FREIGHT (Rs.)']]
        for i, li in enumerate(line_items, 1):
            t = li['trip']
            item_rows.append([str(i), cell(t['lr_number']), t['date'] or '', cell(t['vehicle_no']),
                               cell(t['from_loc']), cell(t['to_loc']), cell(t['material']),
                               f"{t['quantity'] or 0:.3f}", f"{li['freight']:,.2f}"])
        item_rows.append(['', '', '', '', '', '', 'TOTAL', f"{total_weight:.3f}", f"{total_freight:,.2f}"])
        item_table = Table(item_rows, colWidths=[0.25*inch, 0.75*inch, 0.65*inch, 0.75*inch, 1.1*inch, 1.1*inch, 0.8*inch, 0.65*inch, 0.85*inch], repeatRows=1)
        item_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), BLACK), ('TEXTCOLOR', (0,0), (-1,0), colors.white), ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'), ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#EEEEEE')),
            ('FONTSIZE', (0,0), (-1,-1), 7.3), ('ALIGN', (7,0), (8,-1), 'RIGHT'), ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#DDDDDD')),
            ('TOPPADDING', (0,0), (-1,-1), 4), ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ]))
        story.append(item_table)
        story.append(Spacer(1, 12))

        multi_additional_total = sub_total - total_freight
        addl_rows = [['DESCRIPTION', 'BASIS', 'AMOUNT (Rs.)']]
        for label, val in [('Loading Charges', total_loading), ('Unloading Charges', total_unloading),
                            ('Permit Charges', total_permit), ('Toll Charges (Actual)', total_toll),
                            ('Weighment Charges', total_weighment), ('Driver Bata', total_driver_bata),
                            ('GPS Charges', total_gps), ('Other Charges', total_other)]:
            if val:
                addl_rows.append([label, 'Per Trip', f"{val:,.2f}"])
        for it in extra_items:
            addl_rows.append([cell(it['desc']), 'Deduction' if it['item_type'] == 'deduction' else 'Charge',
                               f"- {it['amount']:,.2f}" if it['item_type'] == 'deduction' else f"{it['amount']:,.2f}"])
        if len(addl_rows) > 1:
            addl_left = Table(addl_rows, colWidths=[2.2*inch, 1*inch, COMBO_L - 3.2*inch])
            addl_left.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,0), colors.HexColor('#EEEEEE')), ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
                ('FONTSIZE',(0,0),(-1,-1),8), ('ALIGN',(2,0),(2,-1),'RIGHT'), ('GRID',(0,0),(-1,-1),0.3,colors.HexColor('#DDDDDD')),
                ('TOPPADDING',(0,0),(-1,-1),3.5), ('BOTTOMPADDING',(0,0),(-1,-1),3.5),
            ]))
            addl_head = Table([[Paragraph('<b>ADDITIONAL CHARGES (SUMMARY)</b>', section_head_style)]], colWidths=[COMBO_L])
            addl_head.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),LIGHTBG), ('BOX',(0,0),(-1,-1),0.5,LINE),
                                            ('TOPPADDING',(0,0),(-1,-1),5), ('BOTTOMPADDING',(0,0),(-1,-1),5), ('LEFTPADDING',(0,0),(-1,-1),8)]))
            addl_tot = Table([['TOTAL ADDITIONAL CHARGES', f"{multi_additional_total:,.2f}"]], colWidths=[COMBO_L - 1*inch, 1*inch])
            addl_tot.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),LIGHTBG), ('FONTNAME',(0,0),(-1,-1),'Helvetica-Bold'),
                                           ('FONTSIZE',(0,0),(-1,-1),8.5), ('ALIGN',(1,0),(1,-1),'RIGHT'), ('BOX',(0,0),(-1,-1),0.5,LINE),
                                           ('TOPPADDING',(0,0),(-1,-1),4), ('BOTTOMPADDING',(0,0),(-1,-1),4)]))
            left_col = [addl_head, addl_left, addl_tot]
        else:
            left_col = ''

        r_rows = [['Total Freight', f"Rs. {total_freight:,.2f}"], ['Total Additional Charges', f"Rs. {multi_additional_total:,.2f}"],
                   ['Sub Total', f"Rs. {sub_total:,.2f}"]]
        if invoice_type == 'tax':
            r_rows.append([f'CGST ({cgst_rate:g}%)', f"Rs. {cgst_amount:,.2f}"])
            r_rows.append([f'SGST ({sgst_rate:g}%)', f"Rs. {sgst_amount:,.2f}"])
        if tds_rate:
            r_rows.append([f'TDS ({tds_rate:g}%)', f"- Rs. {tds_amount:,.2f}"])
        if invoice_type == 'tax' or tds_rate:
            r_rows.append(['Round Off', f"Rs. {round_off:,.2f}"])
        r_rows.append(['GRAND TOTAL (Rs.)', f"{grand_total:,.2f}"])
        # Same single-box, every-row-visible pattern as the single-trip layout above.
        right_stack = Table(r_rows, colWidths=[COMBO_R - 1.1*inch, 1.1*inch])
        right_stack.setStyle(TableStyle([
            ('FONTSIZE',(0,0),(-1,-1),8.3), ('ALIGN',(1,0),(1,-1),'RIGHT'),
            ('TOPPADDING',(0,0),(-1,-1),4), ('BOTTOMPADDING',(0,0),(-1,-1),4), ('LEFTPADDING',(0,0),(-1,-1),8),
            ('BOX',(0,0),(-1,-1),0.5,LINE), ('LINEBELOW',(0,0),(-1,-2),0.3,colors.HexColor('#DDDDDD')),
            ('BACKGROUND',(0,-1),(-1,-1),colors.HexColor('#E8F5E9')), ('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold'),
            ('FONTSIZE',(0,-1),(-1,-1),10.5), ('TEXTCOLOR',(1,-1),(1,-1),colors.HexColor('#2E7D32')),
        ]))
        combo = Table([[left_col, '', right_stack]], colWidths=[COMBO_L, COMBO_GAP, COMBO_R])
        combo.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'TOP'), ('LEFTPADDING',(1,0),(1,-1),0), ('RIGHTPADDING',(1,0),(1,-1),0)]))
        story.append(combo)

    if invoice_type == 'vehicle_owner' and total_already_given:
        deduction_rows = [['Gross Bill Amount', f"Rs. {grand_total:,.2f}"]]
        if total_fuel_given:
            deduction_rows.append(['Less: Fuel Advance Given', f"Rs. {total_fuel_given:,.2f}"])
        if total_driver_adv_given:
            deduction_rows.append(['Less: Driver Advance Given', f"Rs. {total_driver_adv_given:,.2f}"])
        if total_paid_to_owner:
            deduction_rows.append(['Less: Amount Already Paid', f"Rs. {total_paid_to_owner:,.2f}"])
        deduction_rows.append(['NET PAYABLE TO OWNER (Rs.)', f"{net_payable:,.2f}"])
        # Same bordered-bar treatment as "TOTAL FREIGHT & CHARGES" above (BOX + row rules +
        # highlighted final row) instead of plain floating text, so this reads as one continuous
        # pattern with the totals above it rather than a visually distinct, "distorted" block.
        deduction_table = Table(deduction_rows, colWidths=[5.6*inch, 1.3*inch])
        deduction_table.setStyle(TableStyle([
            ('FONTSIZE',(0,0),(-1,-1),8.5), ('ALIGN',(1,0),(1,-1),'RIGHT'),
            ('TOPPADDING',(0,0),(-1,-1),4), ('BOTTOMPADDING',(0,0),(-1,-1),4),
            ('LEFTPADDING',(0,0),(-1,-1),8), ('BOX',(0,0),(-1,-1),0.5,LINE),
            ('LINEBELOW',(0,0),(-1,-2),0.3,colors.HexColor('#DDDDDD')),
            ('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold'), ('FONTSIZE',(0,-1),(-1,-1),10.5),
            ('BACKGROUND',(0,-1),(-1,-1),LIGHTBG),
        ]))
        story.append(Spacer(1, 8))
        story.append(deduction_table)

    story.append(Spacer(1, 8))
    # Constrained to CONTENT_W (not the full page frame, which is wider) so this line's left/right
    # edges match every box above it instead of running further right and reading "out of line".
    words_table = Table([[Paragraph(f"<b>Amount In Words:</b> Rupees {number_to_words_inr(net_payable)} Only", label_style)]],
                         colWidths=[CONTENT_W])
    words_table.setStyle(TableStyle([('LEFTPADDING',(0,0),(-1,-1),0), ('RIGHTPADDING',(0,0),(-1,-1),0),
                                      ('TOPPADDING',(0,0),(-1,-1),0), ('BOTTOMPADDING',(0,0),(-1,-1),0)]))
    story.append(words_table)
    story.append(Spacer(1, 16))

    # ---------- Footer: Payment Details + Terms (left) / Signature (right) ----------
    # A fixed 2-column layout — left content always starts flush left, signature always sits
    # flush right — instead of the previous 3-column table that padded a blank cell onto the
    # front whenever Payment Details didn't apply, which pushed Terms & Conditions into the
    # middle column and left the true left edge empty.
    left_blocks = []
    if invoice_type == 'tax' and s['bank_name']:
        bank_lines = ["<b>PAYMENT DETAILS</b>", f"Bank Name : {esc(s['bank_name'])}", f"A/C Number : {esc(s['account_number'])}",
                      f"IFSC Code : {esc(s['ifsc_code'])}", f"Branch : {esc(s['branch'])}"]
        left_blocks.append(Paragraph('<br/>'.join(bank_lines), label_style))
        left_blocks.append(Spacer(1, 10))
    terms_lines = ["<b>TERMS &amp; CONDITIONS</b>", "1. Payment should be made within due date.",
                   "2. Interest @ 18% p.a. will be charged on overdue.", "3. All disputes subject to Rourkela Jurisdiction."]
    left_blocks.append(Paragraph('<br/>'.join(terms_lines), label_style))
    signature_block = Paragraph(f"<b>FOR {esc(s['company_name'])}</b><br/><br/><br/>Authorised Signatory", label_style)
    footer_table = Table([[left_blocks, signature_block]], colWidths=[4.6*inch, 2.3*inch])
    footer_table.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP'), ('LINEABOVE', (0,0), (-1,0), 0.7, LINE),
                                       ('TOPPADDING', (0,0), (-1,-1), 8), ('LEFTPADDING', (0,0), (0,0), 0)]))
    story.append(footer_table)

    if invoice_type == 'tax' and s['rcm_clause']:
        story.append(Spacer(1, 10))
        story.append(Paragraph(f"<b>RCM Note:</b> {esc(s['rcm_clause'])}", label_style))
    if remarks:
        story.append(Spacer(1, 8))
        story.append(Paragraph(f"<b>Remarks:</b> {esc(remarks)}", label_style))

    story.append(Spacer(1, 14))
    story.append(Paragraph("Thank you for your business!", ParagraphStyle('TY', parent=styles['Normal'], fontSize=9, alignment=1, fontName='Helvetica-Oblique')))

    doc.build(story)
    buf.seek(0)
    return buf

@app.route('/invoice-center/generate', methods=['POST'])
def invoice_center_generate():
    from flask import send_file
    import datetime

    f = request.form
    mode = f.get('mode') or 'generate'  # 'generate' | 'draft' | 'preview'
    trip_ids = f.getlist('trip_ids')
    invoice_type = f.get('invoice_type')
    party_id = f.get('party_id') or None
    vendor_id = f.get('vendor_id') or None
    invoice_date = f.get('invoice_date') or datetime.datetime.now().strftime('%Y-%m-%d')
    due_date = f.get('due_date') or ''
    payment_terms = f.get('payment_terms') or ''
    place_of_supply = f.get('place_of_supply') or ''
    remarks = f.get('remarks') or ''
    payment_status = f.get('payment_status') or 'PENDING'
    extra_loading = float(f.get('loading_charges') or 0)
    extra_other = float(f.get('other_charges') or 0)
    tds_rate = float(f.get('tds_rate') or 0)
    gst_rate_input = f.get('gst_rate')

    conn = get_db()
    placeholders = ','.join('?' * len(trip_ids))
    trips = conn.execute(f"""SELECT t.*, v.vehicle_no FROM trips t LEFT JOIN vehicles v ON t.vehicle_id=v.id
                             WHERE t.id IN ({placeholders})""", trip_ids).fetchall() if trip_ids else []

    # Each selected trip's own "Others" line items (real per-trip charges/deductions, e.g.
    # Detention, an advance already given) — every item checkbox is checked (included) by default
    # in the UI, so a plain form submit with JS disabled still includes everything; unchecking one
    # removes it from included_item_ids. Kept as named entries (not folded into a number) so each
    # one is printed on the invoice by its own description instead of disappearing into one total.
    extra_items = []
    included_item_ids = {int(x) for x in f.getlist('included_item_ids') if x}
    if trip_ids:
        for it in conn.execute(f"""SELECT id, description, amount, item_type FROM invoice_items WHERE trip_id IN ({placeholders})""", trip_ids).fetchall():
            if it['id'] not in included_item_ids:
                continue
            extra_items.append({'desc': it['description'] or 'Other', 'amount': it['amount'] or 0, 'item_type': it['item_type']})

    # Ad-hoc charges/deductions typed directly into Invoice Center at generation time — kept as
    # named entries too (once the batch exists, saved as real invoice_batch_items so they're still
    # visible/editable from Generated Invoices afterwards, same as items added there).
    extra_descs = f.getlist('extra_item_desc')
    extra_amounts = f.getlist('extra_item_amount')
    extra_types = f.getlist('extra_item_type')
    new_items = []
    for i, desc in enumerate(extra_descs):
        desc = (desc or '').strip()
        if not desc:
            continue
        try:
            amt = float(extra_amounts[i]) if i < len(extra_amounts) and extra_amounts[i] else 0
        except ValueError:
            amt = 0
        item_type = extra_types[i] if i < len(extra_types) and extra_types[i] in ('charge', 'deduction') else 'charge'
        new_items.append((desc, amt, item_type))
        extra_items.append({'desc': desc, 'amount': amt, 'item_type': item_type})

    entity = None
    if invoice_type == 'vehicle_owner' and vendor_id:
        entity = conn.execute("SELECT * FROM vendors WHERE id=?", (vendor_id,)).fetchone()
    elif party_id:
        entity = conn.execute("SELECT * FROM parties WHERE id=?", (party_id,)).fetchone()

    s = _get_invoice_settings(conn)
    toll_map = _toll_by_trip(conn, [t['id'] for t in trips])

    if invoice_type == 'tax':
        total_gst_rate = float(gst_rate_input) if gst_rate_input else (float(s['cgst_rate'] or 0) + float(s['sgst_rate'] or 0))
        cgst_rate = sgst_rate = round(total_gst_rate / 2, 4)
    else:
        cgst_rate = sgst_rate = 0

    next_num = int(s['next_invoice_number'] or 1)
    invoice_number = f"{s['invoice_prefix'] or 'ATS/INV'}/{datetime.datetime.now().year}/{next_num:04d}"

    if mode == 'preview':
        # Preview never touches the DB or the invoice-number counter — it's just a look at the PDF.
        buf = _build_invoice_pdf(trips, invoice_type, entity, s, invoice_number, invoice_date, due_date, payment_status, remarks,
                                  cgst_rate=cgst_rate, sgst_rate=sgst_rate, extra_loading=extra_loading, extra_other=extra_other,
                                  tds_rate=tds_rate, extra_items=extra_items, toll_map=toll_map)
        conn.close()
        return send_file(buf, download_name=f'preview-{invoice_number.replace("/","-")}.pdf', mimetype='application/pdf')

    now = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    status = 'draft' if mode == 'draft' else 'generated'
    cur = conn.execute("""INSERT INTO invoice_batches (invoice_number, invoice_type, party_id, vendor_id, invoice_date, due_date,
                          payment_terms, place_of_supply, remarks, gst_rate, tds_rate, loading_charges, other_charges, status, payment_status, created_at)
                          VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                 (invoice_number, invoice_type, party_id, vendor_id, invoice_date, due_date, payment_terms,
                  place_of_supply, remarks, cgst_rate+sgst_rate, tds_rate, extra_loading, extra_other,
                  status, payment_status, now))
    batch_id = cur.lastrowid
    for tid in trip_ids:
        conn.execute("INSERT INTO invoice_batch_trips (invoice_batch_id, trip_id) VALUES (?,?)", (batch_id, tid))
    for desc, amt, item_type in new_items:
        conn.execute("INSERT INTO invoice_batch_items (invoice_batch_id, description, amount, item_type) VALUES (?,?,?,?)",
                     (batch_id, desc, amt, item_type))
    conn.execute("UPDATE settings SET value=? WHERE key='next_invoice_number'", (str(next_num + 1),))
    conn.commit()
    conn.close()

    if mode == 'draft':
        return redirect(url_for('invoice_batches_list'))

    buf = _build_invoice_pdf(trips, invoice_type, entity, s, invoice_number, invoice_date, due_date, payment_status, remarks,
                              cgst_rate=cgst_rate, sgst_rate=sgst_rate, extra_loading=extra_loading, extra_other=extra_other,
                              tds_rate=tds_rate, extra_items=extra_items, toll_map=toll_map)
    return send_file(buf, as_attachment=True, download_name=f'{invoice_number.replace("/","-")}.pdf', mimetype='application/pdf')

@app.route('/invoices/generated')
def invoice_batches_list():
    conn = get_db()
    search_f = request.args.get('search', '')
    type_f = request.args.get('invoice_type', '')
    query = """SELECT ib.id, ib.invoice_number, ib.invoice_type, ib.invoice_date, ib.due_date,
               ib.payment_status, ib.status, ib.created_at, p.name as party_name, v.name as vendor_name,
               (SELECT COUNT(*) FROM invoice_batch_trips ibt WHERE ibt.invoice_batch_id=ib.id) as trip_count
               FROM invoice_batches ib
               LEFT JOIN parties p ON ib.party_id=p.id
               LEFT JOIN vendors v ON ib.vendor_id=v.id
               WHERE 1=1"""
    params = []
    if type_f:
        query += " AND ib.invoice_type=?"; params.append(type_f)
    if search_f:
        query += " AND (p.name LIKE ? OR v.name LIKE ? OR ib.invoice_number LIKE ?)"
        params += [f"%{search_f}%", f"%{search_f}%", f"%{search_f}%"]
    query += " ORDER BY ib.created_at DESC"
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return render_template('invoice_batches_list.html', rows=rows, f_search=search_f, f_type=type_f, active='invoices')

@app.route('/invoices/generated/<int:batch_id>/edit', methods=['GET', 'POST'])
def invoice_batch_edit(batch_id):
    conn = get_db()
    if request.method == 'POST':
        f = request.form
        conn.execute("""UPDATE invoice_batches SET invoice_date=?, due_date=?, payment_terms=?, place_of_supply=?,
                        remarks=?, payment_status=?, gst_rate=?, tds_rate=?, loading_charges=?, other_charges=?
                        WHERE id=?""",
                     (f.get('invoice_date'), f.get('due_date'), f.get('payment_terms'), f.get('place_of_supply'),
                      f.get('remarks'), f.get('payment_status'), float(f.get('gst_rate') or 0), float(f.get('tds_rate') or 0),
                      float(f.get('loading_charges') or 0), float(f.get('other_charges') or 0), batch_id))
        conn.commit()
        conn.close()
        return redirect(url_for('invoice_batches_list'))
    batch = conn.execute("""SELECT ib.*, p.name as party_name, v.name as vendor_name FROM invoice_batches ib
                            LEFT JOIN parties p ON ib.party_id=p.id LEFT JOIN vendors v ON ib.vendor_id=v.id
                            WHERE ib.id=?""", (batch_id,)).fetchone()
    batch_items = conn.execute("SELECT * FROM invoice_batch_items WHERE invoice_batch_id=? ORDER BY id", (batch_id,)).fetchall()
    conn.close()
    return render_template('invoice_batch_edit.html', b=batch, batch_items=batch_items, active='invoices')

@app.route('/invoices/generated/<int:batch_id>/items/add', methods=['POST'])
def invoice_batch_item_add(batch_id):
    conn = get_db()
    f = request.form
    conn.execute("INSERT INTO invoice_batch_items (invoice_batch_id, description, amount, item_type) VALUES (?,?,?,?)",
                 (batch_id, f.get('description'), float(f.get('amount') or 0), f.get('item_type') or 'charge'))
    conn.commit()
    conn.close()
    return redirect(url_for('invoice_batch_edit', batch_id=batch_id))

@app.route('/invoices/generated/items/<int:item_id>/delete', methods=['POST'])
def invoice_batch_item_delete(item_id):
    conn = get_db()
    row = conn.execute("SELECT invoice_batch_id FROM invoice_batch_items WHERE id=?", (item_id,)).fetchone()
    conn.execute("DELETE FROM invoice_batch_items WHERE id=?", (item_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('invoice_batch_edit', batch_id=row['invoice_batch_id'])) if row else redirect(url_for('invoice_batches_list'))

@app.route('/invoices/generated/<int:batch_id>/delete', methods=['POST'])
def invoice_batch_delete(batch_id):
    conn = get_db()
    conn.execute("DELETE FROM invoice_batch_trips WHERE invoice_batch_id=?", (batch_id,))
    conn.execute("DELETE FROM invoice_batch_items WHERE invoice_batch_id=?", (batch_id,))
    conn.execute("DELETE FROM invoice_batches WHERE id=?", (batch_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('invoice_batches_list'))

@app.route('/invoices/generated/<int:batch_id>/pdf')
def invoice_batch_pdf(batch_id):
    from flask import send_file
    conn = get_db()
    batch = conn.execute("SELECT * FROM invoice_batches WHERE id=?", (batch_id,)).fetchone()
    trip_id_rows = conn.execute("SELECT trip_id FROM invoice_batch_trips WHERE invoice_batch_id=?", (batch_id,)).fetchall()
    trip_ids = [str(r['trip_id']) for r in trip_id_rows]
    placeholders = ','.join('?' * len(trip_ids))
    trips = conn.execute(f"""SELECT t.*, v.vehicle_no FROM trips t LEFT JOIN vehicles v ON t.vehicle_id=v.id
                             WHERE t.id IN ({placeholders})""", trip_ids).fetchall() if trip_ids else []

    entity = None
    if batch['invoice_type'] == 'vehicle_owner' and batch['vendor_id']:
        entity = conn.execute("SELECT * FROM vendors WHERE id=?", (batch['vendor_id'],)).fetchone()
    elif batch['party_id']:
        entity = conn.execute("SELECT * FROM parties WHERE id=?", (batch['party_id'],)).fetchone()
    s = _get_invoice_settings(conn)

    # Batch-level custom items (added from the Edit screen) plus each trip's own "Others" items —
    # kept as named entries, same as at generate time, so a regenerated PDF still prints each one
    # by its own description (e.g. "Detention") instead of folding them into one anonymous total.
    extra_items = []
    for it in conn.execute("SELECT description, amount, item_type FROM invoice_batch_items WHERE invoice_batch_id=?", (batch_id,)).fetchall():
        extra_items.append({'desc': it['description'] or 'Other', 'amount': it['amount'] or 0, 'item_type': it['item_type']})
    if trip_ids:
        for it in conn.execute(f"SELECT description, amount, item_type FROM invoice_items WHERE trip_id IN ({placeholders})", trip_ids).fetchall():
            extra_items.append({'desc': it['description'] or 'Other', 'amount': it['amount'] or 0, 'item_type': it['item_type']})
    toll_map = _toll_by_trip(conn, [t['id'] for t in trips])
    conn.close()

    cgst_rate = sgst_rate = round((batch['gst_rate'] or 0) / 2, 4) if batch['invoice_type'] == 'tax' else 0
    buf = _build_invoice_pdf(trips, batch['invoice_type'], entity, s, batch['invoice_number'], batch['invoice_date'],
                              batch['due_date'], batch['payment_status'], batch['remarks'],
                              cgst_rate=cgst_rate, sgst_rate=sgst_rate,
                              extra_loading=batch['loading_charges'] or 0, extra_other=batch['other_charges'] or 0,
                              tds_rate=batch['tds_rate'] or 0, extra_items=extra_items, toll_map=toll_map)
    return send_file(buf, as_attachment=True, download_name=f'{batch["invoice_number"].replace("/","-")}.pdf', mimetype='application/pdf')

@app.route('/invoices')
def invoices_search():
    lr_f = request.args.get('lr_number', '')
    conn = get_db()
    results = []
    if lr_f:
        results = conn.execute("""SELECT t.id, t.date, t.lr_number, v.vehicle_no, p.name as party_name,
                                  t.from_loc, t.to_loc, t.billed_amount
                                  FROM trips t LEFT JOIN vehicles v ON t.vehicle_id=v.id
                                  LEFT JOIN parties p ON t.party_id=p.id
                                  WHERE t.lr_number LIKE ? ORDER BY t.date DESC""", (f"%{lr_f}%",)).fetchall()
    conn.close()
    return render_template('invoices_search.html', results=results, f_lr=lr_f, active='invoices')

@app.route('/invoices/<int:trip_id>')
def invoice_preview(trip_id):
    conn = get_db()
    t = conn.execute("""SELECT t.*, v.vehicle_no, p.name as party_name
                        FROM trips t LEFT JOIN vehicles v ON t.vehicle_id=v.id
                        LEFT JOIN parties p ON t.party_id=p.id WHERE t.id=?""", (trip_id,)).fetchone()
    inv = conn.execute("SELECT * FROM invoices WHERE trip_id=?", (trip_id,)).fetchone()
    extra_items = conn.execute("SELECT * FROM invoice_items WHERE trip_id=?", (trip_id,)).fetchall()
    conn.close()
    charges = [
        ('Driver Payment', t['driver_payment']), ('Detention Charges', t['detention_charges']),
        ('GPS Cost', t['gps_cost']), ('Loading Charge', t['loading_charge']),
        ('Unloading Charge', t['unloading_charge']), ('Police Charges', t['police_charges']),
        ('SIM Tracking', t['sim_tracking']), ('Union Charges', t['union_charges']),
        ('Weight Charges', t['weight_charges']), ('Other Charges', t['other_charges']),
    ]
    charges = [(l, v) for l, v in charges if v]
    charges += [(item['description'], item['amount']) for item in extra_items if item['item_type']=='charge']
    deductions = [
        ('Brokerage', t['brokerage']), ('Builty Commission', t['builty_commission']),
        ('Late Fees', t['late_fees']), ('Material Damage', t['material_damage']),
        ('Shortage Amount', t['shortage_amount']), ('TDS', t['tds']), ('Other Deductions', t['other_deductions']),
    ]
    deductions = [(l, v) for l, v in deductions if v]
    deductions += [(item['description'], item['amount']) for item in extra_items if item['item_type']=='deduction']
    freight = t['fixed_rate_amount'] if t['rate_type'] == 'FIXED' else (t['quantity'] or 0) * (t['rate'] or 0)
    extra_charges_total = sum(item['amount'] or 0 for item in extra_items if item['item_type']=='charge')
    extra_deductions_total = sum(item['amount'] or 0 for item in extra_items if item['item_type']=='deduction')
    invoice_total = (t['billed_amount'] or 0) + extra_charges_total - extra_deductions_total
    amount_paid = (t['payment_received'] or 0) + (t['party_advance'] or 0)
    balance_due = invoice_total - amount_paid
    return render_template('invoice_preview.html', t=t, inv=inv, charges=charges, deductions=deductions, freight=freight,
                            invoice_total=invoice_total, amount_paid=amount_paid, balance_due=balance_due, extra_items=extra_items, active='invoices')

@app.route('/invoices/<int:trip_id>/edit', methods=['GET', 'POST'])
def invoice_edit(trip_id):
    conn = get_db()
    if request.method == 'POST':
        f = request.form
        existing = conn.execute("SELECT id FROM invoices WHERE trip_id=?", (trip_id,)).fetchone()
        if existing:
            conn.execute("UPDATE invoices SET invoice_number=?, due_date=?, notes=? WHERE trip_id=?",
                         (f.get('invoice_number'), f.get('due_date'), f.get('notes'), trip_id))
        else:
            conn.execute("INSERT INTO invoices (trip_id, invoice_number, due_date, notes) VALUES (?,?,?,?)",
                         (trip_id, f.get('invoice_number'), f.get('due_date'), f.get('notes')))
        conn.commit()
        conn.close()
        return redirect(url_for('invoice_preview', trip_id=trip_id))

    t = conn.execute("SELECT lr_number FROM trips WHERE id=?", (trip_id,)).fetchone()
    inv = conn.execute("SELECT * FROM invoices WHERE trip_id=?", (trip_id,)).fetchone()
    extra_items = conn.execute("SELECT * FROM invoice_items WHERE trip_id=?", (trip_id,)).fetchall()
    conn.close()
    default_invoice_no = f"INV-{t['lr_number']}" if not (inv and inv['invoice_number']) else None
    return render_template('invoice_edit.html', t=t, inv=inv, default_invoice_no=default_invoice_no, trip_id=trip_id, extra_items=extra_items, active='invoices')

@app.route('/invoices/<int:trip_id>/items/add', methods=['POST'])
def invoice_item_add(trip_id):
    conn = get_db()
    f = request.form
    conn.execute("INSERT INTO invoice_items (trip_id, description, amount, item_type) VALUES (?,?,?,?)",
                 (trip_id, f.get('description'), float(f.get('amount') or 0), f.get('item_type')))
    conn.commit()
    conn.close()
    return redirect(url_for('invoice_edit', trip_id=trip_id))

@app.route('/invoices/items/<int:item_id>/delete', methods=['POST'])
def invoice_item_delete(item_id):
    conn = get_db()
    trip_id = conn.execute("SELECT trip_id FROM invoice_items WHERE id=?", (item_id,)).fetchone()[0]
    conn.execute("DELETE FROM invoice_items WHERE id=?", (item_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('invoice_edit', trip_id=trip_id))

@app.route('/invoices/<int:trip_id>/pdf')
def invoice_pdf(trip_id):
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from flask import send_file
    import io

    conn = get_db()
    t = conn.execute("""SELECT t.*, v.vehicle_no, p.name as party_name
                        FROM trips t LEFT JOIN vehicles v ON t.vehicle_id=v.id
                        LEFT JOIN parties p ON t.party_id=p.id WHERE t.id=?""", (trip_id,)).fetchone()
    inv = conn.execute("SELECT * FROM invoices WHERE trip_id=?", (trip_id,)).fetchone()
    extra_items = conn.execute("SELECT * FROM invoice_items WHERE trip_id=?", (trip_id,)).fetchall()
    conn.close()

    invoice_number = (inv['invoice_number'] if inv and inv['invoice_number'] else f"INV-{t['lr_number']}")
    due_date = (inv['due_date'] if inv and inv['due_date'] else '')
    inv_notes = (inv['notes'] if inv and inv['notes'] else '')

    freight = t['fixed_rate_amount'] if t['rate_type'] == 'FIXED' else (t['quantity'] or 0) * (t['rate'] or 0)
    charges_total = sum(t[c] or 0 for c in ['driver_payment','detention_charges','gps_cost','loading_charge',
                        'unloading_charge','police_charges','sim_tracking','union_charges','weight_charges','other_charges'])
    deductions_total = sum(t[c] or 0 for c in ['brokerage','builty_commission','late_fees','material_damage',
                            'shortage_amount','tds','other_deductions'])
    extra_charges_total = sum(item['amount'] or 0 for item in extra_items if item['item_type']=='charge')
    extra_deductions_total = sum(item['amount'] or 0 for item in extra_items if item['item_type']=='deduction')
    charges_total += extra_charges_total
    deductions_total += extra_deductions_total
    amount_paid = (t['payment_received'] or 0) + (t['party_advance'] or 0)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.6*inch)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('T', parent=styles['Title'], fontSize=20, textColor=colors.HexColor('#1B2A4A'), alignment=1)
    sub_style = ParagraphStyle('S', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#5A6B8C'), alignment=1)
    label_style = ParagraphStyle('L', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#5A6B8C'))
    story = []

    story.append(Paragraph("ANIL TRANSPORT SERVICE", title_style))
    story.append(Paragraph("Head Off.: Shop No. D/8, Nirmal Market Power House Road, Rourkela - 769001", sub_style))
    story.append(Paragraph("GSTIN No.: 21ABDPL6110E1ZG &nbsp;&nbsp;|&nbsp;&nbsp; Mob. +91 9437246272 &nbsp;&nbsp;|&nbsp;&nbsp; Ph./Fax: 0661-2501272", sub_style))
    story.append(Spacer(1, 6))
    line_table = Table([['']], colWidths=[7*inch], rowHeights=[2])
    line_table.setStyle(TableStyle([('LINEBELOW', (0,0), (-1,-1), 1.5, colors.HexColor('#1B2A4A'))]))
    story.append(line_table)
    story.append(Spacer(1, 14))

    story.append(Paragraph(f"<b>INVOICE</b>", styles['Heading2']))
    story.append(Spacer(1, 8))

    info_data = [
        ['Invoice No.', invoice_number, 'LR Number', t['lr_number'] or ''],
        ['Date', t['date'] or '', 'Due Date', due_date or '—'],
        ['Bill To', t['party_name'] or '', 'Vehicle No', t['vehicle_no'] or ''],
        ['From', t['from_loc'] or '', 'To', t['to_loc'] or ''],
        ['Material', t['material'] or '', 'Quantity (MT)', str(t['quantity'] or '')],
    ]
    info_table = Table(info_data, colWidths=[1*inch, 2.2*inch, 1*inch, 1.8*inch])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'), ('FONTNAME', (2,0), (2,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9), ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 16))

    charge_rows = [['Description', 'Amount (Rs.)']]
    charge_rows.append(['Freight' + (' (Fixed)' if t['rate_type']=='FIXED' else f" ({t['quantity']} MT x Rs.{t['rate']})"), f"{freight:,.0f}"])
    if charges_total:
        charge_rows.append(['Total Charges (+)', f"{charges_total:,.0f}"])
    if deductions_total:
        charge_rows.append(['Total Deductions (-)', f"-{deductions_total:,.0f}"])
    invoice_total = (t['billed_amount'] or 0) + extra_charges_total - extra_deductions_total
    balance_due = invoice_total - amount_paid
    charge_rows.append(['TOTAL BILLED', f"{invoice_total:,.0f}"])
    if amount_paid:
        charge_rows.append(['Amount Already Paid', f"-{amount_paid:,.0f}"])
        charge_rows.append(['BALANCE DUE', f"{balance_due:,.0f}"])

    charge_table = Table(charge_rows, colWidths=[4*inch, 2*inch])
    charge_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1B2A4A')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('LINEABOVE', (0,-1), (-1,-1), 1, colors.HexColor('#1B2A4A')),
        ('ALIGN', (1,0), (1,-1), 'RIGHT'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('TOPPADDING', (0,0), (-1,-1), 7), ('BOTTOMPADDING', (0,0), (-1,-1), 7),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#DDE3EC')),
    ]))
    story.append(charge_table)
    story.append(Spacer(1, 20))
    if inv_notes:
        story.append(Spacer(1, 10))
        story.append(Paragraph(f"<b>Notes:</b> {inv_notes}", label_style))
    story.append(Spacer(1, 20))
    story.append(Paragraph("This is a system-generated invoice.", label_style))

    doc.build(story)
    buf.seek(0)
    safe_lr = "".join(c for c in (t['lr_number'] or 'invoice') if c.isalnum() or c in " _-")[:40]
    return send_file(buf, as_attachment=True, download_name=f'invoice_{safe_lr}.pdf', mimetype='application/pdf')

@app.route('/ledger/<entity_type>/<int:entity_id>/contact', methods=['POST'])
def update_contact(entity_type, entity_id):
    conn = get_db()
    f = request.form
    table = 'parties' if entity_type == 'party' else 'vendors'
    status = f.get('status') if f.get('status') in ('Active', 'Inactive') else 'Active'
    conn.execute(f"UPDATE {table} SET address=?, contact=?, email=?, credit_limit=?, gstin=?, category=?, status=? WHERE id=?",
                 (f.get('address'), f.get('contact'), f.get('email'), f.get('credit_limit') or None,
                  f.get('gstin') or None, f.get('category') or None, status, entity_id))
    conn.commit()
    conn.close()
    if entity_type == 'party':
        return redirect(url_for('party_ledger', party_id=entity_id))
    return redirect(url_for('vendor_ledger', vendor_id=entity_id))

@app.route('/ledger/<entity_type>/<int:entity_id>/opening-balance', methods=['POST'])
def update_opening_balance(entity_type, entity_id):
    conn = get_db()
    f = request.form
    table = 'parties' if entity_type == 'party' else 'vendors'
    conn.execute(f"UPDATE {table} SET opening_balance=?, opening_balance_date=? WHERE id=?",
                 (float(f.get('opening_balance') or 0), f.get('opening_balance_date') or None, entity_id))
    conn.commit()
    conn.close()
    if entity_type == 'party':
        return redirect(url_for('party_ledger', party_id=entity_id))
    return redirect(url_for('vendor_ledger', vendor_id=entity_id))

def get_company_name():
    conn = get_db()
    row = conn.execute("SELECT value FROM settings WHERE key='company_name'").fetchone()
    conn.close()
    return row['value'] if row else 'ANIL TRANSPORT SERVICE'

@app.context_processor
def inject_company_name():
    try:
        return {'company_name': get_company_name()}
    except Exception:
        return {'company_name': 'ANIL TRANSPORT SERVICE'}

@app.context_processor
def inject_current_year():
    """Sidebar footer copyright year — computed, not hardcoded, so it never goes stale."""
    return {'current_year': datetime.datetime.now().year}

@app.before_request
def require_login():
    exempt = ['login', 'static', 'send_login_otp', 'verify_login_otp']
    if request.endpoint in exempt:
        return
    if not session.get('user_id'):
        return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    from werkzeug.security import check_password_hash
    error = None
    if request.method == 'POST':
        f = request.form
        conn = get_db()
        user = conn.execute("SELECT * FROM users WHERE username=?", (f.get('username'),)).fetchone()
        if user and check_password_hash(user['password_hash'], f.get('password') or ''):
            if (user['status'] or 'Active') == 'Inactive':
                conn.close()
                return render_template('login.html', error='This account has been deactivated. Contact an administrator.',
                                        company_name=get_company_name())
            session.permanent = bool(f.get('remember_me'))
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['is_admin'] = user['is_admin']
            now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            conn.execute("UPDATE users SET last_login=? WHERE id=?", (now, user['id']))
            conn.execute("INSERT INTO access_logs (user_id, event, date) VALUES (?,?,?)", (user['id'], 'Login', now))
            conn.commit()
            conn.close()
            return redirect(url_for('dashboard'))
        conn.close()
        error = 'Invalid username or password.'
    return render_template('login.html', error=error, company_name=get_company_name())

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

def _send_otp_sms(s, phone, otp):
    from twilio.rest import Client
    client = Client(s['twilio_account_sid'], s['twilio_auth_token'])
    client.messages.create(
        body=f"Your {s.get('company_name') or 'Fleet App'} login OTP is {otp}. Valid for 5 minutes.",
        from_=s['twilio_from_number'],
        to=phone,
    )

@app.route('/login/otp/send', methods=['POST'])
def send_login_otp():
    import random, datetime, hashlib
    phone = (request.form.get('phone') or '').strip()
    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE phone=?", (phone,)).fetchone()
    s = _get_all_settings(conn)
    conn.close()

    if not phone:
        return render_template('login.html', company_name=get_company_name(), otp_mode=True,
                                otp_error='Enter a mobile number.')
    if not user:
        return render_template('login.html', company_name=get_company_name(), otp_mode=True,
                                otp_error='No account found with that mobile number.')
    if not (s['twilio_account_sid'] and s['twilio_auth_token'] and s['twilio_from_number']):
        return render_template('login.html', company_name=get_company_name(), otp_mode=True,
                                otp_error="SMS login isn't configured yet. Ask an admin to add Twilio details in Settings.")

    otp = f"{random.randint(0, 999999):06d}"
    session['otp_user_id'] = user['id']
    session['otp_phone'] = phone
    session['otp_hash'] = hashlib.sha256(otp.encode()).hexdigest()
    session['otp_expires'] = (datetime.datetime.now() + datetime.timedelta(minutes=5)).timestamp()

    try:
        _send_otp_sms(s, phone, otp)
    except Exception as e:
        return render_template('login.html', company_name=get_company_name(), otp_mode=True,
                                otp_error=f'Could not send OTP: {e}')

    return render_template('login.html', company_name=get_company_name(), otp_mode=True,
                            otp_sent=True, otp_phone=phone)

@app.route('/login/otp/verify', methods=['POST'])
def verify_login_otp():
    import hashlib, datetime
    code = (request.form.get('otp') or '').strip()
    phone = request.form.get('phone') or ''

    if not session.get('otp_user_id') or session.get('otp_phone') != phone:
        return render_template('login.html', company_name=get_company_name(), otp_mode=True,
                                otp_error='Session expired. Please request a new OTP.')
    if datetime.datetime.now().timestamp() > session.get('otp_expires', 0):
        return render_template('login.html', company_name=get_company_name(), otp_mode=True,
                                otp_error='OTP expired. Please request a new one.')
    if hashlib.sha256(code.encode()).hexdigest() != session.get('otp_hash'):
        return render_template('login.html', company_name=get_company_name(), otp_mode=True,
                                otp_sent=True, otp_phone=phone, otp_error='Incorrect OTP.')

    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE id=?", (session['otp_user_id'],)).fetchone()
    for key in ('otp_user_id', 'otp_phone', 'otp_hash', 'otp_expires'):
        session.pop(key, None)
    if not user:
        conn.close()
        return redirect(url_for('login'))
    if (user['status'] or 'Active') == 'Inactive':
        conn.close()
        return render_template('login.html', company_name=get_company_name(),
                                error='This account has been deactivated. Contact an administrator.')
    session['user_id'] = user['id']
    session['username'] = user['username']
    session['is_admin'] = user['is_admin']
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn.execute("UPDATE users SET last_login=? WHERE id=?", (now, user['id']))
    conn.execute("INSERT INTO access_logs (user_id, event, date) VALUES (?,?,?)", (user['id'], 'Login (OTP)', now))
    conn.commit()
    conn.close()
    return redirect(url_for('dashboard'))

ALL_SETTING_KEYS = [
    'company_name', 'address', 'state', 'phone', 'email', 'gstin', 'pan', 'business_type', 'company_since', 'currency',
    'bank_name', 'account_holder', 'account_number', 'ifsc_code', 'branch', 'account_type',
    'default_invoice_type', 'default_payment_terms', 'default_place_of_supply', 'default_due_days',
    'show_company_logo', 'show_bank_details', 'show_signatory', 'print_amount_words', 'logo_path',
    'invoice_prefix', 'next_invoice_number',
    'cgst_rate', 'sgst_rate', 'igst_rate', 'reverse_charge_applicable', 'rcm_on_transport',
    'tds_applicable', 'tds_rate_default', 'eway_bill_mandatory', 'round_off_limit', 'rcm_clause',
    'twilio_account_sid', 'twilio_auth_token', 'twilio_from_number',
    'rc_lookup_api_key'
]

MODULE_LIST = ['Dashboard', 'Trips', 'Maintenance', 'Vehicles', 'Invoices', 'Payments',
               'Ledger', 'Reports', 'Settings', 'User Management', 'Expenses', 'Salaries']
ROLE_SUGGESTIONS = ['Administrator', 'Manager', 'Operations Head', 'Accountant', 'Supervisor', 'Driver']

def _users_with_stats(conn):
    rows = conn.execute("""SELECT * FROM users ORDER BY
                           CASE access_level WHEN 'Full Access' THEN 0 WHEN 'Read Only' THEN 1 ELSE 2 END,
                           username""").fetchall()
    users = [dict(r) for r in rows]
    total = len(users)
    admin_n = sum(1 for u in users if u['access_level'] == 'Full Access')
    readonly_n = sum(1 for u in users if u['access_level'] == 'Read Only')
    limited_n = sum(1 for u in users if u['access_level'] == 'Limited Access')
    inactive_n = sum(1 for u in users if u['status'] == 'Inactive')
    return users, total, admin_n, readonly_n, limited_n, inactive_n

@app.route('/settings', methods=['GET', 'POST'])
def settings_page():
    conn = get_db()
    if request.method == 'POST':
        form_type = request.form.get('form_type')
        field_groups = {
            'company': ['company_name', 'address', 'state', 'phone', 'email', 'gstin', 'pan', 'business_type', 'company_since', 'currency'],
            'banking': ['bank_name', 'account_holder', 'account_number', 'ifsc_code', 'branch', 'account_type'],
            'invoice_prefs': ['default_invoice_type', 'default_payment_terms', 'default_place_of_supply', 'default_due_days',
                               'show_company_logo', 'show_bank_details', 'show_signatory', 'print_amount_words'],
            'invoice_numbering': ['invoice_prefix', 'next_invoice_number'],
            'tax': ['cgst_rate', 'sgst_rate', 'igst_rate', 'reverse_charge_applicable', 'rcm_on_transport',
                    'tds_applicable', 'tds_rate_default', 'eway_bill_mandatory', 'round_off_limit'],
            'rcm': ['rcm_clause'],
            'sms': ['twilio_account_sid', 'twilio_auth_token', 'twilio_from_number'],
            'rc_lookup': ['rc_lookup_api_key'],
        }
        if form_type in field_groups:
            for key in field_groups[form_type]:
                if key in ['show_company_logo', 'show_bank_details', 'show_signatory', 'print_amount_words',
                           'reverse_charge_applicable', 'rcm_on_transport', 'tds_applicable', 'eway_bill_mandatory']:
                    val = '1' if request.form.get(key) == 'on' else '0'
                else:
                    val = request.form.get(key, '')
                conn.execute("UPDATE settings SET value=? WHERE key=?", (val, key))
            conn.commit()

    users, total_users, admin_users, readonly_users, limited_users, inactive_users = _users_with_stats(conn)
    role_counts = {}
    for u in users:
        r = u['role'] or 'Unassigned'
        role_counts[r] = role_counts.get(r, 0) + 1
    access_logs = conn.execute("""SELECT al.date, al.event, u.username, u.full_name FROM access_logs al
                                  LEFT JOIN users u ON al.user_id=u.id ORDER BY al.id DESC LIMIT 50""").fetchall()
    s = _get_all_settings(conn)
    conn.close()

    active_settings_tab = request.args.get('tab', 'company')
    invoice_example = f"{s['invoice_prefix']}/2026/{int(s['next_invoice_number'] or 1):04d}" if s['invoice_prefix'] else ''
    return render_template('settings.html', company_name=s.get('company_name') or get_company_name(),
                            users=users, total_users=total_users, admin_users=admin_users, readonly_users=readonly_users,
                            limited_users=limited_users, inactive_users=inactive_users, role_counts=role_counts,
                            access_logs=access_logs, module_list=MODULE_LIST, role_suggestions=ROLE_SUGGESTIONS,
                            s=s, invoice_example=invoice_example, active='settings', active_settings_tab=active_settings_tab)

def _get_all_settings(conn):
    s = {}
    for key in ALL_SETTING_KEYS:
        row = conn.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
        s[key] = row['value'] if row else ''
    return s

@app.route('/settings/users/add', methods=['POST'])
def add_user():
    from werkzeug.security import generate_password_hash
    import datetime, sqlite3, secrets
    f = request.form
    conn = get_db()
    access_level = f.get('access_level') or 'Read Only'
    modules = ','.join(request.form.getlist('modules')) if access_level == 'Limited Access' else None
    # A password is optional — leave it blank and the user signs in with mobile OTP only;
    # set one here and they can also sign in the normal username+password way.
    pw = f.get('password') or ''
    pw_hash = generate_password_hash(pw) if pw else generate_password_hash(secrets.token_hex(16))
    try:
        conn.execute("""INSERT INTO users (username, password_hash, role, is_admin, phone, full_name, email,
                        access_level, module_access, status, created_at)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                     (f.get('username'), pw_hash, f.get('role'),
                      1 if access_level == 'Full Access' else 0, (f.get('phone') or '').strip() or None,
                      f.get('full_name') or None, f.get('email') or None, access_level, modules,
                      f.get('status') or 'Active', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        conn.commit()
        conn.close()
    except sqlite3.IntegrityError:
        users, total_users, admin_users, readonly_users, limited_users, inactive_users = _users_with_stats(conn)
        role_counts = {}
        for u in users:
            r = u['role'] or 'Unassigned'
            role_counts[r] = role_counts.get(r, 0) + 1
        access_logs = conn.execute("""SELECT al.date, al.event, u.username, u.full_name FROM access_logs al
                                      LEFT JOIN users u ON al.user_id=u.id ORDER BY al.id DESC LIMIT 50""").fetchall()
        s = _get_all_settings(conn)
        invoice_example = f"{s['invoice_prefix']}/2026/{int(s['next_invoice_number'] or 1):04d}" if s['invoice_prefix'] else ''
        conn.close()
        return render_template('settings.html', company_name=s.get('company_name') or get_company_name(),
                                users=users, total_users=total_users, admin_users=admin_users, readonly_users=readonly_users,
                                limited_users=limited_users, inactive_users=inactive_users, role_counts=role_counts,
                                access_logs=access_logs, module_list=MODULE_LIST, role_suggestions=ROLE_SUGGESTIONS,
                                s=s, invoice_example=invoice_example, active='settings', active_settings_tab='users',
                                user_error=f"Username \"{f.get('username')}\" is already taken.", reopen_add_user=True)
    return redirect(url_for('settings_page', tab='users'))

@app.route('/settings/users/<int:user_id>/edit', methods=['POST'])
def edit_user(user_id):
    from werkzeug.security import generate_password_hash
    f = request.form
    conn = get_db()
    access_level = f.get('access_level') or 'Read Only'
    modules = ','.join(request.form.getlist('modules')) if access_level == 'Limited Access' else None
    conn.execute("""UPDATE users SET full_name=?, role=?, phone=?, email=?, access_level=?, module_access=?,
                    status=?, is_admin=? WHERE id=?""",
                 (f.get('full_name') or None, f.get('role'), (f.get('phone') or '').strip() or None,
                  f.get('email') or None, access_level, modules, f.get('status') or 'Active',
                  1 if access_level == 'Full Access' else 0, user_id))
    # Leave the password untouched unless a new one was actually typed in.
    new_pw = f.get('password') or ''
    if new_pw:
        conn.execute("UPDATE users SET password_hash=? WHERE id=?", (generate_password_hash(new_pw), user_id))
    conn.commit()
    conn.close()
    return redirect(url_for('settings_page', tab='users'))

@app.route('/settings/users/<int:user_id>/phone', methods=['POST'])
def update_user_phone(user_id):
    conn = get_db()
    conn.execute("UPDATE users SET phone=? WHERE id=?", ((request.form.get('phone') or '').strip() or None, user_id))
    conn.commit()
    conn.close()
    return redirect(url_for('settings_page', tab='users'))

@app.route('/settings/users/<int:user_id>/delete', methods=['POST'])
def delete_user(user_id):
    conn = get_db()
    conn.execute("DELETE FROM users WHERE id=?", (user_id,))
    conn.execute("DELETE FROM access_logs WHERE user_id=?", (user_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('settings_page', tab='users'))

@app.route('/settings/users/export')
def export_users():
    import openpyxl, io
    from flask import send_file
    conn = get_db()
    users, *_ = _users_with_stats(conn)
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Users'
    ws.append(['Full Name', 'Username', 'Role', 'Access Level', 'Mobile', 'Email', 'Status', 'Last Login'])
    for u in users:
        ws.append([u['full_name'] or '', u['username'], u['role'] or '', u['access_level'] or '',
                   u['phone'] or '', u['email'] or '', u['status'] or '', u['last_login'] or ''])
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='users.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/route-analytics')
def route_analytics():
    conn = get_db()
    tab = request.args.get('tab', 'best-routes')
    date_from = request.args.get('date_from', '2026-04-01')
    date_to = request.args.get('date_to', '2026-07-31')
    from_f = request.args.get('from', '')
    to_f = request.args.get('to', '')
    type_f = request.args.get('vehicle_type', '')

    if date_from > date_to:
        date_from, date_to = date_to, date_from

    # All origins/destinations ever seen (unfiltered), for the From/To dropdowns — independent of
    # the current filters, and independent of each other, so you can filter by From alone, To
    # alone, or both together for one exact route.
    all_routes_rows = conn.execute("SELECT DISTINCT from_loc, to_loc FROM trips").fetchall()
    all_from_list = sorted(set(_clean_loc(r['from_loc']) for r in all_routes_rows if r['from_loc']))
    all_to_list = sorted(set(_clean_loc(r['to_loc']) for r in all_routes_rows if r['to_loc']))

    trip_query = """SELECT id, from_loc, to_loc, type, rate, rate_type, billed_amount, fuel_amount, driver_adv_amount,
                    toll, parking, agent_commission, builty_expense, conductor_expense, fine, labour_charges,
                    puncture, urea, loading_expense, unloading_expense, wear_tear, weighbridge_charges,
                    other_expense, permit_charges, fixed_rate_amount, owner_rate, owner_rate_type, owner_fixed_amount, quantity
                    FROM trips WHERE date>=? AND date<=? AND (lr_number IS NULL OR lr_number NOT LIKE 'Empty%')"""
    params = [date_from, date_to]
    if type_f:
        trip_query += " AND type=?"
        params.append(type_f)
    trips = conn.execute(trip_query, params).fetchall()
    route_toll_map = _toll_by_trip(conn, [t['id'] for t in trips])

    # ---------- Shared per-route revenue/cost/profit (both tabs draw from this) ----------
    groups = {}
    for t in trips:
        cf, ct = _clean_loc(t['from_loc']), _clean_loc(t['to_loc'])
        if (from_f and cf != from_f) or (to_f and ct != to_f):
            continue
        d = groups.setdefault((cf, ct), {'trips': 0, 'revenue': 0, 'cost': 0})
        d['trips'] += 1
        d['revenue'] += t['billed_amount'] or 0
        cost = ((t['fuel_amount'] or 0) + (t['driver_adv_amount'] or 0) + _trip_toll(t, route_toll_map) + (t['parking'] or 0) +
                (t['agent_commission'] or 0) + (t['builty_expense'] or 0) + (t['conductor_expense'] or 0) + (t['fine'] or 0) +
                (t['labour_charges'] or 0) + (t['puncture'] or 0) + (t['urea'] or 0) + (t['loading_expense'] or 0) +
                (t['unloading_expense'] or 0) + (t['wear_tear'] or 0) + (t['weighbridge_charges'] or 0) +
                (t['other_expense'] or 0) + (t['permit_charges'] or 0))
        if t['type'] == 'Market':
            cost += (t['owner_fixed_amount'] or 0) if (t['owner_rate_type'] or 'PER_MT') == 'FIXED' else (t['owner_rate'] or 0) * (t['quantity'] or 0)
        d['cost'] += cost

    route_rows = []
    for (cf, ct), d in groups.items():
        profit = d['revenue'] - d['cost']
        margin = round(profit / d['revenue'] * 100, 1) if d['revenue'] else 0
        route_rows.append({'route': f"{cf} → {ct}", 'clean_from': cf, 'clean_to': ct, 'trips': d['trips'],
                            'revenue': d['revenue'], 'cost': d['cost'], 'profit': profit, 'margin': margin})

    total_routes = len(route_rows)
    total_trips_sel = sum(r['trips'] for r in route_rows)
    total_revenue_sel = sum(r['revenue'] for r in route_rows)
    total_cost_sel = sum(r['cost'] for r in route_rows)
    avg_margin = round((total_revenue_sel - total_cost_sel) / total_revenue_sel * 100, 1) if total_revenue_sel else 0

    rate_values = [t['rate'] for t in trips if t['rate_type'] == 'PER_MT' and (t['rate'] or 0) > 0
                   and (not from_f or _clean_loc(t['from_loc']) == from_f) and (not to_f or _clean_loc(t['to_loc']) == to_f)]
    avg_rate = round(sum(rate_values) / len(rate_values), 0) if rate_values else 0

    top5_by_margin = sorted(route_rows, key=lambda r: r['margin'], reverse=True)[:5]
    top5_by_revenue = sorted(route_rows, key=lambda r: r['revenue'], reverse=True)[:5]
    top_rev_max = max([r['revenue'] for r in top5_by_revenue], default=1) or 1
    for r in top5_by_revenue:
        r['pct'] = round(r['revenue'] / top_rev_max * 100, 1)

    # Route Summary table — same rows, paginated, sorted by revenue.
    route_rows.sort(key=lambda r: r['revenue'], reverse=True)
    page, per_page, total_pages = _paginate(request.args.get('page'), request.args.get('per_page'), total_routes,
                                             per_page_options=(10, 25, 50, 100), default_per_page=10)
    summary_rows = route_rows[(page - 1) * per_page: page * per_page]
    page_tokens = _page_tokens(page, total_pages)

    # ---------- Route Rates tab: rate stats grouped by (from, to, vehicle type) ----------
    # Driver Advance / Fuel Taken are tracked in parallel dicts (same grouping key, same trip set
    # that contributes the rate stats) rather than folded into rr_groups itself, so every existing
    # rr_groups reader below keeps working untouched.
    rr_groups = {}
    rr_driver_adv_groups = {}
    rr_fuel_groups = {}
    for t in trips:
        if t['rate_type'] != 'PER_MT' or not t['rate'] or t['rate'] <= 0:
            continue
        cf, ct = _clean_loc(t['from_loc']), _clean_loc(t['to_loc'])
        if (from_f and cf != from_f) or (to_f and ct != to_f):
            continue
        rr_groups.setdefault((cf, ct, t['type']), []).append(t['rate'])
        if t['driver_adv_amount']:
            rr_driver_adv_groups.setdefault((cf, ct, t['type']), []).append(t['driver_adv_amount'])
        if t['fuel_amount']:
            rr_fuel_groups.setdefault((cf, ct, t['type']), []).append(t['fuel_amount'])

    rr_rows = []
    for (cf, ct, ttype), rates in rr_groups.items():
        da = rr_driver_adv_groups.get((cf, ct, ttype), [])
        fu = rr_fuel_groups.get((cf, ct, ttype), [])
        rr_rows.append({'clean_from': cf, 'clean_to': ct, 'type': ttype, 'highest': max(rates),
                         'average': sum(rates) / len(rates), 'lowest': min(rates), 'trips': len(rates),
                         'da_lowest': min(da) if da else 0, 'da_average': (sum(da) / len(da)) if da else 0, 'da_highest': max(da) if da else 0,
                         'fuel_lowest': min(fu) if fu else 0, 'fuel_average': (sum(fu) / len(fu)) if fu else 0, 'fuel_highest': max(fu) if fu else 0})
    rr_rows.sort(key=lambda r: (r['clean_from'], r['clean_to']))

    # Overall Driver Advance / Fuel Taken KPI cards — every trip in the current filters (not just
    # the PER_MT-rate subset that feeds the rate table), since these are real costs paid regardless
    # of how the party was charged. Zero/blank amounts are excluded — "lowest" should mean the
    # smallest amount actually given, not an untouched trip.
    overall_driver_adv = [t['driver_adv_amount'] for t in trips if t['driver_adv_amount']
                           and (not from_f or _clean_loc(t['from_loc']) == from_f) and (not to_f or _clean_loc(t['to_loc']) == to_f)]
    overall_fuel = [t['fuel_amount'] for t in trips if t['fuel_amount']
                     and (not from_f or _clean_loc(t['from_loc']) == from_f) and (not to_f or _clean_loc(t['to_loc']) == to_f)]
    da_lowest = min(overall_driver_adv) if overall_driver_adv else 0
    da_average = round(sum(overall_driver_adv) / len(overall_driver_adv)) if overall_driver_adv else 0
    da_highest = max(overall_driver_adv) if overall_driver_adv else 0
    fuel_lowest = min(overall_fuel) if overall_fuel else 0
    fuel_average = round(sum(overall_fuel) / len(overall_fuel)) if overall_fuel else 0
    fuel_highest = max(overall_fuel) if overall_fuel else 0

    # Route Rates table — its own page number (rr_page) but shares the same per_page value/param
    # as the Best Routes summary table above, same two-pagers-one-per_page pattern Performance
    # already uses for its Driver/Vehicle tabs (d_page/v_page sharing one per_page).
    rr_total_count = len(rr_rows)
    rr_page, rr_per_page, rr_total_pages = _paginate(request.args.get('rr_page'), request.args.get('per_page'), rr_total_count,
                                                       per_page_options=(10, 25, 50, 100), default_per_page=10)
    rr_rows_page = rr_rows[(rr_page - 1) * rr_per_page: rr_page * rr_per_page]
    rr_page_tokens = _page_tokens(rr_page, rr_total_pages)

    all_rates_flat = [r for rates in rr_groups.values() for r in rates]
    highest_entry = None
    lowest_entry = None
    for (cf, ct, ttype), rates in rr_groups.items():
        h, l = max(rates), min(rates)
        if highest_entry is None or h > highest_entry[2]:
            highest_entry = (cf, ct, h)
        if lowest_entry is None or l < lowest_entry[2]:
            lowest_entry = (cf, ct, l)
    rr_avg_rate_overall = round(sum(all_rates_flat) / len(all_rates_flat), 0) if all_rates_flat else 0
    rr_total_routes = len(rr_groups)

    line_rates = [r for (cf, ct, tt), rates in rr_groups.items() if tt == 'Line' for r in rates]
    local_rates = [r for (cf, ct, tt), rates in rr_groups.items() if tt == 'Local' for r in rates]
    line_avg = round(sum(line_rates) / len(line_rates), 0) if line_rates else 0
    local_avg = round(sum(local_rates) / len(local_rates), 0) if local_rates else 0
    line_pct = round(len(line_rates) / len(all_rates_flat) * 100, 0) if all_rates_flat else 0

    rt_query = """SELECT substr(date,1,7) as month, AVG(rate) as avg_rate FROM trips
                  WHERE rate_type='PER_MT' AND rate > 0 AND date>=? AND date<=?
                  AND (lr_number IS NULL OR lr_number NOT LIKE 'Empty%')"""
    rt_params = [date_from, date_to]
    if type_f:
        rt_query += " AND type=?"
        rt_params.append(type_f)
    rt_query += " GROUP BY month ORDER BY month"
    monthly_trips = conn.execute(rt_query, rt_params).fetchall()
    trend_points = [{'month': m['month'], 'rate': round(m['avg_rate'] or 0)} for m in monthly_trips]
    svg_points = ''
    svg_labels = []
    if trend_points:
        rates = [p['rate'] for p in trend_points]
        rmin, rmax = min(rates), max(rates)
        rrange = (rmax - rmin) or 1
        n = len(trend_points)
        chart_w, chart_h = 400, 140
        coords = []
        for i, p in enumerate(trend_points):
            x = (i / (n - 1) * chart_w) if n > 1 else chart_w / 2
            y = chart_h - ((p['rate'] - rmin) / rrange * (chart_h - 20)) - 10
            coords.append((round(x, 1), round(y, 1)))
        svg_points = ' '.join(f"{x},{y}" for x, y in coords)
        svg_labels = [p['month'][5:] + '/' + p['month'][2:4] for p in trend_points]

    conn.close()

    from urllib.parse import urlencode
    base_params = request.args.to_dict()
    base_params.pop('page', None)
    base_params.pop('per_page', None)
    base_qs = urlencode(base_params)

    rr_base_params = request.args.to_dict()
    rr_base_params.pop('rr_page', None)
    rr_base_params.pop('per_page', None)
    rr_base_qs = urlencode(rr_base_params)

    return render_template('route_analytics.html',
        tab=tab, f_date_from=date_from, f_date_to=date_to, f_from=from_f, f_to=to_f, f_vehicle_type=type_f,
        all_from_list=all_from_list, all_to_list=all_to_list,
        total_routes=total_routes, total_trips_sel=total_trips_sel, total_revenue_sel=total_revenue_sel,
        avg_rate=avg_rate, avg_margin=avg_margin,
        top5_by_margin=top5_by_margin, top5_by_revenue=top5_by_revenue,
        summary_rows=summary_rows, page=page, per_page=per_page, total_pages=total_pages,
        page_tokens=page_tokens, base_qs=base_qs,
        rr_rows=rr_rows_page, rr_total_count=rr_total_count, rr_page=rr_page, rr_per_page=rr_per_page,
        rr_total_pages=rr_total_pages, rr_page_tokens=rr_page_tokens, rr_base_qs=rr_base_qs,
        highest_entry=highest_entry, lowest_entry=lowest_entry, rr_avg_rate_overall=rr_avg_rate_overall,
        rr_total_routes=rr_total_routes, line_avg=line_avg, local_avg=local_avg, line_pct=line_pct,
        da_lowest=da_lowest, da_average=da_average, da_highest=da_highest,
        fuel_lowest=fuel_lowest, fuel_average=fuel_average, fuel_highest=fuel_highest,
        trend_points=trend_points, svg_points=svg_points, svg_labels=svg_labels,
        active='route-analytics')

@app.route('/route-analytics/export')
def export_route_analytics():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from flask import send_file
    import io
    conn = get_db()
    date_from = request.args.get('date_from', '2026-04-01')
    date_to = request.args.get('date_to', '2026-07-31')
    from_f = request.args.get('from', '')
    to_f = request.args.get('to', '')
    type_f = request.args.get('vehicle_type', '')
    if date_from > date_to:
        date_from, date_to = date_to, date_from

    trip_query = """SELECT id, from_loc, to_loc, type, rate, rate_type, billed_amount, fuel_amount, driver_adv_amount,
                    toll, parking, agent_commission, builty_expense, conductor_expense, fine, labour_charges,
                    puncture, urea, loading_expense, unloading_expense, wear_tear, weighbridge_charges,
                    other_expense, permit_charges, fixed_rate_amount, owner_rate, owner_rate_type, owner_fixed_amount, quantity
                    FROM trips WHERE date>=? AND date<=? AND (lr_number IS NULL OR lr_number NOT LIKE 'Empty%')"""
    params = [date_from, date_to]
    if type_f:
        trip_query += " AND type=?"
        params.append(type_f)
    trips = conn.execute(trip_query, params).fetchall()
    route_toll_map = _toll_by_trip(conn, [t['id'] for t in trips])
    conn.close()

    groups = {}
    rr_groups = {}
    rr_driver_adv_groups = {}
    rr_fuel_groups = {}
    for t in trips:
        cf, ct = _clean_loc(t['from_loc']), _clean_loc(t['to_loc'])
        if (from_f and cf != from_f) or (to_f and ct != to_f):
            continue
        d = groups.setdefault((cf, ct), {'trips': 0, 'revenue': 0, 'cost': 0})
        d['trips'] += 1
        d['revenue'] += t['billed_amount'] or 0
        cost = ((t['fuel_amount'] or 0) + (t['driver_adv_amount'] or 0) + _trip_toll(t, route_toll_map) + (t['parking'] or 0) +
                (t['agent_commission'] or 0) + (t['builty_expense'] or 0) + (t['conductor_expense'] or 0) + (t['fine'] or 0) +
                (t['labour_charges'] or 0) + (t['puncture'] or 0) + (t['urea'] or 0) + (t['loading_expense'] or 0) +
                (t['unloading_expense'] or 0) + (t['wear_tear'] or 0) + (t['weighbridge_charges'] or 0) +
                (t['other_expense'] or 0) + (t['permit_charges'] or 0))
        if t['type'] == 'Market':
            cost += (t['owner_fixed_amount'] or 0) if (t['owner_rate_type'] or 'PER_MT') == 'FIXED' else (t['owner_rate'] or 0) * (t['quantity'] or 0)
        d['cost'] += cost
        if t['rate_type'] == 'PER_MT' and (t['rate'] or 0) > 0:
            rr_groups.setdefault((cf, ct, t['type']), []).append(t['rate'])
            if t['driver_adv_amount']:
                rr_driver_adv_groups.setdefault((cf, ct, t['type']), []).append(t['driver_adv_amount'])
            if t['fuel_amount']:
                rr_fuel_groups.setdefault((cf, ct, t['type']), []).append(t['fuel_amount'])

    route_rows = []
    for (cf, ct), d in groups.items():
        profit = d['revenue'] - d['cost']
        margin = round(profit / d['revenue'] * 100, 1) if d['revenue'] else 0
        route_rows.append((f"{cf} → {ct}", d['trips'], d['revenue'], d['cost'], profit, margin))
    route_rows.sort(key=lambda r: r[2], reverse=True)

    rate_rows = []
    for (cf, ct, ttype), rates in rr_groups.items():
        da = rr_driver_adv_groups.get((cf, ct, ttype), [])
        fu = rr_fuel_groups.get((cf, ct, ttype), [])
        rate_rows.append((f"{cf} → {ct}", ttype, max(rates), sum(rates) / len(rates), min(rates), len(rates),
                           min(da) if da else 0, (sum(da) / len(da)) if da else 0, max(da) if da else 0,
                           min(fu) if fu else 0, (sum(fu) / len(fu)) if fu else 0, max(fu) if fu else 0))
    rate_rows.sort(key=lambda r: r[0])

    navy = "1B2A4A"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=navy)

    wb = Workbook()
    ws = wb.active
    ws.title = "Route Summary"
    headers = ["Route", "Trips", "Total Revenue", "Total Cost", "Profit", "Margin %"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h); c.font = header_font; c.fill = header_fill
    for r_idx, row in enumerate(route_rows, 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    ws2 = wb.create_sheet("Route Rates")
    headers2 = ["Route", "Type", "Highest Rate", "Average Rate", "Lowest Rate", "Trips",
                "Driver Adv. Lowest", "Driver Adv. Average", "Driver Adv. Highest",
                "Fuel Lowest", "Fuel Average", "Fuel Highest"]
    for i, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=i, value=h); c.font = header_font; c.fill = header_fill
    for r_idx, row in enumerate(rate_rows, 2):
        for c_idx, val in enumerate(row, 1):
            ws2.cell(row=r_idx, column=c_idx, value=val)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='route_analytics_export.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/idle-tracker')
def idle_tracker():
    args = request.args.to_dict()
    return redirect(url_for('fleet_utilization', **args))

@app.route('/empty-runs')
def empty_runs():
    args = request.args.to_dict()
    return redirect(url_for('fleet_utilization', **args))

@app.route('/fleet-utilization')
def fleet_utilization():
    conn = get_db()
    date_from = request.args.get('date_from', '2026-04-01')
    date_to = request.args.get('date_to', '2026-07-31')
    vehicle_f = request.args.get('vehicle', '')
    type_f = request.args.get('type', '')

    # From Date must never be after To Date — swap rather than let the math go negative.
    if date_from > date_to:
        date_from, date_to = date_to, date_from

    d2 = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()

    # ---------- Idle tracker: active days = union of each trip's start..end_date span ----------
    query = "SELECT id, vehicle_no, type, registration_date FROM vehicles WHERE type IN ('Line','Local')"
    params = []
    if type_f:
        query += " AND type = ?"
        params.append(type_f)
    if vehicle_f:
        query += " AND vehicle_no = ?"
        params.append(vehicle_f)
    own_vehicles = conn.execute(query, params).fetchall()

    rows = []
    for v in own_vehicles:
        active_days, idle_days, total_days_v = _vehicle_active_idle_days(
            conn, v['id'], v['registration_date'], date_from, date_to)
        last_trip = conn.execute("SELECT MAX(date) FROM trips WHERE vehicle_id=?", (v['id'],)).fetchone()[0]
        status = 'Idle' if idle_days >= active_days else 'Active'
        empty_run_count = conn.execute("""SELECT COUNT(*) FROM trips WHERE vehicle_id=? AND lr_number LIKE 'Empty%'
                                          AND date>=? AND date<=?""", (v['id'], date_from, date_to)).fetchone()[0]
        vehicle_trip_count = conn.execute("SELECT COUNT(*) FROM trips WHERE vehicle_id=? AND date>=? AND date<=?",
                                           (v['id'], date_from, date_to)).fetchone()[0]
        empty_pct = round(empty_run_count / vehicle_trip_count * 100, 1) if vehicle_trip_count else 0
        idle_pct = round(idle_days / total_days_v * 100, 1) if total_days_v else 0
        rows.append({'vehicle_no': v['vehicle_no'], 'type': v['type'], 'group': v['type'],
                      'active_days': active_days, 'idle_days': idle_days, 'total_days': total_days_v,
                      'empty_run_count': empty_run_count, 'empty_pct': empty_pct, 'idle_pct': idle_pct,
                      'last_trip': last_trip, 'status': status})
    rows.sort(key=lambda r: r['idle_days'], reverse=True)

    total_vehicles = len(rows)
    active_vehicles = sum(1 for r in rows if r['status'] == 'Active')
    idle_vehicles = sum(1 for r in rows if r['status'] == 'Idle')
    active_pct = round((active_vehicles / total_vehicles * 100), 1) if total_vehicles > 0 else 0
    idle_vehicles_pct = round((idle_vehicles / total_vehicles * 100), 1) if total_vehicles > 0 else 0
    total_active_days = sum(r['active_days'] for r in rows)
    total_idle_days = sum(r['idle_days'] for r in rows)
    total_all_days = total_active_days + total_idle_days
    active_days_pct = round((total_active_days / total_all_days * 100), 1) if total_all_days > 0 else 0
    idle_days_pct = round((total_idle_days / total_all_days * 100), 1) if total_all_days > 0 else 0
    avg_idle_days = round(total_idle_days / total_vehicles, 1) if total_vehicles > 0 else 0
    most_idle = sorted(rows, key=lambda r: r['idle_days'], reverse=True)[:5]

    # ---------- Empty runs (same formulas as the old Empty Runs page — LR number starting with "Empty") ----------
    eq = """SELECT t.id, t.date, v.vehicle_no, t.from_loc, t.to_loc, t.fuel_amount, t.toll
            FROM trips t LEFT JOIN vehicles v ON t.vehicle_id=v.id
            WHERE t.lr_number LIKE 'Empty%'"""
    eparams = []
    if date_from:
        eq += " AND t.date >= ?"; eparams.append(date_from)
    if date_to:
        eq += " AND t.date <= ?"; eparams.append(date_to)
    if vehicle_f:
        eq += " AND v.vehicle_no = ?"; eparams.append(vehicle_f)
    eq += " ORDER BY t.date DESC"
    empty_rows = conn.execute(eq, eparams).fetchall()
    empty_toll_map = _toll_by_trip(conn, [r['id'] for r in empty_rows])
    total_fuel = sum(r['fuel_amount'] or 0 for r in empty_rows)
    total_toll = sum(_trip_toll(r, empty_toll_map) for r in empty_rows)
    total_empty_cost = total_fuel + total_toll
    empty_run_trips = len(empty_rows)

    total_trips_period = conn.execute("SELECT COUNT(*) FROM trips WHERE date>=? AND date<=?", (date_from, date_to)).fetchone()[0]
    empty_run_pct_overall = round(empty_run_trips / total_trips_period * 100, 1) if total_trips_period else 0

    # ---------- 12-week idle trend (own vehicles, same corrected active/idle formula, bucketed weekly) ----------
    trend = []
    for i in range(11, -1, -1):
        w_end = d2 - datetime.timedelta(days=7 * i)
        w_start = w_end - datetime.timedelta(days=6)
        week_idle = 0
        for v in own_vehicles:
            _, w_idle, _ = _vehicle_active_idle_days(conn, v['id'], v['registration_date'],
                                                       w_start.isoformat(), w_end.isoformat())
            week_idle += w_idle
        trend.append({'label': w_start.strftime('%-d %b'), 'idle_days': week_idle})
    trend_max = max([t['idle_days'] for t in trend], default=1) or 1
    chart_w, chart_h, pad_l, pad_r, pad_t, pad_b = 700, 220, 30, 20, 20, 34
    plot_w, plot_h, n = chart_w - pad_l - pad_r, chart_h - pad_t - pad_b, len(trend)
    chart_bottom = pad_t + plot_h
    for i, t in enumerate(trend):
        t['x'] = round(pad_l + (i * plot_w / (n - 1) if n > 1 else 0), 1)
        t['y'] = round(pad_t + plot_h - (t['idle_days'] / trend_max * plot_h if trend_max else 0), 1)
    y_ticks = []
    for k in range(5):
        tick_val = round(trend_max * k / 4)
        y_ticks.append({'value': tick_val, 'y': round(pad_t + plot_h - (tick_val / trend_max * plot_h if trend_max else 0), 1)})

    all_vehicles_list = conn.execute("SELECT vehicle_no FROM vehicles WHERE type IN ('Line','Local') ORDER BY vehicle_no").fetchall()
    conn.close()
    return render_template('fleet_utilization.html',
        rows=rows, most_idle=most_idle, empty_rows=empty_rows, trend=trend, trend_max=trend_max,
        chart_bottom=chart_bottom, y_ticks=y_ticks,
        total_vehicles=total_vehicles, active_vehicles=active_vehicles, idle_vehicles=idle_vehicles,
        active_pct=active_pct, idle_vehicles_pct=idle_vehicles_pct,
        total_active_days=total_active_days, total_idle_days=total_idle_days, active_days_pct=active_days_pct, idle_days_pct=idle_days_pct,
        avg_idle_days=avg_idle_days, empty_run_trips=empty_run_trips, empty_run_pct_overall=empty_run_pct_overall,
        total_fuel=total_fuel, total_toll=total_toll, total_empty_cost=total_empty_cost,
        f_date_from=date_from, f_date_to=date_to, f_vehicle=vehicle_f, f_type=type_f, vehicles=all_vehicles_list,
        active='fleet-utilization')

@app.route('/fleet-utilization/export')
def export_fleet_utilization():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from flask import send_file
    import io
    conn = get_db()
    date_from = request.args.get('date_from', '2026-04-01')
    date_to = request.args.get('date_to', '2026-07-31')
    vehicle_f = request.args.get('vehicle', '')
    type_f = request.args.get('type', '')
    if date_from > date_to:
        date_from, date_to = date_to, date_from

    query = "SELECT id, vehicle_no, type, registration_date FROM vehicles WHERE type IN ('Line','Local')"
    params = []
    if type_f:
        query += " AND type = ?"; params.append(type_f)
    if vehicle_f:
        query += " AND vehicle_no = ?"; params.append(vehicle_f)
    own_vehicles = conn.execute(query, params).fetchall()

    rows = []
    for v in own_vehicles:
        active_days, idle_days, total_days_v = _vehicle_active_idle_days(
            conn, v['id'], v['registration_date'], date_from, date_to)
        last_trip = conn.execute("SELECT MAX(date) FROM trips WHERE vehicle_id=?", (v['id'],)).fetchone()[0]
        status = 'Idle' if idle_days >= active_days else 'Active'
        empty_run_count = conn.execute("""SELECT COUNT(*) FROM trips WHERE vehicle_id=? AND lr_number LIKE 'Empty%'
                                          AND date>=? AND date<=?""", (v['id'], date_from, date_to)).fetchone()[0]
        rows.append((v['vehicle_no'], v['type'], active_days, idle_days, total_days_v, empty_run_count, last_trip or '', status))
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Fleet Utilization"
    headers = ["Vehicle No", "Type", "Active Days", "Idle Days", "Total Days", "Empty Run Trips", "Last Trip Date", "Status"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1B2A4A")
    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='fleet_utilization.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/trips/edit/<int:trip_id>', methods=['GET', 'POST'])
def edit_trip(trip_id):
    conn = get_db()
    if request.method == 'POST':
        f = request.form
        def n(key):
            return float(f.get(key) or 0)
        def get_or_create_vehicle(vno):
            if not vno or not vno.strip():
                return None
            vno = vno.strip()
            row = conn.execute("SELECT id FROM vehicles WHERE vehicle_no = ? COLLATE NOCASE", (vno,)).fetchone()
            if row:
                return row[0]
            cur = conn.execute("INSERT INTO vehicles (vehicle_no) VALUES (?)", (vno,))
            return cur.lastrowid

        vehicle_id = get_or_create_vehicle(f.get('vehicle_no'))
        party_id = get_or_create_party(conn, f.get('party_name'))
        fuel_vendor_id = get_or_create_vendor(conn, f.get('fuel_vendor'))
        driveradv_vendor_id = get_or_create_vendor(conn, f.get('driver_adv_vendor'))
        misc_vendor_id = get_or_create_vendor(conn, f.get('misc_vendor'))

        quantity = n('quantity')
        rate = n('rate')
        rate_type = f.get('rate_type')
        fixed_rate_amount = n('fixed_rate_amount')
        owner_rate_type = f.get('owner_rate_type') or 'PER_MT'
        owner_fixed_amount = n('owner_fixed_amount')
        freight = fixed_rate_amount if rate_type == 'FIXED' else quantity * rate
        # driver_payment ("Driver Bata") no longer has a form field — deliberately left out of both
        # the recompute and the UPDATE below so an existing trip's historical value is never touched.
        total_charges = (n('detention_charges')+n('gps_cost')+n('loading_charge')+
                          n('unloading_charge')+n('police_charges')+n('sim_tracking')+n('union_charges')+
                          n('weight_charges')+n('other_charges'))
        total_deductions = (n('brokerage')+n('builty_commission')+n('late_fees')+n('material_damage')+
                             n('shortage_amount')+n('tds')+n('other_deductions'))
        billed_amount = freight + total_charges - total_deductions

        owner_vendor_id = get_or_create_vendor(conn, f.get('owner_name')) if f.get('owner_name') else None

        # conductor_expense/wear_tear no longer have form fields — left out of the UPDATE entirely
        # (not just zeroed) so an existing trip's historical value is never touched, same pattern as
        # driver_payment above.
        conn.execute("""UPDATE trips SET
            date=?, lr_number=?, vehicle_id=?, type=?, party_id=?, from_loc=?, to_loc=?, quantity=?, rate=?,
            driver_name=?, material=?, rate_type=?, billed_amount=?,
            detention_charges=?, gps_cost=?, loading_charge=?, unloading_charge=?,
            police_charges=?, sim_tracking=?, union_charges=?, weight_charges=?, other_charges=?,
            brokerage=?, builty_commission=?, late_fees=?, material_damage=?, shortage_amount=?, shortage_qty=?, tds=?, other_deductions=?,
            fuel_amount=?, fuel_liters=?, fuel_price=?, driver_adv_amount=?, party_advance=?, payment_received=?, fuel_vendor_id=?, driver_adv_vendor_id=?,
            owner_name=?, fixed_rate_amount=?, owner_rate=?, owner_rate_type=?, owner_fixed_amount=?, paid_to_owner=?, owner_vendor_id=?,
            agent_commission=?, builty_expense=?, fine=?, labour_charges=?, parking=?, puncture=?,
            toll=?, urea=?, loading_expense=?, unloading_expense=?, weighbridge_charges=?, other_expense=?, misc_vendor_id=?,
            lr_received=?, is_empty=?
            WHERE id=?""",
            (f.get('date'), f.get('lr_number'), vehicle_id, f.get('type'), party_id, f.get('from_loc'), f.get('to_loc'),
             quantity, rate, f.get('driver_name'), f.get('material'), rate_type, billed_amount,
             n('detention_charges'), n('gps_cost'), n('loading_charge'), n('unloading_charge'),
             n('police_charges'), n('sim_tracking'), n('union_charges'), n('weight_charges'), n('other_charges'),
             n('brokerage'), n('builty_commission'), n('late_fees'), n('material_damage'), n('shortage_amount'),
             n('shortage_qty'), n('tds'), n('other_deductions'),
             n('fuel_amount'), f.get('fuel_liters') or None, n('fuel_price'), n('driver_adv_amount'), n('party_advance'), n('payment_received'), fuel_vendor_id, driveradv_vendor_id,
             f.get('owner_name'), fixed_rate_amount, n('owner_rate'), owner_rate_type, owner_fixed_amount, n('paid_to_owner'), owner_vendor_id,
             n('agent_commission'), n('builty_expense'), n('fine'), n('labour_charges'),
             n('parking'), n('puncture'), n('toll'), n('urea'), n('loading_expense'), n('unloading_expense'),
             n('weighbridge_charges'), n('other_expense'), misc_vendor_id,
             f.get('lr_received') or None, 1 if f.get('is_empty') else 0, trip_id))
        _save_trip_custom_items(conn, trip_id, f)
        conn.commit()
        conn.close()
        return_to = f.get('return_to')
        return redirect(return_to) if return_to else redirect(url_for('trips_list'))

    trip = conn.execute("""SELECT t.*, v.vehicle_no, p.name as party_name,
                           fv.name as fuel_vendor_name, dv.name as driveradv_vendor_name, mv.name as misc_vendor_name
                           FROM trips t
                           LEFT JOIN vehicles v ON t.vehicle_id=v.id
                           LEFT JOIN parties p ON t.party_id=p.id
                           LEFT JOIN vendors fv ON t.fuel_vendor_id=fv.id
                           LEFT JOIN vendors dv ON t.driver_adv_vendor_id=dv.id
                           LEFT JOIN vendors mv ON t.misc_vendor_id=mv.id
                           WHERE t.id=?""", (trip_id,)).fetchone()
    if not trip:
        conn.close()
        return redirect(url_for('trips_list'))
    custom_item_rows = conn.execute("""SELECT ii.description, ii.amount, ii.item_type, ve.name as vendor_name
                                       FROM invoice_items ii LEFT JOIN vendors ve ON ii.vendor_id=ve.id
                                       WHERE ii.trip_id=?""", (trip_id,)).fetchall()
    custom_items = [{'description': r['description'], 'item_type': r['item_type'], 'rate': r['amount'], 'quantity': 1,
                      'vendor_name': r['vendor_name'] or ''}
                     for r in custom_item_rows]
    conn.close()
    vehicles, parties, vendors, combined_names = _get_autocomplete_lists()
    conn2 = get_db()
    employees = conn2.execute("SELECT name FROM employees ORDER BY name").fetchall()
    conn2.close()
    return render_template('trip_form.html', mode='edit', t=dict(trip), custom_items=custom_items,
                            vehicles=vehicles, parties=parties, vendors=vendors, combined_names=combined_names,
                            employees=employees, return_to=request.args.get('return_to', ''), active='trips')

@app.route('/trips/<int:trip_id>/view')
def trip_view(trip_id):
    """Read-only trip detail — opened from a Ledger row's Trip Bill entry in the slide-in side
    panel (see openDetailPanel in base.html), instead of routing to the editable Trip form. Same
    fields as trip_form.html/edit_trip, just displayed rather than edited; a footer link still
    goes to the real Edit Trip page for anyone who does need to make a change."""
    conn = get_db()
    trip = conn.execute("""SELECT t.*, v.vehicle_no, p.name as party_name,
                           fv.name as fuel_vendor_name, dv.name as driveradv_vendor_name, mv.name as misc_vendor_name
                           FROM trips t
                           LEFT JOIN vehicles v ON t.vehicle_id=v.id
                           LEFT JOIN parties p ON t.party_id=p.id
                           LEFT JOIN vendors fv ON t.fuel_vendor_id=fv.id
                           LEFT JOIN vendors dv ON t.driver_adv_vendor_id=dv.id
                           LEFT JOIN vendors mv ON t.misc_vendor_id=mv.id
                           WHERE t.id=?""", (trip_id,)).fetchone()
    if not trip:
        conn.close()
        return redirect(url_for('trips_list'))
    custom_item_rows = conn.execute("""SELECT ii.description, ii.amount, ii.item_type, ve.name as vendor_name
                                       FROM invoice_items ii LEFT JOIN vendors ve ON ii.vendor_id=ve.id
                                       WHERE ii.trip_id=?""", (trip_id,)).fetchall()
    conn.close()
    t = dict(trip)
    is_market = t.get('type') == 'Market'
    owner_amount = 0
    if is_market:
        if (t.get('owner_rate_type') or 'PER_MT') == 'FIXED':
            owner_amount = t.get('owner_fixed_amount') or 0
        else:
            owner_amount = (t.get('owner_rate') or 0) * (t.get('quantity') or 0)
    return render_template('trip_view.html', t=t, custom_items=custom_item_rows, is_market=is_market,
                            owner_amount=owner_amount, active='accounts')

@app.route('/business-performance')
def business_performance():
    conn = get_db()
    # Default to the last fully completed calendar month (standard CEO-dashboard convention).
    last_month_end = datetime.date.today().replace(day=1) - datetime.timedelta(days=1)
    default_from, default_to = _month_bounds(last_month_end.year, last_month_end.month)
    date_from = request.args.get('date_from') or default_from
    date_to = request.args.get('date_to') or default_to

    curr = _period_financials(conn, date_from, date_to)
    prev_from, prev_to = _shift_period_back(date_from, date_to)
    prev = _period_financials(conn, prev_from, prev_to)
    q_from, q_to = _quarter_bounds(datetime.datetime.strptime(date_to, '%Y-%m-%d').date())
    pq_from, pq_to = _shift_period_back(q_from, q_to)
    curr_q = _period_financials(conn, q_from, q_to)
    prev_q = _period_financials(conn, pq_from, pq_to)
    y_from, y_to = _shift_period_year(date_from, date_to)
    curr_y = _period_financials(conn, y_from, y_to)

    days_in_period = (datetime.datetime.strptime(date_to, '%Y-%m-%d').date() -
                       datetime.datetime.strptime(date_from, '%Y-%m-%d').date()).days + 1

    revenue_growth = _pct_growth(curr['revenue'], prev['revenue'])
    profit_growth = _pct_growth(curr['net_profit'], prev['net_profit'])
    expense_growth = _pct_growth(curr['total_expenses'], prev['total_expenses'])
    profit_growth_qoq = _pct_growth(curr_q['net_profit'], prev_q['net_profit'])
    profit_growth_yoy = _pct_growth(curr['net_profit'], curr_y['net_profit'])
    revenue_growth_qoq = _pct_growth(curr_q['revenue'], prev_q['revenue'])
    revenue_growth_yoy = _pct_growth(curr['revenue'], curr_y['revenue'])

    # Outstanding receivables / payables — all-time snapshot balances, same formula as accounts().
    party_rows = conn.execute("""SELECT p.id, p.name,
        (SELECT COALESCE(SUM(billed_amount),0) FROM trips WHERE party_id=p.id) -
        (SELECT COALESCE(SUM(payment_received)+SUM(party_advance),0) FROM trips WHERE party_id=p.id) -
        (SELECT COALESCE(SUM(amount-COALESCE(allocated_amount,0)),0) FROM payments WHERE party_id=p.id AND payment_type='received') +
        COALESCE(p.opening_balance,0) as balance
        FROM parties p""").fetchall()
    party_balance = {r['id']: (r['balance'] or 0) for r in party_rows}
    party_name = {r['id']: r['name'] for r in party_rows}
    total_receivables = sum(b for b in party_balance.values() if b > 0)

    vendor_rows = conn.execute("""SELECT v.id, v.name,
        (SELECT COALESCE(SUM(m.amount),0) FROM maintenance m WHERE m.vendor_id=v.id) +
        (SELECT COALESCE(SUM(fuel_amount),0) FROM trips WHERE fuel_vendor_id=v.id) +
        (SELECT COALESCE(SUM(driver_adv_amount),0) FROM trips WHERE driver_adv_vendor_id=v.id) +
        (SELECT COALESCE(SUM(CASE WHEN owner_rate_type='FIXED' THEN owner_fixed_amount ELSE owner_rate*quantity END),0) FROM trips WHERE owner_vendor_id=v.id) +
        (SELECT COALESCE(SUM(CASE WHEN item_type='charge' THEN amount ELSE -amount END),0) FROM invoice_items WHERE vendor_id=v.id) -
        (SELECT COALESCE(SUM(m.paid_amount),0) FROM maintenance m WHERE m.vendor_id=v.id) -
        (SELECT COALESCE(SUM(paid_to_owner),0) FROM trips WHERE owner_vendor_id=v.id) -
        (SELECT COALESCE(SUM(amount-COALESCE(allocated_amount,0)),0) FROM payments WHERE vendor_id=v.id AND payment_type='paid') -
        COALESCE(v.opening_balance,0) as balance
        FROM vendors v WHERE v.linked_party_id IS NULL""").fetchall()
    total_payables = sum((r['balance'] or 0) for r in vendor_rows if (r['balance'] or 0) > 0)

    # Cash movement for the selected period.
    cash_collected = conn.execute("SELECT COALESCE(SUM(amount),0) FROM payments WHERE payment_type='received' AND date>=? AND date<=?",
                                   (date_from, date_to)).fetchone()[0]
    cash_collected += sum(t['party_advance'] or 0 for t in curr['trips'])
    cash_paid = conn.execute("SELECT COALESCE(SUM(amount),0) FROM payments WHERE payment_type='paid' AND date>=? AND date<=?",
                              (date_from, date_to)).fetchone()[0]
    cash_paid += curr['maint'] + curr['overheads'] + curr['salaries']

    prev_cash_collected = conn.execute("SELECT COALESCE(SUM(amount),0) FROM payments WHERE payment_type='received' AND date>=? AND date<=?",
                                        (prev_from, prev_to)).fetchone()[0]
    prev_cash_collected += sum(t['party_advance'] or 0 for t in prev['trips'])
    collection_growth = _pct_growth(cash_collected, prev_cash_collected)
    collection_efficiency = round(cash_collected / curr['revenue'] * 100, 2) if curr['revenue'] else 0

    avg_revenue_per_trip = round(curr['revenue'] / curr['trip_count'], 2) if curr['trip_count'] else 0
    avg_profit_per_trip = round(curr['net_profit'] / curr['trip_count'], 2) if curr['trip_count'] else 0
    operating_ratio = round(curr['total_expenses'] / curr['revenue'] * 100, 2) if curr['revenue'] else 0
    ebitda = curr['net_profit']  # no interest/tax/depreciation tracked separately, so this is profit-based estimate

    # Financial Overview / Monthly Profit Trend — trailing 6 months ending at date_to's month.
    end_d = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()
    trailing_months = []
    yy_, mm_ = end_d.year, end_d.month
    for i in range(5, -1, -1):
        mm = mm_ - i
        yy = yy_
        while mm <= 0:
            mm += 12
            yy -= 1
        trailing_months.append((yy, mm))
    monthly_rows = []
    for (yy, mm) in trailing_months:
        mf, mt = _month_bounds(yy, mm)
        mfin = _period_financials(conn, mf, mt)
        monthly_rows.append({'label': calendar.month_abbr[mm], 'revenue': mfin['revenue'],
                              'expenses': mfin['total_expenses'], 'net_profit': mfin['net_profit']})

    # Top customers by revenue (period).
    party_period = {}
    curr_toll_map = _toll_by_trip(conn, [t['id'] for t in curr['trips']])
    for t in curr['trips']:
        if not t['party_id']:
            continue
        d = party_period.setdefault(t['party_id'], {'revenue': 0, 'direct_cost': 0, 'trips': 0})
        d['revenue'] += t['billed_amount'] or 0
        d['trips'] += 1
        direct = (t['fuel_amount'] or 0) + (t['driver_adv_amount'] or 0) + _trip_toll(t, curr_toll_map) + (t['parking'] or 0)
        if t['type'] == 'Market':
            direct += (t['owner_fixed_amount'] or 0) if (t['owner_rate_type'] or 'PER_MT') == 'FIXED' else (t['owner_rate'] or 0) * (t['quantity'] or 0)
        d['direct_cost'] += direct
    top_customers = []
    for pid, d in sorted(party_period.items(), key=lambda kv: kv[1]['revenue'], reverse=True)[:6]:
        top_customers.append({
            'name': party_name.get(pid, '—'), 'revenue': d['revenue'],
            'pct': round(d['revenue'] / curr['revenue'] * 100, 2) if curr['revenue'] else 0,
            'profit': d['revenue'] - d['direct_cost'],
            'outstanding': max(party_balance.get(pid, 0), 0),
        })
    customers_other_revenue = curr['revenue'] - sum(c['revenue'] for c in top_customers)
    customer_growth = _pct_growth(len(party_period), len(prev['party_ids']))

    # Expense breakdown (period).
    exp_cats = [
        ('Fuel', curr['fuel'], '#2a78d6'), ('Driver Advance', curr['driver_adv'], '#eb6834'),
        ('Maintenance', curr['maint'], '#1baf7a'), ('Salary', curr['salaries'], '#eda100'),
        ('Toll / Parking', curr['toll'] + curr['parking'], '#e87ba4'),
        ('Other Expenses', curr['misc'] + curr['overheads'] + curr['owner_cost'], '#008300'),
    ]
    exp_total = sum(v for _, v, _ in exp_cats) or 1
    expense_breakdown = [{'label': l, 'amount': v, 'color': c, 'pct': round(v / exp_total * 100, 2)} for l, v, c in exp_cats]

    # Collections analytics — age each party's *net* balance (party_balance, already netted against
    # opening balance and unallocated payments) off their oldest still-pending trip, so overdue_total
    # stays a subset of total_receivables instead of double-counting against raw per-trip pending.
    all_trips = conn.execute("SELECT id, date, lr_number, party_id, billed_amount, payment_received, party_advance FROM trips").fetchall()
    party_pending_trips = {}
    for t in all_trips:
        if not t['party_id'] or not t['date']:
            continue
        pending = (t['billed_amount'] or 0) - (t['payment_received'] or 0) - (t['party_advance'] or 0)
        if pending <= 0.01:
            continue
        party_pending_trips.setdefault(t['party_id'], []).append(t)

    overdue_total = 0
    oldest_outstanding = None
    for pid, bal in party_balance.items():
        if bal <= 0.01:
            continue
        trips_for_party = party_pending_trips.get(pid)
        if not trips_for_party:
            continue
        oldest_trip = min(trips_for_party, key=lambda t: t['date'])
        t_date = datetime.datetime.strptime(oldest_trip['date'], '%Y-%m-%d').date()
        age = (end_d - t_date).days
        if age > 30:
            overdue_total += bal
        if oldest_outstanding is None or t_date < oldest_outstanding['date']:
            oldest_outstanding = {'date': t_date, 'lr': oldest_trip['lr_number'] or f"Trip #{oldest_trip['id']}", 'age': age}

    alloc_rows = conn.execute("""SELECT pay.date as pay_date, t.date as trip_date FROM payment_allocations pa
                                 JOIN payments pay ON pa.payment_id = pay.id
                                 JOIN trips t ON pa.trip_id = t.id
                                 WHERE pay.date>=? AND pay.date<=?""", (date_from, date_to)).fetchall()
    if not alloc_rows:
        alloc_rows = conn.execute("""SELECT pay.date as pay_date, t.date as trip_date FROM payment_allocations pa
                                     JOIN payments pay ON pa.payment_id = pay.id
                                     JOIN trips t ON pa.trip_id = t.id""").fetchall()
    collection_days_list = []
    for r in alloc_rows:
        if not r['pay_date'] or not r['trip_date']:
            continue
        pd_ = datetime.datetime.strptime(r['pay_date'], '%Y-%m-%d').date()
        td_ = datetime.datetime.strptime(r['trip_date'], '%Y-%m-%d').date()
        collection_days_list.append((pd_ - td_).days)
    avg_collection_days = round(sum(collection_days_list) / len(collection_days_list), 1) if collection_days_list else None

    # Highest / lowest revenue & profit month across all months with data.
    all_months_rows = conn.execute("SELECT DISTINCT substr(date,1,7) as ym FROM trips ORDER BY ym").fetchall()
    month_fin = []
    for r in all_months_rows:
        yy, mm = int(r['ym'][:4]), int(r['ym'][5:7])
        mf, mt = _month_bounds(yy, mm)
        f = _period_financials(conn, mf, mt)
        month_fin.append({'label': f"{calendar.month_abbr[mm]} {yy}", 'revenue': f['revenue'], 'net_profit': f['net_profit']})
    highest_revenue_month = max(month_fin, key=lambda r: r['revenue']) if month_fin else None
    highest_profit_month = max(month_fin, key=lambda r: r['net_profit']) if month_fin else None
    lowest_profit_month = min(month_fin, key=lambda r: r['net_profit']) if month_fin else None

    # Highest collection / expense day within the selected period.
    day_collect = {}
    for r in conn.execute("SELECT date, SUM(amount) as amt FROM payments WHERE payment_type='received' AND date>=? AND date<=? GROUP BY date",
                           (date_from, date_to)).fetchall():
        day_collect[r['date']] = day_collect.get(r['date'], 0) + (r['amt'] or 0)
    for t in curr['trips']:
        if t['party_advance']:
            day_collect[t['date']] = day_collect.get(t['date'], 0) + t['party_advance']
    highest_collection_day = max(day_collect.items(), key=lambda kv: kv[1]) if day_collect else None

    day_expense = {}
    for t in curr['trips']:
        # Toll excluded here on purpose — the maintenance-by-date sum just below already carries
        # Toll Management's real toll cost (category='Toll' rows), dated per entry; adding a
        # per-trip toll figure too would double-count it, same reasoning as Dashboard's total.
        direct = ((t['fuel_amount'] or 0) + (t['driver_adv_amount'] or 0) + (t['parking'] or 0) +
                  (t['agent_commission'] or 0) + (t['builty_expense'] or 0) + (t['conductor_expense'] or 0) + (t['fine'] or 0) +
                  (t['labour_charges'] or 0) + (t['puncture'] or 0) + (t['urea'] or 0) + (t['loading_expense'] or 0) +
                  (t['unloading_expense'] or 0) + (t['wear_tear'] or 0) + (t['weighbridge_charges'] or 0) +
                  (t['other_expense'] or 0) + (t['permit_charges'] or 0))
        if t['type'] == 'Market':
            direct += (t['owner_fixed_amount'] or 0) if (t['owner_rate_type'] or 'PER_MT') == 'FIXED' else (t['owner_rate'] or 0) * (t['quantity'] or 0)
        day_expense[t['date']] = day_expense.get(t['date'], 0) + direct
    for r in conn.execute("SELECT date, SUM(amount) as amt FROM maintenance WHERE date>=? AND date<=? GROUP BY date", (date_from, date_to)).fetchall():
        day_expense[r['date']] = day_expense.get(r['date'], 0) + (r['amt'] or 0)
    for r in conn.execute("SELECT date, SUM(amount) as amt FROM overheads WHERE date>=? AND date<=? GROUP BY date", (date_from, date_to)).fetchall():
        day_expense[r['date']] = day_expense.get(r['date'], 0) + (r['amt'] or 0)
    highest_expense_day = max(day_expense.items(), key=lambda kv: kv[1]) if day_expense else None

    avg_daily_revenue = round(curr['revenue'] / days_in_period, 2) if days_in_period else 0
    avg_daily_profit = round(curr['net_profit'] / days_in_period, 2) if days_in_period else 0

    own_vehicle_count = conn.execute("SELECT COUNT(*) FROM vehicles WHERE type IN ('Line','Local')").fetchone()[0]
    own_revenue = sum(t['billed_amount'] or 0 for t in curr['trips'] if t['type'] in ('Line', 'Local'))
    avg_revenue_per_vehicle = round(own_revenue / own_vehicle_count, 2) if own_vehicle_count else 0
    employee_count = conn.execute("SELECT COUNT(*) FROM employees").fetchone()[0]
    avg_revenue_per_employee = round(curr['revenue'] / employee_count, 2) if employee_count else 0

    invoice_count_curr = conn.execute("SELECT COUNT(*) FROM invoice_batches WHERE invoice_date>=? AND invoice_date<=?",
                                       (date_from, date_to)).fetchone()[0]
    invoice_count_prev = conn.execute("SELECT COUNT(*) FROM invoice_batches WHERE invoice_date>=? AND invoice_date<=?",
                                       (prev_from, prev_to)).fetchone()[0]
    invoice_growth = _pct_growth(invoice_count_curr, invoice_count_prev)
    avg_invoice_value = round(curr['revenue'] / invoice_count_curr, 2) if invoice_count_curr else avg_revenue_per_trip
    trip_growth = _pct_growth(curr['trip_count'], prev['trip_count'])

    # Rule-based ("AI") insights, generated from the numbers above — not a live model call.
    insights = []
    if revenue_growth is not None:
        if revenue_growth >= 0:
            insights.append({'kind': 'good', 'text': f"Revenue increased by {revenue_growth}% compared to last month."})
        else:
            insights.append({'kind': 'bad', 'text': f"Revenue dropped by {abs(revenue_growth)}% compared to last month."})
    if profit_growth is not None:
        if profit_growth >= 0:
            insights.append({'kind': 'good', 'text': f"Profit increased by {profit_growth}% due to better revenue and controlled expenses."})
        else:
            insights.append({'kind': 'bad', 'text': f"Profit fell by {abs(profit_growth)}% — expenses grew faster than revenue."})
    prev_cats = {
        'Fuel': prev['fuel'], 'Driver Advance': prev['driver_adv'], 'Maintenance': prev['maint'],
        'Salary': prev['salaries'], 'Toll / Parking': prev['toll'] + prev['parking'],
        'Other Expenses': prev['misc'] + prev['overheads'] + prev['owner_cost'],
    }
    cat_growth = []
    for cat in expense_breakdown:
        g = _pct_growth(cat['amount'], prev_cats.get(cat['label'], 0))
        if g is not None and g > 0:
            cat_growth.append((cat['label'], g))
    if cat_growth:
        top_cat, top_g = max(cat_growth, key=lambda x: x[1])
        insights.append({'kind': 'warn', 'text': f"{top_cat} expenses increased by {top_g}%. Review costs in this category."})
    if total_receivables > 0:
        insights.append({'kind': 'info', 'text': f"Outstanding receivables stand at ₹{total_receivables:,.0f}. Focus on collections."})
    if avg_collection_days is not None and avg_collection_days > 30:
        insights.append({'kind': 'bad', 'text': f"Average collection period is {avg_collection_days} days. Consider tighter follow-up on pending invoices."})

    # Business Health Score — five sub-scores, each 0-100, averaged.
    def clamp(v, lo=0, hi=100):
        return max(lo, min(hi, v))
    profitability_score = clamp(round(curr['margin'] / 30 * 100)) if curr['revenue'] else 0
    financial_score = clamp(round(200 - 2 * operating_ratio)) if curr['revenue'] else 0
    collection_score = clamp(round(100 - (overdue_total / total_receivables * 100))) if total_receivables > 0 else 100
    cashflow_score = clamp(round(50 + (cash_collected - cash_paid) / curr['revenue'] * 100)) if curr['revenue'] else 50
    growth_score = clamp(round(50 + (revenue_growth or 0) * 2))
    sub_scores = [
        {'label': 'Financial Health', 'score': financial_score},
        {'label': 'Profitability', 'score': profitability_score},
        {'label': 'Collection Health', 'score': collection_score},
        {'label': 'Cash Flow', 'score': cashflow_score},
        {'label': 'Growth', 'score': growth_score},
    ]
    overall_score = round(sum(s['score'] for s in sub_scores) / len(sub_scores))
    if overall_score >= 80:
        health_label, health_kind = 'Healthy', 'good'
    elif overall_score >= 60:
        health_label, health_kind = 'Stable', 'info'
    elif overall_score >= 40:
        health_label, health_kind = 'Needs Attention', 'warn'
    else:
        health_label, health_kind = 'Critical', 'bad'
    weakest = min(sub_scores, key=lambda s: s['score'])
    suggestions = []
    if weakest['score'] < 70:
        tips = {
            'Financial Health': 'Bring operating expenses down relative to revenue to strengthen financial health.',
            'Profitability': 'Review pricing on low-margin routes and trim direct trip costs to lift profitability.',
            'Collection Health': 'Chase overdue receivables past 30 days — this is dragging down collection health.',
            'Cash Flow': 'Cash paid is outpacing cash collected this period — prioritise follow-up on pending payments.',
            'Growth': 'Revenue growth has slowed — look for new routes or repeat business with top customers.',
        }
        suggestions.append(f"{weakest['label']} is your weakest score ({weakest['score']}). {tips.get(weakest['label'], '')}")
    if overall_score >= 80:
        suggestions.append('Great job! Your business is performing well. Keep focusing on collections and cost optimization.')

    fy_map = {'2026-04-01|2027-03-31': ('2026-04-01', '2027-03-31'), '2025-04-01|2026-03-31': ('2025-04-01', '2026-03-31')}
    f_fy = ''
    for key, (fy_from, fy_to) in fy_map.items():
        if date_from == fy_from and date_to == fy_to:
            f_fy = key
            break

    conn.close()
    return render_template('business_performance.html',
        f_date_from=date_from, f_date_to=date_to, f_fy=f_fy,
        total_revenue=curr['revenue'], net_profit=curr['net_profit'], profit_margin=curr['margin'],
        total_receivables=total_receivables, total_payables=total_payables,
        cash_collected=cash_collected, cash_paid=cash_paid, collection_efficiency=collection_efficiency,
        revenue_growth=revenue_growth, expense_growth=expense_growth, profit_growth=profit_growth,
        avg_revenue_per_trip=avg_revenue_per_trip, avg_profit_per_trip=avg_profit_per_trip,
        operating_ratio=operating_ratio, ebitda=ebitda, trip_count=curr['trip_count'],
        monthly_rows=monthly_rows,
        profit_growth_mom=profit_growth, profit_growth_qoq=profit_growth_qoq, profit_growth_yoy=profit_growth_yoy,
        revenue_growth_mom=revenue_growth, revenue_growth_qoq=revenue_growth_qoq, revenue_growth_yoy=revenue_growth_yoy,
        curr_revenue=curr['revenue'], prev_revenue=prev['revenue'], revenue_diff=curr['revenue'] - prev['revenue'],
        curr_q_revenue=curr_q['revenue'], prev_q_revenue=prev_q['revenue'], curr_y_revenue=curr_y['revenue'],
        top_customers=top_customers, customers_other_revenue=customers_other_revenue,
        expense_breakdown=expense_breakdown, exp_total=exp_total,
        overdue_total=overdue_total, oldest_outstanding=oldest_outstanding, avg_collection_days=avg_collection_days,
        customer_growth=customer_growth, trip_growth=trip_growth, invoice_growth=invoice_growth, collection_growth=collection_growth,
        highest_revenue_month=highest_revenue_month, highest_profit_month=highest_profit_month, lowest_profit_month=lowest_profit_month,
        highest_collection_day=highest_collection_day, highest_expense_day=highest_expense_day,
        avg_daily_revenue=avg_daily_revenue, avg_daily_profit=avg_daily_profit, avg_invoice_value=avg_invoice_value,
        avg_revenue_per_vehicle=avg_revenue_per_vehicle, avg_revenue_per_employee=avg_revenue_per_employee,
        insights=insights, sub_scores=sub_scores, overall_score=overall_score, health_label=health_label, health_kind=health_kind,
        suggestions=suggestions,
        active='business-performance')

@app.route('/business-performance/export')
def export_business_performance():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from flask import send_file
    import io
    conn = get_db()
    last_month_end = datetime.date.today().replace(day=1) - datetime.timedelta(days=1)
    default_from, default_to = _month_bounds(last_month_end.year, last_month_end.month)
    date_from = request.args.get('date_from') or default_from
    date_to = request.args.get('date_to') or default_to

    curr = _period_financials(conn, date_from, date_to)

    party_rows = conn.execute("""SELECT p.id, p.name,
        (SELECT COALESCE(SUM(billed_amount),0) FROM trips WHERE party_id=p.id) -
        (SELECT COALESCE(SUM(payment_received)+SUM(party_advance),0) FROM trips WHERE party_id=p.id) -
        (SELECT COALESCE(SUM(amount-COALESCE(allocated_amount,0)),0) FROM payments WHERE party_id=p.id AND payment_type='received') +
        COALESCE(p.opening_balance,0) as balance
        FROM parties p""").fetchall()
    party_balance = {r['id']: (r['balance'] or 0) for r in party_rows}
    party_name = {r['id']: r['name'] for r in party_rows}
    total_receivables = sum(b for b in party_balance.values() if b > 0)

    vendor_rows = conn.execute("""SELECT v.id, v.name,
        (SELECT COALESCE(SUM(m.amount),0) FROM maintenance m WHERE m.vendor_id=v.id) +
        (SELECT COALESCE(SUM(fuel_amount),0) FROM trips WHERE fuel_vendor_id=v.id) +
        (SELECT COALESCE(SUM(driver_adv_amount),0) FROM trips WHERE driver_adv_vendor_id=v.id) +
        (SELECT COALESCE(SUM(CASE WHEN owner_rate_type='FIXED' THEN owner_fixed_amount ELSE owner_rate*quantity END),0) FROM trips WHERE owner_vendor_id=v.id) +
        (SELECT COALESCE(SUM(CASE WHEN item_type='charge' THEN amount ELSE -amount END),0) FROM invoice_items WHERE vendor_id=v.id) -
        (SELECT COALESCE(SUM(m.paid_amount),0) FROM maintenance m WHERE m.vendor_id=v.id) -
        (SELECT COALESCE(SUM(paid_to_owner),0) FROM trips WHERE owner_vendor_id=v.id) -
        (SELECT COALESCE(SUM(amount-COALESCE(allocated_amount,0)),0) FROM payments WHERE vendor_id=v.id AND payment_type='paid') -
        COALESCE(v.opening_balance,0) as balance
        FROM vendors v WHERE v.linked_party_id IS NULL""").fetchall()
    total_payables = sum((r['balance'] or 0) for r in vendor_rows if (r['balance'] or 0) > 0)

    cash_collected = conn.execute("SELECT COALESCE(SUM(amount),0) FROM payments WHERE payment_type='received' AND date>=? AND date<=?",
                                   (date_from, date_to)).fetchone()[0]
    cash_collected += sum(t['party_advance'] or 0 for t in curr['trips'])
    cash_paid = conn.execute("SELECT COALESCE(SUM(amount),0) FROM payments WHERE payment_type='paid' AND date>=? AND date<=?",
                              (date_from, date_to)).fetchone()[0]
    cash_paid += curr['maint'] + curr['overheads'] + curr['salaries']
    collection_efficiency = round(cash_collected / curr['revenue'] * 100, 2) if curr['revenue'] else 0
    operating_ratio = round(curr['total_expenses'] / curr['revenue'] * 100, 2) if curr['revenue'] else 0

    party_period = {}
    export_toll_map = _toll_by_trip(conn, [t['id'] for t in curr['trips']])
    for t in curr['trips']:
        if not t['party_id']:
            continue
        d = party_period.setdefault(t['party_id'], {'revenue': 0, 'direct_cost': 0})
        d['revenue'] += t['billed_amount'] or 0
        direct = (t['fuel_amount'] or 0) + (t['driver_adv_amount'] or 0) + _trip_toll(t, export_toll_map) + (t['parking'] or 0)
        if t['type'] == 'Market':
            direct += (t['owner_fixed_amount'] or 0) if (t['owner_rate_type'] or 'PER_MT') == 'FIXED' else (t['owner_rate'] or 0) * (t['quantity'] or 0)
        d['direct_cost'] += direct
    top_customers = []
    for pid, d in sorted(party_period.items(), key=lambda kv: kv[1]['revenue'], reverse=True):
        top_customers.append((party_name.get(pid, '—'), d['revenue'], d['revenue'] - d['direct_cost'],
                               max(party_balance.get(pid, 0), 0)))

    expense_breakdown = [
        ('Fuel', curr['fuel']), ('Driver Advance', curr['driver_adv']), ('Maintenance', curr['maint']),
        ('Salary', curr['salaries']), ('Toll / Parking', curr['toll'] + curr['parking']),
        ('Other Expenses', curr['misc'] + curr['overheads'] + curr['owner_cost']),
    ]
    conn.close()

    navy = "1B2A4A"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=navy)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Business Performance", f"{date_from} to {date_to}"])
    ws['A1'].font = Font(bold=True, size=13)
    ws.append([])
    summary_pairs = [
        ("Total Revenue", curr['revenue']), ("Net Profit", curr['net_profit']), ("Profit Margin %", curr['margin']),
        ("Outstanding Receivables", total_receivables), ("Outstanding Payables", total_payables),
        ("Cash Collected", cash_collected), ("Cash Paid", cash_paid),
        ("Collection Efficiency %", collection_efficiency), ("Operating Ratio %", operating_ratio),
        ("Total Trips", curr['trip_count']),
    ]
    for label, val in summary_pairs:
        ws.append([label, val])
    for row in ws.iter_rows(min_row=3, max_row=2 + len(summary_pairs), min_col=1, max_col=1):
        for c in row:
            c.font = Font(bold=True)

    ws2 = wb.create_sheet("Top Customers")
    for i, h in enumerate(["Customer", "Revenue", "Profit", "Outstanding"], 1):
        c = ws2.cell(row=1, column=i, value=h); c.font = header_font; c.fill = header_fill
    for r_idx, row in enumerate(top_customers, 2):
        for c_idx, val in enumerate(row, 1):
            ws2.cell(row=r_idx, column=c_idx, value=val)

    ws3 = wb.create_sheet("Expense Breakdown")
    for i, h in enumerate(["Category", "Amount"], 1):
        c = ws3.cell(row=1, column=i, value=h); c.font = header_font; c.fill = header_fill
    for r_idx, row in enumerate(expense_breakdown, 2):
        for c_idx, val in enumerate(row, 1):
            ws3.cell(row=r_idx, column=c_idx, value=val)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='business_performance_export.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/export')
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from flask import send_file
    import io

    conn = get_db()
    wb = Workbook()

    navy = "1B2A4A"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=navy)

    # Trips sheet
    ws = wb.active
    ws.title = "Trips"
    headers = ["Date","LR Number","Vehicle","Party","From","To","Quantity","Rate","Billed Amount",
               "Fuel Amount","Fuel Vendor","Driver Adv Amount","Driver Adv Vendor","Party Advance","Payment Received"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font; c.fill = header_fill
    rows = conn.execute("""SELECT t.date, t.lr_number, v.vehicle_no, p.name, t.from_loc, t.to_loc,
                           t.quantity, t.rate, t.billed_amount, t.fuel_amount, fv.name, t.driver_adv_amount,
                           dv.name, t.party_advance, t.payment_received
                           FROM trips t
                           LEFT JOIN vehicles v ON t.vehicle_id=v.id
                           LEFT JOIN parties p ON t.party_id=p.id
                           LEFT JOIN vendors fv ON t.fuel_vendor_id=fv.id
                           LEFT JOIN vendors dv ON t.driver_adv_vendor_id=dv.id
                           ORDER BY t.date""").fetchall()
    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    # Maintenance sheet
    ws2 = wb.create_sheet("Maintenance")
    headers2 = ["Date","Vehicle","Category","Amount","Paid Amount","Vendor","Notes"]
    for i, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font = header_font; c.fill = header_fill
    rows2 = conn.execute("""SELECT m.date, v.vehicle_no, m.category, m.amount, m.paid_amount, ve.name, m.notes
                            FROM maintenance m
                            LEFT JOIN vehicles v ON m.vehicle_id=v.id
                            LEFT JOIN vendors ve ON m.vendor_id=ve.id
                            ORDER BY m.date""").fetchall()
    for r_idx, row in enumerate(rows2, 2):
        for c_idx, val in enumerate(row, 1):
            ws2.cell(row=r_idx, column=c_idx, value=val)

    # Payments sheet
    ws3 = wb.create_sheet("Payments")
    headers3 = ["Date","Type","Amount","Party","Vendor","Mode"]
    for i, h in enumerate(headers3, 1):
        c = ws3.cell(row=1, column=i, value=h)
        c.font = header_font; c.fill = header_fill
    rows3 = conn.execute("""SELECT pay.date, pay.payment_type, pay.amount, p.name, v.name, pay.mode
                            FROM payments pay
                            LEFT JOIN parties p ON pay.party_id=p.id
                            LEFT JOIN vendors v ON pay.vendor_id=v.id
                            ORDER BY pay.date""").fetchall()
    for r_idx, row in enumerate(rows3, 2):
        for c_idx, val in enumerate(row, 1):
            ws3.cell(row=r_idx, column=c_idx, value=val)

    conn.close()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='fleet_export.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    # Weekly compliance scheduler — opt-in via ENABLE_COMPLIANCE_SCHEDULER=1, not on by default.
    # A background scheduler that starts silently every time `python app.py` is run (which
    # happens constantly during local development, including on every debug-reloader restart)
    # is more surprising than helpful; a real deployment should set this env var deliberately.
    # When it IS enabled: with debug=True, Werkzeug's reloader runs this file in two processes
    # (a watcher + the actual server); only the actual server process has WERKZEUG_RUN_MAIN set,
    # so gating on it stops the job from being registered twice. Under gunicorn (no reloader, no
    # debug), it starts once per worker — fine for a single-worker deployment; a multi-worker
    # production deployment should instead trigger the sync from an external cron hitting a
    # protected endpoint, not this in-process scheduler.
    if os.environ.get('ENABLE_COMPLIANCE_SCHEDULER') == '1' and (os.environ.get('WERKZEUG_RUN_MAIN') == 'true' or not app.debug):
        import logging
        logging.basicConfig(level=logging.INFO)
        from scheduler import start_scheduler
        start_scheduler(get_db)
    app.run(debug=True, port=5050)
