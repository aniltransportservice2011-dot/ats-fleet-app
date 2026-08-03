from openpyxl import load_workbook
import sqlite3
import datetime

wb = load_workbook('/mnt/user-data/outputs/Fleet_Master_Tracker.xlsx', data_only=True)
conn = sqlite3.connect('fleet.db')
cur = conn.cursor()

def fmt_date(d):
    if isinstance(d, datetime.datetime):
        return d.strftime('%Y-%m-%d')
    return None

# Migrate Vehicle types
ws_v = wb['Vehicles']
veh_count = 0
for r in range(6, 100):
    vno = ws_v.cell(row=r, column=2).value
    vtype = ws_v.cell(row=r, column=3).value
    if not vno:
        continue
    vno = str(vno).strip()
    cur.execute("SELECT id FROM vehicles WHERE vehicle_no=?", (vno,))
    row = cur.fetchone()
    if row:
        cur.execute("UPDATE vehicles SET type=? WHERE id=?", (vtype, row[0]))
    else:
        cur.execute("INSERT INTO vehicles (vehicle_no, type) VALUES (?,?)", (vno, vtype))
    veh_count += 1

# Migrate Salaries (months are columns 4 onwards)
ws_s = wb['Salaries']
sal_count = 0
month_headers = {}
for c in range(4, 20):
    h = ws_s.cell(row=6, column=c).value
    if h:
        month_headers[c] = str(h)
for r in range(7, 50):
    emp = ws_s.cell(row=r, column=2).value
    if not emp:
        continue
    for c, month in month_headers.items():
        amt = ws_s.cell(row=r, column=c).value
        if amt:
            cur.execute("INSERT INTO salaries (employee, month, amount) VALUES (?,?,?)", (emp, month, amt))
            sal_count += 1

# Migrate Overheads
ws_o = wb['Overheads']
oh_count = 0
for r in range(6, 100):
    cat = ws_o.cell(row=r, column=3).value
    if not cat:
        continue
    date = fmt_date(ws_o.cell(row=r, column=2).value)
    amt = ws_o.cell(row=r, column=4).value
    notes = ws_o.cell(row=r, column=5).value
    cur.execute("INSERT INTO overheads (date, category, amount, notes) VALUES (?,?,?,?)", (date, cat, amt, notes))
    oh_count += 1

conn.commit()
print(f"Migrated: {veh_count} vehicles updated/added, {sal_count} salary entries, {oh_count} overhead entries")
conn.close()
